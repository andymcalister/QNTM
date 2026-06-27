"""
QNTM — Conviction Factor Model Platform
Futuristic dark design · Financial green · Full platform
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, date
import sys, os, contextlib
sys.path.insert(0, os.path.dirname(__file__))

# Resolve the Q favicon robustly across Render/local working dirs; load as a PIL
# image (the most reliable page_icon input). Fall back to ⚡ only if truly absent.
_page_icon = "⚡"
try:
    from PIL import Image as _PILImage
    for _cand in (
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "qntm_icon.png"),
        os.path.join(os.getcwd(), "qntm_icon.png"),
        "qntm_icon.png",
    ):
        if os.path.exists(_cand):
            _page_icon = _PILImage.open(_cand)
            break
except Exception:
    _page_icon = "⚡"

st.set_page_config(
    page_title="QNTM — Conviction Factor Model",
    page_icon=_page_icon,
    layout="wide",
    initial_sidebar_state="collapsed",
)


def qntm_html(html, *, height=0, scrolling=False, iframe=False):
    """Version-safe successor to st.components.v1.html (deprecated in Streamlit
    1.56, removed in a later release).

    On Streamlit >=1.56:
      - iframe=True with height>0 -> st.iframe (isolated visible content, e.g. a
        self-contained chart document).
      - everything else -> st.html(..., unsafe_allow_javascript=True), which runs
        inline in the main document. This is required for the zero-height JS-only
        payloads (st.iframe rejects height=0) and is also more robust: there's no
        iframe sandbox, so window.parent/parent.document gracefully resolve to the
        top window.
    On the pinned 1.55.x: falls back to components.html. See MIGRATION.md."""
    if hasattr(st, "iframe"):
        if iframe and height and height > 0:
            st.iframe(html, height=height)
        else:
            st.html(html, unsafe_allow_javascript=True)
    else:
        import streamlit.components.v1 as _cv1_compat
        _cv1_compat.html(html, height=height, scrolling=scrolling)


# ── DEV ENVIRONMENT BANNER ────────────────────────────────────────────────────
import os
if os.getenv("ENVIRONMENT") == "dev":
    st.markdown("""
    <div style="background:#7c3aed;color:#fff;text-align:center;padding:6px 0;
         font-family:'DM Mono',monospace;font-size:13px;letter-spacing:.1em;
         position:sticky;top:0;z-index:9999;">
      ⚠ DEV ENVIRONMENT — changes here do not affect production
    </div>
    """, unsafe_allow_html=True)

from db import (register_user, login_user, get_holdings, upsert_holding,
                delete_holding, get_notifications, create_notification,
                mark_notifications_read, generate_totp_secret, verify_totp,
                enable_mfa, disable_mfa, get_user_mfa, update_preferences,
                upgrade_plan, plan_limit, PLAN_LIMITS,
                check_and_notify_signal_changes, save_signal_snapshot,
                get_signal_snapshot, get_unread_count, get_user_by_id,
                request_email_verification, consume_verify_token, is_email_verified)


def _universe_n() -> int:
    """Live size of the scoring universe (len of SECTORS) so every stock count
    shown across the site stays accurate as the universe changes over time."""
    try:
        from model_engine import SECTORS
        return len(SECTORS) or 1400
    except Exception:
        return 1400


@st.cache_data(ttl=900, show_spinner=False)
def _universe_rank_dist_cached():
    """Latest signal_date's full adj_composite distribution. Process-cached for
    15 min and SHARED across all users/sessions (st.cache_data), so the
    percentile lookup hits Supabase at most once per window instead of once per
    session. Returns a list of floats (possibly empty). Never raises."""
    try:
        from data_refresh import _get_supabase, _fetch_all_rows
        sb = _get_supabase()
        if not sb:
            return []
        # Latest signal_date, then ALL of that day's rows (paginated) so the
        # percentile base is the full universe, not just the first ~1,000.
        _latest = (sb.table("signal_log").select("signal_date")
                   .order("signal_date", desc=True).limit(1).execute())
        if not _latest.data:
            return []
        _ld = _latest.data[0].get("signal_date")
        rows = _fetch_all_rows(lambda: sb.table("signal_log")
                               .select("adj_composite,composite")
                               .eq("signal_date", _ld))
        dist = []
        for x in rows:
            v = x.get("adj_composite")
            if v is None:
                v = x.get("composite")
            if v is None:
                continue
            try:
                dist.append(float(v))
            except (TypeError, ValueError):
                continue
        return dist
    except Exception:
        return []


def _universe_rank_dist():
    """Public accessor. If the cached fetch came back empty (transient failure),
    drop the cache entry so the next call retries instead of serving an empty
    distribution for the full 15-min TTL."""
    dist = _universe_rank_dist_cached()
    if not dist:
        try:
            _universe_rank_dist_cached.clear()
        except Exception:
            pass
    return dist


def _pct_rank_of(score):
    """Percentile rank (0–100) of `score` within the latest full universe, or
    None if the distribution is unavailable. Never raises."""
    try:
        dist = _universe_rank_dist()
        if not dist:
            return None
        s = float(score)
        return sum(1 for c in dist if c <= s) / len(dist) * 100.0
    except Exception:
        return None


def _ordinal(n) -> str:
    """Integer with its ordinal suffix: 1->1st, 2->2nd, 62->62nd, 96->96th,
    11/12/13->th. Used for the RANK cell so cards read '62nd', not '62th'."""
    try:
        n = int(round(float(n)))
    except (TypeError, ValueError):
        return f"{n}th"
    suf = "th" if 10 <= (n % 100) <= 20 else {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
    return f"{n}{suf}"


from model_engine import (run_full_scan, detect_hidden_gems, BACKTEST_DATA,
                           ENTRY_THRESHOLD, EXIT_THRESHOLD, SECTORS,
                           fetch_macro_overlay, apply_macro_overlay,
                           MODEL_EPOCH, MODEL_INCEPTION)

# ── SIGNED JWT HELPERS ────────────────────────────────────────────────────────
import hmac, hashlib, base64, json as _json, time as _time
import analytics

# ── Build stamp ─────────────────────────────────────────────────────────────
# Bump APP_BUILD on every deploy-relevant change. The displayed tag prefers the
# real deployed commit (Render sets RENDER_GIT_COMMIT automatically); on
# Streamlit Cloud / local that env var is absent, so it falls back to APP_BUILD.
# Lets us read which code an instance is actually running at a glance, instead
# of inferring it from the numbers.
APP_BUILD = "2026-06-24.price-freshness"

def _build_tag() -> str:
    _sha = (os.environ.get("RENDER_GIT_COMMIT") or "").strip()
    return f"{APP_BUILD} \u00b7 {_sha[:7]}" if _sha else f"{APP_BUILD} \u00b7 local"


def _jwt_secret() -> str:
    """Use ENCRYPTION_KEY as JWT signing secret, fall back to a fixed dev key."""
    try:
        import streamlit as _st
        return _st.secrets.get("ENCRYPTION_KEY", "dev-secret-qntm-2025")
    except Exception:
        return "dev-secret-qntm-2025"

def _sign_token(uid: str, plan: str, days: int = 30) -> str:
    """For now return plain uid — JWT signing to be added once auth is stable."""
    return uid

def _verify_token(token: str):
    """For now treat token as plain uid."""
    return token, None

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&family=Outfit:wght@300;400;500;600&family=Inter:wght@400;500;600;700&display=swap');

*,*::before,*::after{box-sizing:border-box;margin:0;padding:0;}

/* ── Per-card action buttons (Add/Remove) — mobile text fit ── */
@media(max-width:640px){
  a[href*="wl_action"], a[href*="port_action"], a[href*="sim_add"], a[href*="sim_remove"]{
    font-size:13px !important; letter-spacing:.02em !important; padding:8px 4px !important;
  }
}

/* ── Popover trigger buttons — match institutional dark/gold aesthetic ── */
div[data-testid="stPopover"] > button,
button[data-testid="stPopoverButton"],
[data-testid="stPopover"] button[kind],
[data-testid="stPopover"] button {
  background: rgba(255,255,255,.03) !important;
  border: 1px solid rgba(212,168,67,.25) !important;
  color: #d4a843 !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 700 !important;
  font-size:13px !important;
  letter-spacing: .06em !important;
  border-radius: 6px !important;
}
div[data-testid="stPopover"] > button:hover,
[data-testid="stPopover"] button:hover {
  background: rgba(212,168,67,.12) !important;
  border-color: rgba(212,168,67,.5) !important;
}
/* Popover panel body — dark surface, not white */
div[data-testid="stPopoverBody"],
[data-baseweb="popover"] div[data-testid="stPopoverBody"],
div[data-testid="stPopover"] [data-baseweb="popover"] > div {
  background: #0a0b14 !important;
  border: 1px solid rgba(255,255,255,.08) !important;
  border-radius: 10px !important;
}
[data-baseweb="popover"] [data-testid="stPopoverBody"] * {
  color: #cbd5e1 !important;
}

/* ── Kill all horizontal overflow everywhere ── */
html, body {
  overflow-x: hidden !important;
  max-width: 100vw !important;
}

/* ── Dark background — covers all Streamlit containers, old and new selectors */
html, body, [class*="css"],
[data-testid="stAppViewContainer"],
[data-testid="stMain"],
[data-testid="stMainBlockContainer"],
[data-testid="stAppViewBlockContainer"],
section[data-testid="stMain"] > div,
.main, .stApp {
  font-family: 'Inter', sans-serif !important;
  background: #0a0b14 !important;
  color: #e2e8f0 !important;
  overflow-x: hidden !important;
  max-width: 100% !important;
}
.main .block-container,
[data-testid="stMainBlockContainer"] {
  padding: 0 !important;
  max-width: 100% !important;
  width: 100% !important;
  background: #0a0b14 !important;
  overflow-x: hidden !important;
}
/* Clamp Streamlit column containers */
[data-testid="stHorizontalBlock"] {
  max-width: 100% !important;
  width: 100% !important;
  overflow-x: hidden !important;
  flex-wrap: wrap !important;
}
[data-testid="stColumn"] {
  min-width: 0 !important;
  overflow-x: hidden !important;
}

/* Hide all Streamlit chrome */
#MainMenu, footer, header,
[data-testid="stHeader"],
[data-testid="stToolbar"],
[data-testid="stStatusWidget"],
[data-testid="stDecoration"] { display: none !important; }
[data-testid="stSidebar"]        { display: none !important; }
[data-testid="collapsedControl"] { display: none !important; }

::-webkit-scrollbar{width:3px;}
::-webkit-scrollbar-track{background:#0a0b14;}
::-webkit-scrollbar-thumb{background:rgba(52,211,153,.35);border-radius:2px;}

/* ── Mobile responsive: watchlist + model portfolio ── */
@media (max-width: 520px) {
  /* Watchlist: hide desktop table, show cards */
  .wl-table-header { display: none !important; }
  .wl-row           { display: none !important; }
  .wl-card          { display: block !important; }
  /* Model portfolio: hide desktop rows, show cards */
  .mp-row  { display: none !important; }
  .mp-card { display: block !important; }
}

/* Animations */
@keyframes fadeUp{from{opacity:0;transform:translateY(24px)}to{opacity:1;transform:translateY(0)}}
@keyframes fadeIn{from{opacity:0}to{opacity:1}}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.3}}
@keyframes glow{0%,100%{box-shadow:none}50%{box-shadow:none}}
@keyframes scanLine{0%{top:-2px;opacity:.3}100%{top:100%;opacity:0}}
@keyframes ticker{0%{transform:translateX(0)}100%{transform:translateX(-50%)}}
@keyframes qntmspin{to{transform:rotate(360deg)}}
@keyframes qntmload{to{opacity:1}}
@keyframes qntmfill{from{width:4%}to{width:93%}}
@keyframes qntmphrase{0%{opacity:0;transform:translateY(4px)}3%{opacity:1;transform:translateY(0)}17%{opacity:1;transform:translateY(0)}20%{opacity:0;transform:translateY(-4px)}100%{opacity:0}}

/* ── Collapsed card (details/summary) ── */
details summary { list-style: none; }
details summary::-webkit-details-marker { display: none; }
details summary::marker { display: none; }

/* Chevron rotation on open */
details summary .card-chevron { transition: transform .2s ease; display:inline-block; }
details[open] summary .card-chevron { transform: rotate(90deg); }

/* Subtle open state highlight */
details[open] {
  border-color: rgba(255,255,255,.12) !important;
  background: rgba(255,255,255,.035) !important;
}

/* One-at-a-time: when any details is open, siblings get slightly dimmed */
details:not([open]) { opacity: .92; }
details:hover:not([open]) { opacity: 1; }
@keyframes borderAnim{0%,100%{border-color:rgba(52,211,153,.15)}50%{border-color:rgba(52,211,153,.3)}}
@keyframes countUp{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}

/* Typography */
.syne{font-family:'Syne',sans-serif;}
.mono{font-family:'DM Mono',monospace;}

/* ── Platform buttons — dark glass with green accent ── */
@keyframes btn-shimmer {
  0%   { background-position: -200% center; }
  100% { background-position:  200% center; }
}
.stButton > button {
  background: rgba(52,211,153,.06) !important;
  color: #34d399 !important;
  border: 1px solid rgba(52,211,153,.22) !important;
  border-radius: 6px !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 700 !important;
  font-size:13px !important;
  letter-spacing: .06em !important;
  padding: 10px 12px !important;
  text-transform: uppercase !important;
  cursor: pointer !important;
  position: relative !important;
  overflow: hidden !important;
  white-space: nowrap !important;
  text-overflow: ellipsis !important;
  min-height: 42px !important;
  height: 42px !important;
  transition: border-color .18s, background .18s, transform .12s !important;
  box-shadow: none !important;
}
.stButton > button::before {
  content: '' !important;
  position: absolute !important;
  inset: 0 !important;
  background: linear-gradient(105deg,
    transparent 40%,
    rgba(52,211,153,.10) 50%,
    transparent 60%) !important;
  background-size: 200% 100% !important;
  opacity: 0 !important;
  transition: opacity .2s !important;
}
.stButton > button:hover {
  border-color: rgba(52,211,153,.5) !important;
  background: rgba(52,211,153,.1) !important;
  box-shadow: none !important;
  transform: translateY(-1px) !important;
}
.stButton > button:hover::before { opacity: 1 !important; }
.stButton > button:active { transform: translateY(0) !important; }

/* Ghost button variant */
div[data-ghost="1"] .stButton > button {
  background: transparent !important;
  color: rgba(52,211,153,.7) !important;
  border: 1px solid rgba(52,211,153,.2) !important;
  box-shadow: none !important;
}
div[data-ghost="1"] .stButton > button:hover {
  background: rgba(52,211,153,.05) !important;
  border-color: rgba(52,211,153,.4) !important;
  box-shadow: none !important;
  transform: none !important;
}

/* ── Inputs — cover all Streamlit input selectors old + new ── */
.stTextInput input,
.stNumberInput input,
.stDateInput input,
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input,
[data-baseweb="input"] input,
[data-baseweb="base-input"] input {
  background: #0d1117 !important;
  border: 1px solid rgba(255,255,255,.18) !important;
  border-radius: 4px !important;
  color: #e2e8f0 !important;
  font-family: 'Inter', sans-serif !important;
  font-size: 14px !important;
  caret-color: #34d399 !important;
}
.stTextInput input:focus,
.stNumberInput input:focus,
[data-baseweb="input"]:focus-within input,
[data-baseweb="base-input"]:focus-within input {
  border-color: rgba(52,211,153,.5) !important;
  box-shadow: 0 0 0 2px rgba(52,211,153,.1) !important;
  outline: none !important;
  color: #ffffff !important;
}
/* Placeholder text */
.stTextInput input::placeholder,
.stNumberInput input::placeholder,
[data-baseweb="input"] input::placeholder {
  color: rgba(148,163,184,.45) !important;
}
/* Labels */
label,
.stTextInput label,
.stNumberInput label,
.stSelectbox label,
.stDateInput label,
[data-testid="stWidgetLabel"] {
  color: #9fabc0 !important;
  font-size:13px !important;
  letter-spacing: .1em !important;
  text-transform: uppercase !important;
  font-family: 'Inter', sans-serif !important;
}
/* Force dark on baseweb input container */
[data-baseweb="input"],
[data-baseweb="base-input"],
[data-baseweb="input"] > div,
[data-baseweb="base-input"] > div {
  background: #0d1117 !important;
}
/* Select/dropdown */
div[data-baseweb="select"] > div,
[data-baseweb="select"] [data-baseweb="select-value-container"] {
  background: rgba(255,255,255,.05) !important;
  border: 1px solid rgba(255,255,255,.12) !important;
  border-radius: 4px !important;
  color: #e2e8f0 !important;
}
div[data-baseweb="select"] span,
[data-baseweb="select"] [data-baseweb="select-single-value"] {
  color: #e2e8f0 !important;
}
/* Number input spinner buttons */
.stNumberInput [data-baseweb="input"] {
  background: rgba(255,255,255,.05) !important;
}
/* Textarea */
.stTextArea textarea {
  background: rgba(255,255,255,.05) !important;
  border: 1px solid rgba(255,255,255,.12) !important;
  color: #e2e8f0 !important;
  font-family: 'Inter', sans-serif !important;
}

/* Tabs */
.stTabs [data-baseweb="tab-list"]{
  background:rgba(255,255,255,.03);border-radius:3px;
  border:1px solid rgba(255,255,255,.07);padding:3px;gap:2px;
}
.stTabs [data-baseweb="tab"]{
  color:#b3bed0;font-family:'Syne',sans-serif;font-size:13px;
  letter-spacing:.08em;text-transform:uppercase;border-radius:2px;padding:8px 18px;
}
.stTabs [aria-selected="true"]{
  color:#34d399!important;background:rgba(52,211,153,.08)!important;
}
.stTabs [data-baseweb="tab-border"]{display:none!important;}

/* ── Tooltips ── */
.qntm-tip {
    position: relative;
    display: inline-flex;
    align-items: center;
    gap: 4px;
    cursor: help;
}
.qntm-tip .tip-icon {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    width: 14px;
    height: 14px;
    background: rgba(255,255,255,.1);
    border: 1px solid rgba(255,255,255,.15);
    border-radius: 50%;
    font-size:11px;
    color: #9fabc0;
    flex-shrink: 0;
    font-style: normal;
}
.qntm-tip {
    position: relative;
    display: inline-block;
}
.qntm-tip .tip-box {
    display: none;
    position: fixed;
    background: #0d1117;
    border: 1px solid rgba(212,168,67,.4);
    border-radius: 10px;
    padding: 14px 16px;
    width: 260px;
    max-width: calc(100vw - 32px);
    z-index: 99999;
    pointer-events: none;
    box-shadow: 0 8px 40px rgba(0,0,0,.9);
    white-space: normal;
}
.qntm-tip .tip-box.visible {
    display: block;
}
.qntm-tip .tip-box .tip-title {
    font-family: 'Syne', sans-serif;
    font-size: 13px;
    font-weight: 700;
    color: #d4a843;
    margin-bottom: 6px;
}
.qntm-tip .tip-box .tip-body {
    font-size: 13px;
    color: #cbd5e1;
    line-height: 1.6;
}
.qntm-tip .tip-box .tip-weight {
    font-family: 'DM Mono', monospace;
    font-size:13px;
    color: #b3bed0;
    margin-top: 6px;
}
</style>
<style>

/* ── CTA button: gold primary — available on all pages ── */
.land-btn-primary > div > button,
.land-btn-primary button,
div.land-btn-primary .stButton > button,
div.land-btn-primary button[kind="secondary"] {
  background: linear-gradient(135deg,#d4a843 0%,#b8922e 50%,#d4a843 100%) !important;
  color: #0a0b14 !important;
  border: none !important;
  border-radius: 6px !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 800 !important;
  font-size: 13px !important;
  letter-spacing: .08em !important;
  text-transform: uppercase !important;
  min-height: 48px !important;
  height: auto !important;
  cursor: pointer !important;
  box-shadow: 0 2px 12px rgba(212,168,67,.15) !important;
  transition: all .2s !important;
  white-space: normal !important;
}
.land-btn-primary > div > button:hover,
.land-btn-primary button:hover,
div.land-btn-primary .stButton > button:hover {
  background: linear-gradient(135deg,#e0b84e 0%,#c9a03e 50%,#e0b84e 100%) !important;
  box-shadow: 0 4px 20px rgba(212,168,67,.25) !important;
  transform: translateY(-1px) !important;
}
/* ── Ghost button ── */
.land-btn-ghost > div > button,
.land-btn-ghost button {
  background: rgba(255,255,255,.04) !important;
  color: #e2e8f0 !important;
  border: 1px solid rgba(255,255,255,.15) !important;
  border-radius: 6px !important;
  font-family: 'Syne', sans-serif !important;
  font-weight: 700 !important;
  font-size: 14px !important;
  letter-spacing: .06em !important;
  text-transform: uppercase !important;
  min-height: 48px !important;
  cursor: pointer !important;
  transition: all .2s !important;
}
.land-btn-ghost > div > button:hover,
.land-btn-ghost button:hover {
  border-color: rgba(255,255,255,.3) !important;
  background: rgba(255,255,255,.08) !important;
}
/* ── CTA button: gold primary — available on all pages ── */
html, body, [class*="css"], .main, .stApp {
  font-size: 16px !important;
}
.stMarkdown p, [data-testid="stMarkdownContainer"] p {
  color: #cbd5e1 !important;
  font-size: 15px !important;
  line-height: 1.7 !important;
}
.stMarkdown h1, h1 { color: #f1f5f9 !important; font-size: 28px !important; }
.stMarkdown h2, h2 { color: #f1f5f9 !important; font-size: 22px !important; }
.stMarkdown h3, h3 { color: #e2e8f0 !important; font-size: 18px !important; }
.stMarkdown h4, h4 { color: #e2e8f0 !important; font-size: 16px !important; }
label, .stTextInput label, .stNumberInput label,
.stSelectbox label, .stDateInput label,
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] {
  color: #b3bed0 !important;
  font-size: 13px !important;
  letter-spacing: .08em !important;
  text-transform: uppercase !important;
}
.stCheckbox label, .stCheckbox label p,
.stToggle label, .stToggle label p {
  color: #cbd5e1 !important;
  font-size: 15px !important;
  text-transform: none !important;
  letter-spacing: 0 !important;
}
.streamlit-expanderHeader, .streamlit-expanderHeader p {
  color: #cbd5e1 !important;
  font-size: 15px !important;
}
.stTabs [data-baseweb="tab"] {
  color: #b3bed0 !important;
  font-size: 14px !important;
}
.stTabs [aria-selected="true"] {
  color: #34d399 !important;
  background: rgba(52,211,153,.08) !important;
}
div[data-baseweb="select"] span,
[data-baseweb="select"] [data-baseweb="select-single-value"] {
  color: #e2e8f0 !important;
  font-size: 15px !important;
}
[data-testid="stMetricValue"] { color: #34d399 !important; font-size: 28px !important; }
[data-testid="stMetricLabel"] { color: #b3bed0 !important; font-size: 13px !important; }
.stRadio label, .stRadio label p,
.stRadio [data-testid="stMarkdownContainer"] p {
  color: #cbd5e1 !important;
  font-size: 14px !important;
  font-weight: 600 !important;
  text-transform: uppercase !important;
  letter-spacing: .06em !important;
}
[aria-checked="true"] label, [aria-checked="true"] label p {
  color: #34d399 !important;
}

/* Forms */
.stForm{border:1px solid rgba(255,255,255,.07)!important;border-radius:6px!important;padding:1.5rem!important;background:rgba(255,255,255,.02)!important;}

/* Divider */
hr{border-color:rgba(255,255,255,.07)!important;}

/* Autocomplete iframe — collapse when dropdown not showing */
iframe[title="st_components_v1.html-0"], 
[data-testid="stCustomComponentV1"]:has(iframe) {
    transition: height .15s ease;
}
/* Autocomplete component iframe — allow dropdown to overflow */
[data-testid="stCustomComponentV1"] {
    overflow: visible !important;
}
[data-testid="stCustomComponentV1"] iframe {
    overflow: visible !important;
}
/* Hide st.components.v1.html iframes used for JS-only operations */
iframe[height="0"], iframe[style*="height: 0"], 
[data-testid="stCustomComponentV1"] iframe {
    display: none !important;
    height: 0 !important;
    width: 0 !important;
    border: none !important;
    position: absolute !important;
    top: -9999px !important;
}



/* Full Universe filter selects — reduce Streamlit feel */
div[data-testid="stSelectbox"] > div > div {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.1) !important;
    border-radius: 6px !important;
    color: #b3bed0 !important;
    font-family: DM Mono, monospace !important;
    font-size:13px !important;
    min-height: 36px !important;
}
div[data-testid="stSelectbox"] label {
    font-family: DM Mono, monospace !important;
    font-size:13px !important;
    color: #8896ac !important;
    letter-spacing: .08em !important;
    text-transform: uppercase !important;
}
/* Expander — reduce Streamlit feel */
[data-testid="stExpander"] {
    background: rgba(255,255,255,.02) !important;
    border: 1px solid rgba(255,255,255,.07) !important;
    border-radius: 8px !important;
}
[data-testid="stExpander"] summary {
    font-family: Syne, sans-serif !important;
    font-size: 13px !important;
    font-weight: 600 !important;
    color: #b3bed0 !important;
}
/* Number inputs — reduce Streamlit feel */
div[data-testid="stNumberInput"] input {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.1) !important;
    border-radius: 6px !important;
    color: #e2e8f0 !important;
    font-family: DM Mono, monospace !important;
    font-size: 13px !important;
}
/* Date input */
div[data-testid="stDateInput"] input {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.1) !important;
    border-radius: 6px !important;
    color: #e2e8f0 !important;
}
/* Add Position text input */
div[data-testid="stTextInput"] input {
    background: rgba(255,255,255,.03) !important;
    border: 1px solid rgba(255,255,255,.1) !important;
    border-radius: 6px !important;
    color: #e2e8f0 !important;
    font-family: Outfit, sans-serif !important;
}
div[data-testid="stTextInput"][data-key="screener_search_raw"] input {
    border: 1px solid rgba(52,211,153,.3) !important;
}
/* st.spinner — reduce Streamlit feel */
[data-testid="stSpinner"] > div {
    color: #8896ac !important;
    font-size:13px !important;
}

/* ── MOBILE RESPONSIVE ── */
@media (max-width: 640px) {
  /* Collapse platform padding on mobile */
  .block-container { padding: 0 !important; }

  /* Card rows — ensure single line, large tap target */
  .qcard-wrap label { min-height: 48px !important; padding: 10px 12px !important; }
  .qcard-wrap { margin-bottom: 6px !important; }

  /* Top-10 HIGH/LOW conviction grid — stack to one column on phones so the
     expanded card gets full width instead of being clipped in a half column */
  .qntm-conv-grid { grid-template-columns: 1fr !important; gap: 0 !important; }
  .qntm-conv-grid > div { margin-bottom: 10px; }

  /* Expanded-card internals: collapse to 2-up on phones. These main-document
     cards don't receive CARD_IFRAME_TAIL's copy of this rule, so mirror it. */
  .qcard-pillars { grid-template-columns: repeat(2,1fr) !important; }
  .qcard-4box    { grid-template-columns: repeat(2,1fr) !important; }

  /* Hide company name in collapsed card on very small screens */
  .qcard-name-mobile { display: none !important; }

  /* Tabs — scrollable on mobile */
  .stTabs [data-baseweb='tab-list'] { overflow-x: auto !important; flex-wrap: nowrap !important; }
  .stTabs [data-baseweb='tab'] { white-space: nowrap !important; padding: 8px 12px !important; font-size:13px !important; }

  /* Expander — larger tap target */
  [data-testid='stExpander'] summary { min-height: 44px !important; }

  /* Number inputs — prevent zoom on focus */
  input[type='number'], input[type='text'] { font-size: 16px !important; }

  /* Page headers — tighter on mobile */
  .page-header { padding: 8px 16px 4px !important; }
}

@media (max-width: 768px) {
    /* Prevent iOS zoom on inputs */
    .stTextInput input,[data-baseweb="input"] input {
        font-size: 16px !important;
    }
    /* Scale landing hero text */
    h1 { font-size: 28px !important; }
    .land-section { padding: 32px 16px !important; }
    /* Tooltips stay on screen */
    .qntm-tip .tip-box {
        width: 220px !important;
        max-width: 80vw !important;
    }
    /* Our custom HTML data tables — horizontal scroll */
    .qntm-table-scroll {
        overflow-x: auto !important;
        -webkit-overflow-scrolling: touch !important;
    }
}
@media (max-width: 480px) {
    h1 { font-size: 22px !important; }
    .land-section { padding: 24px 12px !important; }
}
/* Viewport meta (Streamlit adds this but ensure scale=1) */

/* ══════════════════════════════════════════════════════════
   TASK 4 — SPACING & BREATHING ROOM PASS
   ══════════════════════════════════════════════════════════ */

/* ── Card toggle — global rules so they work across separate st.markdown calls ── */
input[id^='c']:checked ~ label .qcard-detail { display: block !important; }
input[id^='c']:checked ~ label {
  border-color: rgba(255,255,255,.14) !important;
  background: rgba(255,255,255,.035) !important;
}

/* ── Card spacing — more room between cards ── */
.qcard-wrap { margin-bottom: 10px !important; }

/* ── Expanded card detail padding ── */
.qcard-detail { padding: 0 20px 20px !important; }

/* ── Platform page container — consistent side padding ── */
[data-testid="stMainBlockContainer"] > div > div {
  padding-left: 0 !important;
  padding-right: 0 !important;
}

/* ── Streamlit tab panels — add top breathing room ── */
[data-testid="stTabPanel"] {
  padding-top: 8px !important;
}

/* ── Section gaps between major blocks ── */
[data-testid="stVerticalBlock"] > [data-testid="stVerticalBlock"] {
  gap: 8px !important;
}

/* ── Expander — more padding inside ── */
[data-testid="stExpander"] summary {
  padding: 12px 16px !important;
}
[data-testid="stExpander"] [data-testid="stExpanderDetails"] {
  padding: 8px 16px 16px !important;
}

/* ── Selectbox — more height ── */
div[data-baseweb="select"] > div {
  min-height: 40px !important;
}

/* ── Mobile: larger tap targets, more card padding ── */
@media (max-width: 768px) {
  .qcard-wrap { margin-bottom: 8px !important; }
  /* Mobile: ensure collapsed row is one line, no wrapping */
  .qcard-wrap label { min-height: 44px !important; }
  .stButton > button { min-height: 48px !important; padding: 12px 16px !important; }
  [data-testid="stExpander"] summary { padding: 14px 16px !important; }
  .stTabs [data-baseweb="tab"] { padding: 10px 14px !important; }
}

/* ── 3-Level Visual Hierarchy ──────────────────────────────────────────────
   Level 1 (high emphasis):  conviction badges, regime badge, primary CTAs, active alerts
   Level 2 (medium):         cards, charts, section headers, expanded content
   Level 3 (low):            timestamps, metadata, helper text, secondary metrics
   ────────────────────────────────────────────────────────────────────────── */

/* Level 3 — timestamps, scan dates, metadata — consistently muted */
.qntm-meta, .qntm-ts {
  color: #94a3b8 !important;
  font-size:13px !important;
}

/* Reduce competing card borders — unified subtle treatment */
.qcard-wrap label {
  transition: border-color .15s ease, background .15s ease;
}
.qcard-wrap label:hover {
  border-color: rgba(255,255,255,.1) !important;
  background: rgba(255,255,255,.03) !important;
}

/* Reduce stat card glow — cards are Level 2, not Level 1 */
[data-testid="stMetricValue"] {
  color: #d4a843 !important;
  font-size: 24px !important;
  text-shadow: none !important;
}

/* Tabs — active is Level 1, inactive is Level 3 */
.stTabs [data-baseweb="tab"] {
  color: #8896ac !important;
}
.stTabs [aria-selected="true"] {
  color: #e2e8f0 !important;
  background: rgba(255,255,255,.05) !important;
}

/* Search suggestion buttons */
.qntm-sug-wrap .stButton > button {
  background:#0d1117 !important;
  border:none !important;
  border-top:1px solid rgba(255,255,255,.05) !important;
  border-radius:0 !important;
  color:#e2e8f0 !important;
  font-family:Syne,sans-serif !important;
  font-size:14px !important;
  font-weight:700 !important;
  text-align:left !important;
  padding:10px 14px !important;
  height:auto !important;
  min-height:44px !important;
  width:100% !important;
  letter-spacing:0 !important;
  text-transform:none !important;
  box-shadow:none !important;
}
.qntm-sug-wrap .stButton > button:hover {
  background:rgba(52,211,153,.07) !important;
  border-color:rgba(255,255,255,.05) !important;
  color:#e2e8f0 !important;
  transform:none !important;
  box-shadow:none !important;
}

/* Section dividers — very subtle Level 3 */
.land-divider {
  border-color: rgba(255,255,255,.04) !important;
}
</style>
""", unsafe_allow_html=True)

# ── TOOLTIP + MOBILE JS — injected via components (only way to run JS in Streamlit) ──
qntm_html("""
<script>
(function() {
    function positionTip(tip, box) {
        var rect = tip.getBoundingClientRect();
        var bw = 260, margin = 12, bh = box.offsetHeight || 160;
        var top  = rect.top - bh - 10;
        var left = rect.left + rect.width / 2 - bw / 2;
        if (top < margin) top = rect.bottom + 10;
        if (top + bh > window.innerHeight - margin) top = margin;
        if (left < margin) left = margin;
        if (left + bw > window.innerWidth - margin) left = window.innerWidth - bw - margin;
        box.style.position = 'fixed';
        box.style.top  = top + 'px';
        box.style.left = left + 'px';
        box.style.display = 'block';
    }
    function hideTip(tip) {
        var box = tip.querySelector('.tip-box');
        if (box) box.style.display = 'none';
    }
    function hideAll() {
        parent.document.querySelectorAll('.tip-box').forEach(function(b) { b.style.display='none'; });
    }
    function showTip(tip) {
        hideAll();
        var box = tip.querySelector('.tip-box');
        if (!box) return;
        positionTip(tip, box);
    }
    // Desktop hover
    parent.document.addEventListener('mouseover', function(e) {
        var tip = e.target.closest('.qntm-tip');
        if (tip) showTip(tip); else hideAll();
    });
    // Mobile tap
    parent.document.addEventListener('touchend', function(e) {
        var tip = e.target.closest('.qntm-tip');
        if (tip) {
            var box = tip.querySelector('.tip-box');
            if (box && box.style.display === 'block') {
                box.style.display = 'none';
            } else {
                e.preventDefault();
                showTip(tip);
            }
        } else {
            hideAll();
        }
    }, { passive: false });
})();

// ── CARD TOGGLE — one-at-a-time close ────────────────────────────────────────
(function() {
    function closeOthers(checkedEl) {
        parent.document.querySelectorAll('input[id^="c"]').forEach(function(cb) {
            if (cb !== checkedEl && cb.type === 'checkbox' && cb.checked) {
                cb.checked = false;
            }
        });
    }
    function attachListeners() {
        parent.document.querySelectorAll('input[id^="c"]').forEach(function(cb) {
            if (cb.type === 'checkbox' && !cb._qntmBound) {
                cb._qntmBound = true;
                cb.addEventListener('change', function() {
                    if (cb.checked) closeOthers(cb);
                });
            }
        });
    }
    attachListeners();
    new MutationObserver(attachListeners).observe(
        parent.document.body, { childList: true, subtree: true }
    );
})();

</script>
""", height=0)

# ── SESSION STATE ─────────────────────────────────────────────────────────────
for k, v in {
    "page": "landing",
    "logged_in": False,
    "user": None,
    "mfa_verified": False,
    "pending_mfa_user": None,
    "pending_mfa_secret": None,
    "scan_results": None,
    "cookies_accepted": False,
    "show_mfa_setup": False,
    "totp_secret_temp": None,
    "auth_tab": "signin",
    "nav": "screener",
    "macro_data": {},
    "auto_upgrade": False,
    "remember_me":  False,
    "legal_doc": "privacy",
    "force_mfa_setup": False,   # True after first login if MFA not set up
    "port_period":  "1M",
    "live_refresh_running": False,
    "mfa_recovery_mode": False,
    "signed_out": False,
    "onboarding_done": True,
    "onboarding_step": 0,
    "screener_search_val": "",
    "screener_search_raw": "",
    "_search_live": "",
    "tz_offset_hours": None,  # browser timezone offset, injected on first load
    "tz_name": None,          # IANA timezone name e.g. America/Los_Angeles
}.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ── TIMEZONE DETECTION ────────────────────────────────────────────────────────
# Read ?_tz= query param set by browser JS, store in session state
# Always read it (don't gate on session_state) so re-injections refresh the value.
_tz_param = st.query_params.get("_tz", "")
if _tz_param:
    try:
        st.session_state.tz_offset_hours = float(_tz_param)
        st.query_params.pop("_tz", None)
    except Exception:
        pass

_tz_name_param = st.query_params.get("_tzname", "")
if _tz_name_param:
    st.session_state.tz_name = _tz_name_param
    st.query_params.pop("_tzname", None)

# Timezone detector — write _tz/_tzname to URL via JS so Python can read them.
# Gate on session_state: if we don't have a tz_offset yet, inject the JS.
# Once injected, the JS replaces the URL with _tz params; Python reads them on
# the next rerun, stores in session_state, pops the URL. After that, the gate
# is False and JS never runs again.
if st.session_state.get("tz_offset_hours") is None or not st.session_state.get("tz_name"):
    qntm_html("""
    <script>
    (function() {
        try {
            var url = new URL(window.parent.location.href);
            if (url.searchParams.get('_tz')) return;  // already set
            var offset = -(new Date().getTimezoneOffset() / 60);
            var tzname = Intl.DateTimeFormat().resolvedOptions().timeZone;
            url.searchParams.set('_tz', offset.toString());
            url.searchParams.set('_tzname', tzname);
            window.parent.history.replaceState(null, '', url.toString());
        } catch(e) {
            // Cross-origin restriction or other — try writing to top instead
            try {
                var url2 = new URL(window.top.location.href);
                if (url2.searchParams.get('_tz')) return;
                var offset = -(new Date().getTimezoneOffset() / 60);
                var tzname = Intl.DateTimeFormat().resolvedOptions().timeZone;
                url2.searchParams.set('_tz', offset.toString());
                url2.searchParams.set('_tzname', tzname);
                window.top.history.replaceState(null, '', url2.toString());
            } catch(e2) {}
        }
    })();
    </script>
    """, height=0)


# ── PERSISTENT LOGIN — 7-day localStorage token ───────────────────────────────
# Stores {uid, plan, expires} in browser localStorage on remember-me login.
# Reads it back on every load. Falls back to query params for old sessions.

def _inject_localstorage_reader():
    """Read QNTM auth token from localStorage and restore session via query params."""
    qntm_html("""
    <script>
    (function() {
        try {
            var raw = localStorage.getItem('qntm_auth');
            if (!raw) return;
            var url = new URL(window.parent.location.href);
            if (!url.searchParams.get('uid')) {
                url.searchParams.set('uid', raw);
                url.searchParams.set('plan', 'restore');
                window.parent.location.replace(url.toString());
            }
        } catch(e) {}
    })();
    </script>
    """, height=0)


def _write_localstorage_token(uid: str, plan: str):
    """Write a signed 30-day auth token to localStorage."""
    token = _sign_token(uid, plan, days=30)
    qntm_html(f"""
    <script>
    try {{
        localStorage.setItem('qntm_auth', {_json.dumps(token)});
    }} catch(e) {{}}
    </script>
    """, height=0)


def _clear_localstorage_token():
    """Clear the auth token from localStorage on sign out."""
    qntm_html("""
    <script>
    try { localStorage.removeItem('qntm_auth'); } catch(e) {}
    </script>
    """, height=0)


# ── Auto-restore session from localStorage or query params ────────────────────
if not st.session_state.logged_in:
    params = st.query_params
    if "uid" in params:
        _restore_ok = False
        try:
            saved_uid = params["uid"]
            verified_uid, _ = _verify_token(saved_uid)
            if verified_uid:
                user = get_user_by_id(verified_uid)
                if user:
                    qp_plan = params.get("plan", "")
                    if qp_plan in ("pro", "institutional") and user.get("plan") == "free":
                        user["plan"] = qp_plan
                    st.session_state.logged_in       = True
                    st.session_state.user            = user
                    st.session_state.mfa_verified    = True
                    st.session_state.signed_out      = False
                    st.session_state.page            = "platform"
                    st.session_state.onboarding_done = True
                    _dest = params.get("qnav", "")
                    _VALID = {"screener","gems","backtest","portfolio","simulator","watchlist",
                              "model_portfolio","alerts","account","methodology"}
                    st.session_state.nav = _dest if _dest in _VALID else "screener"
                    _restore_ok = True
                else:
                    # DB returned nothing — build minimal session from query params
                    qp_plan = params.get("plan", "free")
                    st.session_state.logged_in       = True
                    st.session_state.user            = {"id": verified_uid, "plan": qp_plan, "email": "", "full_name": ""}
                    st.session_state.mfa_verified    = True
                    st.session_state.signed_out      = False
                    st.session_state.page            = "platform"
                    st.session_state.onboarding_done = True
                    _dest = params.get("qnav", "")
                    _VALID = {"screener","gems","backtest","portfolio","simulator","watchlist",
                              "model_portfolio","alerts","account","methodology"}
                    st.session_state.nav = _dest if _dest in _VALID else "screener"
                    _restore_ok = True
        except Exception as _e:
            pass
    

    _nav_param = st.query_params.get("nav", "")
    _has_uid   = "uid" in st.query_params
    # Only inject localStorage reader on the landing page as a last resort
    # Injecting it globally causes location.replace() to wipe nav params mid-session

# ── HELPERS ───────────────────────────────────────────────────────────────────
def uid():
    return (st.session_state.user or {}).get("id", "demo")

def is_pro():
    return (st.session_state.user or {}).get("plan", "free") in ("pro", "institutional")

def go(page):
    st.session_state.page = page
    if st.session_state.get("logged_in") and st.session_state.get("user"):
        u = st.session_state.user
        signed = _sign_token(u["id"], u.get("plan", "free"))
        st.query_params["uid"]  = signed
        st.query_params["plan"] = u.get("plan", "free")
    st.rerun()

# ── UI HELPERS ────────────────────────────────────────────────────────────────

def _pin_nav(page_key: str):
    """Pin nav to current page — prevents text input reruns from dropping to screener."""
    st.session_state.nav  = page_key
    st.session_state.page = "platform"

def _back_btn(href: str, label: str = "← Back") -> str:
    """Styled ghost back button as HTML link."""
    return (
        f'<a href="{href}" target="_self" style="'
        f'display:inline-flex;align-items:center;gap:4px;'
        f'padding:7px 10px;border-radius:6px;'
        f'border:1px solid rgba(255,255,255,.12);'
        f'background:rgba(255,255,255,.03);'
        f'font-family:Syne,sans-serif;font-size:13px;font-weight:700;'
        f'letter-spacing:.04em;white-space:nowrap;'
        f'color:#b3bed0;text-decoration:none;">'
        f'← Back</a>'
    )

def _upgrade_url(feature: str, return_nav: str) -> str:
    """Build URL to the upgrade page preserving session. Deliberately omits the
    _n nav param: _n drives the reconnect-recovery handler, which on a fresh
    page-reload session would override page='upgrade' back to the platform tab
    (that's what made the Alerts/Simulator upgrade CTAs appear to do nothing)."""
    from urllib.parse import quote_plus
    _uid  = (st.session_state.user or {}).get("id", "")
    _plan = (st.session_state.user or {}).get("plan", "free")
    return (f"?upgrade_page=1&feature={quote_plus(feature)}"
            f"&return_nav={return_nav}&uid={_uid}&plan={_plan}&ck=1")

# ── ONBOARDING MODAL ──────────────────────────────────────────────────────────
def show_onboarding():
    pass  # disabled


# ══════════════════════════════════════════════════════════════════════════════
# LANDING PAGE
# ══════════════════════════════════════════════════════════════════════════════





# Legal page constants (PRIVACY_POLICY, TERMS_OF_SERVICE, BILLING_POLICY,
# DISCLAIMER_FULL, COOKIE_POLICY) live near page_legal() below — single source
# of truth. Edit them there.


def data_freshness_banner():
    """Show data age pill with actual datetime from signal_log.created_at, in user's local time."""
    try:
        from data_refresh import _get_supabase
        from datetime import datetime, timezone, timedelta
        dt_str = None
        fresh  = True
        tz_offset = st.session_state.get("tz_offset_hours")
        tz_name   = st.session_state.get("tz_name")

        def _fmt(raw: str) -> str:
            dt     = datetime.fromisoformat(raw.replace("Z", "+00:00"))
            dt_utc = dt.astimezone(timezone.utc)
            if tz_name:
                try:
                    from zoneinfo import ZoneInfo
                    dt_local = dt_utc.astimezone(ZoneInfo(tz_name))
                    tz_abbr  = dt_local.strftime("%Z")  # e.g. PDT, EST, GMT
                    return dt_local.strftime(f"%b %d · %H:%M {tz_abbr}")
                except Exception:
                    pass
            if tz_offset is not None:
                dt_local = dt_utc + timedelta(hours=tz_offset)
                sign     = "+" if tz_offset >= 0 else ""
                hrs      = int(tz_offset)
                return dt_local.strftime(f"%b %d · %H:%M UTC{sign}{hrs}")
            # No client TZ info available — default to America/Los_Angeles
            # so US users (the majority) don't see UTC. Override available
            # in Account → Preferences when that ships.
            try:
                from zoneinfo import ZoneInfo
                dt_local = dt_utc.astimezone(ZoneInfo("America/Los_Angeles"))
                tz_abbr  = dt_local.strftime("%Z")
                return dt_local.strftime(f"%b %d · %H:%M {tz_abbr}")
            except Exception:
                return dt_utc.strftime("%b %d · %H:%M UTC")

        try:
            sb = _get_supabase()
            if sb:
                resp = sb.table("fundamentals_cache").select("refreshed_at").order(
                    "refreshed_at", desc=True).limit(1).execute()
                if resp.data and resp.data[0].get("refreshed_at"):
                    raw    = resp.data[0]["refreshed_at"]
                    dt_str = _fmt(raw)
                    dt_utc = datetime.fromisoformat(raw.replace("Z", "+00:00")).astimezone(timezone.utc)
                    fresh  = (datetime.now(timezone.utc) - dt_utc) < timedelta(hours=26)
                else:
                    resp2 = sb.table("signal_log").select("created_at").order(
                        "created_at", desc=True).limit(1).execute()
                    if resp2.data:
                        raw    = resp2.data[0]["created_at"]
                        dt_str = _fmt(raw)
                        dt_utc = datetime.fromisoformat(raw.replace("Z", "+00:00")).astimezone(timezone.utc)
                        fresh  = (datetime.now(timezone.utc) - dt_utc) < timedelta(hours=26)
        except Exception:
            dt_str = None

        label  = f"Last refresh · {dt_str}" if dt_str else ("Data fresh" if fresh else "Stale data")
        color  = "#94a3b8" if fresh else "#9fabc0"
        bg     = "rgba(255,255,255,.03)" if fresh else "rgba(245,158,11,.05)"
        border = "rgba(255,255,255,.08)"  if fresh else "rgba(245,158,11,.15)"
        suffix = "" if fresh else " · Rescan for live scores"
        st.markdown(
            f'<div style="display:inline-flex;align-items:center;gap:6px;'
            f'background:{bg};border:1px solid {border};'
            f'border-radius:20px;padding:5px 14px;font-size:13px;color:{color};'
            f'font-family:DM Mono,monospace;margin-bottom:6px;">'
            f'<span style="width:5px;height:5px;border-radius:50%;background:{color};display:inline-block;opacity:.6;"></span>'
            f'{label}{suffix}</div>',
            unsafe_allow_html=True)
    except Exception:
        pass

def finalize_scores_from_signal_log(results: list, macro_data: dict = None) -> list:
    """
    Single source of truth for displayed scores.

    Trusts adj_composite from signal_log (the nightly cron's macro overlay).
    Computes adj_action, score_delta, and applies sorting + MIN_POSITIONS floor
    without ever recomputing adj_composite. This guarantees every page that
    reads the same signal_log row shows the same score — home, screener, gems,
    watchlist, portfolio, model portfolio, simulator.

    Use this everywhere on read-side. apply_macro_overlay() is reserved for the
    nightly cron (data_refresh.py) where overlay actually needs to be computed.
    """
    from model_engine import ENTRY_THRESHOLD, EXIT_THRESHOLD, MOM_EXIT, MIN_POSITIONS, DYNAMIC_THRESHOLD_HI

    if not results:
        return results

    # Dynamic threshold mirrors apply_macro_overlay so action labels match the cron
    n_above_60 = sum(1 for s in results
                     if float(s.get("composite", 0) or 0) >= ENTRY_THRESHOLD)
    eff_threshold = DYNAMIC_THRESHOLD_HI if n_above_60 > 30 else ENTRY_THRESHOLD

    for s in results:
        try:
            adj   = float(s.get("adj_composite") or s.get("composite") or 50)
            quant = float(s.get("composite") or adj)
        except (TypeError, ValueError):
            adj, quant = 50.0, 50.0

        # Normalize so factor_panel_html and all downstream readers see identical values
        s["adj_composite"] = adj
        s["composite"]     = quant
        s["score_delta"]   = round(adj - quant, 1)

        mom = float(s.get("momentum", 50) or 50)
        if adj >= eff_threshold:
            s["adj_action"] = "BUY"
        elif adj < EXIT_THRESHOLD or mom < MOM_EXIT:
            s["adj_action"] = "SELL"
        else:
            s["adj_action"] = "HOLD"

    # Sort by adjusted composite, same as the cron does
    results.sort(key=lambda x: float(x.get("adj_composite", 0) or 0), reverse=True)

    # MIN_POSITIONS floor: promote top HOLDs to BUY if we don't have enough BUYs.
    # Matches apply_macro_overlay so screener doesn't show a different BUY count
    # than what the cron produced.
    buys = [s for s in results if s.get("adj_action") == "BUY"]
    if len(buys) < MIN_POSITIONS:
        holds = [s for s in results if s.get("adj_action") == "HOLD"]
        needed = MIN_POSITIONS - len(buys)
        for s in holds[:needed]:
            s["adj_action"] = "BUY"
            s["promoted"]   = True

    # Standard card field set for every finalize-based surface (screener, gems,
    # portfolio, simulator pre-load): ensure sector, cap bucket and the valuation
    # band are present. Fill-if-missing and the signal_log fetch is already cached
    # from enrich, so this is a no-op cost on the hot screener path.
    hydrate_card_rows(results)

    return results


@st.cache_data(ttl=300, show_spinner=False)
def _signal_log_map(tickers_key: tuple) -> dict:
    """Cached fetch of the best recent signal_log row per ticker, keyed by the
    (sorted) ticker set. Returns each ticker's most recent CLEAN snapshot — one
    with no null/sentinel-50 pillars — falling back to the newest row only when
    no clean one exists in the lookback window. The scorer writes exactly 50.0
    on a per-ticker data failure, so this carries forward the last good score
    instead of surfacing a broken neutral one (the recurring '50s' problem),
    while genuine near-50 scores (49.7, 50.3, ...) are kept as-is."""
    try:
        from data_refresh import _get_supabase, _fetch_all_rows
        from datetime import date, timedelta
        sb = _get_supabase()
        if not sb or not tickers_key:
            return {}
        _since = (date.today() - timedelta(days=3)).isoformat()
        _wanted = set(tickers_key)
        # signal_log ~= the universe; pull the last few days in full (paginated so
        # the ~1,000-row cap can't truncate it now that the universe is >1,000),
        # then keep only the requested tickers.
        rows_data = _fetch_all_rows(lambda: sb.table("signal_log")
            .select("ticker,signal_date,adj_composite,composite,signal,"
                    "momentum,quality,volume,value,sentiment,price,mktcap,"
                    "is_hidden_gem,hidden_gem_reason,"
                    "val_low,val_high,value_position,val_basis")
            .gte("signal_date", _since)
            .not_.is_("composite", "null")
            .order("signal_date", desc=True))
        if not rows_data:
            return {}
        _PILLARS = ("momentum", "quality", "volume", "value", "sentiment")

        def _clean(row):
            # Clean = no pillar missing or pinned to the exact 50.0 neutral
            # sentinel the scorer writes when a ticker's data fetch fails.
            for k in _PILLARS:
                v = row.get(k)
                if v is None:
                    return False
                try:
                    if abs(float(v) - 50.0) < 0.01:
                        return False
                except (TypeError, ValueError):
                    return False
            return True

        latest, clean = {}, {}
        for row in rows_data:               # newest first
            tk = row["ticker"]
            if tk not in _wanted:
                continue
            if tk not in latest:
                latest[tk] = row
            if tk not in clean and _clean(row):
                clean[tk] = row
        # Prefer the freshest clean snapshot; fall back to newest so a ticker
        # that has only ever been neutral still appears.
        return {tk: clean.get(tk, latest[tk]) for tk in latest}
    except Exception:
        return {}


def hydrate_card_rows(rows: list) -> list:
    """Canonical card-row standardizer — the single contract for every expanded
    stock card in the app (screener, watchlist, portfolio, model portfolio,
    simulator, search). Guarantees each row carries the full display field set:
    sector, market-cap bucket (cap badge), the valuation band (val_low/val_high/
    value_position/val_basis), pillars, composite/adj_composite, price and
    signal_date — by FILLING ANY MISSING field from the latest signal_log
    snapshot (cached via _signal_log_map) and the in-memory sector map.

    Fill-if-absent only: values already present on the row (a surface's live
    price, entry context, recomputed action) are never overwritten, so live and
    position data survive. Best-effort; never raises. Call once per surface on
    the assembled row list (not per card) to keep it to a single cached fetch."""
    try:
        if not rows:
            return rows
        try:
            from model_engine import SECTORS as _SEC
        except Exception:
            _SEC = {}
        tks = tuple(sorted({r["ticker"] for r in rows if r.get("ticker")}))
        m = _signal_log_map(tks) if tks else {}
        _FILL = ("adj_composite", "composite", "momentum", "quality", "volume",
                 "value", "sentiment", "price", "signal_date", "is_hidden_gem",
                 "hidden_gem_reason", "mktcap",
                 "val_low", "val_high", "value_position", "val_basis")
        for r in rows:
            tk = r.get("ticker")
            db = m.get(tk, {}) if tk else {}
            for f in _FILL:
                if r.get(f) is None and db.get(f) is not None:
                    r[f] = db[f]
            if not r.get("sector") and tk:
                _s = _SEC.get(tk)
                if _s:
                    r["sector"] = _s
    except Exception:
        pass
    return rows


def enrich_with_signal_log(results: list) -> list:
    """
    Replaces model-computed scores with latest signal_log values from Supabase.
    This ensures screener always shows nightly cron scores, not local model estimates.
    Falls back to run_full_scan results if signal_log unavailable.
    """
    try:
        if not results:
            return results
        tickers = [r["ticker"] for r in results]
        log_map = _signal_log_map(tuple(sorted(set(tickers))))
        if not log_map:
            return results
        # Merge signal_log scores into results — DB scores take precedence
        for r in results:
            tk = r["ticker"]
            if tk in log_map:
                db = log_map[tk]
                # Always use DB scores — they come from the nightly cron
                for field in ["adj_composite","composite","momentum","quality",
                               "volume","value","sentiment","price","signal_date",
                               "is_hidden_gem","hidden_gem_reason","mktcap",
                               "val_low","val_high","value_position","val_basis"]:
                    if db.get(field) is not None:
                        r[field] = db[field]
                if db.get("signal_date"):
                    r["signal_date"] = str(db["signal_date"])[:10]
    except Exception:
        pass
    return results


@st.cache_data(ttl=900, show_spinner=False)
def _run_full_scan_cached():
    """Process-level cache of the expensive universe scan. Unlike
    st.session_state (which a full-page reload wipes by starting a new
    session), this survives across reloads/sessions — so navigating between
    screens no longer re-scores the full universe from scratch every time."""
    return run_full_scan(use_live_prices=False)


def _cached_full_scan():
    """Deep-copied view of the cached scan so callers can freely mutate it
    (sector fill, macro overlay) without corrupting the shared cache."""
    import copy
    return copy.deepcopy(_run_full_scan_cached())


DELISTED = set()  # intentionally empty: yfinance's "possibly delisted" is a
# generic/rate-limit error, NOT proof a ticker is dead (CTRA, SEE, MAXN are all
# live, held names). Excluding real tickers would corrupt the model-portfolio
# track record. The retry-storm slowness is handled by resilient pricing, not by
# dropping symbols.

def _strip_delisted(tks):
    """Dedupe/clean a ticker list before a yfinance pull. With DELISTED empty
    this only removes blanks/dupes — it never excludes a real holding."""
    return sorted({t for t in (tks or []) if t and t not in DELISTED})


@st.cache_data(ttl=300, show_spinner=False)
def _cached_live_prices(tickers: tuple):
    """Process-level cache of the top-tickers intraday price pull so it isn't
    re-fetched from yfinance on every navigation."""
    out = {}
    tickers = tuple(_strip_delisted(tickers))
    if not tickers:
        return out
    try:
        import yfinance as yf
        _pd = yf.download(list(tickers), period="1d", interval="1m",
                          progress=False, auto_adjust=True, threads=True)
        if not _pd.empty:
            _close = _pd["Close"] if "Close" in _pd else _pd
            for _tk in tickers:
                try:
                    if hasattr(_close, "columns") and _tk in _close.columns:
                        _px = float(_close[_tk].dropna().iloc[-1])
                    else:
                        _px = float(_close.dropna().iloc[-1])
                    if _px > 0:
                        out[_tk] = _px
                except Exception:
                    pass
    except Exception:
        pass
    return out


def scan_health_check():
    """
    Shows last successful nightly scan time pulled from signal_log.
    Green = scanned today. Amber = scanned yesterday. Red = >48h ago or never.
    """
    try:
        from data_refresh import _get_supabase
        sb = _get_supabase()
        if not sb:
            return
        result = sb.table("signal_log") \
            .select("signal_date") \
            .order("signal_date", desc=True) \
            .limit(1) \
            .execute()
        if not result.data:
            st.markdown(
                '<div style="display:inline-flex;align-items:center;gap:6px;'
                'background:rgba(248,113,113,.08);border:1px solid rgba(248,113,113,.2);'
                'border-radius:20px;padding:5px 14px;font-size:13px;color:#f87171;'
                'font-family:DM Mono,monospace;margin-bottom:8px;">'
                '<span style="width:7px;height:7px;border-radius:50%;background:#f87171;display:inline-block;"></span>'
                'No scan data found — run seed_track_record.py</div>',
                unsafe_allow_html=True)
            return
        from datetime import datetime, timezone
        last_date_str = result.data[0]["signal_date"]
        # signal_date may be "YYYY-MM-DD" or ISO datetime string
        try:
            last_dt = datetime.fromisoformat(last_date_str.replace("Z",""))
        except Exception:
            last_dt = datetime.strptime(last_date_str[:10], "%Y-%m-%d")
        now = datetime.now()
        age_h = (now - last_dt).total_seconds() / 3600
        if age_h < 26:
            color, bg, dot, label = "#34d399", "rgba(52,211,153,.08)", "#34d399", f"Nightly scan OK · {last_date_str[:10]}"
        elif age_h < 50:
            color, bg, dot, label = "#f59e0b", "rgba(245,158,11,.08)", "#f59e0b", f"Last scan {last_date_str[:10]} · check GitHub Actions"
        else:
            color, bg, dot, label = "#f87171", "rgba(248,113,113,.08)", "#f87171", f"Scan stale · last run {last_date_str[:10]} · check GitHub Actions"
        st.markdown(
            f'<div style="display:inline-flex;align-items:center;gap:6px;'
            f'background:{bg};border:1px solid {color}40;'
            f'border-radius:20px;padding:5px 14px;font-size:13px;color:{color};'
            f'font-family:DM Mono,monospace;margin-bottom:8px;">'
            f'<span style="width:7px;height:7px;border-radius:50%;background:{dot};display:inline-block;"></span>'
            f'{label}</div>',
            unsafe_allow_html=True)
    except Exception:
        pass


# ── PAGE SUMMARY BANNERS ──────────────────────────────────────────────────────
def page_summary(icon: str, title: str, subtitle: str, pills: list = None):
    """Consistent page header — pills param accepted but ignored (removed from UI)."""
    st.markdown(
        f'<div style="padding:10px 32px 6px;">'
        f'<span style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#e2e8f0;">{icon} {title}</span>'
        + (f'<div style="font-size:13px;color:#94a3b8;margin-top:2px;">{subtitle}</div>' if subtitle else '')
        + f'</div>',
        unsafe_allow_html=True
    )

def nav(section):
    st.session_state.nav = section
    st.rerun()

def signal_color(sig):
    return {"STRONG ALIGN":"#34d399","HIGH ALIGN":"#4ade80","MODERATE":"#fbbf24",
            "LOW ALIGN":"#f97316","WEAK/NEG":"#f87171"}.get(sig,"#9fabc0")

# ── PILLAR TOOLTIPS ──────────────────────────────────────────────────────────
PILLAR_TIPS = {
    "Momentum": {
        "weight": "30%",
        "body": "Price trend, RSI, MACD, MA crossovers, 52-week proximity.",
    },
    "Quality": {
        "weight": "25%",
        "body": "ROE, profit margin, revenue growth, EPS beat rate, FCF yield.",
    },
    "Volume": {
        "weight": "20%",
        "body": "Relative volume, OBV, Chaikin Money Flow, accumulation/distribution.",
    },
    "Value": {
        "weight": "15%",
        "body": "Forward P/E, PEG, EV/EBITDA, Price-to-Sales, FCF yield.",
    },
    "Sentiment": {
        "weight": "10%",
        "body": "Short interest, insider buy ratio, institutional ownership changes.",
    },
}

def get_company_info(ticker: str) -> dict:
    """
    Returns {name, description} for a ticker.
    Common tickers resolve instantly from a built-in map.
    Others pull from yfinance and cache in session state.
    """
    # Instant lookup for most common tickers
    KNOWN = {
        "AAPL":"Apple Inc.","MSFT":"Microsoft Corporation","NVDA":"NVIDIA Corporation",
        "GOOGL":"Alphabet Inc.","GOOG":"Alphabet Inc.","META":"Meta Platforms Inc.",
        "AMZN":"Amazon.com Inc.","TSLA":"Tesla Inc.","NFLX":"Netflix Inc.",
        "AMD":"Advanced Micro Devices","INTC":"Intel Corporation","CSCO":"Cisco Systems",
        "ORCL":"Oracle Corporation","CRM":"Salesforce Inc.","ADBE":"Adobe Inc.",
        "INTU":"Intuit Inc.","QCOM":"Qualcomm Inc.","TXN":"Texas Instruments",
        "AVGO":"Broadcom Inc.","MU":"Micron Technology","AMAT":"Applied Materials",
        "JPM":"JPMorgan Chase & Co.","BAC":"Bank of America","GS":"Goldman Sachs",
        "MS":"Morgan Stanley","V":"Visa Inc.","MA":"Mastercard Inc.",
        "BLK":"BlackRock Inc.","AXP":"American Express","PYPL":"PayPal Holdings",
        "UNH":"UnitedHealth Group","LLY":"Eli Lilly and Company","JNJ":"Johnson & Johnson",
        "ABBV":"AbbVie Inc.","MRK":"Merck & Co.","PFE":"Pfizer Inc.",
        "TMO":"Thermo Fisher Scientific","AMGN":"Amgen Inc.","GILD":"Gilead Sciences",
        "WMT":"Walmart Inc.","COST":"Costco Wholesale","PG":"Procter & Gamble",
        "KO":"The Coca-Cola Company","PEP":"PepsiCo Inc.","HD":"Home Depot",
        "MCD":"McDonald's Corporation","NKE":"Nike Inc.","SBUX":"Starbucks Corporation",
        "XOM":"Exxon Mobil Corporation","CVX":"Chevron Corporation",
        "BRK":"Berkshire Hathaway","PLTR":"Palantir Technologies",
        "COIN":"Coinbase Global","HOOD":"Robinhood Markets",
        "SNOW":"Snowflake Inc.","DDOG":"Datadog Inc.","NET":"Cloudflare Inc.",
        "ZS":"Zscaler Inc.","CRWD":"CrowdStrike Holdings","PANW":"Palo Alto Networks",
        "NOW":"ServiceNow Inc.","WDAY":"Workday Inc.","TEAM":"Atlassian Corporation",
        "UBER":"Uber Technologies","LYFT":"Lyft Inc.","ABNB":"Airbnb Inc.",
        "DASH":"DoorDash Inc.","SPOT":"Spotify Technology",
    }

    cache_key = "company_info_cache"
    if cache_key not in st.session_state:
        st.session_state[cache_key] = {}
    cache = st.session_state[cache_key]
    if ticker in cache:
        return cache[ticker]

    # Use known name if available, skip yfinance for speed
    if ticker in KNOWN:
        result = {"name": KNOWN[ticker], "description": ""}
        cache[ticker] = result
        return result

    # Unknown ticker — try yfinance (only for search results, not bulk universe)
    try:
        import yfinance as yf
        info = yf.Ticker(ticker).info or {}
        name = info.get("longName") or info.get("shortName") or ticker
        desc = info.get("longBusinessSummary") or ""
        if len(desc) > 220:
            desc = desc[:220].rsplit(" ", 1)[0] + "..."
        result = {"name": name, "description": desc}
    except Exception:
        result = {"name": ticker, "description": ""}
    cache[ticker] = result
    return result


    """Render a term with a hover tooltip info icon."""
    title = tip_dict.get("title", label)
    body  = tip_dict.get("body", "")
    weight= tip_dict.get("weight", "")
    weight_html = f'<div class="tip-weight">{weight}</div>' if weight else ""
    return (
        f'<span class="qntm-tip">{label}'
        f'<i class="tip-icon">i</i>'
        f'<span class="tip-box">'
        f'<div class="tip-title">{title}</div>'
        f'<div class="tip-body">{body}</div>'
        f'{weight_html}'
        f'</span></span>'
    )

def score_bar_html(val, width=80):
    col = "#34d399" if val>=65 else "#fbbf24" if val>=50 else "#f87171"
    return f'<div style="width:{width}px;height:4px;background:rgba(255,255,255,.06);border-radius:2px;overflow:hidden;"><div style="width:{val}%;height:100%;background:{col};border-radius:2px;"></div></div>'

@st.cache_data(ttl=300, show_spinner=False)
def _live_macro() -> dict:
    """Single source of truth for the displayed macro regime: read the cron's
    persisted live overlay (macro_state) so the banner and regime always match the
    scoring that produced adj_composite. Falls back to a one-off live scan only if
    macro_state is empty (e.g. before the first macro pass has run).
    Cached 5 min: macro_state only changes on the ~30-min macro cron, and it was
    previously re-fetched several times per render."""
    try:
        from data_refresh import _load_macro_state
        m = _load_macro_state()
        if m:
            return m
    except Exception:
        pass
    try:
        from model_engine import fetch_macro_overlay
        return fetch_macro_overlay(use_live_feeds=True)
    except Exception:
        return {}


@st.cache_data(ttl=300, show_spinner=False)
def _conviction_movers(tickers_key: tuple, lookback_days: int = 10, top_n: int = 24,
                       collapse_macro: bool = True, compare_days: int = 1) -> list:
    """Day-over-day conviction movers for the hero feed — the 'what's new today'
    surface. Compares each ticker's latest CLEAN scored row against the most recent
    clean row at least `compare_days` day(s) earlier (default: the prior scored
    session). Includes small moves (>=1 point) on purpose: a +1/-2 shift in
    conviction is meaningful and gives a reason to check back daily. Reports
    prev->now adj_composite plus what drove it — the pillar that moved most, or the
    macro overlay when the regime shift dominated the move (those uniform-band
    names collapse into one labelled chip so they never flood the feed). Research
    framing only; cached 5 min; one batched signal_log read."""
    try:
        from data_refresh import _get_supabase
        from datetime import date, timedelta
        sb = _get_supabase()
        if not sb or not tickers_key:
            return []
        since = (date.today() - timedelta(days=lookback_days)).isoformat()
        rows = (sb.table("signal_log")
                .select("ticker,signal_date,adj_composite,composite,"
                        "momentum,quality,volume,value,sentiment")
                .in_("ticker", list(tickers_key))
                .gte("signal_date", since)
                .not_.is_("composite", "null")
                .order("signal_date", desc=True)
                .execute()).data or []
    except Exception:
        return []

    _PILL = ("momentum", "quality", "volume", "value", "sentiment")
    _PLAB = {"momentum": "Momentum", "quality": "Quality", "volume": "Volume",
             "value": "Value", "sentiment": "Sentiment"}

    def _f(row, k):
        try:
            return float(row.get(k))
        except (TypeError, ValueError):
            return None

    from datetime import date as _date

    def _pd(s):
        try:
            return _date.fromisoformat(s)
        except (TypeError, ValueError):
            return None

    by = {}
    for r in rows:                       # already date-desc
        by.setdefault(r["ticker"], []).append(r)

    movers = []
    for tk, rs in by.items():
        # distinct scored dates, newest first
        distinct, seen = [], set()
        for r in rs:
            d = r.get("signal_date")
            if d and d not in seen:
                seen.add(d); distinct.append(r)
        if len(distinct) < 2:
            continue
        now = distinct[0]
        nd = _pd(now.get("signal_date"))
        # prev = newest row at least compare_days older than now; else oldest we have
        prev = None
        if nd:
            for r in distinct[1:]:
                rd = _pd(r.get("signal_date"))
                if rd and (nd - rd).days >= compare_days:
                    prev = r
                    break
        if prev is None:
            prev = distinct[-1]
        if prev.get("signal_date") == now.get("signal_date"):
            continue
        a_now, a_prev = _f(now, "adj_composite"), _f(prev, "adj_composite")
        if a_now is None or a_prev is None:
            continue
        delta = round(a_now - a_prev, 1)
        if abs(round(delta)) < 1:        # include small but real moves (+/-1)
            continue

        c_now, c_prev = _f(now, "composite"), _f(prev, "composite")
        comp_delta = (c_now - c_prev) if (c_now is not None and c_prev is not None) else 0.0
        macro_contrib = delta - comp_delta

        drv_p, drv_pd = None, 0.0
        for p in _PILL:
            pn, pp = _f(now, p), _f(prev, p)
            if pn is None or pp is None:
                continue
            dd = pn - pp
            if abs(dd) > abs(drv_pd):
                drv_p, drv_pd = p, dd

        # "Macro-only" = the name's own factors barely moved but the overlay
        # shifted its score (the regime-flip band). Otherwise it's a real quant
        # story, attributed to its biggest-moving pillar even if macro also helped.
        # "Macro-driven" = the overlay shift dominated the move (regime days);
        # otherwise it's the name's own factors. Macro-driven names fold into one
        # collapsed chip so they never flood the daily feed of small quant moves.
        macro_only = abs(macro_contrib) >= 2 and abs(macro_contrib) > abs(comp_delta)
        if macro_only:
            driver, ddelta = "Macro overlay", round(macro_contrib, 1)
        elif drv_p and abs(drv_pd) >= 2:
            driver, ddelta = _PLAB[drv_p], round(drv_pd, 1)
        else:
            driver, ddelta = None, 0.0

        def _tier(v):
            return "HIGH" if v >= 60 else ("MOD" if v >= 45 else "LOW")

        movers.append({"ticker": tk, "now": a_now, "prev": a_prev, "delta": delta,
                       "quant_delta": round(comp_delta, 1), "macro_only": macro_only,
                       "now_tier": _tier(a_now), "prev_tier": _tier(a_prev),
                       "driver": driver, "driver_delta": ddelta})

    # Lead with name-specific (quant) movers so a uniform macro-shift band never
    # dominates the feed; collapse that band into one labelled summary entry.
    quant = sorted([m for m in movers if not m["macro_only"]],
                   key=lambda m: abs(m["quant_delta"]), reverse=True)
    macro = sorted([m for m in movers if m["macro_only"]],
                   key=lambda m: abs(m["delta"]), reverse=True)
    if collapse_macro and len(macro) >= 3:
        ups   = [m for m in macro if m["delta"] > 0]
        downs = [m for m in macro if m["delta"] < 0]
        grp = ups if len(ups) >= len(downs) else downs
        deltas = [m["delta"] for m in grp]
        out = quant[:top_n]
        out.append({"kind": "macro_summary", "count": len(grp),
                    "up": grp[0]["delta"] > 0, "lo": min(deltas), "hi": max(deltas)})
        return out
    return (quant + macro)[:top_n]


def _hero_card_html(macro: dict, results: list, movers: list = None,
                    wl_movers: list = None, has_watchlist: bool = False) -> str:
    """Login hero: today's macro regime, then 'What changed' -- a personal
    watchlist-movers row (when signed in) above the universe's biggest conviction
    changes -- plus the conviction-change alert hook. Falls back to top-conviction
    cards when no movers exist. Research framing only: conviction scores and
    factor attribution, no performance/benchmark claims (those stay gated)."""
    import html as _h
    regime = (macro or {}).get("regime", "NEUTRAL") or "NEUTRAL"
    rlab = regime.replace("_", " ").title()
    ru = regime.upper()
    if   "HIGH VOL" in ru:                      rcol = "#fb923c"
    elif "RISK_OFF" in ru or "RISK OFF" in ru:  rcol = "#fbbf24"
    elif "RISK_ON" in ru or "RISK ON" in ru or "BULL" in ru: rcol = "#34d399"
    else:                                       rcol = "#9fabc0"

    _TIER_COL = {"HIGH": "#34d399", "MOD": "#fbbf24", "LOW": "#f87171"}

    def _chip(m):
        if m.get("kind") == "macro_summary":
            up  = m["up"]
            col = "#34d399" if up else "#f87171"
            arr = "&#9650;" if up else "&#9660;"
            lo, hi = int(round(m["lo"])), int(round(m["hi"]))
            rng = f'{arr}{abs(lo)}' if lo == hi else f'{arr}{abs(lo)}&ndash;{abs(hi)}'
            return (
                f'<span style="display:inline-flex;flex-direction:column;gap:4px;'
                f'padding:8px 13px;margin:0 9px 9px 0;background:rgba(212,168,67,.06);'
                f'border:1px solid rgba(212,168,67,.28);border-radius:10px;'
                f'white-space:nowrap;vertical-align:top;">'
                f'<span style="display:flex;align-items:center;gap:7px;">'
                f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;'
                f'font-weight:600;">MACRO REGIME SHIFT</span></span>'
                f'<span style="display:flex;align-items:center;gap:8px;">'
                f'<span style="font-size:12px;color:#b3bed0;">{m["count"]} names moved together</span>'
                f'<span style="font-family:DM Mono,monospace;font-size:12px;color:{col};">{rng}'
                f'</span></span></span>')
        up  = m["delta"] >= 0
        col = "#34d399" if up else "#f87171"
        arr = "&#9650;" if up else "&#9660;"
        nt  = m.get("now_tier", "MOD")
        pt  = m.get("prev_tier", nt)
        ntc = _TIER_COL.get(nt, "#8896ac")
        if nt != pt:                      # crossed a tier line -> emphasise
            tarr = "&#9650;" if up else "&#9660;"
            tier_badge = (f'<span style="font-family:DM Mono,monospace;font-size:11px;'
                          f'color:{ntc};font-weight:600;border:1px solid {ntc}55;'
                          f'border-radius:5px;padding:1px 6px;">{tarr} {nt}</span>')
        else:
            tier_badge = (f'<span style="font-family:DM Mono,monospace;font-size:11px;'
                          f'color:{ntc};opacity:.85;">{nt}</span>')
        drv = ""
        if m.get("driver"):
            dd = m["driver_delta"]
            drv = (f'<span style="font-size:11px;color:#6b7686;">&middot; '
                   f'{_h.escape(str(m["driver"]))} {"+" if dd >= 0 else ""}{dd:.0f}</span>')
        ci = get_company_info(m["ticker"]) or {}
        nm = _h.escape(str(ci.get("name", "") or "")[:22])
        return (
            f'<span style="display:inline-flex;flex-direction:column;gap:4px;'
            f'padding:8px 13px;margin:0 9px 9px 0;background:rgba(255,255,255,.025);'
            f'border:1px solid rgba(255,255,255,.07);border-radius:10px;'
            f'white-space:nowrap;vertical-align:top;">'
            f'<span style="display:flex;align-items:center;gap:7px;">'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#e7ecf3;'
            f'font-weight:600;">{_h.escape(str(m["ticker"]))}</span>'
            f'<span style="font-size:11px;color:#6b7686;max-width:130px;overflow:hidden;'
            f'text-overflow:ellipsis;">{nm}</span></span>'
            f'<span style="display:flex;align-items:center;gap:8px;">'
            f'<span style="font-family:DM Mono,monospace;font-size:12px;color:#8896ac;">'
            f'{m["prev"]:.0f}&#8594;<span style="color:{ntc};font-weight:600;">{m["now"]:.0f}</span></span>'
            f'<span style="font-family:DM Mono,monospace;font-size:12px;color:{col};">'
            f'{arr}{abs(m["delta"]):.0f}</span>{tier_badge}{drv}</span></span>')

    def _label(text, qualifier=""):
        q = (f'<span style="color:#6b7686;">&middot; {qualifier}</span>') if qualifier else ""
        return (f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#9fabc0;'
                f'letter-spacing:.06em;margin-bottom:8px;">{text} {q}</div>')

    sections = ""

    # ── Personal: watchlist movers (only when signed in with a list) ──────────
    if has_watchlist:
        if wl_movers:
            wl_chips = "".join(_chip(m) for m in wl_movers)
            sections += (
                _label("ON YOUR WATCHLIST", f"{len(wl_movers)} changed")
                + f'<div style="display:flex;flex-wrap:wrap;margin-bottom:14px;">{wl_chips}</div>')
        else:
            sections += (
                _label("ON YOUR WATCHLIST")
                + '<div style="font-size:12px;color:#6b7686;margin-bottom:14px;">'
                  'No conviction changes on your watchlist since last scored.</div>')

    # ── Universe: biggest conviction changes (scrolling feed) ─────────────────
    if movers:
        chips = "".join(_chip(m) for m in movers)
        dur = max(24, len(movers) * 4)
        sections += (
            _label("TODAY&#39;S CONVICTION MOVES", "since last scored")
            + f'<div class="qntm-mv-wrap" style="width:100%;margin-bottom:4px;'
            f'-webkit-mask-image:linear-gradient(90deg,transparent,#000 4%,#000 96%,transparent);'
            f'mask-image:linear-gradient(90deg,transparent,#000 4%,#000 96%,transparent);">'
            f'<div class="qntm-mv" style="display:inline-flex;">'
            f'<span class="qntm-mv-set" style="display:inline-flex;">{chips}</span>'
            f'<span class="qntm-mv-set qntm-mv-dup" aria-hidden="true" '
            f'style="display:inline-flex;">{chips}</span></div></div>'
            f'<style>@keyframes qntm-mv-scroll{{from{{transform:translateX(0)}}'
            f'to{{transform:translateX(-50%)}}}}'
            f'.qntm-mv-wrap{{overflow-x:auto;overflow-y:hidden;-webkit-overflow-scrolling:touch;'
            f'scrollbar-width:none;}}'
            f'.qntm-mv-wrap::-webkit-scrollbar{{display:none;}}'
            f'.qntm-mv-dup{{display:none;}}'
            f'@media (hover:hover) and (pointer:fine){{'
            f'.qntm-mv-wrap{{overflow:hidden;}}'
            f'.qntm-mv-dup{{display:inline-flex;}}'
            f'.qntm-mv{{animation:qntm-mv-scroll {dur}s linear infinite;}}'
            f'.qntm-mv-wrap:hover .qntm-mv{{animation-play-state:paused;}}}}'
            f'@media (prefers-reduced-motion:reduce){{'
            f'.qntm-mv{{animation:none!important;}}'
            f'.qntm-mv-wrap{{overflow-x:auto;}}'
            f'.qntm-mv-dup{{display:none;}}}}'
            f'</style>')
    elif not has_watchlist:
        # No movers and no personal row -> fall back to top conviction names
        top = sorted([r for r in (results or []) if r.get("adj_action", r.get("action")) == "BUY"],
                     key=lambda x: x.get("adj_composite", x.get("composite", 0)), reverse=True)[:3]
        cards = ""
        for r in top:
            score = r.get("adj_composite", r.get("composite", 0)) or 0
            if   score >= 60: lbl, lc = "High",     "#34d399"
            elif score >= 45: lbl, lc = "Moderate", "#fbbf24"
            else:             lbl, lc = "Low",      "#f87171"
            ci = get_company_info(r["ticker"]) or {}
            nm = _h.escape(str(ci.get("name", r["ticker"]))[:26])
            cards += (
                f'<div style="flex:1;min-width:118px;background:rgba(255,255,255,.02);'
                f'border:1px solid rgba(255,255,255,.06);border-radius:8px;padding:10px 12px;">'
                f'<div style="font-family:DM Mono,monospace;font-size:14px;color:#e7ecf3;'
                f'font-weight:600;">{_h.escape(str(r["ticker"]))}</div>'
                f'<div style="font-size:11px;color:#6b7686;margin:2px 0 6px;white-space:nowrap;'
                f'overflow:hidden;text-overflow:ellipsis;">{nm}</div>'
                f'<div style="display:flex;align-items:center;gap:6px;">'
                f'<span style="font-size:11px;color:{lc};font-weight:600;">{lbl} conviction</span>'
                f'<span style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;">{score:.0f}</span>'
                f'</div></div>')
        if not cards:
            cards = ('<div style="font-size:12px;color:#6b7686;">Run the screener to surface '
                     'today\'s movers.</div>')
        sections += f'<div style="display:flex;gap:10px;flex-wrap:wrap;">{cards}</div>'

    return (
        f'<div style="background:linear-gradient(180deg,rgba(212,168,67,.06),rgba(0,0,0,0));'
        f'border:1px solid rgba(212,168,67,.18);border-radius:12px;padding:16px 18px;margin-bottom:14px;">'
        f'<div style="display:flex;align-items:center;justify-content:space-between;'
        f'flex-wrap:wrap;gap:8px;margin-bottom:12px;">'
        f'<span style="font-family:Syne,sans-serif;font-size:12px;letter-spacing:.14em;'
        f'color:#9fabc0;text-transform:uppercase;">Today at a glance</span>'
        f'<span style="font-family:DM Mono,monospace;font-size:12px;color:{rcol};">'
        f'&#9679; Macro regime: {_h.escape(rlab)}</span></div>'
        f'{sections}'
        f'<div style="display:flex;align-items:center;justify-content:space-between;'
        f'flex-wrap:wrap;gap:10px;padding-top:12px;margin-top:12px;'
        f'border-top:1px solid rgba(255,255,255,.06);">'
        f'<span style="font-size:12px;color:#b3bed0;">&#9889; Conviction-change alerts flag the '
        f'moment a name shifts tier (Pro).</span>'
        f'<span style="font-size:12px;color:#6b7686;">New here? Open any stock for its '
        f'plain-English rationale and 5-pillar breakdown.</span></div></div>')


def macro_regime_banner_html(macro: dict) -> str:
    """Renders the macro regime banner with live stats from macro_data."""
    regime    = macro.get("regime","NEUTRAL")
    events    = macro.get("active_events",[])
    source    = macro.get("source","estimated")
    vix       = macro.get("vix")
    oil       = macro.get("oil_price")
    vix_level = vix    # alias used in stats block
    oil_price = oil    # alias used in stats block
    n_hdl     = macro.get("headlines_scanned", 0)

    # Regime-scaled macro weight (matches apply_macro_overlay)
    macro_w = {"RISK_OFF":25,"HIGH VOLATILITY":25,"RISK_ON":15,"MILDLY BULLISH":15,"NEUTRAL":10}.get(regime,25)
    quant_w = 100 - macro_w

    cfg = {
        "RISK_ON":          ("#1D9E75","rgba(29,158,117,.08)","rgba(29,158,117,.25)","●","Macro overlay amplifying high-conviction signals"),
        "MILDLY BULLISH":   ("#4ade80","rgba(74,222,128,.06)","rgba(74,222,128,.2)","◕","Mildly bullish environment — quant signals favoured"),
        "NEUTRAL":          ("#d4a843","rgba(212,168,67,.07)","rgba(212,168,67,.2)","◐","Macro overlay at baseline — minimal sector adjustment"),
        "RISK_OFF":         ("#f87171","rgba(248,113,113,.07)","rgba(248,113,113,.2)","●","Macro dampening active — high-beta exposure reduced"),
        "HIGH VOLATILITY":  ("#f97316","rgba(249,115,22,.07)","rgba(249,115,22,.2)","⚡","High volatility — macro overlay at maximum dampening"),
    }.get(regime, ("#d4a843","rgba(212,168,67,.07)","rgba(212,168,67,.2)","◐","Macro overlay at baseline"))

    color, bg, border, icon, desc = cfg

    events_html = ""
    if events:
        nice = {"tariff_broad":"Tariff Headwinds","tariff_relief":"Tariff Relief",
                "fed_hawkish":"Fed Hawkish","fed_dovish":"Fed Dovish",
                "recession_signal":"Recession Signal","war_escalation":"War Escalation",
                "chip_export_ban":"Chip Export Ban","oil_spike":"Oil Spike"}
        for e in events[:4]:
            events_html += (f'<span style="background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.1);'
                           f'border-radius:3px;padding:2px 8px;font-size:13px;color:#b3bed0;margin-right:6px;">'
                           f'{nice.get(e,e.replace("_"," ").title())}</span>')

    # News read — prefer the explanatory narrative, fall back to the one-line summary
    import html as _html
    summary_txt  = (macro.get("narrative") or macro.get("summary") or "").strip()
    summary_html = ""
    if summary_txt:
        summary_html = (f'<div style="font-size:13px;color:#cbd5e1;margin-top:6px;line-height:1.5;">'
                        f'<span style="color:#9fabc0;">News read:</span> {_html.escape(summary_txt)}</div>')

    # Factor breakdown — how each active driver moved the regime score
    drivers = macro.get("drivers", [])
    breakdown_html = ""
    if drivers:
        rows = ""
        for d in drivers[:6]:
            c = d.get("contribution", 0.0) or 0.0
            if c < 0:   c_col, arrow = "#f87171", "&#9660;"
            elif c > 0: c_col, arrow = "#34d399", "&#9650;"
            else:       c_col, arrow = "#9fabc0", "&ndash;"
            sig = d.get("signals", 0)
            rows += (f'<div style="display:flex;align-items:center;justify-content:space-between;'
                     f'gap:10px;padding:3px 0;border-bottom:1px solid rgba(255,255,255,.04);">'
                     f'<span style="color:#cbd5e1;font-size:13px;">{_html.escape(str(d.get("label","")))}'
                     f'<span style="color:#6b7686;font-size:12px;"> &middot; {sig} signal{"s" if sig != 1 else ""}</span>'
                     f'</span>'
                     f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{c_col};">'
                     f'{arrow} {c:+.2f}</span></div>')
        net = macro.get("regime_score", 0.0) or 0.0
        breakdown_html = (
            f'<div style="margin-top:8px;background:rgba(0,0,0,.15);border-radius:6px;padding:8px 11px;">'
            f'<div style="font-size:12px;color:#9fabc0;letter-spacing:.06em;margin-bottom:4px;">'
            f'WHAT&#39;S MOVING THE REGIME</div>'
            f'{rows}'
            f'<div style="display:flex;justify-content:space-between;margin-top:5px;padding-top:5px;'
            f'border-top:1px solid rgba(255,255,255,.08);">'
            f'<span style="color:#b3bed0;font-size:13px;">Net regime score</span>'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{color};">'
            f'{net:+.2f} &#8594; {macro_w}% macro weight</span></div></div>')
    drivers_or_chips = breakdown_html if breakdown_html else f'<div style="margin-top:6px;">{events_html}</div>'

    # Source badge
    if macro.get("live"):
        src_parts = [f'⚡ Live']
        if n_hdl:  src_parts.append(f'{n_hdl} headlines')
        src_badge = f'<span style="font-size:13px;color:#1D9E75;margin-left:8px;">{" · ".join(src_parts)}</span>'
    else:
        src_badge = '<span style="font-size:13px;color:#b3bed0;margin-left:8px;">Est. · no live feeds</span>'

    # VIX / oil indicators
    indicators_html = ""
    if vix is not None:
        vix_col = "#f87171" if vix >= 30 else "#fbbf24" if vix >= 20 else "#1D9E75"
        indicators_html += (f'<div style="text-align:center;">'
                           f'<div style="font-family:DM Mono,monospace;font-size:16px;font-weight:500;color:{vix_col};">{vix:.1f}</div>'
                           f'<div style="font-size:13px;color:#b3bed0;">VIX</div></div>')
    if oil is not None:
        oil_col = "#f87171" if oil >= 90 else "#fbbf24" if oil >= 75 else "#1D9E75"
        indicators_html += (f'<div style="text-align:center;">'
                           f'<div style="font-family:DM Mono,monospace;font-size:16px;font-weight:500;color:{oil_col};">${oil:.0f}</div>'
                           f'<div style="font-size:13px;color:#b3bed0;">WTI Crude</div></div>')

    return (
        f'<div style="background:{bg};border:1px solid {border};border-radius:8px;'
        f'padding:14px 20px;margin-bottom:16px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:12px;">'
        f'<div style="display:flex;align-items:center;gap:10px;">'
        f'<span style="color:{color};font-size:13px;">{icon}</span>'
        f'<div>'
        f'<div style="display:flex;align-items:center;gap:6px;">'
        f'<span style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;'
        f'color:{color};letter-spacing:.1em;">MACRO REGIME: {regime}</span>'
        f'{src_badge}</div>'
        f'<div style="font-size:14px;color:#b3bed0;margin-top:2px;">{desc}</div>'
        f'{summary_html}'
        f'{drivers_or_chips}'
        f'</div></div>'
        f'<div style="display:flex;gap:20px;flex-wrap:wrap;align-items:flex-start;">'

        # Quant weight
        f'<div class="qntm-tip" style="text-align:center;cursor:help;">'
        f'<div style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;color:#e2e8f0;">{quant_w}%</div>'
        f'<div style="font-size:14px;color:#b3bed0;margin-top:3px;letter-spacing:.04em;">Quant Weight</div>'
        f'<span class="tip-box" style="width:260px;">'
        f'<div class="tip-title">Quant Weight</div>'
        f'<div class="tip-body">The percentage of each stock\'s final score driven purely by the 5-pillar factor model — momentum, quality, volume, value, and sentiment. Higher quant weight = model is doing the heavy lifting.</div>'
        f'</span></div>'

        # Macro weight
        f'<div class="qntm-tip" style="text-align:center;cursor:help;">'
        f'<div style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;color:{color};">{macro_w}%</div>'
        f'<div style="font-size:14px;color:#b3bed0;margin-top:3px;letter-spacing:.04em;">Macro Weight</div>'
        f'<span class="tip-box" style="width:260px;">'
        f'<div class="tip-title">Macro Weight</div>'
        f'<div class="tip-body">The percentage of each score adjusted by the current macro regime. In RISK_OFF this rises to 35% — dampening high-beta exposure. In NEUTRAL it drops to 10% to let quant signals dominate.</div>'
        f'</span></div>'

        # Active events
        f'<div class="qntm-tip" style="text-align:center;cursor:help;">'
        f'<div style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;color:#e2e8f0;">{len(events)}</div>'
        f'<div style="font-size:14px;color:#b3bed0;margin-top:3px;letter-spacing:.04em;">Active Events</div>'
        f'<span class="tip-box" style="width:260px;">'
        f'<div class="tip-title">Active Macro Events</div>'
        f'<div class="tip-body">Number of macro events currently detected — tariffs, Fed stance, geopolitical risk, oil shocks. Each event applies sector-level adjustments to scores. More events = stronger regime signal.</div>'
        f'</span></div>'

        + (
            f'<div class="qntm-tip" style="text-align:center;cursor:help;">'
            f'<div style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;'
            f'color:{"#f87171" if vix_level>=30 else "#fbbf24" if vix_level>=20 else "#1D9E75"};">{vix_level:.1f}</div>'
            f'<div style="font-size:14px;color:#b3bed0;margin-top:3px;letter-spacing:.04em;">VIX</div>'
            f'<span class="tip-box" style="left:auto;right:0;transform:none;">'
            f'<div class="tip-title">VIX — Fear Index</div>'
            f'<div class="tip-body">CBOE Volatility Index. Below 15 = calm market (RISK_ON). 15–25 = elevated uncertainty. Above 30 = fear/panic (forces RISK_OFF regime). VIX above 35 overrides all other regime signals.</div>'
            f'</span></div>'
            if vix_level is not None else ""
        )

        + (
            f'<div class="qntm-tip" style="text-align:center;cursor:help;">'
            f'<div style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;'
            f'color:{"#f87171" if oil_price>=90 else "#fbbf24" if oil_price>=75 else "#1D9E75"};">${oil_price:.0f}</div>'
            f'<div style="font-size:14px;color:#b3bed0;margin-top:3px;letter-spacing:.04em;">WTI Crude</div>'
            f'<span class="tip-box" style="left:auto;right:0;transform:none;">'
            f'<div class="tip-title">WTI Crude Oil Price</div>'
            f'<div class="tip-body">West Texas Intermediate crude price per barrel. Above $90 triggers an oil_spike macro event — bullish for Energy, bearish for Consumer Discretionary and Industrials. Below $65 signals weak demand.</div>'
            f'</span></div>'
            if oil_price is not None else ""
        )

        + f'</div></div></div>'
    )


def _build_why_html(r: dict) -> str:
    """Generate plain-English WHY THIS SCORE explanation from a score dict."""
    mom  = float(r.get("momentum",  50) or 50)
    qual = float(r.get("quality",   50) or 50)
    vol  = float(r.get("volume",    50) or 50)
    val  = float(r.get("value",     50) or 50)
    sent = float(r.get("sentiment", 50) or 50)
    adj  = float(r.get("adj_composite", r.get("composite", 50)) or 50)
    raw  = float(r.get("composite", adj) or adj)
    delta = adj - raw

    pillars = sorted([("MOM",mom),("QUAL",qual),("VOL",vol),("VAL",val),("SENT",sent)],
                     key=lambda x: x[1], reverse=True)
    PILLAR_EXPLAIN = {
        "MOM":  ("price trend and relative strength are strong",  "price trend is weakening"),
        "QUAL": ("earnings quality and balance sheet are solid",  "earnings quality is a concern"),
        "VOL":  ("volume confirms institutional interest",        "volume signal is weak"),
        "VAL":  ("stock looks undervalued vs sector peers",       "stock looks stretched on valuation"),
        "SENT": ("analyst sentiment is improving",                "analyst sentiment is negative"),
    }
    drivers_text = []
    watches_text = []
    for pname, pval in pillars:
        pos, neg = PILLAR_EXPLAIN.get(pname, (pname, pname))
        if pval >= 65:   drivers_text.append(pos)
        elif pval < 45:  watches_text.append(neg)
    why_parts = []
    if drivers_text:
        why_parts.append(f'<span style="color:#b3bed0;">{"; ".join(drivers_text[:2]).capitalize()}.</span>')
    if watches_text:
        why_parts.append(f'<span style="color:#f87171;">Watch: {watches_text[0]}.</span>')
    if abs(delta) >= 2:
        macro_txt = "Macro regime is adding a tailwind." if delta > 0 else "Macro regime is dampening the score."
        why_parts.append(f'<span style="color:{"#34d399" if delta>0 else "#f97316"};">{macro_txt}</span>')
    if not why_parts:
        return ""
    return (
        f'<div style="font-size:13px;line-height:1.6;padding:8px 10px;margin-top:8px;'
        f'background:rgba(255,255,255,.02);border-radius:4px;border-left:2px solid rgba(255,255,255,.08);">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;letter-spacing:.08em;">WHY THIS SCORE · </span>'
        + " ".join(why_parts) +
        f'</div>'
    )


@st.cache_data(ttl=300, show_spinner=False)
def _whats_changed_html(ticker: str, now_sig_date: str = "") -> str:
    """Detail-page 'WHAT'S CHANGED' box: per-pillar deltas and the macro-overlay
    delta between this ticker's two most recent CLEAN scored days. Factual model
    deltas only — no interpretation, no performance claims. Returns '' when there
    is no prior scored day to compare against. Cached per (ticker, date); only
    called on the single search/detail view, never per list card."""
    if not ticker:
        return ""
    try:
        from data_refresh import _get_supabase
        from datetime import date, timedelta
        sb = _get_supabase()
        if not sb:
            return ""
        since = (date.today() - timedelta(days=21)).isoformat()
        rows = (sb.table("signal_log")
                .select("signal_date,adj_composite,composite,"
                        "momentum,quality,volume,value,sentiment")
                .eq("ticker", ticker)
                .gte("signal_date", since)
                .not_.is_("composite", "null")
                .order("signal_date", desc=True)
                .execute()).data or []
    except Exception:
        return ""

    from datetime import date as _date

    def _pd(s):
        try:
            return _date.fromisoformat(s)
        except (TypeError, ValueError):
            return None

    distinct, seen = [], set()
    for rr in rows:
        d = rr.get("signal_date")
        if d and d not in seen:
            seen.add(d); distinct.append(rr)
    if len(distinct) < 2:
        return ""
    now = distinct[0]
    nd = _pd(now.get("signal_date"))
    prev = None
    if nd:
        for rr in distinct[1:]:
            rd = _pd(rr.get("signal_date"))
            if rd and (nd - rd).days >= 1:    # prior scored session
                prev = rr
                break
    if prev is None:
        prev = distinct[-1]
    if prev.get("signal_date") == now.get("signal_date"):
        return ""

    def _f(x, k):
        try:
            return float(x.get(k))
        except (TypeError, ValueError):
            return None

    parts = []
    for key, lab in (("momentum", "Momentum"), ("quality", "Quality"),
                     ("volume", "Volume"), ("value", "Value"), ("sentiment", "Sentiment")):
        pn, pp = _f(now, key), _f(prev, key)
        if pn is None or pp is None:
            continue
        dd = round(pn - pp)
        if abs(dd) < 1:
            continue
        c = "#34d399" if dd > 0 else "#f87171"
        a = "&#9650;" if dd > 0 else "&#9660;"
        parts.append(
            f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:16px;">'
            f'<span style="font-size:13px;color:#b3bed0;">{lab}</span>'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{c};">'
            f'{a}{abs(int(dd))}</span></span>')

    a_now, c_now = _f(now, "adj_composite"), _f(now, "composite")
    a_prev, c_prev = _f(prev, "adj_composite"), _f(prev, "composite")
    if None not in (a_now, c_now, a_prev, c_prev):
        md = round((a_now - c_now) - (a_prev - c_prev))
        if abs(md) >= 1:
            c = "#34d399" if md > 0 else "#f97316"
            a = "&#9650;" if md > 0 else "&#9660;"
            parts.append(
                f'<span style="display:inline-flex;align-items:center;gap:5px;margin-right:16px;">'
                f'<span style="font-size:13px;color:#b3bed0;">Macro overlay</span>'
                f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{c};">'
                f'{a}{abs(int(md))}</span></span>')
        else:
            parts.append('<span style="font-size:13px;color:#6b7686;margin-right:16px;">'
                         'Macro overlay unchanged</span>')

    overall = ""
    if a_now is not None and a_prev is not None and abs(round(a_now - a_prev)) >= 1:
        c = "#34d399" if a_now > a_prev else "#f87171"
        overall = (f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;">'
                   f'Score {a_prev:.0f}&#8594;'
                   f'<span style="color:{c};font-weight:600;">{a_now:.0f}</span></span>')

    pdate = prev.get("signal_date", "last scored")
    has_pillar_move = any("Momentum" in p or "Quality" in p or "Volume" in p
                          or "Value" in p or "Sentiment" in p for p in parts)
    if not has_pillar_move and not overall:
        return (
            f'<div style="font-size:13px;line-height:1.6;padding:8px 10px;margin-top:8px;'
            f'background:rgba(255,255,255,.02);border-radius:4px;'
            f'border-left:2px solid rgba(255,255,255,.08);">'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
            f'letter-spacing:.08em;">WHAT&#39;S CHANGED &middot; </span>'
            f'<span style="color:#6b7686;">Holding steady since {pdate} '
            f'&mdash; no pillar moved a point.</span></div>')
    return (
        f'<div style="font-size:13px;line-height:1.7;padding:8px 10px;margin-top:8px;'
        f'background:rgba(255,255,255,.02);border-radius:4px;'
        f'border-left:2px solid rgba(52,211,153,.18);">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
        f'letter-spacing:.08em;">WHAT&#39;S CHANGED &middot; since {pdate} </span>'
        f'{overall}<div style="margin-top:6px;">{"".join(parts)}</div></div>')


# ── Shared iframe tail for batch card pages ───────────────────────────────────
# Style + click-to-expand + dynamic resize logic. Used wherever multiple cards
# are rendered into a single st.components.v1.html iframe (screener, watchlist,
# portfolio, model portfolio). Resize logic uses ResizeObserver + multiple
# retries so the iframe grows to actual content height even when paint is slow.
CARD_IFRAME_TAIL = """
<style>
@media(max-width:640px){
  .qcard-pillars{grid-template-columns:repeat(2,1fr)!important;}
  .qntm-action-link{font-size:13px!important;letter-spacing:.02em!important;padding:8px 4px!important;}
  .qcard-4box{grid-template-columns:repeat(2,1fr)!important;}
}
body{margin:0;padding-bottom:8px;}
.qntm-action-link:hover{filter:brightness(1.15);}
</style>
<script>
(function(){
  var _lastH=0, _pending=null;
  function _measure(){
    var h=Math.max(
      document.documentElement.scrollHeight,
      document.body.scrollHeight,
      document.documentElement.offsetHeight,
      document.body.offsetHeight
    );
    // Set EXACT height (grow on expand, shrink on collapse) so per-card iframes
    // don't leave a gap after collapsing and never clip when expanded.
    if(h!==_lastH){
      _lastH=h;
      if(window.Streamlit&&window.Streamlit.setFrameHeight){
        window.Streamlit.setFrameHeight(h);
      }
      if(window.parent){
        try{ window.parent.postMessage(
          {type:'streamlit:setFrameHeight',height:h},'*'); }catch(e){}
      }
    }
  }
  function _schedule(){
    if(_pending)cancelAnimationFrame(_pending);
    _pending=requestAnimationFrame(_measure);
  }
  // Click-to-expand (one open at a time)
  document.querySelectorAll('.qcard-header').forEach(function(h){
    h.addEventListener('click',function(){
      var d=h.querySelector('.qcard-detail');
      if(!d)return;
      var open=d.style.display==='block';
      document.querySelectorAll('.qcard-detail').forEach(function(x){x.style.display='none';});
      if(!open)d.style.display='block';
      setTimeout(_schedule,10);
      setTimeout(_schedule,150);
    });
  });
  // Action button (Add/Remove watchlist/portfolio) — navigate the PARENT.
  // window.open(url,'_top') is the call the component-iframe sandbox permits
  // (a plain <a target=_top> and window.top.location assignment are blocked).
  document.querySelectorAll('.qntm-action-link').forEach(function(a){
    a.addEventListener('click',function(e){
      e.preventDefault();
      e.stopPropagation();
      var href=a.getAttribute('data-wlhref');
      if(href){ window.open(href,'_top'); }
    });
  });
  // ResizeObserver picks up font/image loads, async layout shifts
  if(window.ResizeObserver){
    new ResizeObserver(_schedule).observe(document.body);
  }
  // Multi-shot initial measure: covers slow paint on long lists
  [0,100,300,600,1200].forEach(function(t){ setTimeout(_schedule,t); });
  window.addEventListener('load',_schedule);
})();
</script>
"""


def _debug_line(r: dict) -> str:
    """Return a debug strip showing raw score fields. Only renders when
    ?debug=1 is in the URL. Used to diagnose score divergence across pages.
    Append the result to a card's detail HTML."""
    try:
        if st.query_params.get("debug") != "1":
            return ""
    except Exception:
        return ""
    return (
        f'<div style="margin-top:8px;padding:6px 10px;background:rgba(212,168,67,.06);'
        f'border:1px dashed rgba(212,168,67,.3);border-radius:4px;'
        f'font-family:DM Mono,monospace;font-size:13px;color:#d4a843;'
        f'letter-spacing:.04em;line-height:1.5;">'
        f'DEBUG · ticker={r.get("ticker","?")} · '
        f'adj_composite={r.get("adj_composite","?")} · '
        f'composite={r.get("composite","?")} · '
        f'score_delta={r.get("score_delta","?")} · '
        f'adj_action={r.get("adj_action","?")} · '
        f'signal_date={str(r.get("signal_date","?"))[:10]}'
        f'</div>'
    )


def _card_action_button(tk: str, mode: str, nav: str, in_set: set,
                        uid_v: str, pln_v: str, remove_url: str = None) -> str:
    """Build the styled main-document action link under a card. `mode` selects
    the action vocabulary: 'watchlist' | 'portfolio' | 'simulator'."""
    if not uid_v or not tk:
        return ""
    _qp = f"?qnav={nav}&uid={uid_v}&plan={pln_v}&ck=1&_n={nav}"
    if mode == "portfolio":
        # Portfolio holdings: remove only (adding happens via the Add-holding form).
        url = remove_url or (_qp + f"&port_action=remove&port_ticker={tk}")
        bg, bd, col, lbl = "rgba(248,113,113,.07)", "rgba(248,113,113,.28)", "#f87171", f"✕ Remove {tk} from Portfolio"
    elif mode == "simulator":
        if tk in in_set:
            url = _qp + f"&sim_remove={tk}"
            bg, bd, col, lbl = "rgba(248,113,113,.07)", "rgba(248,113,113,.28)", "#f87171", f"✕ Remove {tk} from Simulation"
        else:
            url = _qp + f"&sim_add={tk}"
            bg, bd, col, lbl = "rgba(212,168,67,.08)", "rgba(212,168,67,.3)", "#d4a843", f"☆ Add {tk} to Simulation"
    else:  # watchlist
        if tk in in_set:
            url = remove_url or (_qp + f"&wl_action=remove&wl_ticker={tk}")
            bg, bd, col, lbl = "rgba(248,113,113,.07)", "rgba(248,113,113,.28)", "#f87171", f"✕ Remove {tk} from Watchlist"
        else:
            url = _qp + f"&wl_action=add&wl_ticker={tk}"
            bg, bd, col, lbl = "rgba(212,168,67,.08)", "rgba(212,168,67,.3)", "#d4a843", f"☆ Add {tk} to Watchlist"
    return (
        f'<a href="{url}" target="_self" '
        f'style="display:block;width:100%;text-align:center;'
        f'padding:9px 7px;margin:12px 0 2px;box-sizing:border-box;background:{bg};'
        f'border:1px solid {bd};border-radius:6px;font-family:Syne,sans-serif;'
        f'font-size:13px;font-weight:700;letter-spacing:.04em;text-transform:uppercase;'
        f'color:{col};text-decoration:none;cursor:pointer;line-height:1.3;'
        f'word-break:break-word;">{lbl}</a>'
    )


def build_card_html(r: dict, nav: str = "screener", is_gem: bool = False,
                    company_info: dict = None, in_list: set = None,
                    extra_detail: str = "", remove_url: str = None,
                    mode: str = "watchlist") -> str:
    """Return ONE card's HTML (no st call) so callers can concatenate a whole
    list and render it in a single st.markdown — far less Streamlit overhead
    than one markdown call per card."""
    tk = r.get("ticker", "")
    _uid_v = (st.session_state.user or {}).get("id", "")
    _pln_v = (st.session_state.user or {}).get("plan", "free")
    if in_list is None and mode == "watchlist":
        in_list = {w["ticker"] for w in get_watchlist(_uid_v)} if _uid_v else set()
    _btn = _card_action_button(tk, mode, nav, in_list or set(), _uid_v, _pln_v, remove_url) if (_uid_v and tk) else ""
    return factor_panel_html(r, is_gem, company_info=company_info,
                             wl_btn=(extra_detail + _btn), as_details=True)


def render_cards_batch(cards_html: str):
    """Render a concatenated string of card HTML in one st.markdown call."""
    if cards_html:
        st.markdown(cards_html, unsafe_allow_html=True)


def render_card_with_watchlist(r: dict, nav: str = "screener", is_gem: bool = False,
                               company_info: dict = None, in_list: set = None,
                               extra_detail: str = "", remove_url: str = None,
                               mode: str = "watchlist"):
    """Render ONE stock card directly in the MAIN document via st.markdown
    (NOT an iframe). This sidesteps the whole iframe saga: no fixed-height
    clipping, no resize JS, no sandboxed navigation. Expand/collapse uses a
    native <details> element (pure HTML, no JS), and the action button is a
    plain target=_self link that reaches the router.

    `mode`: 'watchlist' | 'portfolio' | 'simulator'.
    """
    st.markdown(
        build_card_html(r, nav, is_gem, company_info, in_list, extra_detail, remove_url, mode),
        unsafe_allow_html=True,
    )


def _unused_external_button():
    pass


def render_watchlist_actions(tickers: list, nav: str = "screener", in_list: set = None):
    """Legacy chip-row (kept for any caller still using it). Prefer
    render_card_with_watchlist for per-card buttons."""
    if not tickers:
        return
    _uid_v = (st.session_state.user or {}).get("id", "")
    _pln_v = (st.session_state.user or {}).get("plan", "free")
    if not _uid_v:
        return
    if in_list is None:
        in_list = {w["ticker"] for w in get_watchlist(_uid_v)}
    _qp = f"?qnav={nav}&uid={_uid_v}&plan={_pln_v}&ck=1&_n={nav}"
    _chips = ""
    for tk in tickers:
        if tk in in_list:
            _url = _qp + f"&wl_action=remove&wl_ticker={tk}"
            _bg, _bd, _col, _lbl = "rgba(248,113,113,.07)", "rgba(248,113,113,.25)", "#f87171", f"✕ {tk}"
        else:
            _url = _qp + f"&wl_action=add&wl_ticker={tk}"
            _bg, _bd, _col, _lbl = "rgba(212,168,67,.08)", "rgba(212,168,67,.28)", "#d4a843", f"☆ {tk}"
        _chips += (
            f'<a href="{_url}" target="_self" style="text-decoration:none;display:inline-block;'
            f'margin:3px;padding:6px 12px;background:{_bg};border:1px solid {_bd};'
            f'border-radius:6px;font-family:DM Mono,monospace;font-size:13px;font-weight:600;'
            f'color:{_col};white-space:nowrap;">{_lbl}</a>'
        )
    st.markdown(
        '<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
        'letter-spacing:.12em;margin:14px 0 6px;">WATCHLIST · TAP TO ADD / REMOVE</div>'
        f'<div style="display:flex;flex-wrap:wrap;">{_chips}</div>',
        unsafe_allow_html=True,
    )


def factor_panel_html(r: dict, is_gem: bool = False, company_info: dict = None, card_id: str = None, rank: int = 0, suppress_wl_btn: bool = False, wl_btn: str = "", as_details: bool = False, extra_detail: str = "") -> str:
    """
    Collapsed card using radio-button CSS hack for one-at-a-time expand.
    All cards share radio group "qntm_card" — checking one unchecks others.
    No JS, no URL params, no rerun. Works inside Streamlit HTML sandbox.
    """
    import hashlib as _hl
    act    = r.get("adj_action", r.get("action","HOLD"))
    score  = r.get("adj_composite", r.get("composite", 50))
    quant  = r.get("composite", 50)
    # MACRO box = macro overlay impact. signal_log rows don't carry `score_delta`,
    # so fall back to adj_composite - composite (the same values shown as
    # score/quant) instead of defaulting to 0 and always reading "+0.0".
    delta  = r.get("score_delta")
    if delta is None:
        try:
            delta = float(score or 50) - float(quant or 50)
        except Exception:
            delta = 0

    # Percentile rank: use the stored value if present (model-portfolio/page code
    # may stamp one), else rank against the full latest-day universe via the
    # shared session-cached helper, else derive from the current scan session,
    # else fall back to 50. The universe helper is what keeps the RANK cell from
    # silently reading "50th" on pages that don't carry pct_rank.
    _pct_rank = r.get("pct_rank")
    if _pct_rank is None:
        _pct_rank = _pct_rank_of(score)
    if _pct_rank is None:
        try:
            _all = st.session_state.get("scan_results") or []
            _comps = [float(s.get("adj_composite", s.get("composite", 50)) or 50) for s in _all]
            if _comps:
                _sc_v = float(score or 50)
                _pct_rank = sum(1 for c in _comps if c <= _sc_v) / len(_comps) * 100
        except Exception:
            _pct_rank = None
    if _pct_rank is None:
        _pct_rank = 50

    act_colors = {
        "BUY":  ("#34d399", "rgba(52,211,153,.08)", "rgba(52,211,153,.22)"),
        "HOLD": ("#fbbf24", "rgba(251,191,36,.06)", "rgba(251,191,36,.2)"),
        "SELL": ("#f87171", "rgba(248,113,113,.08)",  "rgba(248,113,113,.2)"),
    }
    act_c, act_bg, act_brd = act_colors.get(act, ("#9fabc0","rgba(100,116,139,.1)","rgba(100,116,139,.3)"))

    action_label = "High Conviction" if act=="BUY" else ("Low Conviction" if act=="SELL" else "Moderate")
    action_arrow = "▲" if act=="BUY" else ("▼" if act=="SELL" else "→")
    gem_badge    = " 💎" if is_gem else ""
    # Market-cap category badge — neutral metadata styling (intentionally NOT a
    # conviction colour, so it reads as an attribute, not a signal). Shown on the
    # collapsed summary of every card. Sourced from signal_log.mktcap on the
    # score row; omitted when the row carries no cap (a context that didn't load
    # it), so this never prints "Unknown".
    _cap_raw    = str(r.get("mktcap") or "").strip().lower()
    _CAP_LABELS = {"large": "LARGE CAP", "mid": "MID CAP", "small": "SMALL CAP"}
    cap_badge   = (
        f'<span style="font-family:DM Mono,monospace;font-size:11px;font-weight:600;'
        f'letter-spacing:.08em;color:#8896ac;background:rgba(136,150,172,.10);'
        f'border:1px solid rgba(136,150,172,.20);border-radius:5px;padding:1px 7px;'
        f'white-space:nowrap;flex-shrink:0;">{_CAP_LABELS[_cap_raw]}</span>'
    ) if _cap_raw in _CAP_LABELS else ""
    delta_c      = "#34d399" if delta >= 0 else "#f87171"
    delta_str    = f"+{delta:.2f}" if delta >= 0 else f"{delta:.2f}"
    # A bald "+0.0" in green reads as "macro broken/off" even when the sector
    # overlay is live — the per-name point impact is just small (the overlay is
    # a gentle tilt, especially in RISK_ON where it's down-weighted to 15%).
    # Distinguish three states honestly:
    #   • impact >= 0.005  -> signed 2dp, green/red (a real tilt)
    #   • impact ~0, overlay ACTIVE on this row -> muted "≈0" (on, negligible)
    #   • impact ~0, no overlay on this row      -> muted "—" (macro genuinely off)
    try:
        _ov_tilt = float(r.get("macro_overlay") or 0.0)
    except Exception:
        _ov_tilt = 0.0
    if abs(delta) < 0.005:
        if abs(_ov_tilt) > 1e-9:
            delta_str, delta_c = "≈0", "#8896ac"
        else:
            delta_str, delta_c = "—", "#8896ac"

    ci_name      = (company_info or {}).get("name", "")
    name_display = ci_name if (ci_name and ci_name != r["ticker"]) else ""

    # Unique ID per card — use ticker + score hash for stability
    cid = card_id or ("c" + _hl.md5(f'{r["ticker"]}{score}'.encode()).hexdigest()[:8])

    # ── Pillar bars ────────────────────────────────────────────────────────────
    pillars = [
        ("MOM",  r.get("momentum", 50)),
        ("QUAL", r.get("quality",  50)),
        ("VOL",  r.get("volume",   50)),
        ("VAL",  r.get("value",    50)),
        ("SENT", r.get("sentiment",50)),
    ]
    PILLAR_FULL_NAMES = {"MOM":"Momentum","QUAL":"Quality","VOL":"Volume","VAL":"Value","SENT":"Sentiment"}
    pillar_bars = ""
    for pname, pval in pillars:
        pc   = "#34d399" if pval>=65 else "#fbbf24" if pval>=50 else "#f87171"
        full = PILLAR_FULL_NAMES.get(pname, pname)
        tip  = PILLAR_TIPS.get(full, {})
        tip_weight = tip.get("weight","")
        pillar_bars += (
            f'<div style="flex:1;min-width:72px;">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:3px;">'
            f'<span style="font-size:13px;color:#9fabc0;">{full}</span>'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{pc};font-weight:700;">{pval:.0f}</span>'
            f'</div>'
            f'<div style="background:rgba(255,255,255,.05);border-radius:3px;height:4px;overflow:hidden;">'
            f'<div style="width:{pval}%;height:100%;background:{pc};border-radius:3px;"></div>'
            f'</div></div>'
        )

    sorted_pillars = sorted(pillars, key=lambda x: x[1], reverse=True)
    top2 = [p[0] for p in sorted_pillars[:2]]
    weak = [p[0] for p in sorted_pillars if p[1] < 45]
    driver = f"Driven by {top2[0]} + {top2[1]}"
    if weak: driver += f" — watch {weak[0]}"

    why_html   = _build_why_html(r)

    # ── QNTM Valuation Range → "Value Position" bar ────────────────────────────
    # Descriptive valuation context (not a target/forecast): a green(low)→red(high)
    # track with a marker at today's value_position, low/high range $ as end labels.
    _vr_low, _vr_high = r.get("val_low"), r.get("val_high")
    _vr_pos, _vr_basis = r.get("value_position"), (r.get("val_basis") or "na")
    _valrange_html = ""
    if _vr_basis != "na" and _vr_low is not None and _vr_high is not None:
        # Marker tracks the LIVE price (worker keeps signal_log.price fresh ~90s)
        # against last night's band; fall back to the stored value_position.
        _pos = None
        try:
            _lo, _hi = float(_vr_low), float(_vr_high)
            _pr = float(r.get("price")) if r.get("price") else None
            if _pr is not None and _hi > _lo:
                _pos = max(0.0, min(100.0, (_pr - _lo) / (_hi - _lo) * 100.0))
            elif _vr_pos is not None:
                _pos = max(0.0, min(100.0, float(_vr_pos)))
        except (TypeError, ValueError):
            _pos = max(0.0, min(100.0, float(_vr_pos))) if _vr_pos is not None else None
        if _pos is not None:
            if   _pos <= 20: _zone, _zc = "Lower range",     "#34d399"
            elif _pos <= 40: _zone, _zc = "Lower-mid range", "#86efac"
            elif _pos <= 60: _zone, _zc = "Mid range",       "#fbbf24"
            elif _pos <= 80: _zone, _zc = "Upper-mid range", "#fb923c"
            else:            _zone, _zc = "Upper range",      "#f87171"
            _basis_note = "" if _vr_basis == "valuation" else " · technical range"
            _cur = f'${r["price"]:,.2f}' if r.get("price") else "—"
            _valrange_html = (
                f'<div style="margin-top:14px;padding-top:12px;'
                f'border-top:1px solid rgba(255,255,255,.04);">'
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'margin-bottom:7px;">'
                f'<span style="font-size:13px;color:#8896ac;letter-spacing:.06em;">VALUE POSITION</span>'
                f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{_zc};'
                f'font-weight:700;">{_pos:.0f}% · {_zone}</span></div>'
                f'<div style="position:relative;height:8px;border-radius:5px;'
                f'background:linear-gradient(90deg,#34d399 0%,#fbbf24 50%,#f87171 100%);">'
                f'<div style="position:absolute;top:-3px;left:calc({_pos:.1f}% - 2px);'
                f'width:4px;height:14px;border-radius:2px;background:#e8edf4;'
                f'box-shadow:0 0 4px rgba(0,0,0,.6);"></div></div>'
                f'<div style="display:flex;justify-content:space-between;margin-top:5px;'
                f'font-family:DM Mono,monospace;font-size:12px;color:#7e8aa0;">'
                f'<span>${_vr_low:,.2f}</span>'
                f'<span style="color:#9fabc0;">{_cur} now{_basis_note}</span>'
                f'<span>${_vr_high:,.2f}</span></div></div>'
            )
    price_html = ""
    if r.get("price"):
        price_html = (
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;">'
            f'${r["price"]:,.2f}</span>'
            + (f' <span style="font-size:13px;color:#94a3b8;">· {r["signal_date"]}</span>'
               if r.get("signal_date") else "")
        )
    elif r.get("signal_date"):
        price_html = f'<span style="font-size:13px;color:#94a3b8;">{r["signal_date"]}</span>'

    # ── Watchlist button DISABLED in-iframe ────────────────────────────────────
    # Links inside st.components.v1.html iframes cannot drive the parent session
    # (this was the long-standing add/remove bug). The watchlist action now lives
    # in the MAIN document via st.markdown + target="_self" (handoff rule #1),
    # rendered alongside each card list. factor_panel_html never emits it.
    _wl_btn_html = '' if suppress_wl_btn else wl_btn

    detail_html = (
        f'<div class="qcard-detail" style="display:none;padding:0 20px 20px;'
        f'border-top:1px solid rgba(255,255,255,.05);">'
        + (f'<div style="display:flex;gap:12px;align-items:center;flex-wrap:wrap;'
           f'padding:10px 0 12px;">{price_html}'
           f'<span style="font-size:13px;color:#8896ac;">{r.get("sector","")[:20]}</span>'
           f'<span style="font-size:13px;color:#8896ac;">{driver}</span>'
           f'</div>' if (price_html or r.get("sector")) else "")
        + f'<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;margin-bottom:16px;"'
          f'class="qcard-pillars">{pillar_bars}</div>'
        + f'<div class="qcard-4box" style="display:grid;grid-template-columns:repeat(4,1fr);gap:5px;'
        f'padding-top:10px;border-top:1px solid rgba(255,255,255,.04);">'
        f'<div style="background:rgba(255,255,255,.03);border-radius:4px;padding:6px 10px;">'
        f'<div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:2px;">QUANT</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:14px;color:#b3bed0;">{quant:.1f}</div></div>'
        f'<div style="background:rgba(255,255,255,.03);border-radius:4px;padding:6px 10px;">'
        f'<div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:2px;">MACRO</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:14px;color:{delta_c};">{delta_str}</div></div>'
        f'<div style="background:rgba(255,255,255,.03);border-radius:4px;padding:6px 10px;">'
        f'<div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:2px;">BLEND</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:14px;color:#d4a843;">75/25</div></div>'
        f'<div style="background:rgba(255,255,255,.03);border-radius:4px;padding:6px 10px;">'
        f'<div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:2px;">RANK</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:14px;color:#b3bed0;">'
        f'{_ordinal(_pct_rank)}</div></div>'
        f'</div>'
        + _valrange_html
        + why_html
        + extra_detail
        + _wl_btn_html
        + f'</div>'
    )

    # ── Native <details> card (no iframe, no JS) for st.markdown rendering ────
    if as_details:
        # detail_html starts with '<div class="qcard-detail" style="display:none;...'
        # — strip the display:none so the <details> body shows when open.
        _body = detail_html.replace('display:none;', '', 1)
        # Optional one-line meta strip rendered inside the collapsed summary
        # (under the conviction label). Callers in list contexts — watchlist,
        # portfolio, model portfolio — populate r["_summary_meta_html"] with a
        # short DM-Mono line like "ENT 05/30 · +2.4% today" so a user can scan
        # entry date and intraday move at a glance without expanding the card.
        _summary_meta = r.get("_summary_meta_html", "")
        _mini_chart = r.get("_mini_chart_html", "")
        _summary = (
            f'<summary style="list-style:none;cursor:pointer;display:flex;'
            f'justify-content:space-between;align-items:center;padding:13px 18px;">'
            f'<div style="min-width:0;flex:1;overflow:hidden;">'
            f'<div style="display:flex;align-items:center;gap:6px;">'
            f'<span style="font-family:Syne,sans-serif;font-size:15px;font-weight:800;'
            f'color:#e2e8f0;white-space:nowrap;">{r["ticker"]}{gem_badge}</span>'
            + cap_badge
            + (f'<span style="font-size:13px;color:#94a3b8;overflow:hidden;text-overflow:ellipsis;'
               f'white-space:nowrap;">{name_display}</span>' if name_display else "")
            + f'</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;'
            f'color:{act_c};letter-spacing:.06em;margin-top:1px;">{action_arrow} {action_label}</div>'
            + (_summary_meta or "")
            + f'</div>'
            f'<div style="display:flex;align-items:center;gap:8px;flex-shrink:0;margin-left:8px;">'
            f'<span style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;color:{act_c};">{score:.0f}</span>'
            f'<span style="font-size:14px;color:{act_c};">{action_arrow}</span>'
            f'<span style="font-size:13px;color:#94a3b8;">›</span>'
            f'</div>'
            f'</summary>'
        )
        return (
            f'<details name="qntm-cards" class="qntm-card-details" '
            f'style="margin-bottom:6px;background:rgba(255,255,255,.02);'
            f'border:1px solid rgba(255,255,255,.06);border-left:3px solid {act_c};'
            f'border-radius:8px;overflow:hidden;">'
            + _summary + _body + _mini_chart + '</details>'
        )

    # ── div with data-cid for JS binding ─────────────────────────────────────
    return (
        f'<div class="qcard-wrap" style="margin-bottom:4px;">'
        f'<div class="qcard-header" data-cid="{cid}" style="display:block;background:rgba(255,255,255,.02);'
        f'border:1px solid rgba(255,255,255,.06);border-left:3px solid {act_c};'
        f'border-radius:8px;overflow:hidden;cursor:pointer;'
        f'transition:border-color .15s ease;">'
        # Summary row
        f'<div style="display:flex;justify-content:space-between;align-items:center;'
        f'padding:13px 18px;">'
        f'<div style="min-width:0;flex:1;overflow:hidden;">'
        f'<div style="display:flex;align-items:center;gap:6px;">'
        f'<span style="font-family:Syne,sans-serif;font-size:15px;font-weight:800;'
        f'color:#e2e8f0;white-space:nowrap;">{r["ticker"]}{gem_badge}</span>'
        + cap_badge
        + (f'<span style="font-size:13px;color:#94a3b8;overflow:hidden;text-overflow:ellipsis;'
           f'white-space:nowrap;">{name_display}</span>' if name_display else "")
        + f'</div>'
        f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;'
        f'color:{act_c};letter-spacing:.06em;margin-top:1px;">'
        f'{action_arrow} {action_label}</div>'
        f'</div>'
        f'<div style="display:flex;align-items:center;gap:8px;flex-shrink:0;margin-left:8px;">'
        f'<span style="font-family:DM Mono,monospace;font-size:20px;font-weight:700;color:{act_c};">'
        f'{score:.0f}</span>'
        f'<span style="font-size:14px;color:{act_c};">{action_arrow}</span>'
        f'<span style="font-size:13px;color:#94a3b8;transition:transform .2s;">›</span>'
        f'</div>'
        f'</div>'
        + detail_html
        + '</div>'  # close qcard-header
        + '</div>'  # close qcard-wrap
    )

def signal_history_chart(ticker: str, current_score: float) -> str:
    """
    Fetch last 30 days of adj_composite from signal_log and render an SVG sparkline.
    Returns HTML string — empty string if fewer than 3 data points.
    """
    try:
        from data_refresh import _get_supabase
        sb = _get_supabase()
        if not sb:
            return ""
        resp = sb.table("signal_log") \
            .select("signal_date,adj_composite,composite") \
            .eq("ticker", ticker) \
            .order("signal_date", desc=False) \
            .limit(60) \
            .execute()
        rows = resp.data or []
        # Deduplicate by date, keep most recent per date
        seen = {}
        for row in rows:
            d = str(row.get("signal_date", ""))[:10]
            if d:
                seen[d] = float(row.get("adj_composite") or row.get("composite") or 50)
        # Sort by date, take last 30
        dates  = sorted(seen.keys())[-30:]
        scores = [seen[d] for d in dates]

        if len(scores) < 3:
            return (
                '<div style="font-size:13px;color:#94a3b8;padding:8px 0 4px;">'
                '⏳ Building score history — check back after a few nightly refreshes.</div>'
            )

        # Add current score as the latest point if newer than last stored
        scores.append(current_score)

        # SVG sparkline
        W, H   = 100, 32   # viewBox units — scales with CSS
        lo, hi = min(scores), max(scores)
        rng    = hi - lo if hi != lo else 10
        pad    = 2

        def _x(i):
            return pad + (i / (len(scores) - 1)) * (W - pad * 2)

        def _y(v):
            return H - pad - ((v - lo) / rng) * (H - pad * 2)

        pts = " ".join(f"{_x(i):.1f},{_y(v):.1f}" for i, v in enumerate(scores))

        # Trend
        first5_avg = sum(scores[:5]) / 5
        last5_avg  = sum(scores[-5:]) / 5
        trend_delta = last5_avg - first5_avg
        if trend_delta >= 3:
            trend_label, trend_color, trend_arrow = "Improving", "#34d399", "↑"
        elif trend_delta <= -3:
            trend_label, trend_color, trend_arrow = "Deteriorating", "#f87171", "↓"
        else:
            trend_label, trend_color, trend_arrow = "Stable", "#fbbf24", "→"

        # Conviction zone bands (background)
        hi_y  = _y(60)
        lo_y  = _y(45)
        mid_y = _y(45)

        n_days = len(dates)
        days_label = f"{n_days}d history" if n_days >= 7 else f"{n_days} data points"

        svg = (
            f'<svg viewBox="0 0 {W} {H}" preserveAspectRatio="none" '
            f'xmlns="http://www.w3.org/2000/svg" style="width:100%;height:32px;">'
            # High conviction zone (>=60) — subtle green tint
            f'<rect x="0" y="0" width="{W}" height="{hi_y:.1f}" '
            f'fill="rgba(52,211,153,.04)"/>'
            # Low conviction zone (<45) — subtle red tint
            f'<rect x="0" y="{mid_y:.1f}" width="{W}" height="{H - mid_y:.1f}" '
            f'fill="rgba(248,113,113,.04)"/>'
            # 60 threshold line
            f'<line x1="0" y1="{hi_y:.1f}" x2="{W}" y2="{hi_y:.1f}" '
            f'stroke="rgba(52,211,153,.2)" stroke-width="0.5" stroke-dasharray="2,2"/>'
            # 45 threshold line
            f'<line x1="0" y1="{lo_y:.1f}" x2="{W}" y2="{lo_y:.1f}" '
            f'stroke="rgba(248,113,113,.2)" stroke-width="0.5" stroke-dasharray="2,2"/>'
            # Sparkline
            f'<polyline points="{pts}" fill="none" stroke="{trend_color}" '
            f'stroke-width="1.5" stroke-linejoin="round" stroke-linecap="round"/>'
            # Current score dot
            f'<circle cx="{_x(len(scores)-1):.1f}" cy="{_y(scores[-1]):.1f}" '
            f'r="2" fill="{trend_color}"/>'
            f'</svg>'
        )

        return (
            f'<div style="margin-top:10px;padding:10px 12px;'
            f'background:rgba(255,255,255,.02);border-radius:6px;'
            f'border:1px solid rgba(255,255,255,.06);">'
            f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:6px;">'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;letter-spacing:.08em;">CONVICTION TREND</div>'
            f'<div style="display:flex;align-items:center;gap:8px;">'
            f'<span style="font-size:13px;color:{trend_color};font-weight:700;">{trend_arrow} {trend_label}</span>'
            f'<span style="font-size:13px;color:#94a3b8;">{days_label}</span>'
            f'</div></div>'
            + svg +
            f'</div>'
        )
    except Exception:
        return ""


def resolve_ticker(query: str) -> tuple[str, str]:
    """
    Given a ticker or company name, return (ticker, display_name).
    Tries: exact ticker match → name substring match in KNOWN → yfinance search.
    """
    q = query.strip().upper()
    if not q:
        return "", ""

    # Direct ticker match in universe
    if q in SECTORS:
        ci = get_company_info(q)
        return q, ci.get("name", q)

    # Search by name in KNOWN dict — key match first, then name substring
    q_lower = query.strip().lower()
    KNOWN_INLINE = {
        "AAPL":"Apple Inc.","MSFT":"Microsoft Corporation","NVDA":"NVIDIA Corporation",
        "GOOGL":"Alphabet Inc.","GOOG":"Alphabet Inc.","META":"Meta Platforms Inc.",
        "AMZN":"Amazon.com Inc.","TSLA":"Tesla Inc.","NFLX":"Netflix Inc.",
        "AMD":"Advanced Micro Devices","INTC":"Intel Corporation","CSCO":"Cisco Systems",
        "ORCL":"Oracle Corporation","CRM":"Salesforce Inc.","ADBE":"Adobe Inc.",
        "INTU":"Intuit Inc.","QCOM":"Qualcomm Inc.","TXN":"Texas Instruments",
        "AVGO":"Broadcom Inc.","MU":"Micron Technology","AMAT":"Applied Materials",
        "JPM":"JPMorgan Chase & Co.","BAC":"Bank of America","GS":"Goldman Sachs",
        "MS":"Morgan Stanley","V":"Visa Inc.","MA":"Mastercard Inc.",
        "BLK":"BlackRock Inc.","AXP":"American Express","PYPL":"PayPal Holdings",
        "UNH":"UnitedHealth Group","LLY":"Eli Lilly and Company","JNJ":"Johnson & Johnson",
        "ABBV":"AbbVie Inc.","MRK":"Merck & Co.","PFE":"Pfizer Inc.",
        "TMO":"Thermo Fisher Scientific","AMGN":"Amgen Inc.","GILD":"Gilead Sciences",
        "WMT":"Walmart Inc.","COST":"Costco Wholesale","PG":"Procter & Gamble",
        "KO":"The Coca-Cola Company","PEP":"PepsiCo Inc.","HD":"Home Depot",
        "MCD":"McDonald's Corporation","NKE":"Nike Inc.","SBUX":"Starbucks Corporation",
        "XOM":"Exxon Mobil Corporation","CVX":"Chevron Corporation",
        "BRK":"Berkshire Hathaway","PLTR":"Palantir Technologies",
        "COIN":"Coinbase Global","HOOD":"Robinhood Markets",
        "SNOW":"Snowflake Inc.","DDOG":"Datadog Inc.","NET":"Cloudflare Inc.",
        "ZS":"Zscaler Inc.","CRWD":"CrowdStrike Holdings","PANW":"Palo Alto Networks",
        "NOW":"ServiceNow Inc.","WDAY":"Workday Inc.","TEAM":"Atlassian Corporation",
        "NVIDIA":"NVIDIA Corporation",
        "NVDA":"NVIDIA Corporation",
        "APPLE":"Apple Inc.",
        "MICROSOFT":"Microsoft Corporation",
        "AMAZON":"Amazon.com Inc.",
        "GOOGLE":"Alphabet Inc.",
        "ALPHABET":"Alphabet Inc.",
        "META":"Meta Platforms Inc.",
        "FACEBOOK":"Meta Platforms Inc.",
        "TESLA":"Tesla Inc.",
        "NETFLIX":"Netflix Inc.",
        "PALANTIR":"Palantir Technologies",
        "COINBASE":"Coinbase Global",
        "SNOWFLAKE":"Snowflake Inc.",
        "CLOUDFLARE":"Cloudflare Inc.",
        "CROWDSTRIKE":"CrowdStrike Holdings",
        "UBER":"Uber Technologies",
        "AIRBNB":"Airbnb Inc.",
        "SPOTIFY":"Spotify Technology",
    }
    for ticker, name in KNOWN_INLINE.items():
        # Exact key match (e.g. "nvidia" → NVDA) or name substring
        if q_lower == ticker.lower() or q_lower in name.lower() or q_lower == ticker.lower():
            return ticker, name

    # Try yfinance search as last resort
    try:
        import yfinance as yf
        results = yf.Search(query, max_results=1).quotes
        if results:
            tk = results[0].get("symbol", q)
            nm = results[0].get("longname") or results[0].get("shortname") or tk
            return tk.upper(), nm
    except Exception:
        pass

    # Fall back to treating input as ticker
    return q, q


# ══════════════════════════════════════════════════════════════════════════════
# WATCHLIST HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _default_watchlist_id(user_id: str):
    """Resolve (and auto-create) the user's default watchlist id."""
    def _resolve():
        try:
            from db import get_watchlists
            lists = get_watchlists(user_id)
            if not lists:
                return None
            # Prefer the flagged default, else the first list
            for w in lists:
                if w.get("is_default"):
                    return w["id"]
            return lists[0]["id"]
        except Exception:
            return None
    return _run_once(f"defwl:{user_id}", _resolve)


def get_watchlist(user_id: str) -> list:
    """Back-compat shim: return items from the user's DEFAULT named list,
    shaped like the old flat watchlist rows (ticker, price_at_add, added_at)."""
    def _fetch():
        try:
            from db import get_watchlist_items
            lid = _default_watchlist_id(user_id)
            if not lid:
                return []
            return get_watchlist_items(user_id, lid)
        except Exception:
            return []
    return _run_once(f"wl:{user_id}", _fetch)


def is_valid_universe_ticker(tk) -> bool:
    """True iff `tk` is a real ticker in the QNTM model universe. Use this at
    every user-driven add path (watchlist, holdings, simulator, custom lists)
    before persisting — otherwise arbitrary user-supplied strings get stored
    against rows that later fail price lookups, signal joins, and exports.
    Fails closed: if the universe import is unavailable for any reason, we
    refuse the add rather than letting bad data through."""
    if not tk or not isinstance(tk, str):
        return False
    tk_norm = tk.strip().upper()
    if not tk_norm:
        return False
    # Tickers are alphanumerics + dot/dash (e.g. BRK.B, BF-B). Anything else
    # is definitely garbage; reject before even importing the universe.
    if not all(c.isalnum() or c in ".-" for c in tk_norm):
        return False
    try:
        from universe_data import SECTORS as _UNI_SECTORS
        return tk_norm in _UNI_SECTORS
    except Exception:
        return False


def _market_phase():
    """Current US-market phase by the ET clock: 'pre' (04:00-09:30),
    'regular' (09:30-16:00), 'post' (16:00-20:00), or 'closed' (overnight /
    weekend). Used to decide which extended-hours figure (if any) to surface."""
    try:
        from datetime import datetime as _dt, time as _t
        from zoneinfo import ZoneInfo
        now = _dt.now(ZoneInfo("America/New_York"))
        if now.weekday() >= 5:
            return "closed"
        tm = now.time()
        if tm < _t(4, 0):   return "closed"
        if tm < _t(9, 30):  return "pre"
        if tm < _t(16, 0):  return "regular"
        if tm < _t(20, 0):  return "post"
        return "closed"
    except Exception:
        return "regular"


def _fetch_extended_hours_map(tickers: list, cache_key: str = "_xh_cache") -> dict:
    """Return {ticker: {pre_pct, pre_chg, post_pct, post_chg}} for the current
    pre-market / after-hours session, vs the relevant regular-session close:
      pre_*  = last pre-market print (04:00-09:30 ET) vs the PRIOR regular close
      post_* = last after-hours print (16:00-20:00 ET) vs TODAY's regular close
    A session's keys are present only when it has prints. Pulls one prepost
    1-min download (yfinance's extended-hours feed is flaky, so this degrades to
    {} on any failure) and caches by ticker set."""
    if not tickers:
        return {}
    cache = st.session_state.setdefault(cache_key, {})
    key = ",".join(sorted(set(tickers)))
    if key in cache:
        return cache[key]
    out = {}
    try:
        import yfinance as yf
        import pandas as pd
        import datetime as _d
        from zoneinfo import ZoneInfo
        ET = ZoneInfo("America/New_York")
        df = yf.download(_strip_delisted(tickers), period="2d", interval="5m",
                         prepost=True, progress=False, threads=True)
        if df is None or df.empty or "Close" not in df:
            cache[key] = out
            return out
        close = df["Close"]
        idx = close.index
        try:
            idx = idx.tz_localize("UTC").tz_convert(ET) if idx.tz is None else idx.tz_convert(ET)
        except Exception:
            pass

        def _sess(ts):
            t = ts.time()
            if t < _d.time(9, 30): return "pre"
            if t < _d.time(16, 0): return "reg"
            return "post"
        meta = [(ts.date(), _sess(ts)) for ts in idx]
        day_list = sorted({d for d, _ in meta})
        if not day_list:
            cache[key] = out
            return out
        today = day_list[-1]
        prev  = day_list[-2] if len(day_list) >= 2 else None
        multi = hasattr(close, "columns")
        for tk in set(tickers):
            col = close[tk] if (multi and tk in getattr(close, "columns", [])) else (None if multi else close)
            if col is None:
                continue
            prior_close = today_close = pre_last = post_last = None
            for i in range(len(col)):
                v = col.iloc[i]
                if pd.isna(v):
                    continue
                v = float(v); d, ss = meta[i]
                if   d == today and ss == "reg":  today_close = v
                elif d == today and ss == "pre":  pre_last = v
                elif d == today and ss == "post": post_last = v
                elif prev and d == prev and ss == "reg": prior_close = v
            e = {}
            if pre_last is not None and prior_close:
                e["pre_chg"] = pre_last - prior_close
                e["pre_pct"] = (pre_last / prior_close - 1) * 100
            if post_last is not None and today_close:
                e["post_chg"] = post_last - today_close
                e["post_pct"] = (post_last / today_close - 1) * 100
            if e:
                out[tk] = e
    except Exception:
        pass
    cache[key] = out
    return out


_LIVE_QUOTE_CACHE = {}   # module-level, shared across sessions: {key: (quotes, ts)}


def _live_quotes(tickers):
    """One bounded live current-price fetch (today's last trade) for `tickers`,
    shared across sessions via a short module-level TTL cache. Returns
    {ticker: price}. Lets the page stay live during market hours with a single
    small download instead of re-pulling full history. {} on failure, so the
    caller falls back to the full live path."""
    import time as _t2
    want = sorted({t for t in (tickers or []) if t})
    if not want:
        return {}
    key = ",".join(want)
    ent = _LIVE_QUOTE_CACHE.get(key)
    if ent and (_t2.time() - ent[1] < 45):
        return ent[0]
    out = {}
    try:
        import yfinance as yf
        df = yf.download(_strip_delisted(want), period="1d",
                         auto_adjust=True, progress=False, threads=True)
        if df is not None and not df.empty and "Close" in df:
            close = df["Close"]
            if hasattr(close, "columns"):
                for tk in want:
                    if tk in close.columns:
                        v = close[tk].dropna()
                        if not v.empty:
                            out[tk] = float(v.iloc[-1])
            else:
                v = close.dropna()
                if not v.empty and len(want) == 1:
                    out[want[0]] = float(v.iloc[-1])
    except Exception:
        out = {}
    if out:
        _LIVE_QUOTE_CACHE[key] = (out, _t2.time())
    return out


def _stored_day_change_map(tickers):
    """Day-change map from stored data only — signal_log prices for stocks,
    benchmark_price for SPY. Returns {ticker: {price, prev_close, chg_pct,
    chg_dollar, settled, market_closed, last_bar_date}}. Returns {} unless EVERY
    requested ticker has >=2 stored sessions, so the caller falls back cleanly to
    yfinance on partial coverage rather than showing a mixed/stale set."""
    want = {t for t in (tickers or []) if t}
    if not want:
        return {}
    try:
        from data_refresh import _get_supabase, _fetch_all_rows
        import datetime as _dt
        sb2 = _get_supabase()
        if not sb2:
            return {}
        try:
            from zoneinfo import ZoneInfo
            today_str = _dt.datetime.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
        except Exception:
            today_str = _dt.date.today().isoformat()
        start = (_dt.date.today() - _dt.timedelta(days=12)).isoformat()
        prices = {}
        non_spy = [t for t in want if t != "SPY"]
        if non_spy:
            srows = _fetch_all_rows(lambda: sb2.table("signal_log")
                                    .select("ticker,signal_date,price")
                                    .in_("ticker", non_spy).gte("signal_date", start))
            for r in (srows or []):
                tk = r.get("ticker"); pv = r.get("price")
                d = str(r.get("signal_date") or "")[:10]
                if tk and pv is not None and len(d) == 10:
                    prices.setdefault(tk, {})[d] = float(pv)
        if "SPY" in want:
            brows = _fetch_all_rows(lambda: sb2.table("benchmark_price")
                                    .select("d,close").gte("d", start).order("d", desc=False))
            sd = {str(r["d"])[:10]: float(r["close"]) for r in (brows or [])
                  if r.get("close") is not None}
            if sd:
                prices["SPY"] = sd
        # Common axis. d_prev is the latest SETTLED session (strictly before
        # today); _pon forward-fills a ticker that's missing an exact date.
        if any(not prices.get(tk) for tk in want):
            return {}
        union = sorted({d for tk in want for d in prices[tk].keys()})

        def _pon(m, d):
            if d in m:
                return m[d]
            earlier = [x for x in m if x <= d]
            return m[max(earlier)] if earlier else None

        # Stored-only: current = today's stored row (kept fresh by the intraday
        # price cron), prev = the most recent settled session before it. No live
        # quote fetch on the request path, so every read is a fast Supabase read;
        # price lag is bounded by the cron cadence. Anchored to one common date
        # pair so model and SPY align.
        d_latest = min(max(prices[tk].keys()) for tk in want)
        priors = [d for d in union if d < d_latest]
        if not priors:
            return {}
        d_prev = priors[-1]
        out = {}
        settled = (d_latest != today_str)
        for tk in want:
            m = prices[tk]
            cur, prev = _pon(m, d_latest), _pon(m, d_prev)
            if cur is None or prev is None:
                return {}
            out[tk] = {
                "price": cur, "prev_close": prev,
                "chg_pct": ((cur - prev) / prev * 100) if prev else None,
                "chg_dollar": (cur - prev),
                "settled": settled,
                "market_closed": settled,
                "last_bar_date": d_latest,
            }
        return out
    except Exception:
        return {}


def _stored_spy_hist(start):
    """SPY daily-close history from benchmark_price as a DataFrame with a 'Close'
    column (DatetimeIndex), matching the shape the caller expects from a yfinance
    download. None when benchmark_price has no rows (caller falls back to live)."""
    try:
        import pandas as pd
        from data_refresh import _get_supabase, _fetch_all_rows
        sb2 = _get_supabase()
        if not sb2:
            return None
        rows = _fetch_all_rows(lambda: sb2.table("benchmark_price").select("d,close")
                               .gte("d", start).order("d", desc=False))
        rows = [r for r in (rows or []) if r.get("close") is not None]
        if not rows:
            return None
        idx = pd.to_datetime([str(r["d"])[:10] for r in rows])
        return pd.DataFrame({"Close": [float(r["close"]) for r in rows]}, index=idx)
    except Exception:
        return None


def _fetch_day_change_map(tickers: list, cache_key: str = "_dc_cache",
                          include_extended: bool = True) -> dict:
    """Return {ticker: {chg_pct, chg_dollar, price, prev_close, settled, last_bar_date}}
    for `tickers`. Cached in st.session_state[cache_key] by sorted ticker set so
    a navigation away and back doesn't re-hit yfinance.

    The change is ALWAYS the most recent session's move vs the prior session's
    close — live and running while the market is open, frozen at the closing
    value once it closes, and the last completed session on weekends/holidays.
    `settled` is True when the latest bar is a finished session (not live today),
    so callers can label it "at close" vs intraday. Returns {} on any failure."""
    if not tickers:
        return {}
    cache = st.session_state.setdefault(cache_key, {})
    key = ",".join(sorted(set(tickers))) + ("" if include_extended else "|noext")
    if key in cache:
        return cache[key]
    # Stored-first: build from signal_log + benchmark_price with no live pull.
    # Falls through to the yfinance path below only on partial stored coverage.
    _stored = _stored_day_change_map(tickers)
    if _stored:
        if include_extended and _market_phase() in ("pre", "post", "closed"):
            try:
                for _tk, _e in _fetch_extended_hours_map(tickers).items():
                    if _tk in _stored:
                        _stored[_tk].update(_e)
            except Exception:
                pass
        cache[key] = _stored
        return _stored
    out = {}
    try:
        import yfinance as yf
        from datetime import datetime as _dt
        try:
            from zoneinfo import ZoneInfo
            today_str = _dt.now(ZoneInfo("America/New_York")).strftime("%Y-%m-%d")
        except Exception:
            today_str = _dt.now().strftime("%Y-%m-%d")
        hist = yf.download(_strip_delisted(tickers), period="5d", auto_adjust=True,
                           progress=False, threads=True)
        if hist.empty:
            cache[key] = out
            return out
        try:
            last_bar_date = str(hist.index[-1])[:10]
        except Exception:
            last_bar_date = ""
        # The latest bar is "live" only when it's today's session; otherwise the
        # number we show is a settled close (today after close, or last session).
        settled = (last_bar_date != today_str)
        close = hist["Close"]

        def _entry(cur, prev):
            return {
                "price": cur, "prev_close": prev,
                "chg_pct":    ((cur - prev) / prev * 100) if prev else None,
                "chg_dollar": (cur - prev),
                "settled": settled,
                "market_closed": settled,  # retained for back-compat
                "last_bar_date": last_bar_date,
            }

        if hasattr(close, "columns"):
            for tk in set(tickers):
                if tk in close.columns:
                    vals = close[tk].dropna()
                    if len(vals) >= 2:
                        out[tk] = _entry(float(vals.iloc[-1]), float(vals.iloc[-2]))
        else:
            vals = close.dropna()
            if len(vals) >= 2 and len(set(tickers)) == 1:
                tk = list(set(tickers))[0]
                out[tk] = _entry(float(vals.iloc[-1]), float(vals.iloc[-2]))
    except Exception:
        pass
    # Extended-hours overlay — only outside regular trading, since during the
    # session the move is already reflected in chg_pct. Merged onto existing
    # entries so callers can read pre_pct/pre_chg/post_pct/post_chg.
    if include_extended and _market_phase() in ("pre", "post", "closed"):
        try:
            for _tk, _e in _fetch_extended_hours_map(tickers).items():
                if _tk in out:
                    out[_tk].update(_e)
        except Exception:
            pass
    cache[key] = out
    return out


def _build_summary_meta_html(entry_date=None, day_change_entry: dict = None) -> str:
    """Compact one-line strip rendered inside the collapsed card summary, under
    the conviction label. Shows "ENT MM/DD" (where applicable) and a coloured
    "+x.xx% today" segment so a user can triage a watchlist / portfolio / model
    portfolio without expanding each card. Returns '' if no data is renderable
    (so callers can safely set _summary_meta_html unconditionally)."""
    bits = []
    # Entry date — accept date / datetime / iso string; render MM/DD only
    if entry_date:
        try:
            ed = str(entry_date)[:10]
            yy, mm, dd = ed.split("-")
            bits.append(f'<span style="color:#9fabc0;">ENT&nbsp;{mm}/{dd}</span>')
        except Exception:
            pass
    # Day change — most recent session's move vs the prior close. Shown live
    # while the market is open ("today") and frozen at the close ("at close")
    # afterward / on weekends, rather than blanking out.
    dc = day_change_entry or {}
    if dc.get("chg_pct") is not None:
        pct = float(dc["chg_pct"])
        col = "#34d399" if pct > 0 else ("#f87171" if pct < 0 else "#b3bed0")
        sign = "+" if pct >= 0 else ""
        lbl = "at&nbsp;close" if dc.get("settled") else "today"
        bits.append(f'<span style="color:{col};">{sign}{pct:.2f}%&nbsp;{lbl}</span>')
    # Extended-hours figure — live session only: pre-market before the open,
    # after-hours after the close (nothing extra during regular trading).
    _phase = _market_phase()
    _xk = "pre" if _phase == "pre" else ("post" if _phase in ("post", "closed") else None)
    if _xk and dc.get(f"{_xk}_pct") is not None:
        _xp = float(dc[f"{_xk}_pct"]); _xd = dc.get(f"{_xk}_chg")
        _xc = "#34d399" if _xp > 0 else ("#f87171" if _xp < 0 else "#b3bed0")
        _xlbl = "pre" if _xk == "pre" else "AH"
        _xdollar = (
            f"&nbsp;({'+' if _xd >= 0 else '-'}${abs(_xd):,.2f})"
            if _xd is not None else ""
        )
        bits.append(
            f'<span style="color:{_xc};">{"+" if _xp >= 0 else ""}{_xp:.2f}%'
            f'&nbsp;{_xlbl}{_xdollar}</span>'
        )
    if not bits:
        return ""
    sep = '<span style="color:#1e293b;">·</span>'
    return (
        '<div style="display:flex;align-items:center;gap:6px;'
        'font-family:DM Mono,monospace;font-size:13px;letter-spacing:.04em;'
        'margin-top:3px;">'
        + sep.join(bits)
        + '</div>'
    )


def add_to_watchlist(user_id: str, ticker: str, price_at_add: float = None) -> bool:
    """Add ticker to the user's DEFAULT named list. Returns True on success.
    Rejects tickers outside the model universe so the watchlist never holds
    rows that can't be priced or scored."""
    if not is_valid_universe_ticker(ticker):
        return False
    try:
        from db import add_watchlist_item
        lid = _default_watchlist_id(user_id)
        if not lid:
            return False
        return add_watchlist_item(user_id, lid, ticker.strip().upper(), price_at_add)
    except Exception:
        return False


@st.cache_data(ttl=300, show_spinner=False)
def _preview_score(ticker: str):
    """Fetch the latest signal_log row for a single ticker so the watchlist add
    box can show a conviction preview before the user commits. Cached briefly so
    repeated reruns (every keystroke / Enter) don't re-hit Supabase. Returns the
    row dict or None."""
    try:
        from data_refresh import _get_supabase
        sb = _get_supabase()
        if not sb:
            return None
        resp = (sb.table("signal_log")
                .select("ticker,adj_composite,composite,price,signal,"
                        "momentum,quality,volume,value,sentiment,signal_date,"
                        "val_low,val_high,value_position,val_basis")
                .eq("ticker", ticker)
                .order("signal_date", desc=True)
                .limit(1)
                .execute())
        rows = resp.data or []
        return rows[0] if rows else None
    except Exception:
        return None


# ── Standard stock search: shared across every search box on the platform ──────
# There is no universe-wide company-name table (names only exist for well-known
# tickers), so the standard is: partial TICKER match across the whole universe +
# company-NAME match across this curated map. Used by Screener, Watchlist,
# Portfolio, and Simulator so search behaves identically everywhere.
_SEARCH_NAMES = {
    "AAPL":"Apple","MSFT":"Microsoft","NVDA":"NVIDIA","GOOGL":"Alphabet (Google)",
    "GOOG":"Alphabet (Google)","META":"Meta Platforms","AMZN":"Amazon","TSLA":"Tesla",
    "NFLX":"Netflix","AMD":"Advanced Micro Devices","INTC":"Intel","CSCO":"Cisco",
    "ORCL":"Oracle","CRM":"Salesforce","ADBE":"Adobe","INTU":"Intuit","QCOM":"Qualcomm",
    "TXN":"Texas Instruments","AVGO":"Broadcom","MU":"Micron","AMAT":"Applied Materials",
    "JPM":"JPMorgan Chase","BAC":"Bank of America","GS":"Goldman Sachs","MS":"Morgan Stanley",
    "WFC":"Wells Fargo","C":"Citigroup","V":"Visa","MA":"Mastercard","BLK":"BlackRock",
    "AXP":"American Express","PYPL":"PayPal","SQ":"Block (Square)","COF":"Capital One",
    "UNH":"UnitedHealth","LLY":"Eli Lilly","JNJ":"Johnson & Johnson","ABBV":"AbbVie",
    "MRK":"Merck","PFE":"Pfizer","TMO":"Thermo Fisher","AMGN":"Amgen","GILD":"Gilead Sciences",
    "BMY":"Bristol Myers Squibb","CVS":"CVS Health","WMT":"Walmart","COST":"Costco",
    "PG":"Procter & Gamble","KO":"Coca-Cola","PEP":"PepsiCo","HD":"Home Depot",
    "LOW":"Lowe's","TGT":"Target","MCD":"McDonald's","NKE":"Nike","SBUX":"Starbucks",
    "DIS":"Disney","XOM":"Exxon Mobil","CVX":"Chevron","COP":"ConocoPhillips","OXY":"Occidental",
    "SLB":"Schlumberger","BRK.B":"Berkshire Hathaway","BRK":"Berkshire Hathaway",
    "PLTR":"Palantir","COIN":"Coinbase","HOOD":"Robinhood","SNOW":"Snowflake","DDOG":"Datadog",
    "NET":"Cloudflare","ZS":"Zscaler","CRWD":"CrowdStrike","PANW":"Palo Alto Networks",
    "FTNT":"Fortinet","NOW":"ServiceNow","WDAY":"Workday","TEAM":"Atlassian","SHOP":"Shopify",
    "UBER":"Uber","LYFT":"Lyft","ABNB":"Airbnb","DASH":"DoorDash","SPOT":"Spotify",
    "BKNG":"Booking","MAR":"Marriott","BA":"Boeing","CAT":"Caterpillar","DE":"Deere",
    "GE":"GE Aerospace","HON":"Honeywell","LMT":"Lockheed Martin","RTX":"RTX (Raytheon)",
    "NOC":"Northrop Grumman","GD":"General Dynamics","F":"Ford","GM":"General Motors",
    "T":"AT&T","VZ":"Verizon","TMUS":"T-Mobile","CMCSA":"Comcast","DINO":"HF Sinclair",
    "KOS":"Kosmos Energy","APA":"APA Corp","VLO":"Valero Energy","MPC":"Marathon Petroleum",
    "CHRD":"Chord Energy","CRGY":"Crescent Energy","SM":"SM Energy","PSX":"Phillips 66",
    "ET":"Energy Transfer","KMI":"Kinder Morgan","WMB":"Williams","MMM":"3M","IBM":"IBM",
    "QCOM ":"Qualcomm","ARM":"Arm Holdings","SMCI":"Super Micro","MRVL":"Marvell",
    "DELL":"Dell","HPQ":"HP","WBD":"Warner Bros Discovery","PARA":"Paramount",
}


def _stock_suggestions(query, limit: int = 6, exclude=None):
    """Standard matcher: returns up to `limit` (ticker, name) tuples for a partial
    ticker OR company-name query. Prefix matches (ticker or name) rank first, then
    substring matches. `exclude` hides tickers already added to the current list."""
    q = (query or "").strip().lower()
    if not q:
        return []
    excl = {t.upper() for t in (exclude or [])}
    starts, contains = [], []
    for tk in SECTORS.keys():
        if tk in excl:
            continue
        tl  = tk.lower()
        nm  = _SEARCH_NAMES.get(tk, "")
        nml = nm.lower()
        if tl.startswith(q) or (nml and nml.startswith(q)):
            starts.append((tk, nm))
        elif (q in tl) or (nml and q in nml):
            contains.append((tk, nm))
        if len(starts) >= limit:
            break
    out = starts + [c for c in contains if c not in starts]
    return out[:limit]


def _render_suggestions(query, key_prefix: str, pick_state_key: str, exclude=None,
                        limit: int = 6):
    """Standard suggestion dropdown. Renders clickable ticker+name rows; on click
    sets st.session_state[pick_state_key] = ticker and reruns. Returns the list of
    suggestions shown (so callers can decide what else to render)."""
    sugg = _stock_suggestions(query, limit=limit, exclude=exclude)
    if not sugg:
        if (query or "").strip():
            st.markdown('<div style="font-size:13px;color:#94a3b8;padding:4px 2px;">'
                        'No matches — try the ticker symbol.</div>', unsafe_allow_html=True)
        return sugg
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;'
                'letter-spacing:.1em;padding:8px 2px 2px;">SUGGESTIONS</div>',
                unsafe_allow_html=True)
    for _tk, _nm in sugg:
        _label = f"{_tk}   ·   {_nm}" if _nm else _tk
        if st.button(_label, key=f"{key_prefix}_{_tk}", use_container_width=True):
            st.session_state[pick_state_key] = _tk
            st.rerun()
    return sugg


def _render_stock_preview(ticker: str) -> bool:
    """Render the standard compact conviction preview card (label + score + price +
    5 pillar bars) for a ticker. Returns True if a score was found and rendered,
    False otherwise (caller can show a fallback)."""
    _pv = _preview_score(ticker)
    if not _pv:
        st.caption(f"No score on file for {ticker} yet — scored on the nightly refresh.")
        return False
    _adj = float(_pv.get("adj_composite") or _pv.get("composite") or 0)
    if   _adj >= 60: _lbl, _col = "High Conviction",     "#34d399"
    elif _adj >= 45: _lbl, _col = "Moderate Conviction", "#fbbf24"
    else:            _lbl, _col = "Low Conviction",       "#f87171"
    _px = _pv.get("price")
    _px_str = f"${_px:,.2f}" if _px else "—"

    def _pv_bar(v):
        v = float(v or 0)
        _c = "#34d399" if v >= 60 else ("#f59e0b" if v >= 45 else "#f87171")
        return ('<div style="height:4px;border-radius:2px;background:rgba(255,255,255,.08);">'
                f'<div style="width:{max(4,int(v))}%;height:100%;background:{_c};'
                'border-radius:2px;"></div></div>')

    _pillars = [("MOM", _pv.get("momentum")), ("QUAL", _pv.get("quality")),
                ("VOL", _pv.get("volume")), ("VAL", _pv.get("value")),
                ("SENT", _pv.get("sentiment"))]
    _pcols = "".join(
        f'<div style="text-align:center;">'
        f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#9fabc0;margin-bottom:3px;">{_nm}</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#cbd5e1;margin-bottom:3px;">{float(_v or 0):.0f}</div>'
        f'{_pv_bar(_v)}</div>'
        for _nm, _v in _pillars
    )
    _nm_disp = _SEARCH_NAMES.get(ticker, "")
    _nm_html = (f'<span style="font-size:13px;color:#8896ac;margin-left:8px;">{_nm_disp}</span>'
                if _nm_disp else "")
    st.markdown(
        f'<div style="background:rgba(255,255,255,.03);border:1px solid {_col}33;'
        f'border-left:3px solid {_col};border-radius:10px;padding:14px 16px;margin-top:8px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:10px;">'
        f'<div><span style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#e2e8f0;">{ticker}</span>{_nm_html}'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:{_col};margin-top:2px;">{_lbl} · {_adj:.0f}</div></div>'
        f'<div style="font-family:DM Mono,monospace;font-size:15px;color:#cbd5e1;">{_px_str}</div></div>'
        f'<div style="display:grid;grid-template-columns:repeat(5,1fr);gap:8px;">{_pcols}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    return True


def _render_rel_spy_chart(ticker: str, days: int = 20):
    """Render a small sparkline of the stock's % return vs SPY over the last
    `days` trading sessions, rebased to 0% at the window start (gold = stock,
    slate = SPY). Reuses the same vs-SPY drawing as the watchlist/portfolio
    cards. Silently renders nothing if data is unavailable."""
    try:
        # Fetch enough calendar days to comfortably cover `days` sessions.
        start = _trail_start(int(days * 1.6) + 7)
        pm, sm = _mini_price_data((ticker,), start)
        series = (pm or {}).get(ticker) or {}
        if len(series) < 2:
            return
        sdates = sorted(series.keys())[-days:]          # last N trading sessions
        if len(sdates) < 2:
            return
        stock_pairs = [(d, series[d]) for d in sdates]
        spy_pairs   = [(d, sm[d]) for d in sdates if d in sm]
        svg = _mini_vs_spy_svg(stock_pairs, spy_pairs)
        if not svg:
            return
        s_ret = (stock_pairs[-1][1] / stock_pairs[0][1] - 1) * 100 if stock_pairs[0][1] else 0
        k_ret = (spy_pairs[-1][1] / spy_pairs[0][1] - 1) * 100 if spy_pairs and spy_pairs[0][1] else 0
        ss = "+" if s_ret >= 0 else ""; ks = "+" if k_ret >= 0 else ""
        st.markdown(
            '<div style="padding:10px 18px 14px;border-top:1px solid rgba(255,255,255,.05);">'
            '<div style="display:flex;gap:14px;align-items:center;font-family:DM Mono,monospace;'
            'font-size:11px;margin-bottom:6px;">'
            f'<span style="color:#d4a843;">\u2014 {ticker} {ss}{s_ret:.1f}%</span>'
            f'<span style="color:#7c8aa0;">\u2014 SPY {ks}{k_ret:.1f}%</span>'
            f'<span style="color:#8896ac;margin-left:auto;">{len(sdates)}-day vs SPY</span></div>'
            f'{svg}</div>',
            unsafe_allow_html=True,
        )
    except Exception:
        return


def _render_stock_result(ticker: str, nav: str = "screener", wl_actions: bool = True):
    """Standard rich search-result card used everywhere a stock is searched:
    full factor panel (detail open) with price + fundamentals, the signal-history
    mini chart, and the Add/Remove-watchlist action row. Scores the ticker live so
    the card matches the Screener exactly. Returns {'ticker','price'} on success,
    else None (and renders a 'not found' note)."""
    from model_engine import score_stock, fetch_price_data
    resolved_tk, resolved_name = resolve_ticker(ticker)
    if not resolved_tk:
        return None
    _ok = None
    _not_found = (
        f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);'
        f'border-radius:8px;padding:20px 24px;">'
        f'<div style="font-family:Syne,sans-serif;font-size:15px;font-weight:700;color:#b3bed0;margin-bottom:6px;">'
        f'"{ticker}" not found</div>'
        f'<div style="font-size:13px;color:#8896ac;line-height:1.6;">'
        f'Could not retrieve data. Try the exact ticker symbol — e.g. '
        f'<strong style="color:#b3bed0;">AAPL</strong>, <strong style="color:#b3bed0;">NVDA</strong>.</div>'
        f'</div>')
    with st.spinner(f"Scoring {resolved_tk}..."):
        try:
            price_data = fetch_price_data([resolved_tk], period="1y")
            hist = price_data.get(resolved_tk, [])
            if not hist or len(hist) < 10:
                st.markdown(_not_found, unsafe_allow_html=True)
            else:
                scored = score_stock(resolved_tk, hist)
                scored["sector"] = SECTORS.get(resolved_tk, "Unknown")
                macro = st.session_state.get("macro_data") or _live_macro()
                sr = apply_macro_overlay([scored], macro)[0]
                hydrate_card_rows([sr])   # standard card fields: sector, cap, band
                if sr.get("promoted"):
                    regime = macro.get("regime", "NEUTRAL")
                    eff = 62 if regime in ("RISK_OFF", "HIGH VOLATILITY") else 60
                    adj = float(sr.get("adj_composite", sr.get("composite", 50)))
                    sr["adj_action"] = "BUY" if adj >= eff else ("SELL" if adj < EXIT_THRESHOLD else "HOLD")
                    sr["promoted"] = False
                # pct_rank intentionally not set here — factor_panel_html ranks
                # this card against the full universe via _pct_rank_of().
                ci = get_company_info(resolved_tk)
                _html = factor_panel_html(sr, False, company_info=ci, suppress_wl_btn=True,
                                          extra_detail=_whats_changed_html(resolved_tk, sr.get("signal_date", "")))
                _html = _html.replace('class="qcard-detail" style="display:none;',
                                      'class="qcard-detail" style="display:block;')
                st.markdown(_html, unsafe_allow_html=True)
                _render_rel_spy_chart(resolved_tk, 20)
                _ok = {"ticker": resolved_tk, "price": sr.get("price")}
                if resolved_tk not in SECTORS:
                    st.markdown('<div style="font-size:13px;color:#8896ac;margin-bottom:8px;">'
                                '⚠ Not in core universe — scored from live price data. '
                                'Fundamental data may be limited.</div>', unsafe_allow_html=True)
        except Exception:
            _ok = None
            st.markdown(_not_found, unsafe_allow_html=True)
    # Watchlist add/remove toggle — rendered outside the try so a button error
    # can never surface as "not found". Suppressed (wl_actions=False) on the
    # Watchlist page, which renders its own prominent Add/Remove button instead.
    if _ok and wl_actions:
        try:
            from db import get_watchlist_items as _gwi, get_watchlists as _gws
            _u = uid()
            _lists = _gws(_u)
            _did = next((l["id"] for l in _lists if l.get("is_default")),
                        _lists[0]["id"] if _lists else None)
            _wtk = {w["ticker"] for w in _gwi(_u, _did)} if _did else set()
            render_watchlist_actions([_ok["ticker"]], nav=nav, in_list=_wtk)
        except Exception:
            pass
    return _ok


def remove_from_watchlist(user_id: str, ticker: str) -> bool:
    """Remove ticker from the user's DEFAULT named list."""
    try:
        from db import remove_watchlist_item
        lid = _default_watchlist_id(user_id)
        if not lid:
            return False
        return remove_watchlist_item(user_id, lid, ticker)
    except Exception:
        return False


def cookie_banner():
    """No-op — cookie consent is now handled as a dedicated page in the router."""
    pass


def _cta_gold(label: str, href: str, full_width: bool = True) -> str:
    """Gold primary CTA — HTML link styled as gold button."""
    w = "width:100%;display:block;" if full_width else "display:inline-block;"
    return (
        f'<a href="{href}" target="_self" style="{w}text-align:center;padding:12px 20px;'
        f'background:linear-gradient(135deg,#d4a843 0%,#b8922e 50%,#d4a843 100%);'
        f'border:none;border-radius:6px;font-family:Syne,sans-serif;font-size:13px;font-weight:800;'
        f'letter-spacing:.06em;text-transform:uppercase;color:#0a0b14;text-decoration:none;'
        f'box-sizing:border-box;margin-top:4px;">{label}</a>'
    )


def _cta_ghost(label: str, href: str, full_width: bool = True) -> str:
    """Ghost secondary CTA — HTML link styled as ghost button."""
    w = "width:100%;display:block;" if full_width else "display:inline-block;"
    return (
        f'<a href="{href}" target="_self" style="{w}text-align:center;padding:12px 20px;'
        f'background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.18);'
        f'border-radius:6px;font-family:Syne,sans-serif;font-size:13px;font-weight:700;'
        f'letter-spacing:.06em;text-transform:uppercase;color:#e2e8f0;text-decoration:none;'
        f'box-sizing:border-box;margin-top:4px;">{label}</a>'
    )


def _render_checkout_button(url: str):
    """Gold 'Continue to secure checkout' button that navigates the SAME tab to
    Stripe. Uses the proven component-iframe pattern: window.open(url,'_top') is
    the one call the sandbox permits to drive the parent window (plain
    <a target=_top> and window.top.location are blocked). Renders inside a
    components.v1.html iframe."""
    _safe = (url or "").replace("\\", "\\\\").replace('"', '\\"')
    try:
        if not st.session_state.get("_checkout_evt_sent"):
            st.session_state["_checkout_evt_sent"] = True
            analytics.capture("stripe_checkout_started", user=st.session_state.get("user"))
    except Exception:
        pass
    qntm_html(
        f'''<div style="font-family:Syne,sans-serif;">
        <button id="qntm-checkout-btn" style="width:100%;box-sizing:border-box;
            padding:16px;border:none;border-radius:8px;cursor:pointer;
            background:linear-gradient(135deg,#d4a843,#b8922e);
            color:#0a0b14;font-family:Syne,sans-serif;font-weight:800;font-size:16px;">
          Continue to secure checkout →
        </button>
        <div style="text-align:center;font-size:13px;color:#9fabc0;margin-top:8px;">
          Opens Stripe's secure checkout in this tab.
        </div></div>
        <script>
          document.getElementById("qntm-checkout-btn").addEventListener("click",function(){{
            window.open("{_safe}","_top");
          }});
        </script>''',
        height=90,
    )


# ── DISCLAIMER ────────────────────────────────────────────────────────────────
DISCLAIMER = """<div style="display:flex;align-items:center;gap:8px;padding:6px 0;margin-bottom:8px;">
<span style="font-size:13px;color:#94a3b8;">ℹ</span>
<span style="font-size:13px;color:#94a3b8;">Quantitative research tool — not investment advice.
<a href="?legal=disclaimer" target="_self" style="color:#8896ac;text-decoration:underline;">Learn more</a></span>
</div>"""

# ══════════════════════════════════════════════════════════════════════════════
# LANDING PAGE
# ══════════════════════════════════════════════════════════════════════════════





# ══════════════════════════════════════════════════════════════════════════════
# LEGAL PAGES — content matches QNTM_POLICIES_FINAL.md (draft for attorney review)
# Update both files together when policies change.
# ══════════════════════════════════════════════════════════════════════════════
PRIVACY_POLICY = """
## Privacy Policy
**Effective Date: June 6, 2026** · QNTM LLC (California)

### What We Collect

**You give us:**
- Email address and password (passwords are bcrypt-hashed — we cannot see them)
- Optional 2FA secret (encrypted at rest)
- Optional full name
- Payment method (processed and stored by Stripe — we never see your card number)
- Your watchlist, portfolio, and simulator data

**We collect automatically:**
- Login times and session data
- IP address (for security only — not tracked across sites, not sold)
- Anonymous, aggregated feature usage (used to improve the product)

### What We Don't Do
- We don't sell your data
- We don't use your data for advertising
- We don't share your data except as described below
- We don't track you across other websites

### Who We Share Data With

| Provider | Purpose | Their privacy policy |
|----------|---------|---------------------|
| Stripe, Inc. | Payment processing | stripe.com/privacy |
| Supabase, Inc. | Database hosting and authentication | supabase.com/privacy |
| Render (Render Services, Inc.) | Application hosting | render.com/privacy |
| SendGrid (Twilio) | Account, security, and notification emails — e.g. email verification, password resets, and optional signal alerts | sendgrid.com/policies/privacy |
| GitHub, Inc. | Scheduled background jobs (data refresh) | github.com/site/privacy |

Market data is sourced from public providers (Yahoo Finance, FRED) and is not personal data.

We will share data with law enforcement if legally required, and may disclose data to protect QNTM's rights or the safety of users.

### Email Communications

We use your email address to send two kinds of messages:

- **Account and security emails** — for example, email verification, password resets, and billing notices. These are necessary to operate your account, so you can't opt out of them while your account is active.
- **Optional signal alerts** — for example, an email when a stock on your watchlist or in your portfolio moves to LOW conviction. These are **off by default** and entirely optional. You can turn them on or off anytime in Account → Notifications, and every alert email includes an unsubscribe reminder and our mailing address.

We send email through SendGrid (Twilio). We don't use your email address for advertising, and we don't sell it.

### Security
- Passwords: bcrypt-hashed (cost factor 12)
- Sensitive personal fields (email, name, 2FA secret): encrypted at rest with authenticated symmetric encryption (Fernet — AES-128-CBC with HMAC-SHA256)
- Email-hash lookup so we can find your account without decrypting your email
- Two-factor authentication available and strongly recommended
- Login events are logged for security review
- HTTPS for all traffic

### How Long We Keep Your Data
- **Active accounts:** kept while your account exists
- **Deleted accounts:** personal data is removed within 30 days. Founding Member status is forfeited at deletion and cannot be restored — see Terms §6.
- **Payment records:** kept 7 years as required by tax and financial-record law
- **Anonymous, aggregated usage statistics:** kept indefinitely

### Your Rights
Email privacy@qntm.live to:
- See what data we hold about you
- Correct inaccurate data
- Delete your account and personal data
- Export your data in machine-readable form

We will respond within 30 days.

California residents have additional rights under the California Consumer Privacy Act (CCPA), including the right to know what personal information we collect, the right to delete it, and the right to non-discrimination for exercising these rights. We do not sell personal information as defined by CCPA.

### Children
QNTM is not directed to anyone under 18. We do not knowingly collect data from anyone under 18. If you believe a child has provided us data, contact privacy@qntm.live and we will delete it.

### Contact
privacy@qntm.live
"""

TERMS_OF_SERVICE = """
## Terms of Service
**Effective Date: June 6, 2026** · QNTM LLC (California)

### 1. Who We Are and What We Do
QNTM is a quantitative research platform operated by QNTM LLC, a California limited liability company. We score stocks using a five-factor model, apply a live macro regime overlay, and surface conviction signals. Everything we produce is algorithmic research — not investment advice.

### 2. Agreement
By creating an account or using QNTM, you agree to these Terms. If you do not agree, do not use the platform.

### 3. Not Investment Advice
**QNTM is not a registered investment adviser, broker-dealer, or financial planner.**

HIGH, MODERATE, and LOW conviction signals are algorithmic outputs — not recommendations to buy or sell any security. You make your own investment decisions. QNTM is not responsible for investment losses. Past model performance does not guarantee future results. Consult a qualified financial adviser before investing.

QNTM LLC holds no securities. Its principals may personally hold or trade securities the model scores; see the Conflicts of Interest section of our Investment Disclaimer.

### 4. Who Can Use QNTM
You must be 18 or older. By using QNTM you confirm this. You are responsible for complying with the financial research laws in your country.

### 5. Your Account
- Use accurate information when signing up
- Keep your password and 2FA credentials secure
- You are responsible for all activity on your account
- Report unauthorized access immediately to security@qntm.live
- By creating an account, you agree to receive account and security emails (such as email verification, password resets, and billing notices), which are necessary to operate your account. Optional signal-alert emails are off by default and can be enabled or disabled anytime in Account → Notifications.

### 6. Subscriptions

**Plans**

| Plan | Price | Notes |
|------|-------|-------|
| Free | $0/month | Limited features |
| Pro | $29/month | Full access — first 7 days free |
| Founding Member | $0 forever | First 50 accounts, Pro access, tied to original account |

**Free Trial**
- Pro subscriptions include a 7-day free trial. You will not be charged until the trial ends.
- One free trial per customer (per person, per household, per payment method). We reserve the right to decline a trial to any account where we reasonably believe a trial has already been used.
- Cancel before the trial ends and you pay nothing.

**Billing**
- Pro is billed monthly on the date your trial ends (your billing anniversary).
- Payments are processed by Stripe.
- By subscribing you authorize recurring monthly charges until you cancel.
- You will receive an email receipt for every charge.

**Cancellation**

You can cancel anytime from Account Settings → Subscription with a single click. Clicking Cancel stops your next charge immediately. You keep Pro access until the end of your current paid period, after which your account converts to Free.

What happens when you cancel:
1. Pro access continues until the end of your current billing period.
2. You will not be charged again.
3. **No refunds are issued for partial months.** The price you pay is for a full month of access; canceling mid-month does not produce a refund.
4. At the end of the period, your account automatically converts to Free. Your data is preserved.

You can reactivate Pro at any time from Account Settings. Reactivation begins a new billing period.

**Failed Payments**
1. Stripe retries the charge up to 3 times over 7 days
2. You will receive an email notification each retry
3. After 3 failures, your subscription is paused — you keep your account and data but lose Pro access
4. Restore Pro access anytime by updating your payment method in Account Settings
5. No data is deleted during a paused subscription

**Price Changes**

We will give you between 7 and 30 days' email notice before any price increase. The notice will state the new price, the date it takes effect, and how to cancel. Existing Pro subscribers continue at their current price for the remainder of the billing period in which the change is announced. Founding Members are not subject to price changes.

**Founding Members**
- The first 50 customer accounts receive Pro access free, for as long as the account remains active.
- Tied to the original account. Founding Member status is not transferable.
- If you delete your QNTM account, you forfeit Founding Member status. Re-registering with a new email does not restore it.
- One per person. We may decline Founding Member status if we reasonably believe an account is a duplicate of a previous Founding Member.
- 90 days' notice before any material change to the Founding Member program.

### 7. What You Can't Do
- Scrape or redistribute model outputs commercially
- Share your account with others
- Try to reverse-engineer the scoring model
- Use bots or automated tools against the platform
- Use QNTM for anything illegal

Violations may result in immediate account termination without refund.

### 8. Our Intellectual Property
The QNTM model, methodology, source code, and platform are proprietary to QNTM LLC. You receive a personal, non-exclusive, non-transferable license to access the platform for your own use during your subscription.

Market data is provided by third parties (including Yahoo Finance and the Federal Reserve Bank of St. Louis / FRED) under their respective terms of service. QNTM does not claim ownership of market data or company identifiers. Ticker symbols and company names are used for identification only.

### 9. Our Liability to You
QNTM is provided "as is" and "as available." We do not guarantee the platform will always be available or that model scores will be accurate, current, or complete.

To the maximum extent permitted by law, our total liability to you for any claim is capped at the greater of (a) the amount you paid us in the 12 months before the claim, or (b) $100. We are not liable for investment losses, market data errors, indirect or consequential damages, or any decisions you make based on our outputs.

### 10. Disputes

**Informal resolution.** We would prefer to resolve issues directly. Email legal@qntm.live first. If we cannot resolve it within 30 days, then arbitration applies.

**Arbitration.** Any dispute that we cannot resolve informally will be resolved by binding individual arbitration administered by JAMS under its Streamlined Arbitration Rules and Procedures. The arbitration will take place in Orange County, California, or by video conference at your option. Judgment on the arbitration award may be entered in any court of competent jurisdiction.

**Class action waiver.** You and QNTM agree to resolve disputes only on an individual basis. Class actions, collective actions, and representative actions are waived to the extent permitted by law.

**Right to opt out of arbitration.** You can opt out of this arbitration agreement (including the class action waiver) by sending written notice to legal@qntm.live within 30 days of first creating your QNTM account. Opting out does not affect any other part of these Terms. If you opt out, disputes will be resolved in the state or federal courts located in Orange County, California, and both parties consent to personal jurisdiction and venue there.

**Governing law.** These Terms are governed by the laws of the State of California, without regard to conflict-of-laws principles. The Federal Arbitration Act governs the interpretation and enforcement of the arbitration provision.

### 11. Changes to These Terms
We will give you 14 days' email notice before material changes take effect. The notice will explain how to cancel: Account Settings → Subscription → Cancel. Continued use after the effective date means you accept the updated Terms. If you do not agree, you can cancel your subscription and delete your account.

### 12. Contact
- General: hello@qntm.live
- Legal: legal@qntm.live
- Security: security@qntm.live
- Billing: billing@qntm.live
- Privacy: privacy@qntm.live
"""

BILLING_POLICY = """
## Billing & Refund Policy
**Effective Date: June 6, 2026** · QNTM LLC (California)

### Plans

| Plan | Price | Trial |
|------|-------|-------|
| Free | $0 | No trial needed |
| Pro | $29/month | 7 days free, one per customer |
| Founding Member | Free forever | First 50 accounts (see Terms §6) |

### Free Trial
- Pro includes a 7-day free trial
- No charge until the trial ends
- One trial per customer (per person, per household, per payment method)
- Cancel before the trial ends: no charge

### Billing
- Billed monthly on the day your trial ends (your billing anniversary)
- Payments processed via Stripe (credit and debit cards)
- Email receipt sent for every charge

### Cancellations

You can cancel anytime in Account Settings → Subscription with a single click. Clicking Cancel stops your next charge immediately.

What happens when you cancel:
1. Pro access continues until the end of your current billing period.
2. You will not be charged again.
3. **We do not issue refunds for partial months.** The monthly fee is for one full month of access; canceling mid-cycle does not generate a refund.
4. At the end of the period, your account converts to Free. Your data is preserved.
5. You can reactivate Pro at any time. Reactivation begins a new billing period.

### Failed Payments
- Stripe retries 3 times over 7 days
- Email notification sent each attempt
- After 3 failures: subscription paused (account and data preserved, Pro access suspended)
- Reactivate anytime by updating your payment method
- No late fees

### Exceptional Refunds
We may issue a refund at our discretion in cases of duplicate charges, billing errors on our side, or extended platform outages. Refunds in those cases are processed within 5–10 business days via Stripe to the original payment method.

### Founding Members
- First 50 customer accounts get Pro free, for as long as the account remains active.
- Founding Member status is tied to the original account and is not transferable. If the account is deleted, the status is forfeited.
- 90 days' notice before any material change to Founding Member benefits.
- See Terms §6 for full Founding Member terms.

### Billing Disputes
Email billing@qntm.live within 60 days of a charge. Include your account email and the date of the charge. We will respond within 5 business days. If you dispute a charge directly with your card issuer (chargeback) before contacting us, your account may be suspended pending resolution.
"""

DISCLAIMER_FULL = """
## Investment Disclaimer
**Applies to all QNTM content and outputs.**

### QNTM is a Research Tool, Not a Financial Adviser
QNTM LLC is not registered with the SEC, FINRA, or any state securities regulator. We are not a registered investment adviser, broker-dealer, or financial planner.

### What Our Signals Mean
HIGH, MODERATE, and LOW conviction signals are produced by an algorithm. They reflect quantitative patterns — not a judgment about whether you personally should buy or sell a stock. The same signal means different things for different investors depending on their situation, risk tolerance, tax position, and goals.

**A HIGH conviction signal is not a buy recommendation.**
**A LOW conviction signal is not a sell recommendation.**
**A MODERATE signal is not a hold recommendation.**

This applies equally to any alerts or emails we send. If you enable signal alerts, an email or in-app notice telling you that a stock has moved to LOW conviction is an automated, algorithmic notification — not advice and not a recommendation to sell. Intraday alerts in particular are based on partial-day data and are more reactive and noisier than the end-of-day signal. You remain solely responsible for your own investment decisions.

## Conflicts of Interest

QNTM LLC does not hold, buy, or sell any securities. The company takes no position in any stock the model scores.

The founder and any principals of QNTM LLC may personally own and trade securities the model scores, including securities carrying a HIGH, MODERATE, or LOW signal. To manage this conflict:

- We act only on signals after they are published to subscribers, using the same information available to every user at the same time.
- We do not trade ahead of a signal we know is about to publish or change.
- Scoring is fully algorithmic. Personal holdings do not influence how any security is scored, ranked, or labeled.

The model applies the same rules to every security in its universe, regardless of whether anyone connected to QNTM holds it.

### Model Portfolio
The Model Portfolio shown in QNTM is a hypothetical illustration of how the model's signals would translate into a position book. It uses equal-weighted notional sizing, ignores slippage, taxes, brokerage commissions, and dividend treatment. It is not a real portfolio, no securities are held on your behalf, and no trades are executed. It is provided for transparency about the model's behavior, not as a recommendation.

### Past Performance
QNTM does not currently publish a historical backtest. The only performance we show is the live Model Portfolio, which is marked daily on the model's rules-based signals from its inception date forward. It is a hypothetical illustration (see Model Portfolio above) and does not account for all real-world costs, taxes, or execution constraints. A live track record is short by nature, and past model performance is not a guarantee of future results.

### You Can Lose Money
All equity investments carry the risk of loss, including loss of your entire investment.

### Get Professional Advice
Before making any investment decision, speak with a qualified financial adviser who understands your complete financial picture.
"""

COOKIE_POLICY = """
## Cookie Policy
**Effective Date: June 6, 2026**

### Cookies We Use

| Cookie | Required | Purpose | Expires |
|--------|----------|---------|---------|
| Session token | Yes | Keeps you logged in | End of session |
| CSRF token | Yes | Security protection | End of session |
| Auth state | Yes | Remembers your login across visits | 30 days |
| Cookie consent | Yes | Records your cookie preference | 12 months |

We do not currently use analytics, advertising, or third-party tracking cookies. If we add analytics in the future, we will update this policy and require fresh consent.

### No Ad Tracking
We do not use advertising cookies. We do not track you across other websites. We do not share cookie data with advertisers.

### Stripe Cookies
When you visit the upgrade or checkout flow, Stripe sets its own cookies to process payment and prevent fraud. See stripe.com/cookies.

### Managing Cookies
Required cookies cannot be turned off — the platform will not function without them. To remove all QNTM cookies, clear your browser storage for qntm.live. This will log you out.

### Contact
privacy@qntm.live
"""


def page_legal(doc_key: str = "privacy"):
    docs = {
        "privacy":    ("Privacy Policy",        PRIVACY_POLICY),
        "terms":      ("Terms of Service",      TERMS_OF_SERVICE),
        "billing":    ("Billing & Refund Policy", BILLING_POLICY),
        "disclaimer": ("Investment Disclaimer", DISCLAIMER_FULL),
        "cookies":    ("Cookie Policy",         COOKIE_POLICY),
    }
    title, text = docs.get(doc_key, docs["privacy"])

    st.markdown("""
    <style>
    .legal-body { max-width: 800px; margin: 0 auto; padding: 40px 32px; }
    .legal-body h2 { font-family:'Syne',sans-serif;font-size:28px;font-weight:800;
                     color:#e2e8f0;margin-bottom:6px; }
    .legal-body h3 { font-family:'Syne',sans-serif;font-size:16px;font-weight:700;
                     color:#d4a843;margin:24px 0 8px; }
    .legal-body p,.legal-body li { font-size:14px;color:#b3bed0;line-height:1.8; }
    .legal-body strong { color:#e2e8f0; }
    .legal-body table { width:100%;border-collapse:collapse;margin:12px 0; }
    .legal-body th { font-size:14px;color:#b3bed0;text-align:left;padding:8px 12px;
                     border-bottom:1px solid rgba(255,255,255,.08); }
    .legal-body td { font-size:13px;color:#b3bed0;padding:8px 12px;
                     border-bottom:1px solid rgba(255,255,255,.04); }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    div[data-testid='stButton'][data-key='legal_back_btn'] button {
        display:inline-flex !important; align-items:center !important;
        background:rgba(255,255,255,.03) !important;
        border:1px solid rgba(255,255,255,.12) !important;
        border-radius:6px !important; color:#b3bed0 !important;
        font-family:Syne,sans-serif !important; font-size:13px !important;
        font-weight:700 !important; letter-spacing:.04em !important;
        padding:7px 10px !important; white-space:nowrap !important;
        text-transform:uppercase !important; width:auto !important;
        min-width:0 !important; box-shadow:none !important;
    }
    </style>
    """, unsafe_allow_html=True)
    if st.button("← Back", key="legal_back_btn"):
        st.session_state.page = "landing"
        st.session_state.signed_out = True
        # Clear legal param so routing doesn't re-set page=legal on rerun
        st.query_params.clear()
        st.rerun()
    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

    st.markdown('<div class="legal-body">', unsafe_allow_html=True)
    st.markdown(text)
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# COOKIE CONSENT PAGE — full page, 100% reliable buttons
# ══════════════════════════════════════════════════════════════════════════════
def page_cookie_consent():
    """No-op — cookie banner is now shown inline at bottom of landing page."""
    pass


def _cookie_banner():
    """Slim informational bottom banner — no click required, implied consent on use."""
    # Auto-accept on display — user is informed by seeing the banner
    if not st.session_state.get("cookies_accepted"):
        st.session_state.cookies_accepted = True
        st.query_params["ck"] = "1"

    st.markdown(
        '<style>'
        '#qntm-cookie-banner{'
        'position:fixed;bottom:0;left:0;right:0;z-index:9999;'
        'background:rgba(8,10,18,.95);backdrop-filter:blur(16px);'
        'border-top:1px solid rgba(255,255,255,.06);'
        'padding:12px 24px;}'
        '</style>'
        '<div id="qntm-cookie-banner">'
        '<div style="font-size:13px;color:#8896ac;line-height:1.5;max-width:900px;">'
        'QNTM uses essential cookies for login and session management and anonymous analytics to improve the platform. '
        'By using QNTM you agree to our '
        '<a href="?legal=privacy" style="color:#9fabc0;text-decoration:underline;">Privacy Policy</a> and '
        '<a href="?legal=terms" style="color:#9fabc0;text-decoration:underline;">Terms of Service</a>. '
        'QNTM is a quantitative research tool — not investment advice.'
        '</div>'
        '</div>',
        unsafe_allow_html=True
    )



# ══════════════════════════════════════════════════════════════════════════════
# PUBLIC MODEL PORTFOLIO PAGE — no auth required, shareable link
def page_landing():

    # Returning user with no uid in URL — try localStorage to restore session
    if not st.session_state.logged_in and not st.session_state.get("signed_out") and "uid" not in st.query_params:
        _inject_localstorage_reader()

    # ── Global landing CSS — overrides Streamlit defaults completely ─────────────
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Syne:wght@400;600;700;800&family=DM+Mono:wght@400;500&family=Outfit:wght@300;400;500;600&family=Inter:wght@400;500;600;700&display=swap');

    /* ── Prevent white flash on page transitions ── */
html { background-color: #0a0b14 !important; }
body { background-color: #0a0b14 !important; }
[data-testid="stAppViewContainer"] { background-color: #0a0b14 !important; }
[data-testid="stApp"] { background-color: #0a0b14 !important; }
.main { background-color: #0a0b14 !important; }

/* ── Hard reset Streamlit to dark theme + kill all horizontal scroll ── */
    html, body { overflow-x: hidden !important; max-width: 100vw !important; }
    html, body, [class*="css"], .main, .block-container,
    [data-testid="stAppViewContainer"], [data-testid="stMain"],
    [data-testid="stMainBlockContainer"] {
        background-color: #0a0b14 !important;
        color: #e2e4f0 !important;
        font-family: 'Inter', sans-serif !important;
        overflow-x: hidden !important;
        max-width: 100% !important;
    }
    [data-testid="stAppViewContainer"] > section > div {
        background-color: #0a0b14 !important;
        overflow-x: hidden !important;
    }
    /* Remove Streamlit default padding */
    .main .block-container {
        padding: 0 !important;
        max-width: 100% !important;
        overflow-x: hidden !important;
    }
    /* Clamp all Streamlit column blocks to viewport */
    [data-testid="stHorizontalBlock"] {
        max-width: 100vw !important;
        overflow-x: hidden !important;
    }
    /* Hide hamburger, header, footer */
    #MainMenu, header[data-testid="stHeader"], footer { display: none !important; }

    /* ── Layout helpers ── */
    .land-section { padding: 28px clamp(16px,4vw,48px) 16px; max-width: 1200px; margin: 0 auto; }
    .land-divider  { border-top: 1px solid rgba(255,255,255,.06); }

    /* ── Animations ── */
    @keyframes land-pulse  { 0%,100%{opacity:1} 50%{opacity:.3} }
    @keyframes land-ticker { 0%{transform:translateX(0)} 100%{transform:translateX(-50%)} }

    /* ── Streamlit columns on landing ── */
    section[data-testid="stMain"] [data-testid="stHorizontalBlock"] {
        gap: 6px !important;
        background: transparent !important;
        max-width: 100vw !important;
        overflow-x: hidden !important;
    }

    /* ── CTA button styles — liquid gold primary, glass ghost ── */
    @keyframes land-gold-shimmer {
      0%   { background-position: -300% center; }
      100% { background-position:  300% center; }
    }
    @keyframes land-float {
      0%,100% { transform: translateY(0px); box-shadow: 0 2px 12px rgba(212,168,67,.15); }
      50%      { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(212,168,67,.2); }
    }

    .land-btn-primary > div > button,
    .land-btn-primary button {
        background: linear-gradient(
          105deg,
          #c8973a 0%,
          #e8be5a 25%,
          #d4a843 45%,
          #f0cc6a 55%,
          #d4a843 70%,
          #b8822a 100%
        ) !important;
        background-size: 300% 100% !important;
        color: #000 !important;
        border: none !important;
        border-radius: 6px !important;
        font-family: 'Syne', sans-serif !important;
        font-weight: 800 !important;
        font-size:13px !important;
        letter-spacing: .18em !important;
        padding: 13px 28px !important;
        text-transform: uppercase !important;
        width: 100% !important;
        cursor: pointer !important;
        position: relative !important;
        overflow: hidden !important;
        box-shadow: 0 6px 24px rgba(212,168,67,.3), 0 2px 6px rgba(0,0,0,.5),
                    inset 0 1px 0 rgba(255,255,255,.25) !important;
        transition: box-shadow .25s, transform .2s, background-position .4s !important;
        animation: land-float 4s ease-in-out infinite !important;
    }
    .land-btn-primary > div > button::after,
    .land-btn-primary button::after {
        content: '' !important;
        position: absolute !important;
        top: 0 !important; left: -100% !important;
        width: 60% !important; height: 100% !important;
        background: linear-gradient(
          90deg, transparent,
          rgba(255,255,255,.25),
          transparent) !important;
        transform: skewX(-20deg) !important;
        transition: left .5s !important;
    }
    .land-btn-primary > div > button:hover::after,
    .land-btn-primary button:hover::after {
        left: 160% !important;
    }
    .land-btn-primary > div > button:hover,
    .land-btn-primary button:hover {
        background-position: 100% center !important;
        box-shadow: 0 12px 40px rgba(212,168,67,.5), 0 4px 12px rgba(0,0,0,.6),
                    inset 0 1px 0 rgba(255,255,255,.3) !important;
        animation-play-state: paused !important;
        transform: translateY(-3px) !important;
    }
    .land-btn-primary > div > button:active,
    .land-btn-primary button:active {
        transform: translateY(0) !important;
        animation-play-state: paused !important;
    }

    .land-btn-ghost > div > button,
    .land-btn-ghost button {
        background: rgba(212,168,67,.04) !important;
        color: #d4a843 !important;
        border: 1px solid rgba(212,168,67,.35) !important;
        border-radius: 6px !important;
        font-family: 'Syne', sans-serif !important;
        font-weight: 700 !important;
        font-size:13px !important;
        letter-spacing: .18em !important;
        padding: 13px 28px !important;
        text-transform: uppercase !important;
        width: 100% !important;
        cursor: pointer !important;
        backdrop-filter: blur(8px) !important;
        box-shadow: inset 0 1px 0 rgba(212,168,67,.12),
                    0 2px 8px rgba(0,0,0,.3) !important;
        transition: background .2s, border-color .2s, box-shadow .2s, transform .2s !important;
    }
    .land-btn-ghost > div > button:hover,
    .land-btn-ghost button:hover {
        background: rgba(212,168,67,.10) !important;
        border-color: rgba(212,168,67,.65) !important;
        box-shadow: 0 0 16px rgba(212,168,67,.15),
                    inset 0 1px 0 rgba(212,168,67,.2),
                    0 4px 12px rgba(0,0,0,.4) !important;
        transform: translateY(-2px) !important;
        color: #e8be5a !important;
    }

    /* ── Scrollbar ── */
    ::-webkit-scrollbar { width: 3px; }
    ::-webkit-scrollbar-track { background: #0a0b14; }
    ::-webkit-scrollbar-thumb { background: #d4a843; border-radius: 2px; }
    </style>
    """, unsafe_allow_html=True)

    # ── NAV BAR ───────────────────────────────────────────────────────────────
    # Inject sticky nav CSS
    st.markdown("""
    <style>
    /* Sticky nav wrapper */
    .qntm-nav {
        position: sticky; top: 0; z-index: 999;
        background: rgba(10,11,20,.97);
        backdrop-filter: blur(12px);
        border-bottom: 1px solid rgba(255,255,255,.06);
        padding: 0 clamp(16px,4vw,48px);
        height: 60px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        width: 100%;
        box-sizing: border-box;
    }
    .qntm-nav-logo {
        font-family: 'Syne', sans-serif;
        font-size: 22px;
        font-weight: 800;
        letter-spacing: .18em;
        color: #e2e4f0;
    }
    .qntm-nav-logo span { color: #d4a843; }
    .qntm-nav-links { display: flex; gap: 10px; align-items: center; }
    </style>
    <div class="qntm-nav">
      <div class="qntm-nav-logo">Q<span>NTM</span></div>
      <div class="qntm-nav-links" id="qntm-nav-btns">
        <!-- Streamlit buttons injected below via columns -->
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Nav buttons — pure HTML links, no st.columns, no layout issues
    st.markdown("""
    <style>
    .qntm-nav-btns { display:flex; gap:8px; align-items:center; }
    .qntm-nav-btns a {
        font-family:'Syne',sans-serif; font-size:13px; font-weight:700;
        letter-spacing:.12em; text-transform:uppercase; text-decoration:none;
        padding:8px 14px; border-radius:6px; white-space:nowrap;
    }
    .qntm-nav-btn-ghost { color:#d4a843 !important; border:1px solid rgba(212,168,67,.4); background:rgba(212,168,67,.04); }
    .qntm-nav-btn-primary { color:#000 !important; background:#d4a843; }
    @media (max-width:600px) {
        .qntm-nav-btns a { font-size:13px; padding:7px 12px; letter-spacing:.04em; }
    }
    </style>
    <div style="display:flex;justify-content:flex-end;padding:0 16px;margin-top:-52px;position:relative;z-index:1000;height:52px;align-items:center;">
      <div class="qntm-nav-btns">
        <a href="?nav=signin" target="_self" class="qntm-nav-btn-ghost">Sign In</a>
        <a href="?nav=register" target="_self" class="qntm-nav-btn-primary">Join Free</a>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── HERO — two-column layout ──────────────────────────────────────────────

    # Fetch macro regime for right-side intelligence panel
    try:
        from model_engine import fetch_macro_overlay
        _macro_now = _live_macro()
    except Exception:
        _macro_now = {}
    _regime       = _macro_now.get("regime", "NEUTRAL")
    # Normalise regime — model can return MILDLY BULLISH / HIGH VOLATILITY too
    if _regime in ("RISK_ON","MILDLY BULLISH"):   _regime_norm = "RISK_ON"
    elif _regime in ("RISK_OFF","HIGH VOLATILITY"): _regime_norm = "RISK_OFF"
    else:                                           _regime_norm = "NEUTRAL"
    _regime_label = {"RISK_ON":"Risk On","RISK_OFF":"Risk Off","NEUTRAL":"Neutral"}.get(_regime_norm,"Neutral")
    _regime_c     = {"RISK_ON":"#34d399","RISK_OFF":"#f87171","NEUTRAL":"#fbbf24"}.get(_regime_norm,"#fbbf24")
    _regime_bg    = {"RISK_ON":"rgba(52,211,153,.08)","RISK_OFF":"rgba(248,113,113,.08)","NEUTRAL":"rgba(251,191,36,.08)"}.get(_regime_norm,"rgba(251,191,36,.08)")
    _regime_brd   = {"RISK_ON":"rgba(52,211,153,.25)","RISK_OFF":"rgba(248,113,113,.25)","NEUTRAL":"rgba(251,191,36,.25)"}.get(_regime_norm,"rgba(251,191,36,.25)")
    _regime_icon  = {"RISK_ON":"▲","RISK_OFF":"▼","NEUTRAL":"─"}.get(_regime_norm,"─")
    _vix          = _macro_now.get("vix", None)
    _vix_str      = f"{_vix:.1f}" if _vix else "—"
    _events       = _macro_now.get("active_events", [])
    _evt_label    = _events[0].replace("_"," ").title() if _events else "None active"

    # Fetch top 5 signals for right panel
    try:
        from data_refresh import _get_supabase as _hero_sb
        _sb2 = _hero_sb()
        _top5 = []
        if _sb2:
            _r5 = _sb2.table("signal_log") \
                .select("ticker,adj_composite,composite,signal,momentum,quality,volume,value,sentiment,price,signal_date,val_low,val_high,value_position,val_basis") \
                .gte("adj_composite", 60) \
                .order("signal_date", desc=True) \
                .order("adj_composite", desc=True) \
                .limit(100) \
                .execute()
            _seen5 = {}
            for _row in (_r5.data or []):
                if _row["ticker"] not in _seen5:
                    _adj = float(_row.get("adj_composite") or _row.get("composite") or 0)
                    _row["adj_action"] = "BUY"
                    _row["score_delta"] = round(_adj - float(_row.get("composite") or _adj), 1)
                    _seen5[_row["ticker"]] = _row
            _top5 = sorted(_seen5.values(), key=lambda x: float(x.get("adj_composite",0) or 0), reverse=True)[:5]
    except Exception:
        _top5 = []

    # Signal rows for right panel — simple scan rows, not full cards
    _n_high_total = len(_top5)  # will update after full count fetch below
    _signal_rows = ""
    for _sr in _top5:
        _stk   = _sr.get("ticker","")
        _ssc   = float(_sr.get("adj_composite",0) or 0)
        _signal_rows += (
            '<div style="display:flex;justify-content:space-between;align-items:center;'
            'padding:9px 0;border-bottom:1px solid rgba(255,255,255,.04);">'
            f'<span style="font-family:Syne,sans-serif;font-size:14px;font-weight:800;color:#e2e8f0;">{_stk}</span>'
            f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#34d399;font-weight:700;">{_ssc:.0f} ▲</span>'
            '</div>'
        )
    if not _signal_rows:
        _signal_rows = '<div style="font-size:13px;color:#94a3b8;padding:12px 0;">Signals loading...</div>'


    # Dynamic founding spots count
    _spots_remaining = 50
    try:
        from data_refresh import _get_supabase as _fs_sb
        _sb_fs = _fs_sb()
        if _sb_fs:
            _uc = _sb_fs.table("users").select("id", count="exact").execute()
            _user_count = _uc.count if hasattr(_uc, 'count') and _uc.count else len(_uc.data or [])
            _spots_remaining = max(0, 50 - _user_count)
    except Exception:
        pass

    hero_html = (
        '<style>'
        '.qntm-hero2{'
        '  display:grid;grid-template-columns:1fr 1fr;gap:40px;align-items:start;'
        '  padding:52px clamp(20px,5vw,64px) 40px;max-width:1200px;margin:0 auto;'
        '  background:radial-gradient(ellipse 60% 70% at 20% 0%,rgba(212,168,67,.07) 0%,transparent 65%);'
        '}'
        '@media(max-width:768px){'
        '  .qntm-hero2{grid-template-columns:1fr!important;gap:28px!important;padding:32px 16px 28px!important;}'
        '  .qntm-hero2-right{display:none!important;}'  # hide right panel on mobile — shows below
        '}'
        '.qntm-trust-strip{display:flex;gap:6px;flex-wrap:wrap;padding:0 clamp(20px,5vw,64px);'
        'max-width:1200px;margin:0 auto 0;}'
        '</style>'
        '<div class="qntm-hero2">'

        # ── LEFT: headline + subtext + CTAs ──────────────────────────────────
        '<div>'
        '<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:18px;">'
        '<div style="display:inline-flex;align-items:center;gap:6px;background:rgba(212,168,67,.08);'
        'border:1px solid rgba(212,168,67,.2);border-radius:100px;padding:5px 14px;">'
        '<div style="width:6px;height:6px;background:#00ff87;border-radius:50%;'
        'animation:land-pulse 2s infinite;flex-shrink:0;"></div>'
        '<span style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;">'
        'MODEL LIVE · UPDATED DAILY</span></div>'
        '<div style="display:inline-flex;align-items:center;background:rgba(52,211,153,.05);'
        'border:1px solid rgba(52,211,153,.15);border-radius:100px;padding:5px 14px;">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#34d399;letter-spacing:.08em;">'
        f'🎯 {_spots_remaining} FOUNDING SPOTS · FREE TODAY</span></div>'
        '</div>'
        '<h1 style="font-family:Syne,sans-serif;font-size:clamp(36px,4vw,60px);'
        'font-weight:800;line-height:1.0;letter-spacing:-.02em;color:#ffffff;margin-bottom:18px;">'
        'Know where<br>conviction is<br>'
        '<span style="color:#d4a843;">strongest.</span></h1>'
        '<p style="font-size:15px;color:#b3bed0;max-width:400px;line-height:1.75;margin-bottom:32px;">'
        f'A multi-factor quantitative model scoring {_universe_n()} stocks daily — blended with a live macro regime overlay.'
        '</p>'
        # CTAs inline in left column
        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;max-width:360px;">'
    )

    hero_html += (
        _cta_gold("Join Free →", "?nav=register&plan=pro")
        + _cta_ghost("Sign In", "?nav=signin")
        + '</div>'  # end CTA grid
        + '<div style="margin-top:10px;font-size:13px;color:#94a3b8;letter-spacing:.02em;">'
        + 'No credit card · cancel anytime · free tier always available'
        + '</div>'
        '</div>'    # end left col

        # ── RIGHT: intelligence panel ─────────────────────────────────────────
        + '<div class="qntm-hero2-right" style="'
        'background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.07);'
        'border-radius:12px;padding:20px 22px;backdrop-filter:blur(8px);">'

        # Regime header
        + f'<div style="background:{_regime_bg};border:1px solid {_regime_brd};border-radius:8px;'
        f'padding:14px 16px;margin-bottom:16px;">'
        f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.12em;margin-bottom:5px;">MARKET REGIME</div>'
        f'<div style="display:flex;justify-content:space-between;align-items:center;">'
        f'<span style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:{_regime_c};">'
        f'{_regime_icon} {_regime_label}</span>'
        f'<div style="text-align:right;">'
        f'<div style="font-family:DM Mono,monospace;font-size:16px;color:#d4a843;">75/25</div>'
        f'<div style="font-size:13px;color:#8896ac;">quant/macro</div>'
        f'</div></div>'
        f'<div style="display:flex;gap:12px;margin-top:8px;flex-wrap:wrap;">'
        f'<span style="font-size:13px;color:#9fabc0;">VIX {_vix_str}</span>'
        f'<span style="font-size:13px;color:#9fabc0;">Event: {_evt_label}</span>'
        f'</div></div>'

        # Top signals label
        + '<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.12em;margin-bottom:2px;">TOP SIGNALS TODAY</div>'
        + _signal_rows

        # High conviction count CTA
        + f'<a href="?nav=register&amp;plan=pro" target="_self" style="display:block;margin-top:10px;padding:7px 12px;'
        f'background:rgba(212,168,67,.08);border:1px solid rgba(212,168,67,.2);border-radius:6px;'
        f'font-family:DM Mono,monospace;font-size:13px;color:#d4a843;text-decoration:none;'
        f'letter-spacing:.06em;text-align:center;">'
        f'VIEW ALL HIGH CONVICTION SIGNALS →</a>'
        # Hidden gems callout
        + '<div style="margin-top:10px;padding:8px 12px;background:rgba(52,211,153,.04);'
        'border:1px solid rgba(52,211,153,.12);border-radius:6px;display:flex;align-items:center;gap:8px;">'
        '<span style="font-size:14px;">💎</span>'
        '<div><div style="font-family:DM Mono,monospace;font-size:11px;color:#34d399;letter-spacing:.1em;">HIDDEN GEMS</div>'
        '<div style="font-size:13px;color:#9fabc0;">Low-coverage stocks with high conviction scores</div></div>'
        '</div>'
        # Compact stats strip at bottom of panel — factual, no performance claims
        + '<div style="display:grid;grid-template-columns:1fr 1fr;gap:6px;margin-top:14px;padding-top:12px;border-top:1px solid rgba(255,255,255,.05);">'
        + '<div><div style="font-size:11px;color:#8896ac;letter-spacing:.08em;">UNIVERSE</div>'
        + f'<div style="font-family:Syne,sans-serif;font-size:17px;font-weight:800;color:#d4a843;">~{_universe_n()}</div></div>'
        + '<div><div style="font-size:11px;color:#8896ac;letter-spacing:.08em;">FACTORS</div>'
        + '<div style="font-family:Syne,sans-serif;font-size:17px;font-weight:800;color:#e2e8f0;">5</div></div>'
        + '<div><div style="font-size:11px;color:#8896ac;letter-spacing:.08em;">REFRESH</div>'
        + '<div style="font-family:Syne,sans-serif;font-size:17px;font-weight:800;color:#34d399;">Daily</div></div>'
        + '<div><div style="font-size:11px;color:#8896ac;letter-spacing:.08em;">SIGNALS</div>'
        + '<div style="font-family:Syne,sans-serif;font-size:17px;font-weight:800;color:#e2e8f0;">Live</div></div>'
        + '</div>'

        + '</div>'  # end right panel
        + '</div>'  # end hero grid
    )

    st.markdown(hero_html, unsafe_allow_html=True)

    # ── TODAY IN QNTM — transition strip ─────────────────────────────────────
    _n_high = len(_top5)
    _n_sell = 0
    try:
        # Use session cache from screener if available — exact same numbers
        _n_high = st.session_state.get("_high_count", 0)
        _n_sell = st.session_state.get("_low_count", 0)
        # Fall back to platform_stats from DB
        if not _n_high and not _n_sell:
            _ps = _sb2.table("platform_stats").select("n_high,n_low") \
                .eq("stat_key", "daily_summary").limit(1).execute()
            if _ps.data:
                _n_high = _ps.data[0].get("n_high", 0)
                _n_sell = _ps.data[0].get("n_low", 0)
    except Exception:
        pass

    _today_items = []
    # Regime is primary — larger and brighter
    _today_items.insert(0, f'<span style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:{_regime_c};">{_regime_icon} {_regime_label}</span>')
    if _n_high:  _today_items.append(f'<span style="color:#9fabc0;"><b style="color:#34d399;">{_n_high}</b> high</span>')
    if _n_sell:  _today_items.append(f'<span style="color:#8896ac;"><b style="color:#f87171;">{_n_sell}</b> low</span>')
    # Gem count — read from platform_stats table (written by cron after every refresh)
    _n_gems = st.session_state.get("_gem_count", None)
    if _n_gems is None:
        try:
            _ps = _sb2.table("platform_stats") \
                .select("n_gems,n_high,n_low,n_total,regime") \
                .eq("stat_key", "daily_summary") \
                .limit(1) \
                .execute()
            if _ps.data:
                _n_gems = _ps.data[0].get("n_gems", 0)
                # Also update high/low counts from DB if not already set
                if not _n_high:
                    _n_high = _ps.data[0].get("n_high", 0)
                if not _n_sell:
                    _n_sell = _ps.data[0].get("n_low", 0)
        except Exception:
            pass
    _gems_display = _n_gems if _n_gems is not None else "—"
    _gem_word = "gem" if _n_gems == 1 else "gems"
    _today_items.append(f'<span style="color:#34d399;font-weight:600;">💎 {_gems_display} hidden {_gem_word}</span>')
    _today_items.append(f'<span style="color:#9fabc0;">{_universe_n()} stocks scored</span>')

    st.markdown(
        '<div style="padding:14px clamp(20px,5vw,64px);max-width:1200px;margin:0 auto;'
        'border-top:1px solid rgba(255,255,255,.04);border-bottom:1px solid rgba(255,255,255,.04);'
        'display:flex;gap:20px;flex-wrap:wrap;align-items:center;">'
        '<span style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;white-space:nowrap;">TODAY IN QNTM</span>'
        + ' <span style="color:#94a3b8;">·</span> '.join(_today_items)
        + '</div>',
        unsafe_allow_html=True
    )

    # ── TICKER TAPE — live from latest signal_log ─────────────────────────────
    tape_scores = st.session_state.get("scan_results") or []
    if not tape_scores:
        # Try fetching latest signal_log for tape — works for all visitors
        try:
            from data_refresh import _get_supabase as _tape_sb
            _sb = _tape_sb()
            if _sb:
                _latest = _sb.table("signal_log") \
                    .select("ticker,adj_composite,signal") \
                    .order("signal_date", desc=True) \
                    .order("adj_composite", desc=True) \
                    .limit(200) \
                    .execute()
                tape_scores = _latest.data or []
        except Exception:
            pass

    _static_tape = [
            ("NVDA","HIGH","#34d399"),("META","HIGH","#34d399"),
            ("AVGO","HIGH","#34d399"),("JPM","HIGH","#34d399"),
            ("NFLX","HIGH","#34d399"),("COST","HIGH","#34d399"),
            ("GS","HIGH","#34d399"),("WMT","HIGH","#34d399"),
            ("MA","HIGH","#34d399"),("MSFT","HIGH","#34d399"),
            ("TSLA","MOD","#d4a843"),
            ("UNH","LOW","#E24B4A"),("NKE","LOW","#E24B4A"),
            ("PFE","LOW","#E24B4A"),("SNAP","LOW","#E24B4A"),
        ]
    if tape_scores:
        buys  = sorted([s for s in tape_scores if s.get("signal","") in ("BUY","HIGH")],
                       key=lambda x: float(x.get("adj_composite",0) or 0), reverse=True)[:10]
        sells = sorted([s for s in tape_scores if s.get("signal","") in ("SELL","LOW")],
                       key=lambda x: float(x.get("adj_composite",100) or 100))[:5]
        tape_items = (
            [(s["ticker"],"HIGH","#34d399") for s in buys] +
            [(s["ticker"],"LOW","#E24B4A")  for s in sells]
        )
        if not tape_items:
            tape_items = _static_tape
    else:
        tape_items = _static_tape

    def tape_span(ticker, action, color):
        return f'<span style="color:{color};">{ticker} {action}</span> &middot; '

    tape_html = " &middot; ".join(tape_span(*i).rstrip(" &middot; ") for i in tape_items)
    # Duplicate for seamless scroll — string concat avoids f-string single-quote conflicts
    _sp = "font-family:DM Mono,monospace;font-size:13px;padding:0 24px;"
    tape_block = (
        '<div style="overflow:hidden;max-width:100vw;background:rgba(52,211,153,.04);'
        'border-top:1px solid rgba(52,211,153,.12);border-bottom:1px solid rgba(52,211,153,.12);'
        'padding:13px 0;margin-top:8px;">'
        '<div style="display:inline-flex;animation:land-ticker 45s linear infinite;white-space:nowrap;will-change:transform;">'
        '<span style="' + _sp + '">' + tape_html + '</span>'
        '<span style="' + _sp + '">' + tape_html + '</span>'
        '</div></div>'
    )
    st.markdown(tape_block, unsafe_allow_html=True)

    st.markdown(f"""
    <div class="land-divider"></div>
    <div class="land-section">
      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;letter-spacing:.2em;margin-bottom:14px;">&mdash; WHY QNTM</div>
      <h2 style="font-family:'Syne',sans-serif;font-size:clamp(28px,4vw,42px);font-weight:800;
           color:#fff;margin-bottom:10px;line-height:1.1;">
        A live model.<br><span style="color:#d4a843;">Transparent by design.</span>
      </h2>
      <p style="color:#9fabc0;margin-bottom:24px;font-size:13px;">
        Every score is computed daily and shown with the reasoning behind it &mdash; no black box, and no
        cherry-picked history. The track record we show is the live Model Portfolio, reported as it happens.
      </p>
    </div>
    <div style="width:100%;box-sizing:border-box;padding:0 16px;margin-bottom:24px;">
      <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;">
        <div style="background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px;min-width:0;">
          <div style="font-family:Syne,sans-serif;font-size:clamp(18px,4.5vw,26px);font-weight:800;color:#d4a843;line-height:1;">~{_universe_n()} stocks</div>
          <div style="font-size:13px;color:#b3bed0;margin-top:6px;">Russell 1000 + top Russell 2000 small-caps, rescored daily</div>
        </div>
        <div style="background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px;min-width:0;">
          <div style="font-family:Syne,sans-serif;font-size:clamp(18px,4.5vw,26px);font-weight:800;color:#34d399;line-height:1;">5-factor model</div>
          <div style="font-size:13px;color:#b3bed0;margin-top:6px;">Momentum, Quality, Volume, Value, Sentiment</div>
        </div>
        <div style="background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px;min-width:0;">
          <div style="font-family:Syne,sans-serif;font-size:clamp(18px,4.5vw,26px);font-weight:800;color:#d4a843;line-height:1;">Plain-English</div>
          <div style="font-size:13px;color:#b3bed0;margin-top:6px;">A written rationale behind every conviction score</div>
        </div>
        <div style="background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px;min-width:0;">
          <div style="font-family:Syne,sans-serif;font-size:clamp(18px,4.5vw,26px);font-weight:800;color:#34d399;line-height:1;">Live portfolio</div>
          <div style="font-size:13px;color:#b3bed0;margin-top:6px;">Rules-based entries &amp; exits, marked daily vs SPY</div>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

        # ── THE MODEL ─────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="land-divider" style="margin-top:16px;"></div>
    <div class="land-section">
      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;letter-spacing:.2em;margin-bottom:14px;">&mdash; THE MODEL</div>
      <h2 style="font-family:'Syne',sans-serif;font-size:clamp(28px,4vw,42px);font-weight:800;
           color:#fff;margin-bottom:12px;line-height:1.1;">
        Five pillars.<br><span style="color:#d4a843;">One conviction score.</span>
      </h2>
      <p style="color:#b3bed0;max-width:520px;margin-bottom:36px;line-height:1.7;">
        36 factors scored weekly across 5 research-backed pillars — plus a 75/25 macro overlay.
        The model tells you exactly what to enter, maintain, or exit. And why.
      </p>
    </div>
    """, unsafe_allow_html=True)

    pillars_html = ""
    for name, weight, desc, color in [
        ("Momentum",  "30%", "Price trend, RSI, MACD, MA crossovers, 52-week proximity",      "#d4a843"),
        ("Quality",   "25%", "ROE, profit margin, revenue growth, EPS beat rate, FCF yield",   "#1D9E75"),
        ("Volume",    "20%", "Relative volume, OBV, Chaikin Money Flow, accumulation/dist.",   "#34d399"),
        ("Value",     "15%", "Forward P/E, PEG ratio, EV/EBITDA, Price-to-Sales, FCF yield",  "#f59e0b"),
        ("Sentiment", "10%", "Short interest, insider buy ratio, institutional ownership",      "#f97316"),
    ]:
        pillars_html += (
            f'<div style="background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:18px 14px;">'
            f'<div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;color:{color};margin-bottom:4px;">{weight}</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:#e2e4f0;margin-bottom:8px;">{name}</div>'
            f'<div style="font-size:13px;color:#b3bed0;line-height:1.6;">{desc}</div>'
            f'</div>'
        )
    st.markdown(
        f'<div style="width:100%;overflow-x:auto;-webkit-overflow-scrolling:touch;padding:0 16px;box-sizing:border-box;">'
        f'<div style="display:grid;grid-template-columns:repeat(5,minmax(140px,1fr));gap:10px;min-width:600px;">'
        f'{pillars_html}</div></div>',
        unsafe_allow_html=True)

    # Signal boxes — pure CSS grid, no st.columns
    st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
    signals_html = ""
    for label, score, desc, color, brd in [
        ("▲ HIGH",        "Score ≥ 60", "Strongest factor profile in the universe. Historically associated with multi-month relative outperformance. In high-volatility or risk-off regimes the model tightens this threshold to 62 to stay selective. Not a recommendation to buy.", "#1D9E75", "rgba(29,158,117,.3)"),
        ("─ MODERATE",    "Score 45–59", "Mixed factor profile — neither strong nor deteriorating on the model's measures. Not a recommendation to hold.",           "#f59e0b", "rgba(245,158,11,.25)"),
        ("▼ LOW",         "Score < 45",  "Weakest factor profile. The model flagged UNH here at month 3, ahead of a −49% full-year drawdown. Not a recommendation to sell.",                "#E24B4A", "rgba(226,75,74,.25)"),
    ]:
        signals_html += (
            f'<div style="background:#0e0f1a;border:1px solid {brd};border-radius:8px;padding:22px;">'
            f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:{color};letter-spacing:.1em;margin-bottom:8px;">{label}</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:22px;font-weight:500;color:{color};margin-bottom:10px;">{score}</div>'
            f'<div style="font-size:15px;color:#cbd5e1;line-height:1.8;">{desc}</div>'
            f'</div>'
        )
    st.markdown(
        f'<div style="width:100%;box-sizing:border-box;padding:0 16px;"><div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(220px,1fr));gap:10px;">{signals_html}</div></div>',
        unsafe_allow_html=True)


    # ── COMPETITOR MATRIX ─────────────────────────────────────────────────────
    st.markdown("""
    <div class="land-divider" style="margin-top:16px;"></div>
    <div class="land-section">
      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;letter-spacing:.2em;margin-bottom:14px;">&mdash; VS THE MARKET</div>
      <h2 style="font-family:'Syne',sans-serif;font-size:clamp(28px,4vw,42px);font-weight:800;
           color:#fff;margin-bottom:10px;line-height:1.1;">
        Institutional tools.<br><span style="color:#d4a843;">Retail price.</span>
      </h2>
      <p style="color:#b3bed0;margin-bottom:32px;">Institutional-grade quant signals &mdash; at roughly 1% of a Bloomberg terminal.</p>
    </div>
    """, unsafe_allow_html=True)

    def chk(v):
        if v == 1:  return '<span style="color:#1D9E75;font-size:15px;">&#10003;</span>'
        if v == 0:  return '<span style="color:#E24B4A;font-size:15px;">&#10007;</span>'
        return '<span style="color:#f59e0b;font-size:13px;">partial</span>'

    matrix_rows = [
        ("Price / mo*",                  ["$29","$17","$25","$21","$30","$2,665"]),
        ("Quant factor model",           [1, 0, 1, "p", "p", 1]),
        ("Live macro regime overlay",    [1, 0, 0, 0, 0, 1]),
        ("Multi-factor conviction score",[1, 0, 1, "p", "p", 1]),
        ("Plain-English signal rationale",[1, 0, "p", "p", "p", 0]),
        ("Hidden-gem detection",         [1, "p", 0, 0, 0, 0]),
        ("Portfolio simulator",          [1, 0, 0, "p", "p", 1]),
        ("Live model portfolio",         [1, "p", "p", 0, 0, 0]),
        ("Daily signal refresh",         [1, 0, 1, "p", "p", 1]),
        ("Full-universe screener",       [1, 0, 1, 1, 1, 1]),
        ("Free tier",                    [1, 0, "p", "p", "p", 0]),
        ("Mobile optimized",             [1, 1, 1, 1, 1, "p"]),
    ]

    cols_h = ["", "QNTM", "Motley Fool", "Seeking Alpha", "Morningstar", "TipRanks", "Bloomberg"]
    col_w  = ["35%", "11%", "11%", "11%", "11%", "11%", "10%"]

    header_html = "".join([
        f'<th style="width:{col_w[i]};padding:8px 6px;font-family:DM Mono,monospace;font-size:13px;'
        f'color:{"#d4a843" if c=="QNTM" else "#9fabc0"};letter-spacing:.06em;'
        f'text-align:{"left" if i==0 else "center"};border-bottom:1px solid rgba(255,255,255,.08);">'
        f'{c}</th>'
        for i,c in enumerate(cols_h)
    ])

    rows_html = ""
    for ri, (label, vals) in enumerate(matrix_rows):
        bg = "rgba(212,168,67,.04)" if ri % 2 == 0 else "transparent"
        row = f'<tr style="background:{bg};">'
        row += f'<td style="padding:8px 6px;font-size:13px;color:#b3bed0;">{label}</td>'
        for ci, v in enumerate(vals):
            is_qntm = ci == 0
            if isinstance(v, str) and v.startswith("$"):
                cell = f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{"#d4a843" if is_qntm else "#8896ac"};">{v}</span>'
            elif v == "p":
                cell = chk("p")
            else:
                cell = chk(v)
            fw = "font-weight:700;" if is_qntm else ""
            row += f'<td style="text-align:center;padding:8px 4px;{fw}">{cell}</td>'
        row += "</tr>"
        rows_html += row

    matrix_html = (
        f'<div style="width:100%;box-sizing:border-box;padding:0 12px;margin-bottom:8px;overflow-x:auto;-webkit-overflow-scrolling:touch;">'
        f'<table style="width:100%;min-width:580px;border-collapse:collapse;background:#0a0b14;'
        f'border:1px solid rgba(255,255,255,.07);border-radius:8px;overflow:hidden;">'
        f'<thead><tr>{header_html}</tr></thead>'
        f'<tbody>{rows_html}</tbody>'
        f'</table>'
        f'<div style="font-size:13px;color:#8896ac;margin-top:8px;padding:0 2px;line-height:1.6;">'
        f'*Monthly equivalent &mdash; most plans bill annually: Motley Fool $199/yr, Seeking Alpha $299/yr, '
        f'Morningstar $249/yr, TipRanks $360/yr (Premium); Bloomberg \u2248$31,980/yr per terminal. '
        f'Features &amp; pricing verified June 2026 from public sources. Partial = limited or higher-tier only.</div>'
        f'</div>'
    )
    st.markdown(matrix_html, unsafe_allow_html=True)

    # ── PRICING ───────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="land-divider" style="margin-top:16px;"></div>
    <div class="land-section">
      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;letter-spacing:.2em;margin-bottom:14px;">&mdash; PRICING</div>
      <h2 style="font-family:'Syne',sans-serif;font-size:clamp(28px,4vw,42px);font-weight:800;
           color:#fff;margin-bottom:10px;line-height:1.1;">
        Two tiers.<br><span style="color:#d4a843;">Both built for serious investors.</span>
      </h2>
      <p style="color:#b3bed0;margin-bottom:36px;">First 50 users get Founding Member access free — unlimited everything.</p>
    </div>
    """, unsafe_allow_html=True)

    def feat_row(text, highlight=False):
        dot = "●" if highlight else "○"
        dc  = "#1D9E75" if highlight else "#8896ac"
        tc  = "#e2e4f0" if highlight else "#9fabc0"
        return f'<div style="display:flex;align-items:flex-start;gap:6px;padding:3px 0;font-size:13px;"><span style="color:{dc};flex-shrink:0;">{dot}</span><span style="color:{tc};">{text}</span></div>'


    def card_style(highlight=False):
        if highlight:
            return "background:rgba(212,168,67,.04);border:2px solid rgba(212,168,67,.5);border-radius:10px;padding:16px 12px;min-width:0;overflow:hidden;"
        return "background:#0e0f1a;border:1px solid rgba(255,255,255,.07);border-radius:10px;padding:16px 12px;min-width:0;overflow:hidden;"

    free_card = f"""
      <div style="{card_style()}">
        <div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:#b3bed0;letter-spacing:.08em;margin-bottom:8px;">FREE</div>
        <div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;color:#e2e4f0;line-height:1;">$0</div>
        <div style="font-size:13px;color:#b3bed0;margin-bottom:14px;margin-top:3px;">forever · no card needed</div>
        <div style="border-top:1px solid rgba(255,255,255,.06);padding-top:12px;">
          {feat_row(f"Screener — top 50 of {_universe_n()}")}
          {feat_row("HIGH / MOD / LOW conviction signals")}
          {feat_row("5-pillar score breakdown")}
          {feat_row("Live macro regime overlay")}
          {feat_row("Top 10 daily picks")}
          {feat_row("Portfolio tracking (10 positions)")}
          {feat_row("Live model portfolio (read only)")}
          {feat_row("Transparent 5-factor model")}
        </div>
      </div>"""

    founding_card = f"""
      <div style="{card_style(True)}">
        <div style="background:#d4a843;color:#000;font-family:Syne,sans-serif;font-size:8px;font-weight:700;letter-spacing:.08em;padding:2px 8px;border-radius:2px;display:inline-block;margin-bottom:8px;">MOST POPULAR</div>
        <div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:#b3bed0;letter-spacing:.08em;margin-bottom:8px;">PRO</div>
        <div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;color:#d4a843;line-height:1;">$29<span style="font-size:14px;font-weight:500;color:#b3bed0;">/mo</span></div>
        <div style="font-size:13px;color:#b3bed0;margin-bottom:14px;margin-top:3px;">first 50 users get it free</div>
        <div style="border-top:1px solid rgba(255,255,255,.06);padding-top:12px;">
          {feat_row("Everything in Free", True)}
          {feat_row(f"Full {_universe_n()}-stock screener", True)}
          {feat_row("Unlimited portfolio positions", True)}
          {feat_row("Hidden Gems detection", True)}
          {feat_row("Portfolio Simulator (risk profiles)", True)}
          {feat_row("Signal change alerts", True)}
          {feat_row("Email notifications", True)}
          {feat_row("Founding member badge", True)}
        </div>
      </div>"""

    st.markdown(
        f'<div style="width:100%;box-sizing:border-box;padding:0 12px;margin-bottom:16px;">'
        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;max-width:760px;margin:0 auto;">'
        f'{free_card}{founding_card}'
        f'</div></div>',
        unsafe_allow_html=True)

    st.markdown(
        '<div style="width:100%;box-sizing:border-box;padding:0 12px;margin-bottom:8px;max-width:760px;margin-left:auto;margin-right:auto;">'
        '<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px;">'
        + _cta_ghost("Start Free →", "?nav=register")
        + _cta_gold("Join Free — Founding Member →", "?nav=register&plan=pro")
        + '</div>'
        + '<div style="text-align:center;margin-top:10px;font-size:13px;color:#d4a843;">'
        '⚡ Founding member pricing — <span style="text-decoration:line-through;color:#8896ac;">$29/mo</span>'
        ' free while spots last &nbsp;·&nbsp; <span style="color:#e2e8f0;">Limited availability</span>'
        '</div></div>',
        unsafe_allow_html=True
    )

    # ── FOOTER ────────────────────────────────────────────────────────────────
    st.markdown("""
    <div class="land-divider" style="margin-top:16px;"></div>
    <div style="background:#080910;padding:48px clamp(16px,4vw,48px) 40px;">
      <div style="max-width:1200px;margin:0 auto;display:flex;justify-content:space-between;
           align-items:flex-start;flex-wrap:wrap;gap:32px;margin-bottom:32px;">
        <div>
          <div style="font-family:'Syne',sans-serif;font-size:20px;font-weight:800;
               color:#e2e4f0;margin-bottom:6px;">Q<span style="color:#d4a843;">NTM</span></div>
          <div style="font-size:14px;color:#b3bed0;line-height:1.7;max-width:280px;">
            Quantitative conviction factor model platform.<br>
            Institutional-grade research for retail investors.
          </div>
        </div>
        <div style="display:flex;gap:48px;flex-wrap:wrap;">
          <div>
            <div style="font-family:'DM Mono',monospace;font-size:13px;color:#9fabc0;letter-spacing:.12em;margin-bottom:12px;">LEGAL</div>
            <div style="font-size:13px;color:#b3bed0;line-height:2.2;">
              <a href="?legal=privacy" style="color:#b3bed0;text-decoration:none;display:block;">Privacy Policy</a>
              <a href="?legal=terms" style="color:#b3bed0;text-decoration:none;display:block;">Terms of Service</a>
              <a href="?legal=billing" style="color:#b3bed0;text-decoration:none;display:block;">Billing & Refunds</a>
              <a href="?legal=disclaimer" style="color:#b3bed0;text-decoration:none;display:block;">Investment Disclaimer</a>
              <a href="?legal=cookies" style="color:#b3bed0;text-decoration:none;display:block;">Cookie Policy</a>
            </div>
          </div>
          <div>
            <div style="font-family:'DM Mono',monospace;font-size:13px;color:#9fabc0;letter-spacing:.12em;margin-bottom:12px;">CONTACT</div>
            <div style="font-size:13px;color:#b3bed0;line-height:2.2;">
              <span style="color:#94a3b8;font-size:13px;font-family:DM Mono,monospace;letter-spacing:.06em;">COMING SOON</span>
            </div>
          </div>
        </div>
      </div>
      <div style="background:rgba(212,168,67,.05);border:1px solid rgba(212,168,67,.15);
           border-radius:8px;padding:18px 22px;margin-bottom:28px;max-width:1200px;margin-left:auto;margin-right:auto;">
        <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;letter-spacing:.12em;margin-bottom:8px;">IMPORTANT DISCLAIMER</div>
        <div style="font-size:13px;color:#9fabc0;line-height:1.8;">
          QNTM is a <strong style="color:#b3bed0;">quantitative research and factor analysis tool</strong>
          for informational and educational purposes only. It does <strong style="color:#b3bed0;">not</strong>
          constitute investment advice, a recommendation to buy or sell any security, or a guarantee of
          future performance. Past model performance does not predict future results. All investments
          involve risk including possible loss of principal. Always consult a qualified financial adviser.
        </div>
      </div>
      <div style="max-width:1200px;margin:0 auto;display:flex;justify-content:space-between;
           align-items:center;flex-wrap:wrap;gap:12px;padding-top:20px;
           border-top:1px solid rgba(255,255,255,.05);">
        <div style="font-size:13px;color:#b3bed0;">&copy; 2025 QNTM. All rights reserved.</div>
        <div style="font-size:13px;color:#b3bed0;">
          Not investment advice &middot; Quantitative research tool only
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    cookie_banner()


def page_auth():
    # Full-page background gradient
    st.markdown("""
    <div style="position:fixed;inset:0;background:radial-gradient(ellipse 70% 50% at 50% -10%,
         rgba(52,211,153,.06) 0%,transparent 65%);pointer-events:none;z-index:0;"></div>
    <style>
    /* Auth tab buttons — full width, equal size, active state highlighted */
    div[data-testid="column"] .stButton > button {
        border-radius: 0 !important;
        border: none !important;
        border-bottom: 2px solid rgba(255,255,255,.08) !important;
        background: transparent !important;
        color: #9fabc0 !important;
        font-size:13px !important;
        font-weight: 700 !important;
        letter-spacing: .04em !important;
        text-transform: uppercase !important;
        height: 44px !important;
        min-height: 44px !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: clip !important;
        padding: 0 8px !important;
    }
    div[data-testid="column"] .stButton > button:hover {
        background: rgba(52,211,153,.05) !important;
        border-bottom-color: rgba(52,211,153,.4) !important;
        color: #b3bed0 !important;
        transform: none !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    div[data-testid="stButton"][data-key="auth_home_btn"] button {
        display:inline-flex !important;
        align-items:center !important;
        gap:4px !important;
        background:rgba(255,255,255,.03) !important;
        border:1px solid rgba(255,255,255,.12) !important;
        border-radius:6px !important;
        color:#b3bed0 !important;
        font-family:Syne,sans-serif !important;
        font-size:13px !important;
        font-weight:700 !important;
        letter-spacing:.04em !important;
        padding:7px 10px !important;
        white-space:nowrap !important;
        text-transform:uppercase !important;
        width:auto !important;
        min-width:0 !important;
        box-shadow:none !important;
        transform:none !important;
    }
    div[data-testid="stButton"][data-key="auth_home_btn"] button:hover {
        background:rgba(255,255,255,.06) !important;
        border-color:rgba(255,255,255,.2) !important;
        color:#e2e8f0 !important;
        transform:none !important;
    }
    </style>
    """, unsafe_allow_html=True)
    # Back button rendered before columns — full width available, no cramping
    if st.button("← Back", key="auth_home_btn"):
        st.session_state.page = "landing"
        st.session_state.signed_out = True
        st.rerun()

    _, col_center, _ = st.columns([1, 4, 1])
    with col_center:
        st.markdown("""
        <div style="text-align:center;padding:24px 0 20px;">
          <div style="font-family:'Syne',sans-serif;font-size:32px;font-weight:800;
               letter-spacing:.15em;color:#e2e4f0;">Q<span style="color:#34d399;">NTM</span></div>
          <div style="font-size:13px;color:#9fabc0;letter-spacing:.2em;margin-top:6px;">
            CONVICTION FACTOR MODEL
          </div>
        </div>
        """, unsafe_allow_html=True)

        # Tab selection — stored in session so it survives reruns
        if "auth_tab" not in st.session_state:
            st.session_state.auth_tab = "signin"

        t1_label = "▶ Sign In" if st.session_state.auth_tab == "signin" else "Sign In"
        t2_label = "▶ Join Free" if st.session_state.auth_tab == "register" else "Join Free"
        tc1, tc2 = st.columns(2)
        with tc1:
            if st.button(t1_label, key="tab_signin_btn", use_container_width=True):
                st.session_state.auth_tab = "signin"
                st.rerun()
        with tc2:
            if st.button(t2_label, key="tab_register_btn", use_container_width=True):
                st.session_state.auth_tab = "register"
                st.rerun()

        if st.session_state.auth_tab == "signin":
            st.markdown('<div style="height:4px;"></div>', unsafe_allow_html=True)

            si_email = st.text_input("Email address", key="si_email",
                                     placeholder="you@example.com")
            si_pass  = st.text_input("Password", type="password", key="si_pass",
                                     placeholder="••••••••")
            st.markdown('<div style="height:10px;"></div>', unsafe_allow_html=True)

            if st.button("Sign In →", key="si_btn", use_container_width=True):
                if not si_email or not si_pass:
                    st.error("Enter your email and password")
                else:
                    with st.spinner("Authenticating..."):
                        res = login_user(si_email, si_pass)
                    if res["success"]:
                        user = res["user"]
                        mfa  = get_user_mfa(user["id"])
                        if mfa.get("mfa_enabled") and mfa.get("totp_secret"):
                            st.session_state.pending_mfa_user   = user
                            st.session_state.pending_mfa_secret = mfa["totp_secret"]
                            go("mfa")
                        else:
                            st.session_state.logged_in    = True
                            st.session_state.user         = user
                            st.session_state.mfa_verified = True
                            st.session_state.scan_results = None
                            st.session_state.show_welcome = True
                            # Only prompt MFA if never offered before
                            if not user.get("mfa_offered"):
                                st.session_state.force_mfa_setup = True
                            # Always persist — signed 30-day token
                            _signed = _sign_token(user["id"], user.get("plan","free"))
                            st.query_params["uid"]  = _signed
                            st.query_params["plan"] = user.get("plan","free")
                            _write_localstorage_token(user["id"], user.get("plan","free"))
                            st.session_state.nav = "screener"

                            go("platform")
                    else:
                        st.error(res.get("error", "Invalid email or password"))

            st.markdown("""
            <div style="text-align:center;margin-top:20px;">
              <span style="font-size:14px;color:#b3bed0;">
                No account? Hit <strong style="color:#34d399;">Join Free</strong> above.
              </span>
            </div>
            """, unsafe_allow_html=True)

            with st.expander("Forgot your password?"):
                fp_email = st.text_input("Your account email", key="fp_email",
                                         placeholder="you@example.com")
                if st.button("Send reset link", key="fp_btn", use_container_width=True):
                    from db import request_password_reset
                    request_password_reset(fp_email)
                    st.success("If an account exists for that email, we've sent a reset link. "
                               "Check your inbox (and spam folder) — the link is valid for 30 minutes.")

        # ── REGISTER ──────────────────────────────────────────────────────────
        else:
            st.markdown('<div style="height:4px;"></div>', unsafe_allow_html=True)

            # Plan selection
            st.markdown("""
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px;">
              <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.08);
                   border-radius:6px;padding:14px;text-align:center;">
                <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;
                     color:#b3bed0;letter-spacing:.08em;margin-bottom:4px;">FREE</div>
                <div style="font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:#e2e4f0;">$0</div>
                <div style="font-size:13px;color:#b3bed0;margin-top:4px;">forever</div>
                <div style="font-size:13px;color:#b3bed0;margin-top:8px;line-height:1.6;">
                  Screener · HIGH/MODERATE/LOW conviction signals<br>Up to 10 portfolio positions<br>Live model portfolio
                </div>
              </div>
              <div style="background:rgba(212,168,67,.05);border:1px solid rgba(212,168,67,.4);
                   border-radius:6px;padding:14px;text-align:center;">
                <div style="background:#d4a843;color:#000;font-size:13px;font-weight:700;
                     letter-spacing:.1em;padding:2px 8px;border-radius:2px;display:inline-block;
                     margin-bottom:4px;">FOUNDING MEMBER</div>
                <div style="font-family:'Syne',sans-serif;font-size:22px;font-weight:800;color:#d4a843;">$0</div>
                <div style="font-size:13px;color:#b3bed0;margin-top:4px;">first 50 users · then $29/mo</div>
                <div style="font-size:13px;color:#b3bed0;margin-top:8px;line-height:1.6;">
                  Unlimited holdings · Hidden Gems<br>Signal alerts · Email notifications<br>Priority support
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            rg_name  = st.text_input("Full name",         key="rg_name",  placeholder="Your name")
            rg_email = st.text_input("Email address",     key="rg_email", placeholder="you@example.com")
            rg_pass  = st.text_input("Password",          key="rg_pass",  placeholder="Min 8 characters",
                                     type="password")
            rg_pass2 = st.text_input("Confirm password",  key="rg_pass2", placeholder="Repeat password",
                                     type="password")
            rg_agree = st.checkbox(
                "I understand QNTM is a quantitative research tool, not investment advice",
                key="rg_agree"
            )
            st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

            if st.button("Create Free Account →", key="rg_btn", use_container_width=True):
                if not rg_agree:
                    st.error("Please acknowledge the disclaimer to continue")
                elif not rg_name or not rg_name.strip():
                    st.error("Enter your full name")
                elif not rg_email or "@" not in rg_email:
                    st.error("Enter a valid email address")
                elif rg_pass != rg_pass2:
                    st.error("Passwords don't match")
                elif len(rg_pass) < 8:
                    st.error("Password must be at least 8 characters")
                else:
                    with st.spinner("Creating account..."):
                        res = register_user(rg_email, rg_pass, rg_name)
                    if res["success"]:
                        try:
                            analytics.capture("signup_completed", distinct_id=rg_email,
                                              props={"plan": "pro" if st.session_state.get("auto_upgrade") else "free"})
                        except Exception:
                            pass
                        # Fire off the email-confirmation link (soft gate)
                        try:
                            request_email_verification(rg_email)
                        except Exception:
                            pass
                        # Auto-upgrade if came from Founding Member CTA
                        if st.session_state.get("auto_upgrade"):
                            upgrade_plan(res["user_id"], "pro")
                            st.session_state.auto_upgrade = False
                            msg = ("✓ Founding Member spot claimed! Full Pro access is active. "
                                   "Check your email to confirm your address, then sign in above.")
                            tag = "🏆 Founding Member"
                        else:
                            msg = ("✓ Account created. We've emailed you a link to confirm your "
                                   "address — then sign in above to continue.")
                            tag = ""
                        st.markdown(f"""
                        <div style="background:rgba(52,211,153,.06);border:1px solid rgba(52,211,153,.25);
                             border-radius:6px;padding:14px 16px;font-size:13px;color:#34d399;margin-top:8px;">
                          {msg}
                          {'<div style="font-size:13px;color:#d4a843;margin-top:4px;">' + tag + ' — unlimited holdings, hidden gems &amp; alerts are live.</div>' if tag else ''}
                        </div>
                        """, unsafe_allow_html=True)
                    else:
                        st.error(res.get("error", "Registration failed — please try again"))

        st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
        st.markdown(DISCLAIMER, unsafe_allow_html=True)

    # Legal footer — 3 rows of 2 cols, always fits any screen width
    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
    r1c1, r1c2 = st.columns(2)
    r2c1, r2c2 = st.columns(2)
    r3c1, r3c2 = st.columns(2)
    link_style = (
        "display:block;text-align:center;font-family:'DM Mono',monospace;"
        "font-size:13px;letter-spacing:.07em;color:#9fabc0;text-decoration:none;"
        "border:1px solid rgba(100,116,139,.2);border-radius:4px;padding:8px 4px;"
    )
    with r1c1:
        st.markdown(f'<a href="?legal=privacy" style="{link_style}">PRIVACY POLICY</a>', unsafe_allow_html=True)
    with r1c2:
        st.markdown(f'<a href="?legal=terms" style="{link_style}">TERMS OF SERVICE</a>', unsafe_allow_html=True)
    with r2c1:
        st.markdown(f'<a href="?legal=billing" style="{link_style}">BILLING &amp; REFUNDS</a>', unsafe_allow_html=True)
    with r2c2:
        st.markdown(f'<a href="?legal=disclaimer" style="{link_style}">INVESTMENT DISCLAIMER</a>', unsafe_allow_html=True)
    with r3c1:
        st.markdown(f'<a href="?legal=cookies" style="{link_style}">COOKIE POLICY</a>', unsafe_allow_html=True)
    with r3c2:
        st.markdown('<div></div>', unsafe_allow_html=True)
    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

    cookie_banner()


# ══════════════════════════════════════════════════════════════════════════════
# MFA PAGE
# ══════════════════════════════════════════════════════════════════════════════
def page_mfa():
    _, col, _ = st.columns([1,2,1])
    with col:
        st.markdown("""
        <div style="text-align:center;padding:80px 0 40px;">
          <div style="font-family:'Syne',sans-serif;font-size:28px;font-weight:800;">
            Q<span style="color:#34d399;">NTM</span>
          </div>
          <div style="margin-top:16px;font-size:36px;">🔐</div>
          <h2 style="font-family:'Syne',sans-serif;font-size:24px;font-weight:700;margin-top:12px;">Two-Factor Auth</h2>
          <p style="color:#b3bed0;margin-top:8px;">Enter the 6-digit code from your authenticator app</p>
        </div>
        """, unsafe_allow_html=True)

        # ── Normal MFA verification ───────────────────────────────────────────
        if not st.session_state.get("mfa_recovery_mode"):
            code = st.text_input("Authentication Code", max_chars=6, placeholder="000000", key="mfa_code")
            if st.button("Verify & Enter →", key="mfa_verify", use_container_width=True):
                if verify_totp(st.session_state.pending_mfa_secret, code):
                    user = st.session_state.pending_mfa_user
                    st.session_state.logged_in    = True
                    st.session_state.user         = user
                    st.session_state.mfa_verified = True
                    st.session_state.show_welcome = True
                    # Always persist — signed 30-day token
                    _signed = _sign_token(user["id"], user.get("plan","free"))
                    st.query_params["uid"]  = _signed
                    st.query_params["plan"] = user.get("plan","free")
                    _write_localstorage_token(user["id"], user.get("plan","free"))
                    st.session_state.nav = "screener"

                    go("platform")
                else:
                    st.error("Invalid code — check your app and try again")

            st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
            st.markdown(_back_btn("?nav=signin", "← Back to Sign In"), unsafe_allow_html=True)

            # Recovery option
            st.markdown("""
            <div style="margin-top:24px;padding-top:20px;border-top:1px solid rgba(255,255,255,.06);
                 text-align:center;">
              <p style="font-size:14px;color:#b3bed0;">Lost access to your authenticator app?</p>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Reset 2FA with password →", key="mfa_recovery_btn", use_container_width=True):
                st.session_state.mfa_recovery_mode = True
                st.rerun()

        # ── MFA Recovery — verify password, then re-enroll ───────────────────
        else:
            st.markdown("""
            <div style="background:rgba(212,168,67,.06);border:1px solid rgba(212,168,67,.2);
                 border-radius:6px;padding:12px 16px;margin-bottom:16px;font-size:14px;color:#b3bed0;">
              Verify your password to disable 2FA and set up a new authenticator.
            </div>
            """, unsafe_allow_html=True)

            recovery_pw = st.text_input("Your Password", type="password", key="mfa_recovery_pw")

            if st.button("Verify Password & Reset 2FA", key="mfa_recovery_verify", use_container_width=True):
                if recovery_pw:
                    # Re-authenticate with password
                    user_data = st.session_state.get("pending_mfa_user", {})
                    email     = user_data.get("email", "")
                    result    = login_user(email, recovery_pw)
                    if result.get("success"):
                        # Disable MFA so they can re-enroll
                        disable_mfa(user_data.get("id",""))
                        # Log them in
                        st.session_state.logged_in          = True
                        st.session_state.user               = user_data
                        st.session_state.mfa_verified       = True
                        st.session_state.mfa_recovery_mode  = False
                        st.session_state.show_mfa_setup     = True   # trigger re-enroll flow
                        st.success("2FA reset. Setting up new authenticator...")
                        import time; time.sleep(1)
                        st.session_state.nav = "screener"

                        go("platform")
                    else:
                        st.error("Incorrect password — try again")
                else:
                    st.warning("Enter your password to continue")

            st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
            if st.button("← Back", key="mfa_recovery_back"):
                st.session_state.mfa_recovery_mode = False
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# PLATFORM — TOP NAV
# ══════════════════════════════════════════════════════════════════════════════
HELP_CONTENT = {
    "screener": ("The Screener",
        "Your home base — every stock in the universe, ranked by conviction.",
        [("Conviction ranking", "Stocks are sorted High → Low by the model's composite score. Start at the top and work down."),
         ("Open any stock", "Tap a row to expand its plain-English rationale and the five pillar scores behind the signal."),
         ("Macro overlay", "The strip up top shows the current market regime and how it's shaping the scores."),
         ("Search & filter", "Narrow by ticker or sector to find what you're after."),
         ("Add to Watchlist", "Star a stock to keep an eye on it from the Watchlist page.")]),
    "watchlist": ("Your Watchlist",
        "The names you're following, organized however you like.",
        [("Multiple lists", "Keep several separate watchlists — group names by theme, sector, or whatever suits you, and switch between them."),
         ("Tracked vs. SPY", "Each stock shows a mini chart of its move since you added it, against the S&P 500."),
         ("Live conviction", "Scores move with the model, so you see when a name strengthens or weakens."),
         ("Remove anytime", "Unstar a stock to drop it from the list.")]),
    "gems": ("Hidden Gems",
        "Strong-scoring names that fly under Wall Street's radar.",
        [("What qualifies", "High conviction plus light analyst coverage and mid- or small-cap size — quality without the crowd."),
         ("Since flagged", "The mini chart tracks each gem from the date the model first flagged it."),
         ("Regime-aware", "In high-volatility regimes the bar rises, so only the strongest names surface.")]),
    "portfolio": ("Your Portfolio",
        "See how the names you hold rate against the model.",
        [("Add holdings", "Enter the stocks you hold to see each one's current conviction."),
         ("Conviction check", "Spot at a glance which of your names the model rates High, Moderate, or Low."),
         ("Track over time", "Watch how your set moves across different periods.")]),
    "simulator": ("Portfolio Simulator",
        "A what-if tool for exploring the model.",
        [("Build a mix", "Pick stocks and weights to assemble a hypothetical set."),
         ("Score it", "See the blended conviction of your selection."),
         ("Hypothetical only", "This is a sandbox for exploring the model — nothing here is a recommendation.")]),
    "model_portfolio": ("Portfolio & Track Record",
        "The model's own rules-based portfolio, marked daily.",
        [("Live equity curve", "Real performance since inception, charted against the S&P 500."),
         ("Active positions", "Every open position the model holds, with entry and current marks."),
         ("Rules-based", "Entries and exits follow fixed model rules — no discretion, no hindsight.")]),
    "alerts": ("Alerts",
        "Get a heads-up when conviction changes.",
        [("Conviction shifts", "A flag when a name you follow crosses a conviction threshold."),
         ("Stay current", "Check in here for the latest model-driven changes.")]),
    "account": ("Account",
        "Your plan and settings.",
        [("Plan & access", "See your current plan and what's included."),
         ("Security", "Manage sign-in and two-factor settings.")]),
    "methodology": ("How QNTM Works",
        "The full methodology, in plain English.",
        [("Start here", "The Getting Started section walks you through where to go and what each page does."),
         ("The model", "How the five-pillar score and the macro overlay are built."),
         ("What it is — and isn't", "QNTM is a research tool: quantitative rankings, not advice.")]),
}


def _render_verify_banner():
    """Soft email-verification nag shown on in-app pages until confirmed.
    Non-blocking: the user keeps full access; this only reminds and offers a
    resend. Reads the cached session flag; while still unverified it confirms
    live (cheap, and stops the moment the address is confirmed)."""
    u = st.session_state.get("user")
    if not u:
        return
    if u.get("email_verified") is not True:
        try:
            if is_email_verified(u.get("id")):
                u["email_verified"] = True
        except Exception:
            return  # transient read error — don't nag
    if u.get("email_verified") is True:
        return
    import html as _html
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;flex-wrap:wrap;'
        'background:rgba(212,168,67,.08);border:1px solid rgba(212,168,67,.30);'
        'border-radius:10px;padding:11px 15px;margin:0 0 12px;">'
        '<span style="font-size:15px;line-height:1;">✉️</span>'
        '<span style="font-family:Inter,sans-serif;font-size:13.5px;color:#e7d6a8;line-height:1.45;">'
        'Please confirm your email address — we sent a link to '
        f'<b>{_html.escape(u.get("email",""))}</b>. Check your inbox or spam folder.</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    if st.button("Resend confirmation email", key="resend_verify_btn"):
        try:
            _vr = request_email_verification(u.get("email", ""))
            if _vr.get("delivered"):
                st.success("Sent — check your inbox (and spam folder).")
            else:
                st.error("We couldn't send the confirmation email right now. "
                         "Please try again shortly, or contact hello@qntm.live "
                         "if it keeps failing.")
        except Exception:
            st.error("Couldn't send right now. Please try again in a moment.")


def _render_help_popup(page_key):
    """Per-page help popup. Pure-CSS checkbox toggle (no JS): the '?' button
    opens it, the ✕ or a click on the backdrop closes it."""
    content = HELP_CONTENT.get(page_key)
    if not content:
        return
    title, intro, items = content
    items_html = "".join(
        '<div style="margin-bottom:11px;">'
        f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:700;color:#34d399;'
        f'letter-spacing:.02em;">{t}</div>'
        f'<div style="font-family:Inter,sans-serif;font-size:13px;color:#cbd5e1;line-height:1.45;'
        f'margin-top:2px;">{d}</div></div>'
        for t, d in items
    )
    cb = f"qntm-help-{page_key}"
    st.markdown(
        "<style>"
        ".qntm-help-cb{display:none;}"
        ".qntm-help-fab{position:fixed;bottom:22px;right:22px;z-index:1400;width:44px;height:44px;"
        "border-radius:50%;display:flex;align-items:center;justify-content:center;"
        "background:rgba(52,211,153,.14);border:1px solid rgba(52,211,153,.5);color:#34d399;"
        "font-family:Syne,sans-serif;font-size:22px;font-weight:800;cursor:pointer;"
        "box-shadow:0 6px 20px rgba(0,0,0,.45);transition:background .15s ease;user-select:none;}"
        ".qntm-help-fab:hover{background:rgba(52,211,153,.26);}"
        ".qntm-help-ov{display:none;position:fixed;inset:0;z-index:1500;background:rgba(2,4,8,.72);"
        "backdrop-filter:blur(3px);-webkit-backdrop-filter:blur(3px);}"
        ".qntm-help-cb:checked ~ .qntm-help-ov{display:block;}"
        ".qntm-help-modal{display:none;position:fixed;top:50%;left:50%;transform:translate(-50%,-50%);"
        "z-index:1501;width:min(440px,92vw);max-height:82vh;overflow-y:auto;"
        "background:rgba(10,13,20,.99);border:1px solid rgba(52,211,153,.28);border-radius:16px;"
        "box-shadow:0 24px 64px rgba(0,0,0,.8);}"
        ".qntm-help-cb:checked ~ .qntm-help-modal{display:block;}"
        ".qntm-help-x{position:absolute;top:14px;right:16px;cursor:pointer;color:#94a3b8;font-size:18px;"
        "width:28px;height:28px;display:flex;align-items:center;justify-content:center;"
        "border-radius:6px;text-decoration:none;}"
        ".qntm-help-x:hover{background:rgba(255,255,255,.08);color:#e2e8f0;}"
        "</style>"
        f'<input type="checkbox" id="{cb}" class="qntm-help-cb">'
        f'<label for="{cb}" class="qntm-help-fab" title="How this page works">?</label>'
        f'<label for="{cb}" class="qntm-help-ov"></label>'
        f'<div class="qntm-help-modal"><div style="padding:24px 26px;position:relative;">'
        f'<label for="{cb}" class="qntm-help-x">✕</label>'
        f'<div style="font-family:Syne,sans-serif;font-size:19px;font-weight:800;color:#e2e8f0;'
        f'margin-bottom:6px;padding-right:30px;">{title}</div>'
        f'<div style="font-family:Inter,sans-serif;font-size:13px;color:#b3bed0;line-height:1.5;'
        f'margin-bottom:16px;">{intro}</div>'
        f'{items_html}'
        '</div></div>',
        unsafe_allow_html=True,
    )


def platform_nav():
    user  = st.session_state.user or {}
    plan  = user.get("plan","free")
    n_count = get_unread_count(uid()) if plan in ("pro","institutional") else 0
    plan_color = "#00ff87" if plan in ("pro","institutional") else "#b3bed0"
    plan_rgb = "0,255,135" if plan=="pro" else "249,115,22" if plan=="institutional" else "148,163,184"
    display_name = (user.get("full_name") or "").split()[0] if user.get("full_name") else ""
    if not display_name:
        em = user.get("email","")
        display_name = em[:14] + ("..." if len(em) > 14 else "")

    cur_nav = st.session_state.get("nav","screener")
    nav_items = [
        ("screener",        "📊", "Screener"),
        ("watchlist",       "★",  "Watchlist"),
        ("gems",            "💎", "Hidden Gems"),
        ("portfolio",       "💼", "Portfolio"),
        ("simulator",       "🧮", "Simulator"),
        ("model_portfolio", "🏆", "Track Record"),
        ("alerts",          "🔔", "Alerts"),
        ("account",         "⚙️", "Account"),
        ("methodology",     "📖", "How It Works"),
    ]
    try:
        if analytics.is_admin():
            nav_items.append(("analytics", "📊", "Analytics"))
    except Exception:
        pass
    cur_em    = next((e for k,e,l in nav_items if k==cur_nav), "📊")
    cur_label = next((l for k,e,l in nav_items if k==cur_nav), "Screener")

    # Session params to preserve across navigation
    # Always read uid from session state — query params may be empty after pop
    _uid_val  = (st.session_state.user or {}).get("id", "")
    _plan_val = user.get("plan","free")
    qp_suffix = f"&plan={_plan_val}&ck=1"
    if _uid_val:
        qp_suffix = f"&uid={_uid_val}" + qp_suffix

    # Build the 3-col grid of box buttons
    grid_html = '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;padding:14px 16px 16px;">'
    for key, em, label in nav_items:
        href = f"?qnav={key}{qp_suffix}"
        if key == cur_nav:
            btn_style = (
                'display:flex;flex-direction:column;align-items:center;justify-content:center;'
                'gap:6px;padding:14px 8px;text-decoration:none;border-radius:8px;'
                'background:linear-gradient(135deg,rgba(52,211,153,.14),rgba(52,211,153,.04));'
                'border:1px solid rgba(52,211,153,.5);'
                'box-shadow:0 0 12px rgba(52,211,153,.1);'
            )
            em_style  = 'font-size:20px;line-height:1;'
            lbl_style = ('font-family:Syne,sans-serif;font-size:10px;font-weight:700;'
                         'letter-spacing:.02em;text-transform:uppercase;color:#34d399;'
                         'line-height:1.2;text-align:center;white-space:normal;word-break:break-word;'
                         'display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;max-width:100%;')
        else:
            btn_style = (
                'display:flex;flex-direction:column;align-items:center;justify-content:center;'
                'gap:6px;padding:14px 8px;text-decoration:none;border-radius:8px;'
                'background:rgba(255,255,255,.07);'
                'border:1px solid rgba(255,255,255,.15);'
                'transition:all .18s ease;'
            )
            em_style  = 'font-size:22px;line-height:1;'
            lbl_style = ('font-family:Syne,sans-serif;font-size:10px;font-weight:700;'
                         'letter-spacing:.02em;text-transform:uppercase;color:#b3bed0;'
                         'line-height:1.2;text-align:center;white-space:normal;word-break:break-word;'
                         'display:-webkit-box;-webkit-line-clamp:2;-webkit-box-orient:vertical;overflow:hidden;max-width:100%;')

        badge = (
            f'<span style="position:absolute;top:6px;right:6px;background:#ef4444;color:#fff;'
            f'border-radius:50%;width:14px;height:14px;display:flex;align-items:center;'
            f'justify-content:center;font-size:8px;font-weight:700;">{n_count}</span>'
        ) if (key == "alerts" and n_count > 0) else ""

        grid_html += (
            f'<a href="{href}" target="_self" style="position:relative;min-width:0;overflow:hidden;{btn_style}">'
            f'<span style="{em_style}">{em}</span>'
            f'<span style="{lbl_style}">{label}</span>'
            f'{badge}</a>'
        )

    # Sign out button
    grid_html += (
        f'<a href="?qnav=signout" target="_self" style="'
        f'display:flex;flex-direction:column;align-items:center;justify-content:center;'
        f'gap:6px;padding:14px 8px;text-decoration:none;border-radius:8px;'
        f'background:linear-gradient(135deg,rgba(239,68,68,.08),rgba(239,68,68,.02));'
        f'border:1px solid rgba(239,68,68,.2);">'
        f'<span style="font-size:20px;line-height:1;opacity:.7;">🚪</span>'
        f'<span style="font-family:Syne,sans-serif;font-size:13px;font-weight:600;'
        f'letter-spacing:.05em;text-transform:uppercase;color:#ef4444;">Sign Out</span>'
        f'</a>'
    )
    grid_html += '</div>'

    _notif_url = f'?qnav=alerts&uid={uid()}&plan={plan}&ck=1'
    notif_dot = (
        f'<a href="{_notif_url}" target="_self" style="'
        f'background:#ef4444;color:#fff;border-radius:50%;width:18px;height:18px;'
        f'display:inline-flex;align-items:center;justify-content:center;font-size:13px;font-weight:700;'
        f'text-decoration:none;cursor:pointer;">'
        f'{n_count}</a>'
    ) if n_count > 0 else ""

    is_dev = os.getenv("ENVIRONMENT") == "dev"
    dd_top = "88px" if is_dev else "56px"

    nav_html = (
        '<style>'
        '#qntm-toggle{display:none;}'
        '#qntm-dd{'
        f'position:fixed;top:{dd_top};left:50%;transform:translateX(-50%);width:340px;'
        'background:rgba(7,10,18,.99);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);'
        'border:1px solid rgba(255,255,255,.1);border-radius:12px;z-index:1000;'
        'max-height:0;overflow:hidden;opacity:0;pointer-events:none;'
        'transition:max-height .3s cubic-bezier(.4,0,.2,1),opacity .22s ease;'
        'box-shadow:0 24px 64px rgba(0,0,0,.8);}'
        '#qntm-toggle:checked ~ #qntm-dd{'
        'max-height:600px;opacity:1;pointer-events:all;}'
        '#qntm-ov{display:none;position:fixed;inset:0;z-index:999;}'
        '#qntm-toggle:checked ~ #qntm-ov{display:block;}'
        '.qntm-menu-trigger{'
        'display:flex;align-items:center;gap:8px;cursor:pointer;'
        'background:rgba(255,255,255,.05);border:1px solid rgba(255,255,255,.12);'
        'border-radius:8px;padding:8px 14px;'
        'font-family:Syne,sans-serif;font-size:13px;font-weight:600;color:#e2e8f0;'
        'transition:border-color .2s,background .2s;user-select:none;'
        'min-width:170px;justify-content:space-between;}'
        '#qntm-toggle:checked ~ div label.qntm-menu-trigger{'
        'border-color:rgba(52,211,153,.4);background:rgba(52,211,153,.06);}'
        '.qntm-chevron{width:14px;height:14px;opacity:.5;transition:transform .2s;flex-shrink:0;}'
        '#qntm-toggle:checked ~ div label.qntm-menu-trigger .qntm-chevron{'
        'transform:rotate(180deg);}'
        'a[href*="qnav"]:hover{background:linear-gradient(135deg,rgba(52,211,153,.1),rgba(52,211,153,.03))!important;'
        'border-color:rgba(52,211,153,.35)!important;}'
        '</style>'
        '<input type="checkbox" id="qntm-toggle">'
        f'<div id="qntm-dd">'
        '<div style="padding:12px 16px 8px;border-bottom:1px solid rgba(255,255,255,.06);">'
        '<span style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;letter-spacing:.14em;">MENU</span>'
        '</div>'
        + grid_html +
        '</div>'
        '<label for="qntm-toggle" id="qntm-ov"></label>'
        '<div style="background:rgba(2,4,8,.97);backdrop-filter:blur(12px);'
        'border-bottom:1px solid rgba(255,255,255,.07);'
        'padding:0 20px;height:56px;display:flex;align-items:center;justify-content:space-between;">'
        '<span style="font-family:Syne,sans-serif;font-size:20px;font-weight:800;'
        'letter-spacing:.15em;color:#e2e8f0;">Q<span style="color:#34d399;">NTM</span></span>'
        '<label for="qntm-toggle" class="qntm-menu-trigger">'
        '<span>☰  MENU</span>'
        '<svg class="qntm-chevron" viewBox="0 0 24 24" fill="none" stroke="#b3bed0" stroke-width="2.5">'
        '<polyline points="6 9 12 15 18 9"/></svg>'
        '</label>'
        '<div style="display:flex;align-items:center;gap:10px;">'
        + notif_dot
        + f'<a href="?qnav=account&uid={user.get("id","")}&plan={plan}&ck=1&_n=account&acct_focus=billing" '
        f'target="_self" title="Manage your plan" style="text-decoration:none;">'
        f'<span style="background:rgba({plan_rgb},.15);color:{plan_color};'
        f'border:1px solid {plan_color}44;border-radius:4px;padding:3px 9px;'
        f'font-size:13px;font-weight:700;letter-spacing:.1em;font-family:Syne,sans-serif;cursor:pointer;">'
        f'{plan.upper()}</span></a>'
        f'<span style="font-size:13px;color:#9fabc0;font-family:DM Mono,monospace;">{display_name}</span>'
        '</div></div>'
    )

    st.markdown(nav_html, unsafe_allow_html=True)

    # ── Dark navigation curtain ───────────────────────────────────────────────
    # Internal nav is via <a href="?qnav=..."> which triggers a full page reload.
    # Browsers show a white gap during that reload before the dark theme paints.
    # This drops a branded dark overlay the instant a nav link is clicked, so the
    # browser holds a dark frame (not white) through the reload. Runs inline in
    # the main document via qntm_html, so `document` is the top window.
    qntm_html(
        "<script>(function(){"
        "try{var doc=window.parent.document;}catch(e){return;}"
        "if(!doc||doc.getElementById('qntm-curtain'))return;"
        "var c=doc.createElement('div');c.id='qntm-curtain';"
        "c.style.cssText='position:fixed;inset:0;z-index:2147483647;background:#0a0b14;"
        "display:none;flex-direction:column;align-items:center;justify-content:center;gap:18px;';"
        "c.innerHTML='<div style=\\\"width:48px;height:48px;border-radius:50%;"
        "border:3px solid rgba(52,211,153,.18);border-top-color:#34d399;"
        "animation:qntmspin .8s linear infinite;\\\"></div>"
        "<div style=\\\"font-family:Syne,sans-serif;font-size:24px;font-weight:800;"
        "letter-spacing:.2em;color:#e2e8f0;\\\">Q<span style=\\\"color:#34d399;\\\">NTM</span></div>"
        "<div style=\\\"font-family:DM Mono,monospace;font-size:11px;letter-spacing:.18em;"
        "color:#94a3b8;text-transform:uppercase;\\\">Loading</div>';"
        "doc.body.appendChild(c);"
        "if(!doc.getElementById('qntm-curtain-kf')){var s=doc.createElement('style');"
        "s.id='qntm-curtain-kf';s.textContent='@keyframes qntmspin{to{transform:rotate(360deg)}}';"
        "doc.head.appendChild(s);}"
        "doc.addEventListener('click',function(e){"
        "var a=e.target.closest&&e.target.closest('a[href*=\\\"qnav\\\"],a[href*=\\\"legal=\\\"],a[href*=\\\"nav=\\\"]');"
        "if(a){var h=a.getAttribute('href')||'';if(h&&h.charAt(0)!=='#'){c.style.display='flex';}}"
        "},true);"
        "window.addEventListener('pageshow',function(){c.style.display='none';});"
        "})();</script>",
        height=0,
    )

    _render_help_popup(cur_nav)

    _render_verify_banner()

    # First-load nudge toward the per-page helper, shown once after login on the
    # Screener (same show_welcome flag; clears on navigation, auto-fades after 9s).
    if cur_nav == "screener" and st.session_state.get("show_welcome"):
        st.markdown(
            '<div style="position:fixed;bottom:76px;right:14px;z-index:1402;max-width:232px;'
            'padding:11px 14px;border-radius:12px;background:rgba(10,13,20,.99);'
            'border:1px solid rgba(52,211,153,.4);box-shadow:0 8px 24px rgba(0,0,0,.55);'
            'font-family:Inter,sans-serif;font-size:12.5px;color:#cbd5e1;line-height:1.4;'
            'animation:helphintout .5s ease 20s forwards;">'
            '<b style="color:#34d399;">Need a hand?</b> Tap the '
            '<span style="color:#34d399;font-weight:700;">?</span> on any page for a quick guide to '
            'what\'s on it.'
            '<span style="position:absolute;bottom:-7px;right:18px;width:12px;height:12px;'
            'background:rgba(10,13,20,.99);border-right:1px solid rgba(52,211,153,.4);'
            'border-bottom:1px solid rgba(52,211,153,.4);transform:rotate(45deg);"></span>'
            '</div>'
            '<style>@keyframes helphintout{to{opacity:0;visibility:hidden;}}</style>',
            unsafe_allow_html=True,
        )


@st.fragment
def _render_screener_cards(filtered, gem_tickers, filter_key):
    """Pagination + card rendering for the screener, isolated in a fragment so
    Prev/Next (and re-renders) re-run ONLY this block via st.rerun(scope=
    "fragment") instead of the entire 10k-line app. Inputs are passed in and
    preserved across fragment reruns; the filter controls live OUTSIDE the
    fragment, so changing a filter triggers a normal full rerun that re-invokes
    this with a fresh `filtered` list."""
    _PAGE_SIZE = 50
    _total_pages = max(1, (len(filtered) + _PAGE_SIZE - 1) // _PAGE_SIZE)
    if "_fu_page" not in st.session_state:
        st.session_state._fu_page = 0
    # Reset to page 0 when the filter selection changes
    if st.session_state.get("_fu_filter_key") != filter_key:
        st.session_state._fu_page = 0
        st.session_state._fu_filter_key = filter_key
    _page = min(st.session_state._fu_page, _total_pages - 1)
    _page_items = filtered[_page * _PAGE_SIZE:(_page + 1) * _PAGE_SIZE]

    # Prev / Next — scoped reruns keep paging snappy
    if _total_pages > 1:
        _pn1, _pn2, _pn3 = st.columns([1, 2, 1])
        with _pn1:
            if _page > 0 and st.button("← Prev", key="fu_prev", use_container_width=True):
                st.session_state._fu_page = _page - 1
                st.rerun(scope="fragment")
        with _pn2:
            st.markdown(f'<div style="text-align:center;font-family:DM Mono,monospace;'
                        f'font-size:13px;color:#8896ac;padding:8px 0;">'
                        f'Page {_page+1} of {_total_pages}</div>', unsafe_allow_html=True)
        with _pn3:
            if _page < _total_pages - 1 and st.button("Next →", key="fu_next", use_container_width=True):
                st.session_state._fu_page = _page + 1
                st.rerun(scope="fragment")

    _show_sparkline = len(_page_items) <= 20
    _ci_cache = st.session_state.get("company_info_cache", {})
    _fu_prog = st.progress(0, text="Loading cards...")
    _fu_wl_now = {w["ticker"] for w in get_watchlist(uid())} if uid() else set()
    _fu_html = ""
    for _fu_i, r in enumerate(_page_items):
        ci = _ci_cache.get(r["ticker"])
        _extra = ""
        if _show_sparkline:
            _sc = float(r.get("adj_composite", r.get("composite", 50)) or 50)
            _ch = signal_history_chart(r["ticker"], _sc)
            if _ch:
                _extra = _ch
        _fu_html += build_card_html(r, nav="screener",
                                    is_gem=(r["ticker"] in gem_tickers),
                                    company_info=ci, in_list=_fu_wl_now,
                                    extra_detail=_extra)
        _fu_prog.progress(int((_fu_i+1)/len(_page_items)*100), text=f"Loading {_fu_i+1}/{len(_page_items)}...")
    _fu_prog.empty()
    render_cards_batch(_fu_html)


def page_screener():
    _pin_nav("screener")
    from model_engine import (MACRO_EVENT_INFO, score_stock, fetch_price_data,
                               SECTORS as ALL_SECTORS, fetch_macro_overlay, apply_macro_overlay)

    # First-run welcome card — shown once after login, cleared on any navigation
    # (a full-page reload starts a fresh session, so the flag naturally resets).
    if st.session_state.get("show_welcome"):
        _wu = st.session_state.user or {}
        _wcta = f"?qnav=screener&uid={_wu.get('id','')}&plan={_wu.get('plan','free')}&ck=1"
        st.markdown(
            '<div style="margin:14px 32px 0;padding:22px 26px;border-radius:14px;'
            'background:linear-gradient(135deg,rgba(52,211,153,.08),rgba(52,211,153,.02));'
            'border:1px solid rgba(52,211,153,.28);">'
            '<div style="font-family:Syne,sans-serif;font-size:20px;font-weight:800;'
            'letter-spacing:.04em;color:#e2e8f0;margin-bottom:4px;">Welcome to '
            '<span style="color:#34d399;">QNTM</span></div>'
            '<div style="font-family:Inter,sans-serif;font-size:14px;color:#b3bed0;'
            'line-height:1.5;margin-bottom:14px;">Here\'s the quickest way to find your footing:</div>'
            '<div style="display:grid;gap:8px;font-family:Inter,sans-serif;font-size:13px;'
            'color:#cbd5e1;line-height:1.45;">'
            '<div><b style="color:#34d399;">1 · Screener</b> — you\'re here. Every stock ranked by '
            'conviction; start at the top and work down.</div>'
            '<div><b style="color:#34d399;">2 · Open a stock</b> — tap any row for its plain-English '
            'rationale and the five pillar scores behind the signal.</div>'
            '<div><b style="color:#34d399;">3 · Watchlist</b> — star names to track them against the S&amp;P 500.</div>'
            '<div><b style="color:#34d399;">4 · Hidden Gems</b> — strong scorers flying under Wall Street\'s radar.</div>'
            '<div><b style="color:#34d399;">5 · Simulator &amp; Track Record</b> — test an allocation, '
            'then see the model\'s live performance.</div>'
            '</div>'
            f'<a href="{_wcta}" target="_self" style="display:inline-block;margin-top:16px;'
            'padding:10px 22px;border-radius:8px;text-decoration:none;background:#34d399;color:#06070f;'
            'font-family:Syne,sans-serif;font-size:13px;font-weight:700;letter-spacing:.06em;'
            'text-transform:uppercase;">Start exploring →</a>'
            '<div style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;margin-top:10px;">'
            'Full walkthrough anytime under How It Works.</div>'
            '</div>',
            unsafe_allow_html=True,
        )

    # Compact header — title + refresh inline, no wasted vertical space.
    # The "updated" date only changes once nightly, so cache it per session
    # rather than querying signal_log on every screener render.
    _fresh_html = ""
    _cached_fresh = st.session_state.get("_hdr_fresh_date")
    if _cached_fresh:
        _fresh_html = f' · updated {_cached_fresh}'
    else:
        try:
            from data_refresh import _get_supabase as _hdr_sb
            from datetime import datetime, timezone, timedelta
            _sb_h = _hdr_sb()
            if _sb_h:
                _hr = _sb_h.table("signal_log").select("signal_date").order("signal_date", desc=True).limit(1).execute()
                if _hr.data:
                    st.session_state._hdr_fresh_date = _hr.data[0]["signal_date"]
                    _fresh_html = f' · updated {_hr.data[0]["signal_date"]}'
        except Exception:
            pass
    st.markdown(
        f'<div style="padding:10px 32px 4px;">'
        f'<span style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#e2e8f0;">📊 Market Screener</span>'
        f'<div style="font-size:13px;color:#94a3b8;margin-top:2px;">{_universe_n()} stocks · 5-pillar quant · macro overlay{_fresh_html}</div>'
        f'</div>',
        unsafe_allow_html=True
    )
    data_freshness_banner()
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    # ── Search box with live suggestions ───────────────────────────────────────
    _ac_pick = st.query_params.get("ac_pick", "")
    if _ac_pick:
        st.session_state.screener_search_val = _ac_pick.upper()
        st.query_params.pop("ac_pick", None)

    _sq_default = st.session_state.get("screener_search_val", "")
    _sq_param = st.query_params.get("sq", "")
    if _sq_param and not _sq_default:
        _sq_default = _sq_param
        st.session_state.screener_search_val = _sq_param
        st.query_params.pop("sq", None)

    # Styled search input
    st.markdown("""
    <style>
    div[data-testid="stTextInput"][data-key="screener_search"] input {
        background: rgba(255,255,255,.04) !important;
        border: 1px solid rgba(52,211,153,.3) !important;
        border-radius: 8px !important; color: #e2e8f0 !important;
        font-size: 15px !important; padding: 13px 20px !important;
        height: 50px !important; transition: border-color .2s, box-shadow .2s !important;
    }
    div[data-testid="stTextInput"][data-key="screener_search"] input:focus {
        border-color: rgba(52,211,153,.6) !important;
        box-shadow: 0 0 0 3px rgba(52,211,153,.08) !important; outline: none !important;
    }
    div[data-testid="stTextInput"][data-key="screener_search"] input::placeholder {
        color: #94a3b8 !important;
    }
    /* Suggestion buttons */
    div[data-testid="stHorizontalBlock"] .stButton > button {
        background: #0d1117 !important;
        border: 1px solid rgba(255,255,255,.1) !important;
        border-radius: 6px !important; color: #e2e8f0 !important;
        font-family: Syne, sans-serif !important; font-size: 13px !important;
        font-weight: 700 !important; padding: 8px 14px !important;
        text-align: left !important; width: 100% !important;
        transition: background .15s !important;
    }
    div[data-testid="stHorizontalBlock"] .stButton > button:hover {
        background: rgba(52,211,153,.07) !important;
        border-color: rgba(52,211,153,.25) !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # Search input with suggestions
    def _on_search_change():
        val = st.session_state.get("screener_search_raw","").strip().upper()
        st.session_state._search_live = val
        # If the submitted value is an exact ticker in the universe, surface its
        # score card right away — Enter now behaves like clicking a suggestion.
        # Otherwise clear the selected stock so suggestions show while typing.
        if val in SECTORS:
            st.session_state.screener_search_val = val
        elif st.session_state.get("screener_search_val","") != val:
            st.session_state.screener_search_val = ""

    # When suggestion selected, delete widget key so it reinits with new value
    if "_sug_just_picked" in st.session_state:
        _picked = st.session_state.pop("_sug_just_picked")
        if "screener_search_raw" in st.session_state:
            del st.session_state["screener_search_raw"]
        _sq_default = _picked
        st.session_state._search_live = _picked

    if "screener_search_raw" not in st.session_state:
        st.session_state.screener_search_raw = _sq_default
    if "_search_live" not in st.session_state:
        st.session_state._search_live = _sq_default.strip().upper()

    st.text_input(
        "Search ticker",
        value=st.session_state.screener_search_raw,
        placeholder="🔍  Search ticker or company — AAPL, Tesla, Nvidia...",
        key="screener_search_raw",
        label_visibility="collapsed",
        on_change=_on_search_change
    )

    _live_q = st.session_state.get("_search_live","").strip().upper()

    # Suggestions — standard ticker/company search (shared platform-wide).
    # Clicking one sets screener_search_val, which renders the score card below.
    if _live_q:
        _render_suggestions(_live_q, "sug", "screener_search_val")

    # Only show stock card when user explicitly clicks a suggestion
    # screener_search_val is set by suggestion click OR ac_pick URL param — not by typing
    search_ticker = st.session_state.get("screener_search_val", "").strip().upper()
    if search_ticker:
        _rl = st.session_state.get("recent_searches", [])
        if search_ticker not in _rl:
            st.session_state.recent_searches = ([search_ticker] + [r for r in _rl if r != search_ticker])[:5]

    if search_ticker:
        # Resolve company name → ticker first
        resolved_tk, resolved_name = resolve_ticker(search_ticker)
        display_query = f"{resolved_name} ({resolved_tk})" if resolved_name and resolved_name != resolved_tk else resolved_tk

        st.markdown(f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;margin:6px 0 6px;">SCORE FOR {display_query}</div>', unsafe_allow_html=True)
        _sr_ok = None
        with st.spinner(f"Scoring {resolved_tk}..."):
            try:
                price_data = fetch_price_data([resolved_tk], period="1y")
                hist = price_data.get(resolved_tk, [])
                if not hist or len(hist) < 10:
                    st.markdown(
                        f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);'
                        f'border-radius:8px;padding:20px 24px;">'
                        f'<div style="font-family:Syne,sans-serif;font-size:15px;font-weight:700;color:#b3bed0;margin-bottom:6px;">'
                        f'"{search_ticker}" not found</div>'
                        f'<div style="font-size:13px;color:#8896ac;line-height:1.6;">'
                        f'No price data available. Try the exact ticker symbol — e.g. <strong style="color:#b3bed0;">AAPL</strong>, '
                        f'<strong style="color:#b3bed0;">NVDA</strong>, <strong style="color:#b3bed0;">TSLA</strong>.</div>'
                        f'</div>',
                        unsafe_allow_html=True)
                else:
                    scored = score_stock(resolved_tk, hist)
                    scored["sector"] = ALL_SECTORS.get(resolved_tk, "Unknown")
                    macro = st.session_state.get("macro_data") or _live_macro()
                    scored_list = apply_macro_overlay([scored], macro)
                    sr = scored_list[0]
                    hydrate_card_rows([sr])   # standard card fields: sector, cap, band
                    if sr.get("promoted"):
                        from model_engine import EXIT_THRESHOLD
                        regime = macro.get("regime","NEUTRAL")
                        eff_threshold = 62 if regime in ("RISK_OFF","HIGH VOLATILITY") else 60
                        adj = float(sr.get("adj_composite", sr.get("composite", 50)))
                        sr["adj_action"] = "BUY" if adj >= eff_threshold else ("SELL" if adj < EXIT_THRESHOLD else "HOLD")
                        sr["promoted"] = False
                    # pct_rank intentionally not set — factor_panel_html ranks
                    # this card against the full universe via _pct_rank_of().
                    ci = get_company_info(resolved_tk)
                    # Show search result pre-expanded — always one card, no toggle needed
                    _sr_html = factor_panel_html(sr, False, company_info=ci, suppress_wl_btn=True,
                                                 extra_detail=_whats_changed_html(resolved_tk, sr.get("signal_date", "")))
                    # Force detail open
                    _sr_html = _sr_html.replace('class="qcard-detail" style="display:none;', 'class="qcard-detail" style="display:block;')
                    st.markdown(_sr_html, unsafe_allow_html=True)
                    # 20-day price vs SPY sparkline
                    _render_rel_spy_chart(resolved_tk, 20)
                    # Flag this ticker as successfully scored so the watchlist
                    # button can render AFTER the try block (never masked by except).
                    _sr_ok = {"ticker": resolved_tk, "price": sr.get("price")}
                    if resolved_tk not in ALL_SECTORS:
                        st.markdown('<div style="font-size:13px;color:#8896ac;margin-bottom:16px;">⚠ Not in core universe — scored from live price data. Fundamental data may be limited.</div>', unsafe_allow_html=True)
            except Exception:
                _sr_ok = None
                st.markdown(
                    f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.08);'
                    f'border-radius:8px;padding:20px 24px;">'
                    f'<div style="font-family:Syne,sans-serif;font-size:15px;font-weight:700;color:#b3bed0;margin-bottom:6px;">'
                    f'"{search_ticker}" not found</div>'
                    f'<div style="font-size:13px;color:#8896ac;">Could not retrieve data. Check the symbol and try again.</div>'
                    f'</div>',
                    unsafe_allow_html=True)

        # ── Watchlist add/remove for the searched stock — OUTSIDE the try so a
        # button error can never show as "not found", and it always renders. ──
        if _sr_ok:
            _srtk = _sr_ok["ticker"]
            from db import (get_watchlist_items as _gwi, add_watchlist_item as _awi,
                            remove_watchlist_item as _rwi, get_watchlists as _gws,
                            get_price_on_date_latest as _gpl)
            _wl_uid_s = uid()
            _def_lists = _gws(_wl_uid_s)
            _def_id = next((l["id"] for l in _def_lists if l.get("is_default")),
                           _def_lists[0]["id"] if _def_lists else None)
            _wl_tickers = {w["ticker"] for w in _gwi(_wl_uid_s, _def_id)} if _def_id else set()
            # Use the shared main-document action row (target=_self, mobile-safe)
            render_watchlist_actions([_srtk], nav="screener", in_list=_wl_tickers)
        st.markdown('<div style="height:1px;background:rgba(255,255,255,.05);margin:8px 0 12px;"></div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # Invalidate cache if signal_log has newer data than what we have cached.
    # Signals only change once nightly, so checking on EVERY rerun (e.g. after a
    # watchlist button click) is wasteful — it adds a Supabase round-trip to each
    # page load. Throttle to at most once every 5 minutes per session.
    if st.session_state.scan_results is not None:
        import time as _time_fc
        _last_fc = st.session_state.get("_last_freshness_check", 0)
        if _time_fc.time() - _last_fc > 300:
            st.session_state._last_freshness_check = _time_fc.time()
            try:
                from data_refresh import _get_supabase as _sc_sb
                _sb_sc = _sc_sb()
                if _sb_sc:
                    _latest_date = _sb_sc.table("signal_log").select("signal_date") \
                        .order("signal_date", desc=True).limit(1).execute()
                    if _latest_date.data:
                        _db_date = str(_latest_date.data[0]["signal_date"])[:10]
                        _cached_date = str((st.session_state.scan_results[0] or {}).get("signal_date",""))[:10]
                        if _db_date > _cached_date:
                            st.session_state.scan_results = None  # force reload
                            st.session_state._live_prices_fetched = False
                            st.session_state.pop("_hdr_fresh_date", None)  # refresh header label too
            except Exception:
                pass

    if st.session_state.scan_results is None:
        # Auto-trigger if data is old
        stale_msg = "Loading universe scores..."
        try:
            from data_refresh import cache_is_fresh
            if not cache_is_fresh():
                stale_msg = "Data is stale — loading estimated scores. Hit Rescan for live data."
        except Exception:
            pass
        with st.spinner(stale_msg):
            raw   = _cached_full_scan()
            macro = _live_macro()
            for r in raw:
                if not r.get("sector") or r.get("sector") == "Unknown":
                    r["sector"] = ALL_SECTORS.get(r["ticker"], "Unknown")
            # Score with overlay once for any tickers that don't have a signal_log row
            results = apply_macro_overlay(raw, macro)
            # Replace anything that has a signal_log row with the cron's values —
            # that's our single source of truth
            enriched = enrich_with_signal_log(results)
            # Normalize action/sort/floor from those values WITHOUT recomputing
            # adj_composite again (which would diverge from what the home page,
            # gems, watchlist, portfolio, etc. all read from signal_log).
            st.session_state.scan_results = finalize_scores_from_signal_log(enriched, macro)
            st.session_state.macro_data   = macro

    results = st.session_state.scan_results

    # Refresh macro every 15 min — keeps VIX, WTI, regime current
    import time as _time
    _macro_age = _time.time() - st.session_state.get("_macro_fetched_at", 0)
    if not st.session_state.get("macro_data") or _macro_age > 900:
        try:
            macro = _live_macro()
            st.session_state.macro_data      = macro
            st.session_state._macro_fetched_at = _time.time()
        except Exception:
            macro = st.session_state.get("macro_data", {})
    else:
        macro = st.session_state.get("macro_data", {})

    gems = detect_hidden_gems(results, macro_data=macro)

    # ── Live price refresh for top visible tickers ─────────────────────────
    # Fetch current prices for top 30 tickers (buys + sells) via yfinance
    # Cached per session so it only fires once per load, not on every rerun
    if not st.session_state.get("_live_prices_fetched"):
        try:
            _top_tks = (
                [r["ticker"] for r in sorted(results, key=lambda x: float(x.get("adj_composite",0) or 0), reverse=True)[:15]] +
                [r["ticker"] for r in sorted(results, key=lambda x: float(x.get("adj_composite",50) or 50))[:10]]
            )
            _top_tks = list(dict.fromkeys(_top_tks))[:25]  # dedupe, cap at 25
            _live_px = _cached_live_prices(tuple(_top_tks))
            # Inject live prices into results
            _price_map = {r["ticker"]: r for r in results}
            for _tk, _px in _live_px.items():
                if _tk in _price_map:
                    _price_map[_tk]["price"] = _px
            st.session_state._live_prices_fetched = True
        except Exception:
            pass
    gem_tickers = {g["ticker"] for g in gems}

    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    # ── Compact breadth strip — search → breadth → regime ─────────────────────
    buys  = sum(1 for r in results if r.get("adj_action",r.get("action"))=="BUY")
    holds = sum(1 for r in results if r.get("adj_action",r.get("action"))=="HOLD")
    sells = sum(1 for r in results if r.get("adj_action",r.get("action"))=="SELL")
    _n_gems_strip = len(gems)
    st.session_state._gem_count  = _n_gems_strip
    st.session_state._high_count = buys
    st.session_state._low_count  = sells
    # Write to platform_stats so landing page always has fresh counts
    try:
        from data_refresh import _get_supabase as _ps_sb
        _ps_client = _ps_sb()
        if _ps_client:
            _ps_client.table("platform_stats").upsert({
                "stat_key": "daily_summary",
                "n_high":   buys,
                "n_low":    sells,
                "n_gems":   _n_gems_strip,
                "n_total":  len(results),
                "regime":   macro.get("regime", "NEUTRAL"),
                "updated_at": "now()",
            }, on_conflict="stat_key").execute()
    except Exception:
        pass
    st.markdown(
        f'<div style="display:flex;gap:16px;flex-wrap:wrap;align-items:center;'
        f'padding:6px 0 10px;margin-bottom:4px;border-bottom:1px solid rgba(255,255,255,.04);">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;">'
        f'<span style="color:#34d399;">HIGH {buys}</span>'
        f'<span style="color:#94a3b8;"> · </span>'
        f'<span style="color:#8896ac;">MOD {holds}</span>'
        f'<span style="color:#94a3b8;"> · </span>'
        f'<span style="color:#8896ac;">LOW {sells}</span>'
        f'<span style="color:#94a3b8;"> · </span>'
        f'<span style="color:#34d399;">💎 {_n_gems_strip}</span>'
        f'<span style="color:#94a3b8;"> · </span>'
        f'<span style="color:#94a3b8;">UNIV {len(results)}</span>'
        f'</span></div>',
        unsafe_allow_html=True
    )

    # ── Login hero: regime + watchlist movers + universe conviction changes ────
    _hero_movers = _conviction_movers(tuple(sorted({r["ticker"] for r in results}))) if results else []
    _wl_tks = sorted({w["ticker"] for w in get_watchlist(uid())}) if uid() else []
    _wl_movers = _conviction_movers(tuple(_wl_tks), top_n=12, collapse_macro=False) if _wl_tks else []
    st.markdown(_hero_card_html(macro, results, _hero_movers, _wl_movers, bool(_wl_tks)),
                unsafe_allow_html=True)

    # ── Macro Regime Banner ────────────────────────────────────────────────────
    from model_engine import MACRO_EVENT_INFO
    st.markdown(macro_regime_banner_html(macro), unsafe_allow_html=True)
    # Active events — single collapsed expander
    active_evts = macro.get("active_events", [])
    if active_evts:
        _evt_labels = [MACRO_EVENT_INFO[e]['label'] for e in active_evts if e in MACRO_EVENT_INFO]
        _evt_title  = f"{len(active_evts)} active macro driver{'s' if len(active_evts)>1 else ''}: {', '.join(_evt_labels[:3])}"
        with st.expander(_evt_title, expanded=False):
            for evt in active_evts:
                info = MACRO_EVENT_INFO.get(evt)
                if not info: continue
                st.markdown(
                    f'<div style="padding:14px 16px;margin-bottom:8px;'
                    f'background:rgba(255,255,255,.02);'
                    f'border:1px solid rgba(255,255,255,.06);'
                    f'border-left:2px solid rgba(212,168,67,.5);'
                    f'border-radius:6px;">'
                    f'<div style="font-family:Syne,sans-serif;font-size:13px;'
                    f'font-weight:700;color:#d4a843;letter-spacing:.08em;'
                    f'text-transform:uppercase;margin-bottom:8px;">{info["label"]}</div>'
                    f'<div style="font-family:Inter,sans-serif;font-size:13px;'
                    f'color:#9fabc0;line-height:1.65;margin-bottom:10px;">{info["detail"]}</div>'
                    f'<div style="display:flex;gap:14px;flex-wrap:wrap;'
                    f'padding-top:8px;border-top:1px solid rgba(255,255,255,.04);">'
                    f'<span style="font-family:DM Mono,monospace;font-size:13px;'
                    f'color:#f87171;letter-spacing:.04em;">▼ {info["impact"]}</span>'
                    f'<span style="font-family:DM Mono,monospace;font-size:13px;'
                    f'color:#34d399;letter-spacing:.04em;">▲ {info["bullish"]}</span>'
                    f'</div></div>',
                    unsafe_allow_html=True)


    st.markdown(DISCLAIMER, unsafe_allow_html=True)

    buys_ranked  = sorted([r for r in results if r.get("adj_action",r.get("action"))=="BUY"],
                          key=lambda x: x.get("adj_composite",x.get("composite",0)), reverse=True)
    sells_ranked = sorted([r for r in results if r.get("adj_action",r.get("action"))=="SELL"],
                          key=lambda x: x.get("adj_composite",x.get("composite",100)))

    scr_tab1, scr_tab2, scr_tab3 = st.tabs(["⭐ TOP 10 SIGNALS", "🔍 FULL UNIVERSE", "📈 SECTOR BREAKDOWN"])

    # ── TAB 1: TOP 10 — scan mode ───────────────────────────────────────────
    with scr_tab1:
        # Responsive CSS grid (not st.columns, which won't stack on mobile and
        # clips the expanded card on narrow screens). 2 cols on desktop, 1 col
        # on phones via the .qntm-conv-grid media query in the main-doc <style>.
        _wl_now = {w["ticker"] for w in get_watchlist(uid())} if uid() else set()

        # Trailing vs-SPY mini charts for the top/bottom 10 — these lists have no
        # stored "entered the list" date, so we show a fixed trailing window.
        _scr_trail = _trail_start(30)
        _scr_tks = [r["ticker"] for r in buys_ranked[:10]] + [r["ticker"] for r in sells_ranked[:10]]
        _scr_pm, _scr_sm = ({}, {})
        if _scr_tks:
            _scr_pm, _scr_sm = _mini_price_data(tuple(sorted(set(_scr_tks))), _scr_trail)

        def _conv_col(label, color, ranked):
            out = (f'<div style="font-family:DM Mono,monospace;font-size:13px;color:{color};'
                   f'letter-spacing:.12em;margin:0 0 6px;padding-bottom:4px;'
                   f'border-bottom:1px solid rgba(255,255,255,.05);">{label}</div>')
            for r in ranked:
                ci     = get_company_info(r["ticker"])
                is_gem = r["ticker"] in gem_tickers
                # Ensure action matches list — signal_log BUY/SELL may not match adj
                if color == "#f87171" and r.get("adj_action",r.get("action")) != "SELL":
                    r = dict(r); r["adj_action"] = "SELL"
                elif color == "#34d399" and r.get("adj_action",r.get("action")) != "BUY":
                    r = dict(r); r["adj_action"] = "BUY"
                r["_mini_chart_html"] = _build_mini_chart_html(
                    r["ticker"], _scr_trail, _scr_pm, _scr_sm, since_label="vs SPY · 20d")
                out += build_card_html(r, nav="screener", is_gem=is_gem,
                                       company_info=ci, in_list=_wl_now)
            return out

        _high_html = _conv_col("▲ HIGH CONVICTION", "#34d399", buys_ranked[:10])
        _low_html  = _conv_col("▼ LOW CONVICTION",  "#f87171", sells_ranked[:10])
        st.markdown(
            '<div class="qntm-conv-grid" style="display:grid;'
            'grid-template-columns:1fr 1fr;gap:16px;align-items:start;">'
            f'<div>{_high_html}</div><div>{_low_html}</div>'
            '</div>',
            unsafe_allow_html=True)

    # ── TAB 2: FULL UNIVERSE ───────────────────────────────────────────────────
    with scr_tab2:
        fc1, fc2, fc3 = st.columns(3)
        with fc1:
            filter_sec = st.selectbox("Sector", ["All"]+sorted(set(SECTORS.values())), key="f_sec")
        with fc2:
            filter_act = st.selectbox("Conviction", ["All","High","Moderate","Low"], key="f_act")
        with fc3:
            filter_min = st.selectbox("Min Score", ["All","60+","70+","80+"], key="f_min")
        # Rescan — small right-aligned, below filters
        _uid_v = (st.session_state.user or {}).get('id','')
        _pln_v = (st.session_state.user or {}).get('plan','free')
        _rscan_url = f'?qnav=screener&uid={_uid_v}&plan={_pln_v}&ck=1&rescan=1'
        st.markdown(
            f'<div style="display:flex;justify-content:flex-end;margin:6px 0 4px;">'
            f'<a href="{_rscan_url}" target="_self" style="'
            f'padding:6px 14px;border-radius:4px;border:1px solid rgba(255,255,255,.12);'
            f'background:rgba(255,255,255,.03);font-family:Syne,sans-serif;font-size:13px;'
            f'font-weight:700;letter-spacing:.08em;color:#8896ac;text-decoration:none;'
            f'text-transform:uppercase;">↺ Rescan</a></div>',
            unsafe_allow_html=True
        )


        filtered = results
        if filter_sec != "All": filtered = [r for r in filtered if r.get("sector")==filter_sec]
        if filter_act != "All":
            act_map = {"High":"BUY","Moderate":"HOLD","Low":"SELL"}
            filtered = [r for r in filtered if r.get("adj_action",r.get("action"))==act_map.get(filter_act)]
        if filter_min != "All":
            min_score = int(filter_min.replace("+",""))
            filtered = [r for r in filtered if float(r.get("adj_composite",r.get("composite",0)) or 0) >= min_score]

        # Free tier: show top 50 results only
        _user_plan = (st.session_state.user or {}).get("plan", "free")
        FREE_LIMIT = 50
        RENDER_LIMIT = 200  # cap render to 200 cards max for performance
        _total_filtered = len(filtered)
        if _user_plan == "free" and _total_filtered > FREE_LIMIT:
            filtered = filtered[:FREE_LIMIT]
            _show_gate = True
        else:
            _show_gate = False
        # Cap render for performance — prompt user to filter
        _show_render_cap = not _show_gate and len(filtered) > RENDER_LIMIT
        if _show_render_cap:
            filtered = filtered[:RENDER_LIMIT]

        st.markdown(
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;'
            f'letter-spacing:.1em;margin:8px 0 12px;">{len(filtered)} STOCKS · 💎 = HIDDEN GEM'
            + (f' · Showing {FREE_LIMIT} of {_total_filtered}' if _show_gate else
               f' · Showing {RENDER_LIMIT} of {_total_filtered} — filter to see all' if _show_render_cap else '') +
            f'</div>',
            unsafe_allow_html=True)
        # Paginate + render cards in a fragment so Prev/Next re-runs ONLY this
        # block, not the whole app (keeps in-page paging snappy).
        _fu_filter_key = f"{filter_sec}_{filter_act}_{filter_min}"
        _render_screener_cards(filtered, gem_tickers, _fu_filter_key)

        if _show_gate:
            st.markdown(
                f'<div style="background:rgba(212,168,67,.06);border:1px solid rgba(212,168,67,.2);'
                f'border-radius:10px;padding:24px;text-align:center;margin-top:16px;">'
                f'<div style="font-family:Syne,sans-serif;font-size:16px;font-weight:700;color:#d4a843;margin-bottom:8px;">'
                f'🔒 {_total_filtered - FREE_LIMIT} more stocks available on Pro</div>'
                f'<div style="font-size:13px;color:#b3bed0;margin-bottom:16px;">'
                f'Free accounts see the top {FREE_LIMIT} signals. Upgrade for the full {_total_filtered}-stock view, '
                f'Hidden Gems, alerts, and unlimited portfolio tracking.</div>'
                f'</div>',
                unsafe_allow_html=True)
            if st.session_state.get("logged_in"):
                st.markdown(_cta_gold(f"Unlock Full Universe — {_total_filtered - FREE_LIMIT} more stocks", _upgrade_url("Full Universe", "screener")), unsafe_allow_html=True)
            else:
                st.markdown(_cta_gold(f"Upgrade to Pro — see all {_total_filtered} stocks", _upgrade_url("Full Universe Access","screener")), unsafe_allow_html=True)

    # ── TAB 3: SECTOR BREAKDOWN ────────────────────────────────────────────────
    with scr_tab3:
        sector_counts = {}
        for r in results:
            sec = r.get("sector","Other")
            act = r.get("adj_action", r.get("action","HOLD"))
            if sec not in sector_counts:
                sector_counts[sec] = {"BUY":0,"HOLD":0,"SELL":0}
            sector_counts[sec][act] = sector_counts[sec].get(act,0)+1

        st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;letter-spacing:.1em;margin:16px 0 10px;">SIGNAL BREAKDOWN BY SECTOR</div>', unsafe_allow_html=True)
        for sec, counts in sorted(sector_counts.items()):
            total = sum(counts.values()) or 1
            b,h,s = counts.get("BUY",0),counts.get("HOLD",0),counts.get("SELL",0)
            bp,hp,sp = b/total*100, h/total*100, s/total*100
            st.markdown(
                f'<div style="display:flex;align-items:center;gap:10px;margin-bottom:7px;">'
                f'<div style="font-size:13px;color:#9fabc0;width:170px;flex-shrink:0;">{sec}</div>'
                f'<div style="flex:1;display:flex;border-radius:4px;overflow:hidden;height:20px;">'
                f'<div style="width:{bp:.0f}%;background:rgba(52,211,153,.6);"></div>'
                f'<div style="width:{hp:.0f}%;background:rgba(251,191,36,.4);"></div>'
                f'<div style="width:{sp:.0f}%;background:rgba(248,113,113,.5);"></div>'
                f'</div>'
                f'<div style="font-size:14px;color:#b3bed0;width:130px;flex-shrink:0;">'
                f'<span style="color:#34d399;">{b} HIGH</span> '
                f'<span style="color:#fbbf24;">{h} MOD</span> '
                f'<span style="color:#f87171;">{s} LOW</span>'
                f'</div></div>',
                unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


def page_watchlist():
    _pin_nav("watchlist")
    page_summary("★", "Watchlist", "Your tracked stocks · conviction scores updated daily")
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    from db import (get_watchlists, create_watchlist, rename_watchlist,
                    delete_watchlist, get_watchlist_items, remove_watchlist_item,
                    add_watchlist_item, get_price_on_date_latest)
    _wl_uid = uid()

    # ── Handle list-management actions via query params ───────────────────────
    _list_action = st.query_params.get("wl_list_action", "")
    if _list_action:
        _tgt = st.query_params.get("wl_list_id", "")
        if _list_action == "select" and _tgt:
            st.session_state.active_wl_list = _tgt
        elif _list_action == "delete" and _tgt:
            delete_watchlist(_wl_uid, _tgt)
            st.session_state.pop("active_wl_list", None)
        elif _list_action == "remove_item":
            _rm_tk = st.query_params.get("wl_rm_ticker", "")
            if _tgt and _rm_tk:
                remove_watchlist_item(_wl_uid, _tgt, _rm_tk)
                st.session_state.active_wl_list = _tgt
                st.session_state.pop("_wl_daychange_cache", None)
        st.query_params.pop("wl_list_action", None)
        st.query_params.pop("wl_list_id", None)
        st.query_params.pop("wl_rm_ticker", None)

    _all_lists = get_watchlists(_wl_uid)   # auto-creates default if none
    if not _all_lists:
        st.warning("Could not load your watchlists. Try refreshing.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # Resolve active list — session pick, else default, else first
    _active_id = st.session_state.get("active_wl_list")
    if _active_id not in {w["id"] for w in _all_lists}:
        _active_id = next((w["id"] for w in _all_lists if w.get("is_default")), _all_lists[0]["id"])
        st.session_state.active_wl_list = _active_id
    _active_list = next(w for w in _all_lists if w["id"] == _active_id)

    # ── List selector bar — real buttons (URL links don't survive reliably) ──
    st.markdown("""
    <style>
    div[data-testid='stButton'][data-key^='wltab_'] > div > button {
        background: rgba(255,255,255,.025) !important;
        border: 1px solid rgba(255,255,255,.07) !important;
        color: #9fabc0 !important;
        font-family: 'Syne', sans-serif !important;
        font-weight: 700 !important; font-size:13px !important;
        letter-spacing: .04em !important;
        border-radius: 8px !important; padding: 9px 16px !important;
        transition: all .18s ease !important;
        box-shadow: none !important;
    }
    div[data-testid='stButton'][data-key^='wltab_'] > div > button:hover {
        background: rgba(255,255,255,.05) !important;
        border-color: rgba(212,168,67,.3) !important;
        color: #cbd5e1 !important;
        transform: translateY(-1px) !important;
    }
    div[data-testid='stButton'][data-key^='wltabactive_'] > div > button {
        background: linear-gradient(135deg, rgba(212,168,67,.22), rgba(212,168,67,.08)) !important;
        border: 1px solid rgba(212,168,67,.55) !important;
        color: #f0c668 !important;
        font-family: 'Syne', sans-serif !important;
        font-weight: 800 !important; font-size:13px !important;
        letter-spacing: .04em !important;
        border-radius: 8px !important; padding: 9px 16px !important;
        box-shadow: 0 0 0 1px rgba(212,168,67,.12), 0 2px 12px rgba(212,168,67,.15) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
                'letter-spacing:.12em;margin-bottom:8px;">LISTS</div>', unsafe_allow_html=True)
    _tab_cols = st.columns(max(len(_all_lists), 1))
    for _i, w in enumerate(_all_lists):
        _is_act = w["id"] == _active_id
        with _tab_cols[_i]:
            _key = f"{'wltabactive' if _is_act else 'wltab'}_{w['id']}"
            _lbl = f"●  {w['name']}" if _is_act else w["name"]
            if st.button(_lbl, key=_key, use_container_width=True):
                st.session_state.active_wl_list = w["id"]
                st.session_state.nav = "watchlist"
                st.rerun()

    # New / rename / delete controls
    _c_new, _c_ren, _c_del = st.columns(3)
    with _c_new:
        with st.popover("➕ New list", use_container_width=True):
            _nm = st.text_input("List name", key="wl_new_name", placeholder="e.g. Tech Watch")
            if st.button("Create", key="wl_new_go", use_container_width=True):
                _n = (_nm or "").strip()
                if not _n:
                    st.warning("Enter a name.")
                else:
                    _created = create_watchlist(_wl_uid, _n)
                    if _created:
                        st.session_state.active_wl_list = _created["id"]
                        st.rerun()
                    else:
                        st.error("That name may already exist.")
    with _c_ren:
        with st.popover("✏️ Rename", use_container_width=True):
            _rn = st.text_input("New name", key="wl_ren_name", value=_active_list["name"])
            if st.button("Save name", key="wl_ren_go", use_container_width=True):
                if rename_watchlist(_wl_uid, _active_id, _rn):
                    st.rerun()
                else:
                    st.error("Rename failed (name may be taken).")
    with _c_del:
        _can_del = len(_all_lists) > 1
        with st.popover("🗑 Delete", use_container_width=True, disabled=not _can_del):
            st.markdown(f"Delete **{_active_list['name']}** and all its tickers?")
            if st.button("Confirm delete", key="wl_del_go", use_container_width=True):
                if delete_watchlist(_wl_uid, _active_id):
                    st.session_state.pop("active_wl_list", None)
                    st.rerun()
                else:
                    st.error("Can't delete your only list.")
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)

    watchlist = get_watchlist_items(_wl_uid, _active_id)
    _ln = (_active_list.get("name") or "").replace("<", "&lt;").replace(">", "&gt;")
    st.markdown(
        '<div style="display:flex;align-items:center;gap:10px;margin:4px 0 12px;flex-wrap:wrap;">'
        '<span style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;'
        'letter-spacing:.14em;text-transform:uppercase;">Viewing</span>'
        f'<span style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#f0c668;'
        f'letter-spacing:.02em;">{_ln}</span>'
        f'<span style="font-family:DM Mono,monospace;font-size:12px;color:#94a3b8;">· '
        f'{len(watchlist)} {"stock" if len(watchlist)==1 else "stocks"}</span>'
        '</div>',
        unsafe_allow_html=True,
    )
    scan      = st.session_state.get("scan_results") or []
    score_map = {r["ticker"]: r for r in scan}

    # ── Native add/remove controls (iframe links are sandbox-blocked) ──────────
    st.markdown("""
    <style>
    div[data-testid='stButton'][class*='wl_native_add'] button,
    .st-key-wl_native_add button {
        background: linear-gradient(135deg,#d4a843,#b8922e) !important;
        color:#0a0b14 !important; border:none !important;
        font-family:Syne,sans-serif !important; font-weight:800 !important;
        font-size:13px !important; letter-spacing:.06em !important;
        border-radius:8px !important; text-transform:uppercase !important;
        box-shadow:0 2px 12px rgba(212,168,67,.2) !important;
    }
    .st-key-wl_native_rm button {
        background:rgba(248,113,113,.08) !important;
        border:1px solid rgba(248,113,113,.3) !important;
        color:#f87171 !important;
        font-family:Syne,sans-serif !important; font-weight:700 !important;
        font-size:13px !important; border-radius:8px !important;
        text-transform:uppercase !important; letter-spacing:.06em !important;
    }
    </style>
    """, unsafe_allow_html=True)
    def _on_wl_search():
        v = st.session_state.get("wl_native_add_tk", "").strip().upper()
        st.session_state.wl_sel_tk = v if v in SECTORS else ""

    st.text_input("Add ticker", key="wl_native_add_tk",
                  placeholder="🔍  Search ticker or company — NVDA, Apple…",
                  label_visibility="collapsed", on_change=_on_wl_search)

    _wl_excl = [w["ticker"] for w in watchlist]
    _wl_live = st.session_state.get("wl_native_add_tk", "").strip()
    if _wl_live and not st.session_state.get("wl_sel_tk"):
        _render_suggestions(_wl_live, "wlsug", "wl_sel_tk", exclude=_wl_excl)

    _wl_sel = st.session_state.get("wl_sel_tk", "").strip().upper()
    if _wl_sel:
        _render_stock_result(_wl_sel, nav="watchlist", wl_actions=False)
        _wl_have = {w["ticker"] for w in watchlist}
        if _wl_sel in _wl_have:
            if st.button(f"✕ Remove {_wl_sel} from Watchlist", key="wl_sel_remove",
                         use_container_width=True):
                if remove_watchlist_item(_wl_uid, _active_id, _wl_sel):
                    st.session_state.pop("_wl_daychange_cache", None)
                    st.session_state.pop("wl_native_add_tk", None)
                    st.session_state.wl_sel_tk = ""
                    st.toast(f"Removed {_wl_sel}")
                    st.rerun()
        else:
            if st.button(f"☆ Add {_wl_sel} to Watchlist", key="wl_sel_add",
                         use_container_width=True, type="primary"):
                _add_px = ((score_map.get(_wl_sel) or {}).get("price")) or get_price_on_date_latest(_wl_sel)
                if add_watchlist_item(_wl_uid, _active_id, _wl_sel, _add_px):
                    st.session_state.pop("_wl_daychange_cache", None)
                    st.session_state.pop("wl_native_add_tk", None)
                    st.session_state.wl_sel_tk = ""
                    st.toast(f"Added {_wl_sel}")
                    st.rerun()
                else:
                    st.toast(f"Could not add {_wl_sel}")
    if watchlist:
        _rm_c, _rmbtn_c = st.columns([3, 1])
        with _rm_c:
            _rm_pick = st.selectbox("Remove ticker", [w["ticker"] for w in watchlist],
                                    key="wl_native_rm_pick", label_visibility="collapsed")
        with _rmbtn_c:
            if st.button("✕ Remove", key="wl_native_rm", use_container_width=True):
                if remove_watchlist_item(_wl_uid, _active_id, _rm_pick):
                    st.session_state.pop("_wl_daychange_cache", None)
                    st.toast(f"Removed {_rm_pick}")
                    st.rerun()
    st.markdown('<div style="height:10px"></div>', unsafe_allow_html=True)

    if not watchlist:
        st.markdown(
            '<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
            'border-radius:10px;padding:40px 24px;text-align:center;margin-top:16px;">'
            '<div style="font-size:32px;margin-bottom:12px;">★</div>'
            '<div style="font-family:Syne,sans-serif;font-size:16px;font-weight:700;color:#9fabc0;margin-bottom:8px;">'
            'Your watchlist is empty</div>'
            '<div style="font-size:13px;color:#94a3b8;line-height:1.6;">'
            'Search any stock on the Screener and hit <strong style="color:#b3bed0;">Add to Watchlist</strong> '
            'to track its conviction score here.</div>'
            '</div>',
            unsafe_allow_html=True
        )
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # If no scan loaded, show scores from signal_log
    if not score_map:
        try:
            from data_refresh import _get_supabase
            sb = _get_supabase()
            if sb:
                tickers = [w["ticker"] for w in watchlist]
                resp = sb.table("signal_log") \
                    .select("ticker,adj_composite,composite,price,signal,momentum,quality,volume,value,sentiment,val_low,val_high,value_position,val_basis") \
                    .in_("ticker", tickers) \
                    .order("signal_date", desc=True) \
                    .limit(len(tickers) * 3) \
                    .execute()
                seen = set()
                _rows = []
                for row in (resp.data or []):
                    tk = row["ticker"]
                    if tk not in seen:
                        seen.add(tk)
                        _rows.append(row)
                # Normalize action labels, score_delta etc. so MACRO box and
                # conviction label match what the screener shows for the same tickers
                finalize_scores_from_signal_log(_rows, st.session_state.get("macro_data"))
                for r in _rows:
                    score_map[r["ticker"]] = r
        except Exception:
            pass

    from model_engine import EXIT_THRESHOLD, ENTRY_THRESHOLD, SECTORS as _WL_SECTORS

    # ── Watchlist alerts — conviction changes ─────────────────────────────────
    # Compare current score to entry score stored in DB — surface big moves
    alerts = []
    for w in watchlist:
        tk = w["ticker"]
        sc = score_map.get(tk, {})
        cur = float(sc.get("adj_composite", sc.get("composite", 0)) or 0)
        entry_score = float(w.get("entry_score") or w.get("adj_composite") or 0)
        if cur == 0 or entry_score == 0:
            continue
        delta = cur - entry_score
        # Conviction level change
        def _level(s):
            return "High" if s >= 60 else ("Low" if s < 45 else "Moderate")
        cur_level   = _level(cur)
        entry_level = _level(entry_score)
        if cur_level != entry_level:
            color = "#34d399" if cur_level == "High" else ("#f87171" if cur_level == "Low" else "#fbbf24")
            arrow = "▲" if cur > entry_score else "▼"
            alerts.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:8px 12px;background:rgba(255,255,255,.02);border-radius:4px;margin-bottom:4px;">'
                f'<div><span style="font-family:Syne,sans-serif;font-weight:700;color:#e2e8f0;">{tk}</span>'
                f' <span style="font-size:13px;color:#9fabc0;">conviction changed</span></div>'
                f'<span style="font-size:13px;font-weight:700;color:{color};">{arrow} {entry_level} → {cur_level}</span>'
                f'</div>'
            )
        # Large score move (10+) without level change
        elif abs(delta) >= 10:
            color = "#34d399" if delta > 0 else "#f87171"
            sign  = "+" if delta > 0 else ""
            alerts.append(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'padding:8px 12px;background:rgba(255,255,255,.02);border-radius:4px;margin-bottom:4px;">'
                f'<div><span style="font-family:Syne,sans-serif;font-weight:700;color:#e2e8f0;">{tk}</span>'
                f' <span style="font-size:13px;color:#9fabc0;">score moved significantly</span></div>'
                f'<span style="font-size:13px;font-weight:700;color:{color};">{sign}{delta:.0f} pts → {cur:.0f}</span>'
                f'</div>'
            )

    if alerts:
        st.markdown(
            f'<div style="background:rgba(212,168,67,.05);border:1px solid rgba(212,168,67,.2);'
            f'border-radius:8px;padding:12px 16px;margin-bottom:16px;">'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;margin-bottom:8px;">🔔 CONVICTION UPDATES</div>'
            + "".join(alerts) +
            f'</div>',
            unsafe_allow_html=True
        )

    wl_trend = {}  # populated below — pre-init so summary can reference it safely
    # Fetch live prices + prev close for day change via yfinance
    wl_tickers = [w["ticker"] for w in watchlist]
    day_change  = {}   # ticker -> {price, prev_close, chg_pct, chg_dollar}
    # Cache by sorted ticker set so toggling back to a loaded list is instant.
    _dc_cache = st.session_state.setdefault("_wl_daychange_cache", {})
    _cache_key = ",".join(sorted(wl_tickers))
    if wl_tickers and _cache_key in _dc_cache:
        day_change = _dc_cache[_cache_key]
    elif wl_tickers:
        with st.spinner(f"Loading live prices for {len(wl_tickers)} stocks…"):
            try:
                import yfinance as yf
                from datetime import datetime as _dt_dc
                try:
                    from zoneinfo import ZoneInfo as _ZI
                    _today_str = _dt_dc.now(_ZI("America/New_York")).strftime("%Y-%m-%d")
                except Exception:
                    _today_str = _dt_dc.now().strftime("%Y-%m-%d")
                hist = yf.download(_strip_delisted(wl_tickers), period="5d", auto_adjust=True,
                                   progress=False, threads=True)
                if not hist.empty:
                    # Latest session is "live" only when it's today; otherwise
                    # the move we show is a settled close (today after close, or
                    # the last completed session on a weekend/holiday).
                    try:
                        _last_bar_date = str(hist.index[-1])[:10]
                    except Exception:
                        _last_bar_date = ""
                    _settled = (_last_bar_date != _today_str)
                    close = hist["Close"]
                    if hasattr(close, "columns"):
                        for tk in wl_tickers:
                            if tk in close.columns:
                                vals = close[tk].dropna()
                                if len(vals) >= 2:
                                    cur  = float(vals.iloc[-1])
                                    prev = float(vals.iloc[-2])
                                    day_change[tk] = {
                                        "price":      cur,
                                        "prev_close": prev,
                                        "chg_pct":    ((cur - prev) / prev * 100) if prev else None,
                                        "chg_dollar": (cur - prev),
                                        "settled": _settled,
                                        "market_closed": _settled,  # back-compat
                                    }
                    else:
                        vals = close.dropna()
                        if len(vals) >= 2 and len(wl_tickers) == 1:
                            cur  = float(vals.iloc[-1])
                            prev = float(vals.iloc[-2])
                            day_change[wl_tickers[0]] = {
                                "price":      cur,
                                "prev_close": prev,
                                "chg_pct":    ((cur - prev) / prev * 100) if prev else None,
                                "chg_dollar": (cur - prev),
                                "settled": _settled,
                                "market_closed": _settled,  # back-compat
                            }
                _dc_cache[_cache_key] = day_change
            except Exception:
                pass

    # Also fetch entry prices from watchlist record if stored, else use first observed price
    wl_entry = {w["ticker"]: w.get("entry_price") or w.get("price_at_add") for w in watchlist}

    # Batch fetch last 2 signal_log rows per ticker for trend arrows
    # wl_trend already initialised above
    # ── Intelligence summary — improving/weakening/posture ─────────────────
    n        = len(watchlist)
    n_hi     = sum(1 for w in watchlist if float((score_map.get(w["ticker"]) or {}).get("adj_composite",0) or 0) >= 60)
    n_lo     = sum(1 for w in watchlist if float((score_map.get(w["ticker"]) or {}).get("adj_composite",50) or 50) < 45)
    scores_all = [float((score_map.get(w["ticker"]) or {}).get("adj_composite",0) or 0) for w in watchlist]
    avg_score  = sum(scores_all) / len(scores_all) if scores_all else 0
    avg_label  = 'High' if avg_score >= 60 else ('Low' if avg_score < 45 else 'Moderate')
    avg_color  = '#34d399' if avg_score >= 60 else ('#f87171' if avg_score < 45 else '#fbbf24')
    # Count improving vs weakening from wl_trend
    n_improving = sum(1 for w in watchlist if (wl_trend.get(w['ticker']) or ('',))[0] == '\u2191')
    n_weakening = sum(1 for w in watchlist if (wl_trend.get(w['ticker']) or ('',))[0] == '\u2193')
    # Sector posture — dominant sector among high conviction
    hi_sectors = [_WL_SECTORS.get(w['ticker'],'') for w in watchlist
                  if float((score_map.get(w['ticker']) or {}).get('adj_composite',0) or 0) >= 60]
    from collections import Counter
    top_sector = Counter(hi_sectors).most_common(1)[0][0] if hi_sectors else ''
    top_sector_html = f'<span style="color:#8896ac;">· {top_sector} leading</span>' if top_sector else ''

    _impr_html = f'<span style="color:#34d399;">↑ {n_improving} improving</span>' if n_improving else ''
    _weak_html = f'<span style="color:#f87171;">↓ {n_weakening} weakening</span>' if n_weakening else ''
    _sep = '<span style="color:#1e293b;"> · </span>'
    _parts = [p for p in [_impr_html, _weak_html, top_sector_html] if p]

    try:
        from data_refresh import _get_supabase as _wl_sb
        _sb = _wl_sb()
        if _sb and wl_tickers:
            _trend_resp = _sb.table("signal_log") \
                .select("ticker,signal_date,adj_composite,composite") \
                .in_("ticker", wl_tickers) \
                .order("signal_date", desc=True) \
                .limit(len(wl_tickers) * 5) \
                .execute()
            # Group by ticker, keep last 2 dates
            _by_tk = {}
            for row in (_trend_resp.data or []):
                tk2 = row["ticker"]
                if tk2 not in _by_tk:
                    _by_tk[tk2] = []
                if len(_by_tk[tk2]) < 2:
                    _by_tk[tk2].append(float(row.get("adj_composite") or row.get("composite") or 50))
            for tk2, scores in _by_tk.items():
                if len(scores) >= 2:
                    delta = scores[0] - scores[1]
                    if delta >= 3:
                        wl_trend[tk2] = ("↑", "#34d399", f"+{delta:.0f}")
                    elif delta <= -3:
                        wl_trend[tk2] = ("↓", "#f87171", f"{delta:.0f}")
                    else:
                        wl_trend[tk2] = ("→", "#fbbf24", f"{delta:+.0f}")
    except Exception:
        pass

    st.markdown(
        f'<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;'
        f'padding:8px 0 10px;margin-bottom:8px;border-bottom:1px solid rgba(255,255,255,.04);">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;">{n} tracked</span>'
        f'<span style="color:#1e293b;">·</span>'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{avg_color};">avg {avg_score:.0f} {avg_label}</span>'
        + (_sep + _sep.join(_parts) if _parts else '')
        + f'</div>',
        unsafe_allow_html=True
    )

    # Watchlist — collapsed card pattern, no table header needed
    _uid_wl  = (st.session_state.user or {}).get("id","")
    _pln_wl  = (st.session_state.user or {}).get("plan","free")

    # Detect current hidden gems with the SAME regime/data the screener uses.
    # detect_hidden_gems sets its thresholds from the macro regime (much stricter
    # in HIGH VOLATILITY); if macro_data is missing it falls back to the looser
    # NEUTRAL thresholds and over-flags gems — that's why a name could show a 💎
    # here while the screener had already dropped it. Load macro if absent.
    _wl_gems = set()
    try:
        _wl_macro = st.session_state.get("macro_data")
        if not _wl_macro:
            try:
                _wl_macro = _live_macro()
                st.session_state.macro_data = _wl_macro
            except Exception:
                _wl_macro = {}
        _wl_gem_list = detect_hidden_gems(
            st.session_state.get("scan_results") or list(score_map.values()),
            macro_data=_wl_macro or {},
        )
        _wl_gems = {g["ticker"] for g in _wl_gem_list}
    except Exception:
        pass

    _wl_active_tickers = {w["ticker"] for w in watchlist}
    # Prefetch daily closes for all watchlist names since the earliest add date,
    # so each expanded card can show a vs-SPY mini chart since it was added.
    _wl_dates = [str(w.get("added_at") or w.get("created_at") or "")[:10] for w in watchlist]
    _wl_dates = [d for d in _wl_dates if d]
    _wl_pm, _wl_sm = ({}, {})
    if _wl_dates:
        _wl_pm, _wl_sm = _mini_price_data(tuple(sorted(_wl_active_tickers)), min(_wl_dates))
    _cards_html = ""
    for w in watchlist:
        tk  = w["ticker"]
        sc  = dict(score_map.get(tk, {}) or {})
        adj = float(sc.get("adj_composite", sc.get("composite", 0)) or 0)
        if sc:
            quant = float(sc.get("composite", adj) or adj)
            sc["adj_action"]    = "BUY" if adj >= 60 else ("SELL" if adj < 45 else "HOLD")
            sc["adj_composite"] = adj
            sc["composite"]     = quant
            # Always recompute macro impact so the MACRO box matches the screener.
            sc["score_delta"]   = round(adj - quant, 1)
        else:
            sc = {"ticker":tk,"adj_action":"N/A","adj_composite":0,"composite":0,
                  "momentum":0,"quality":0,"volume":0,"value":0,"sentiment":0,"score_delta":0}
        # Standard card field set: stamp sector (signal_log carries no sector
        # column) so the meta line reads "$price · date · Sector" like the screener.
        sc["ticker"] = tk
        sc["sector"] = sc.get("sector") or _WL_SECTORS.get(tk, "")
        ci = get_company_info(tk)
        # ── P&L strip: always two columns (SINCE ADDED + TODAY) for a uniform
        #    layout across every card; missing data shows "—" rather than
        #    dropping the column (which made cards look inconsistent).
        _entry_px = w.get("price_at_add")
        _dc = day_change.get(tk) or {}
        _cur_px = _dc.get("price") or sc.get("price")

        # SINCE ADDED
        _since_inner = '<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:#8896ac;">—</div>'
        try:
            if _entry_px and _cur_px:
                _ep = float(_entry_px); _cp = float(_cur_px)
                if _ep > 0:
                    _ret = (_cp - _ep) / _ep * 100
                    _rc = "#34d399" if _ret > 0 else ("#f87171" if _ret < 0 else "#b3bed0")
                    _sign = "+" if _ret >= 0 else ""
                    _since_inner = (
                        f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:{_rc};">{_sign}{_ret:.1f}%</div>'
                        f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;">${_ep:,.2f} → ${_cp:,.2f}</div>'
                    )
        except Exception:
            pass
        _seg_since = (
            f'<div style="flex:1;text-align:center;">'
            f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.08em;margin-bottom:2px;">SINCE ADDED</div>'
            f'{_since_inner}</div>'
        )

        # TODAY — close-to-close move; live intraday, frozen at the close.
        _today_inner = '<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:#8896ac;">—</div>'
        try:
            if _dc.get("chg_pct") is not None:
                _tc_pct = float(_dc["chg_pct"]); _tc_d = float(_dc.get("chg_dollar", 0))
                _tcc = "#34d399" if _tc_pct > 0 else ("#f87171" if _tc_pct < 0 else "#b3bed0")
                _tsign = "+" if _tc_pct >= 0 else ""
                _tsub = f'{_tsign}${_tc_d:,.2f}' + (' · at close' if _dc.get("settled") else '')
                _today_inner = (
                    f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:{_tcc};">{_tsign}{_tc_pct:.2f}%</div>'
                    f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;">{_tsub}</div>'
                )
        except Exception:
            pass
        _seg_today = (
            f'<div style="flex:1;text-align:center;border-left:1px solid rgba(255,255,255,.05);">'
            f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.08em;margin-bottom:2px;">TODAY</div>'
            f'{_today_inner}</div>'
        )
        # Extended-hours segment — live session only (pre before open, AH after
        # close). Same merged source as the per-stock line; its own column.
        _seg_xh = ""
        _xphase = _market_phase()
        _xkk = "pre" if _xphase == "pre" else ("post" if _xphase in ("post", "closed") else None)
        if _xkk and _dc.get(f"{_xkk}_pct") is not None:
            _xhp = float(_dc[f"{_xkk}_pct"]); _xhd = _dc.get(f"{_xkk}_chg")
            _xhc = "#34d399" if _xhp > 0 else ("#f87171" if _xhp < 0 else "#b3bed0")
            _xhlbl = "PRE&nbsp;MKT" if _xkk == "pre" else "AFTER&nbsp;HRS"
            _xhsign = "+" if _xhp >= 0 else ""
            _xhsub = (f'{"+" if (_xhd or 0) >= 0 else "-"}${abs(_xhd):,.2f}'
                      if _xhd is not None else '')
            _seg_xh = (
                f'<div style="flex:1;text-align:center;border-left:1px solid rgba(255,255,255,.05);">'
                f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.08em;margin-bottom:2px;">{_xhlbl}</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:{_xhc};">{_xhsign}{_xhp:.2f}%</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#94a3b8;">{_xhsub}</div></div>'
            )
        _segments = [_seg_since, _seg_today] + ([_seg_xh] if _seg_xh else [])
        _since_html = ""
        if _segments:
            _since_html = (
                '<div style="display:flex;align-items:stretch;gap:0;'
                'padding:8px 4px;margin:-4px 0 4px 0;background:rgba(255,255,255,.015);'
                'border-radius:0 0 6px 6px;">'
                + "".join(_segments) + '</div>'
            )
        # Per-card: card + P&L strip + Remove button, accumulated for one render.
        _wl_rm_url = (f"?qnav=watchlist&uid={st.query_params.get('uid','')}"
                      f"&plan={st.query_params.get('plan','free')}&ck=1&_n=watchlist"
                      f"&wl_list_action=remove_item&wl_list_id={_active_id}&wl_rm_ticker={tk}")
        # Collapsed-card meta strip — surfaces "added on / today's %" at a glance
        # without making the user expand. The added_at field is set when the
        # watchlist item is created; day_change is the same dict already used by
        # the expanded P&L strip above.
        sc["_summary_meta_html"] = _build_summary_meta_html(
            entry_date=w.get("added_at") or w.get("created_at"),
            day_change_entry=day_change.get(tk),
        )
        sc["_mini_chart_html"] = _build_mini_chart_html(
            tk, w.get("added_at") or w.get("created_at"), _wl_pm, _wl_sm, since_label="since added")
        _cards_html += build_card_html(sc, nav="watchlist", is_gem=(tk in _wl_gems),
                                       company_info=ci, in_list=_wl_active_tickers,
                                       extra_detail=_since_html, remove_url=_wl_rm_url)

    render_cards_batch(_cards_html)

    st.markdown(
        '<div style="padding:8px 14px;background:#050a0f;border:1px solid rgba(255,255,255,.07);'
        'border-radius:0 0 6px 6px;font-size:13px;color:#94a3b8;">'
        'Scores updated daily via nightly refresh · Add stocks via Screener search</div>',
        unsafe_allow_html=True
    )
    st.markdown('</div>', unsafe_allow_html=True)


def _gem_why_tags(r: dict) -> list:
    """Short reason tags for why this stock was surfaced as a Hidden Gem."""
    tags, comp = [], float(r.get("adj_composite", r.get("composite", 50)) or 50)
    mom  = float(r.get("momentum",  50) or 50)
    qual = float(r.get("quality",   50) or 50)
    val  = float(r.get("value",     50) or 50)
    sent = float(r.get("sentiment", 50) or 50)
    vol  = float(r.get("volume",    50) or 50)
    delta = float(r.get("score_delta", 0) or 0)
    if mom  >= 65: tags.append("Momentum leadership")
    if qual >= 65: tags.append("Quality improving")
    if val  >= 65: tags.append("Undervalued")
    if sent >= 60: tags.append("Positive sentiment")
    if vol  >= 65: tags.append("Volume confirming")
    if delta > 2:  tags.append("Macro tailwind")
    if comp >= 68: tags.append("Rising conviction")
    reason = r.get("hidden_gem_reason", "")
    if "coverage" in reason.lower():  tags.append("Low analyst coverage")
    if "insider"  in reason.lower():  tags.append("Insider activity")
    return tags[:3]


def page_gems():
    _pin_nav("gems")
    page_summary(
        "💎", "Hidden Gems",
        "Mid-cap stocks with high conviction scores flying under Wall Street's radar"
    )

    if not is_pro():
        st.markdown("""
        <div style="margin:0 32px;background:rgba(52,211,153,.04);border:1px solid rgba(52,211,153,.2);
             border-radius:8px;padding:28px 24px;text-align:center;">
          <div style="font-size:48px;margin-bottom:16px;">🔒</div>
          <div style="font-family:'Syne',sans-serif;font-size:24px;font-weight:800;
               color:#34d399;margin-bottom:12px;">Founding Member Feature</div>
          <div style="color:#9fabc0;max-width:480px;margin:0 auto;line-height:1.7;margin-bottom:24px;">
            Hidden Gem detection is free for the first 50 founding members.
            These are mid- and small-cap stocks with institutional-grade factor scores that
            fly under Wall Street's radar — the ones that show up before the crowd notices.
          </div>
          <div style="background:rgba(52,211,153,.08);border:1px solid rgba(52,211,153,.3);
               border-radius:6px;padding:16px 24px;display:inline-block;margin-bottom:24px;">
            <div style="font-family:'DM Mono',monospace;font-size:13px;color:#34d399;">
              🎯 Preview: CELH — Revenue +62% YoY · Earnings +148% · Beat 4/4 quarters
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.session_state.get("logged_in"):
            st.markdown(_cta_gold("Join Founding Members — Unlock Now", _upgrade_url("Hidden Gems", "gems")), unsafe_allow_html=True)
        else:
            st.markdown(_cta_gold("Join Free — First 50 Spots", _upgrade_url("Hidden Gems","screener")), unsafe_allow_html=True)
        return

    # Use exactly same data pipeline as screener — guarantees matching gem count
    if st.session_state.scan_results is None:
        st.markdown(
            '<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
            'letter-spacing:.08em;margin-bottom:8px;">LOADING UNIVERSE SCORES</div>',
            unsafe_allow_html=True)
        _gems_prog = st.progress(0, text="Fetching scores...")
        try:
            from model_engine import fetch_macro_overlay, apply_macro_overlay
            from model_engine import SECTORS as _GEM_SECTORS
            _gems_prog.progress(15, text="Running factor model...")
            _raw = _cached_full_scan()
            _gems_prog.progress(60, text="Applying macro overlay...")
            _mac = _live_macro()
            for _r in _raw:
                if not _r.get("sector") or _r.get("sector") == "Unknown":
                    _r["sector"] = _GEM_SECTORS.get(_r["ticker"], "Unknown")
            _res = apply_macro_overlay(_raw, _mac)
            _gems_prog.progress(85, text="Detecting gems...")
            _enriched = enrich_with_signal_log(_res)
            # Single source of truth — finalize from signal_log values, no recompute
            st.session_state.scan_results = finalize_scores_from_signal_log(_enriched, _mac)
            st.session_state.macro_data   = _mac
            _gems_prog.progress(100, text="Done")
            _gems_prog.empty()
        except Exception as _ge:
            _gems_prog.empty()
            st.error(f"Failed to load scores: {_ge}")
            return

    _macro_gems = st.session_state.get("macro_data") or {}
    gems = detect_hidden_gems(st.session_state.scan_results, macro_data=_macro_gems)

    # Resolve the macro regime up front so BOTH the empty-state branch and the
    # normal render path can reference it. Previously `regime` was only assigned
    # below the empty-state block; because Python scopes it as a function-local
    # for the whole function, the empty-state reference raised UnboundLocalError
    # whenever `gems` came back empty.
    if not st.session_state.get("macro_data"):
        try:
            st.session_state.macro_data = _live_macro()
        except Exception:
            st.session_state.macro_data = {}
    regime = st.session_state.get("macro_data", {}).get("regime", "NEUTRAL")
    regime_colors = {"RISK_OFF":"#f87171","HIGH VOLATILITY":"#f97316","RISK_ON":"#34d399","MILDLY BULLISH":"#4ade80","NEUTRAL":"#d4a843"}
    regime_color = regime_colors.get(regime, "#d4a843")

    if not gems:
        st.markdown(
            '<div style="padding:48px 32px;text-align:center;">'
            '<div style="font-size:40px;margin-bottom:16px;">💎</div>'
            '<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:700;color:#8896ac;margin-bottom:8px;">'
            'No hidden gems today</div>'
            '<div style="font-size:13px;color:#94a3b8;max-width:320px;margin:0 auto;line-height:1.7;">'
            'The model detected no mid- or small-cap stocks clearing the high-conviction threshold in the current macro regime. '
            'Check back after the next nightly scan or visit the Screener to explore all signals.'
            '</div>'
            '<div style="margin-top:20px;font-family:DM Mono,monospace;font-size:13px;color:#1e293b;letter-spacing:.08em;">'
            f'THRESHOLD: {"67+" if regime in ("RISK_OFF","HIGH VOLATILITY") else "62+"} IN {regime} REGIME'
            '</div></div>',
            unsafe_allow_html=True)
        return

    st.markdown(f'<div style="padding:0 32px;">', unsafe_allow_html=True)
    st.markdown(f"""
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:16px;">
      <div style="font-size:13px;color:#b3bed0;">{len(gems)} hidden {'gem' if len(gems)==1 else 'gems'} identified</div>
      <div style="font-size:13px;color:{regime_color};font-family:DM Mono,monospace;">
        Regime: {regime} · {"Threshold 67+" if regime in ("RISK_OFF","HIGH VOLATILITY") else "Threshold 60+" if regime in ("RISK_ON","MILDLY BULLISH") else "Threshold 62+"}
      </div>
    </div>
    """, unsafe_allow_html=True)

    # Load current watchlist to know which gems are already added
    wl_tickers = {w["ticker"] for w in get_watchlist(uid())}

    # When each current gem was first flagged is_hidden_gem in signal_log — used
    # as the start date for a vs-SPY mini chart in the expanded card.
    _gem_first = {}
    _gem_tks = [g.get("ticker", "") for g in gems if g.get("ticker")]
    try:
        from data_refresh import _get_supabase as _gs
        _sbg = _gs()
        if _sbg and _gem_tks:
            _gr = (_sbg.table("signal_log").select("ticker,signal_date")
                   .in_("ticker", _gem_tks).eq("is_hidden_gem", True)
                   .order("signal_date", desc=False).execute())
            for _row in (_gr.data or []):
                _t = _row.get("ticker"); _d = str(_row.get("signal_date") or "")[:10]
                if _t and _d and _t not in _gem_first:   # asc order → earliest wins
                    _gem_first[_t] = _d
    except Exception:
        pass
    _gem_pm, _gem_sm = ({}, {})
    _gem_trail = _trail_start(30)
    if _gem_tks:
        _starts = list(_gem_first.values()) + [_gem_trail]
        _gem_pm, _gem_sm = _mini_price_data(tuple(sorted(set(_gem_tks))), min(_starts))

    # Gems use same collapsed card pattern — batched into one markdown render
    _gems_html = ""
    for g in gems:
        tk = g.get("ticker", "")
        try:
            adj   = float(g.get("adj_composite") or g.get("composite") or 0)
            raw   = float(g.get("composite") or 0)
            g["adj_action"]  = "BUY"
            g["score_delta"] = round(adj - raw, 1)
            if not g.get("sector") or g.get("sector") == "Unknown":
                g["sector"] = SECTORS.get(tk, "")
            ci = get_company_info(tk)
            _gstart = _gem_first.get(tk) or _gem_trail
            _glabel = "since flagged" if _gem_first.get(tk) else "vs SPY · 20d"
            g["_mini_chart_html"] = _build_mini_chart_html(
                tk, _gstart, _gem_pm, _gem_sm, since_label=_glabel)
            _gems_html += build_card_html(g, nav="gems", is_gem=True,
                                          company_info=ci, in_list=wl_tickers)
        except Exception:
            pass
    render_cards_batch(_gems_html)


# ══════════════════════════════════════════════════════════════════════════════
# BACKTEST PAGE
# ══════════════════════════════════════════════════════════════════════════════
def _stored_close_frame(sb, tickers, inception):
    """Build the daily-close price frame for the equity curve from data already
    in Supabase — held-ticker prices from signal_log, SPY from benchmark_price —
    so the Model Portfolio page needs no live yfinance pull. Returns a forward-
    filled DataFrame (index = trading days, columns = SPY + held tickers) or None
    when stored coverage is insufficient (caller then falls back to yfinance).
    The downstream ledger replay is identical regardless of source."""
    try:
        import pandas as pd
        from data_refresh import _fetch_all_rows
        # SPY benchmark is the date spine.
        brows = _fetch_all_rows(lambda: sb.table("benchmark_price")
                                .select("d,close").gte("d", inception)
                                .order("d", desc=False))
        spy = {str(r["d"])[:10]: float(r["close"]) for r in (brows or [])
               if r.get("close") is not None}
        if len(spy) < 2:
            return None
        dates = sorted(spy.keys())
        # Held-ticker daily closes (signal_log is keyed ticker,signal_date).
        srows = _fetch_all_rows(lambda: sb.table("signal_log")
                                .select("ticker,signal_date,price")
                                .in_("ticker", list(tickers))
                                .gte("signal_date", inception))
        cols = {"SPY": spy}
        for r in (srows or []):
            tk = r.get("ticker"); pv = r.get("price")
            d = str(r.get("signal_date") or "")[:10]
            if tk and pv is not None and len(d) == 10:
                cols.setdefault(tk, {})[d] = float(pv)
        frame = pd.DataFrame(index=pd.to_datetime(dates))
        for col, m in cols.items():
            frame[col] = [m.get(d) for d in dates]
        frame = frame.ffill()
        # Stored-only: the frame already includes today's row (signal_log +
        # benchmark_price are kept fresh by the intraday price cron), so the
        # curve ends on today's stored mark with no live quote fetch on the
        # request path.
        return frame
    except Exception:
        return None


def _track_record_data(sb):
    """Build the live Model Portfolio equity curve vs SPY from real data.

    Ledger replay over SPY trading days: $100K notional, $2,000/position, daily
    mark-to-market using signal_log closes (forward-filled across gaps); realized
    proceeds from exits are held as cash until redeployed. SPY benchmark is $100K
    invested at inception and marked daily. Returns a dict, or None on any
    failure / no positions (caller shows the honest fallback)."""
    if not sb:
        return None
    try:
        import yfinance as yf
        try:
            from universe_data import SECTORS, HELD_SECTOR_OVERRIDES
            # Merge held-but-dropped overrides so out-of-universe holdings (e.g.
            # STX/CTRA, or any May-seed name not yet back in the universe) resolve
            # to their real sector instead of "Unknown". No key overlap with
            # SECTORS by construction, so the universe values are never shadowed.
            _SEC = {**SECTORS, **HELD_SECTOR_OVERRIDES}
        except Exception:
            _SEC = {}

        resp = sb.table("model_portfolio_positions").select("*") \
            .eq("epoch", MODEL_EPOCH) \
            .order("entry_date", desc=False).execute()
        rows = resp.data or []
        if not rows:
            return None
        # Dedup exact (ticker, entry_date, exit_date); newest id wins
        seen = {}
        for p in rows:
            k = (p["ticker"], str(p.get("entry_date") or "")[:10], str(p.get("exit_date") or "")[:10])
            if k not in seen or (p.get("id") or 0) > (seen[k].get("id") or 0):
                seen[k] = p
        positions = list(seen.values())

        POS_SIZE = 2000.0
        inception = min(str(p["entry_date"])[:10] for p in positions)

        # Daily close history for SPY + every held ticker in ONE yfinance pull.
        # Same price source the Model Portfolio page uses, so the two pages
        # reconcile (signal_log can lag a session behind when the batch gate is
        # holding). Forward-filled across any gaps; SPY index is the date axis.
        tickers = sorted({p["ticker"] for p in positions})
        # Prefer the stored price history (signal_log + benchmark_price) so the
        # page does no live network pull. Fall back to a live yfinance download
        # only when stored coverage is insufficient (e.g. brand-new cohort before
        # benchmark_price has two sessions). Downstream replay is identical.
        close = _stored_close_frame(sb, tickers, inception)
        if close is None:
            dl = yf.download(_strip_delisted(tickers) + ["SPY"], start=inception,
                             progress=False, auto_adjust=True)
            if dl.empty:
                return None
            close = dl["Close"]
            if not hasattr(close, "columns") or "SPY" not in close.columns:
                return None
            close = close.ffill()
        spy_close = close["SPY"]
        dates = [d.date().isoformat() for d in close.index]
        if len(dates) < 2:
            # Same-day cohort (inception == today): the daily ledger needs two
            # sessions to replay, so instead of bailing (which blanks the cards'
            # SPY and the whole chart), build a minimal point-in-time result from
            # the latest vs prior close. The full daily ledger takes over
            # automatically once a second session is recorded.
            import datetime as _dt
            _td = _dt.date.today().isoformat()
            try:
                _rec = yf.download(_strip_delisted(tickers) + ["SPY"], period="5d",
                                   progress=False, auto_adjust=True)
                _rc = _rec["Close"].ffill()
            except Exception:
                _rc = None

            def _cur(tk):
                try:    return float(_rc[tk].iloc[-1])
                except Exception: return None

            def _prev(tk):
                try:    return float(_rc[tk].iloc[-2])
                except Exception: return None

            n_pos = len(positions)
            cash  = 100000.0 - n_pos * POS_SIZE
            cur_val, pm = cash, {}
            for p in positions:
                ep = p.get("entry_price")
                cp = _cur(p["ticker"]) or (float(ep) if ep else None)
                if ep and float(ep) > 0 and cp:
                    cur_val += (POS_SIZE / float(ep)) * cp
                    pm[p["ticker"]] = {_td: cp}
            model_ret = (cur_val / 100000.0 - 1) * 100
            _sc, _sp = _cur("SPY"), _prev("SPY")
            spy_ret  = ((_sc / _sp - 1) * 100) if (_sc and _sp) else 0.0
            spy_val  = 100000.0 * (1 + spy_ret / 100.0)
            _sect = {}
            for p in positions:
                if p.get("is_active"):
                    _s = _SEC.get(p["ticker"], "Unknown")
                    _sect[_s] = _sect.get(_s, 0) + 1
            return {
                "inception": inception,
                "model_series": [(inception, 100000.0), (_td, cur_val)],
                "spy_series":   [(inception, 100000.0), (_td, spy_val)],
                "model_ret": model_ret, "spy_ret": spy_ret,
                "day_model": model_ret, "day_spy": spy_ret,
                "model_value": cur_val, "spy_value": spy_val,
                "exits": [], "basis": 100000.0,
                "sector_counts": sorted(_sect.items(), key=lambda x: x[1], reverse=True),
                "price_map": pm, "spy_map": ({_td: _sc} if _sc else {}),
                "n_active": sum(1 for p in positions if p.get("is_active")),
                "n_exited": 0, "n_sessions": 1,
            }

        price_map = {}
        for tk in tickers:
            if tk in close.columns:
                price_map[tk] = {d.date().isoformat(): float(v)
                                 for d, v in close[tk].items() if v == v}

        def price_on(tk, d, fallback):
            m = price_map.get(tk)
            if not m:
                return fallback
            if d in m:
                return m[d]
            prior = [dt for dt in m if dt <= d]
            return m[max(prior)] if prior else fallback

        # TRUE equity curve — a real portfolio ledger. Start with $100K; each
        # entry deploys $2K (shares = $2K / entry price); each exit returns
        # shares x exit price to cash (so realized P&L, e.g. BORR's loss, is
        # permanently reflected); every session the whole book is marked to the
        # close. Endpoint = $100K + realized P&L + unrealized P&L = the true
        # account value. The poisoned-batch exits were reversed in the data, so
        # they carry no exit record here and are correctly excluded.
        BASE = 100000.0
        entries_by_date, exits_by_date = {}, {}
        for p in positions:
            entries_by_date.setdefault(str(p["entry_date"])[:10], []).append(p)
            xd = str(p.get("exit_date") or "")[:10]
            if len(xd) == 10:
                exits_by_date.setdefault(xd, []).append(p)

        cash = BASE
        open_lots = {}
        model_series = []
        for d in dates:
            for p in exits_by_date.get(d, []):
                lot = open_lots.pop(p.get("id"), None)
                if lot:
                    xp = p.get("exit_price") or price_on(lot["ticker"], d, lot["entry_price"])
                    cash += lot["shares"] * float(xp)
            for p in entries_by_date.get(d, []):
                ep = p.get("entry_price")
                if not ep or float(ep) <= 0:
                    continue
                open_lots[p.get("id")] = {"ticker": p["ticker"],
                                          "shares": POS_SIZE / float(ep),
                                          "entry_price": float(ep)}
                cash -= POS_SIZE
            mtm = cash + sum(lot["shares"] * price_on(lot["ticker"], d, lot["entry_price"])
                             for lot in open_lots.values())
            model_series.append((d, mtm))

        basis = BASE
        spy0 = float(spy_close.iloc[0])
        spy_series = [(d, BASE * float(v) / spy0)
                      for d, v in zip(dates, spy_close.values)]

        m_last, s_last = model_series[-1][1], spy_series[-1][1]
        model_ret = (m_last / BASE - 1) * 100
        spy_ret   = (s_last / BASE - 1) * 100
        day_model = ((model_series[-1][1] / model_series[-2][1] - 1) * 100
                     if model_series[-2][1] else 0.0)
        day_spy   = ((spy_series[-1][1] / spy_series[-2][1] - 1) * 100
                     if spy_series[-2][1] else 0.0)

        exits = []
        for p in positions:
            xd = str(p.get("exit_date") or "")[:10]
            if len(xd) == 10 and p.get("exit_price") and p.get("entry_price"):
                try:
                    ret = (float(p["exit_price"]) / float(p["entry_price"]) - 1) * 100
                except Exception:
                    ret = 0.0
                exits.append({"ticker": p["ticker"], "sector": _SEC.get(p["ticker"], "\u2014"),
                              "entry_date": str(p["entry_date"])[:10], "exit_date": xd,
                              "ret": ret, "reason": p.get("exit_reason") or "\u2014"})
        exits.sort(key=lambda x: x["exit_date"], reverse=True)

        sect = {}
        for p in positions:
            if p.get("is_active"):
                s = _SEC.get(p["ticker"], "Unknown")
                sect[s] = sect.get(s, 0) + 1
        sector_counts = sorted(sect.items(), key=lambda x: x[1], reverse=True)

        # Daily close paths for per-stock mini charts (stock vs SPY since entry).
        spy_map = {d: float(v) for d, v in zip(dates, spy_close.values) if v == v}

        return {
            "inception": inception, "model_series": model_series, "spy_series": spy_series,
            "model_ret": model_ret, "spy_ret": spy_ret, "day_model": day_model, "day_spy": day_spy,
            "model_value": m_last, "spy_value": s_last, "exits": exits, "basis": basis,
            "sector_counts": sector_counts, "price_map": price_map, "spy_map": spy_map,
            "n_active": sum(1 for p in positions if p.get("is_active")),
            "n_exited": len(exits), "n_sessions": len(dates),
        }
    except Exception:
        return None


def _mini_vs_spy_svg(stock_pairs, spy_pairs):
    """Tiny sparkline for an expanded position card: the holding's % return
    (gold) vs SPY (slate) over the same window, normalized to 0% at the start.
    Inputs are [(date, price), ...]. Returns '' if there isn't enough data."""
    if not stock_pairs or len(stock_pairs) < 2 or not spy_pairs or len(spy_pairs) < 2:
        return ""
    def norm(pairs):
        base = pairs[0][1]
        if not base:
            return None
        return [(d, (v / base - 1) * 100) for d, v in pairs]
    s, k = norm(stock_pairs), norm(spy_pairs)
    if not s or not k:
        return ""
    W, H, P = 260, 58, 5
    allv = [v for _, v in s] + [v for _, v in k]
    lo, hi = min(allv), max(allv)
    if hi == lo:
        hi = lo + 1
    pad = (hi - lo) * 0.14
    lo -= pad; hi += pad
    def X(i, n): return P + (W - 2 * P) * (i / (n - 1))
    def Y(v): return P + (H - 2 * P) * (1 - (v - lo) / (hi - lo))
    def path(arr): return "M " + " L ".join(f"{X(i, len(arr)):.1f},{Y(v):.1f}" for i, (_, v) in enumerate(arr))
    zero = ""
    if lo <= 0 <= hi:
        y0 = Y(0)
        zero = (f'<line x1="{P}" y1="{y0:.1f}" x2="{W-P}" y2="{y0:.1f}" '
                f'stroke="rgba(255,255,255,.10)" stroke-dasharray="2,3"/>')
    return (f'<svg viewBox="0 0 {W} {H}" width="100%" style="display:block;max-width:300px;">{zero}'
            f'<path d="{path(k)}" fill="none" stroke="#7c8aa0" stroke-width="1.4" stroke-linejoin="round"/>'
            f'<path d="{path(s)}" fill="none" stroke="#d4a843" stroke-width="2" stroke-linejoin="round"/></svg>')


def _trail_start(days: int = 30):
    """ISO date ~`days` calendar days back — ~20 trading sessions at 30 days,
    used as the window for trailing vs-SPY mini charts (top/bottom 10, and the
    fallback for gems with no recorded flag date)."""
    from datetime import date, timedelta
    return (date.today() - timedelta(days=days)).isoformat()


@st.cache_data(ttl=1800, show_spinner=False)
def _mini_price_data(tickers_tuple: tuple, start: str):
    """Batch-download daily closes for an arbitrary set of tickers + SPY since
    `start`, for per-stock vs-SPY mini charts outside the model portfolio
    (watchlist, hidden gems). Returns (price_map, spy_map). Cached per arg set."""
    try:
        import yfinance as yf
        import pandas as pd
        syms = sorted({t for t in tickers_tuple if t and t not in DELISTED} | {"SPY"})
        if len(syms) < 2 or not start:
            return {}, {}
        raw = yf.download(syms, start=start, auto_adjust=True, progress=False)
        if raw is None or len(raw) == 0:
            return {}, {}
        close = raw["Close"] if isinstance(raw.columns, pd.MultiIndex) and "Close" in raw.columns.get_level_values(0) else raw
        if isinstance(close, pd.Series):
            close = close.to_frame(name=syms[0])
        price_map = {}
        for tk in syms:
            if tk in close.columns:
                col = close[tk].dropna()
                if len(col):
                    price_map[tk] = {d.date().isoformat(): float(v) for d, v in col.items() if v == v}
        return price_map, price_map.get("SPY", {})
    except Exception:
        return {}, {}


def _build_mini_chart_html(ticker, entry_date, price_map, spy_map, since_label="since entry"):
    """Wrap _mini_vs_spy_svg with a label for injection into an expanded card.
    Slices each series from the position's entry date so the comparison covers
    the holding period. Returns '' on insufficient data."""
    try:
        pm = (price_map or {}).get(ticker) or {}
        sm = spy_map or {}
        e = str(entry_date)[:10]
        sdates = sorted(d for d in pm if d >= e)
        if len(sdates) < 2:
            return ""
        stock_pairs = [(d, pm[d]) for d in sdates]
        spy_pairs   = [(d, sm[d]) for d in sdates if d in sm]
        svg = _mini_vs_spy_svg(stock_pairs, spy_pairs)
        if not svg:
            return ""
        s_ret = (stock_pairs[-1][1] / stock_pairs[0][1] - 1) * 100 if stock_pairs[0][1] else 0
        k_ret = (spy_pairs[-1][1] / spy_pairs[0][1] - 1) * 100 if spy_pairs and spy_pairs[0][1] else 0
        ss = "+" if s_ret >= 0 else ""; ks = "+" if k_ret >= 0 else ""
        return (
            '<div style="padding:10px 18px 14px;border-top:1px solid rgba(255,255,255,.05);">'
            '<div style="display:flex;gap:14px;align-items:center;font-family:DM Mono,monospace;font-size:11px;margin-bottom:6px;">'
            f'<span style="color:#d4a843;">\u2014 {ticker} {ss}{s_ret:.1f}%</span>'
            f'<span style="color:#7c8aa0;">\u2014 SPY {ks}{k_ret:.1f}%</span>'
            '<span style="color:#8896ac;margin-left:auto;">' + since_label + '</span></div>'
            f'{svg}</div>'
        )
    except Exception:
        return ""


def _tr_line_chart_svg(model_series, spy_series, intraday=False):
    """Polished two-line area chart (gold = model, slate = SPY) of $ book value
    over time, with gradient fills, value gridlines + $ labels, date ticks, a
    $100K reference, and labeled end-points. When intraday=True, x-axis shows
    clock times and every point carries a hover tooltip (timestamp · value)."""
    if not model_series or len(model_series) < 2:
        return ""
    dates  = [d for d, _ in model_series]
    m_vals = [v for _, v in model_series]
    s_vals = [v for _, v in spy_series]
    W, H = 760, 300
    PL, PR, PT, PB = 52, 70, 16, 30
    lo, hi = min(min(m_vals), min(s_vals)), max(max(m_vals), max(s_vals))
    if hi == lo:
        hi = lo + 1
    pad = (hi - lo) * 0.10
    lo -= pad; hi += pad
    n = len(m_vals)
    BASE = H - PB
    def X(i): return PL + (W - PL - PR) * (i / (n - 1))
    def Y(v): return PT + (BASE - PT) * (1 - (v - lo) / (hi - lo))
    def line(vals): return "M " + " L ".join(f"{X(i):.1f},{Y(v):.1f}" for i, v in enumerate(vals))
    def area(vals):
        return (f"M {X(0):.1f},{BASE:.1f} L "
                + " L ".join(f"{X(i):.1f},{Y(v):.1f}" for i, v in enumerate(vals))
                + f" L {X(n-1):.1f},{BASE:.1f} Z")
    def fmtk(v): return f"${v/1000:.1f}K"
    p = [
        '<defs>'
        '<linearGradient id="trGold" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0%" stop-color="#d4a843" stop-opacity="0.28"/>'
        '<stop offset="100%" stop-color="#d4a843" stop-opacity="0"/></linearGradient>'
        '<linearGradient id="trSpy" x1="0" y1="0" x2="0" y2="1">'
        '<stop offset="0%" stop-color="#9fabc0" stop-opacity="0.12"/>'
        '<stop offset="100%" stop-color="#9fabc0" stop-opacity="0"/></linearGradient>'
        '</defs>'
    ]
    # value gridlines + $ labels
    for k in range(4):
        tv = lo + (hi - lo) * k / 3.0
        y = Y(tv)
        p.append(f'<line x1="{PL}" y1="{y:.1f}" x2="{W-PR}" y2="{y:.1f}" stroke="rgba(255,255,255,.045)"/>')
        p.append(f'<text x="{PL-8}" y="{y+3:.1f}" text-anchor="end" font-family="DM Mono,monospace" font-size="10" fill="#8896ac">{fmtk(tv)}</text>')
    # $100K reference
    if lo <= 100000.0 <= hi:
        y0 = Y(100000.0)
        p.append(f'<line x1="{PL}" y1="{y0:.1f}" x2="{W-PR}" y2="{y0:.1f}" stroke="rgba(255,255,255,.15)" stroke-dasharray="2,4"/>')
    # x-axis ticks (dates for daily windows, clock times for intraday)
    if intraday:
        _tick_idx = sorted(set(round(i * (n - 1) / 4) for i in range(5)))
        def _xlab(i): return dates[i]
    else:
        _tick_idx = (0, n // 2, n - 1)
        def _xlab(i): return dates[i][5:].replace("-", "/")
    for i in _tick_idx:
        anc = "start" if i == 0 else ("end" if i == n - 1 else "middle")
        p.append(f'<text x="{X(i):.1f}" y="{H-8}" text-anchor="{anc}" font-family="DM Mono,monospace" font-size="10" fill="#8896ac">{_xlab(i)}</text>')
    # area fills
    p.append(f'<path d="{area(s_vals)}" fill="url(#trSpy)" stroke="none"/>')
    p.append(f'<path d="{area(m_vals)}" fill="url(#trGold)" stroke="none"/>')
    # lines (rounded)
    p.append(f'<path d="{line(s_vals)}" fill="none" stroke="#7c8aa0" stroke-width="1.6" stroke-linejoin="round" stroke-linecap="round"/>')
    p.append(f'<path d="{line(m_vals)}" fill="none" stroke="#d4a843" stroke-width="2.6" stroke-linejoin="round" stroke-linecap="round"/>')
    # intraday: a dot + hover tooltip (timestamp · value) at every mark
    if intraday:
        for i, (d, v) in enumerate(spy_series):
            p.append(f'<circle cx="{X(i):.1f}" cy="{Y(v):.1f}" r="2" fill="#7c8aa0" opacity="0.7"><title>{d} · {fmtk(v)} · SPY</title></circle>')
        for i, (d, v) in enumerate(model_series):
            p.append(f'<circle cx="{X(i):.1f}" cy="{Y(v):.1f}" r="2.4" fill="#d4a843"><title>{d} · {fmtk(v)} · Model</title></circle>')
    # period % change over the visible window (end vs start) — shown at each endpoint
    def _pct(v): return (v[-1] / v[0] - 1) * 100 if v and v[0] else 0.0
    def _pcol(x): return "#34d399" if x >= 0 else "#f87171"
    m_pct, s_pct = _pct(m_vals), _pct(s_vals)
    _pd = 2 if intraday else 1  # DAY window: 2 decimals so it matches the TODAY card
    # start-point dots + value labels (left)
    sx = X(0)
    smy, ssy = Y(m_vals[0]), Y(s_vals[0])
    bm, bs = smy, ssy
    if abs(bm - bs) < 13:
        if bm <= bs: bm -= 7; bs += 7
        else:        bm += 7; bs -= 7
    p.append(f'<circle cx="{sx:.1f}" cy="{ssy:.1f}" r="3" fill="#7c8aa0"/>')
    p.append(f'<circle cx="{sx:.1f}" cy="{smy:.1f}" r="3.5" fill="#d4a843"/>')
    p.append(f'<text x="{sx+8:.1f}" y="{bs+3:.1f}" font-family="DM Mono,monospace" font-size="10" fill="#8896ac">{fmtk(s_vals[0])}</text>')
    p.append(f'<text x="{sx+8:.1f}" y="{bm+3.5:.1f}" font-family="DM Mono,monospace" font-size="10" font-weight="700" fill="#caa23f">{fmtk(m_vals[0])}</text>')
    # end-points: dots + value labels + period % (nudge apart if colliding)
    my, sy = Y(m_vals[-1]), Y(s_vals[-1])
    lm, ls = my, sy
    if abs(lm - ls) < 26:
        _mid = (lm + ls) / 2.0
        if lm <= ls: lm, ls = _mid - 13, _mid + 13
        else:        lm, ls = _mid + 13, _mid - 13
    ex = X(n - 1)
    p.append(f'<circle cx="{ex:.1f}" cy="{sy:.1f}" r="3" fill="#7c8aa0"/>')
    p.append(f'<circle cx="{ex:.1f}" cy="{my:.1f}" r="3.5" fill="#d4a843"/>')
    p.append(f'<text x="{ex+8:.1f}" y="{ls+3:.1f}" font-family="DM Mono,monospace" font-size="10" fill="#b3bed0">{fmtk(s_vals[-1])}</text>')
    p.append(f'<text x="{ex+8:.1f}" y="{ls+15:.1f}" font-family="DM Mono,monospace" font-size="9" fill="{_pcol(s_pct)}">{s_pct:+.{_pd}f}%</text>')
    p.append(f'<text x="{ex+8:.1f}" y="{lm+3.5:.1f}" font-family="DM Mono,monospace" font-size="11" font-weight="700" fill="#d4a843">{fmtk(m_vals[-1])}</text>')
    p.append(f'<text x="{ex+8:.1f}" y="{lm+15.5:.1f}" font-family="DM Mono,monospace" font-size="9" fill="{_pcol(m_pct)}">{m_pct:+.{_pd}f}%</text>')
    return f'<svg viewBox="0 0 {W} {H}" width="100%" style="display:block;">' + "".join(p) + '</svg>'


@st.cache_data(ttl=60, show_spinner=False)
def _track_record_cached():
    """Process-wide cache of the model-portfolio ledger. The portfolio is
    identical for every user, so this computes ONCE per TTL for the whole server
    (shared across all sessions and reruns) instead of once per session — first
    hit per window pays, everyone else is instant, and it cuts yfinance calls.
    Raises on an empty result so a transient miss is NOT cached as 'no data'
    (st.cache_data does not store a value when the call raises)."""
    from data_refresh import _get_supabase
    sb = _get_supabase()
    pt = _track_record_data(sb) if sb else None
    if not pt:
        raise RuntimeError("track-record unavailable")
    return pt


def _portfolio_truth(sb, ttl=300):
    """Single source of truth for portfolio totals, shared by the Model
    Portfolio and Track Record pages so they report identical numbers. Backed by
    a process-wide cache (see _track_record_cached) so all users and reruns share
    one computation. On a transient miss, tries once uncached, then gives up to
    the page's honest fallback. `ttl` is retained for signature compatibility;
    the effective TTL now lives on the cache decorator."""
    try:
        return _track_record_cached()
    except Exception:
        try:
            return _track_record_data(sb)
        except Exception:
            return None


def page_backtest():
    # "Track Record" was merged into the combined Portfolio & Track Record
    # page. Kept as an alias so existing ?qnav=backtest links still resolve.
    return page_model_portfolio()


def _make_excel(rows: list, headers: list, sheet_name: str = "Export") -> bytes:
    """Generate an in-memory Excel file from a list of dicts. Returns bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    header_fill = PatternFill("solid", start_color="0D1117")
    header_font = Font(name="Arial", bold=True, color="D4A843", size=10)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    row_font  = Font(name="Arial", size=10)
    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(headers, 1):
            val = row.get(h, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = row_font
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", start_color="0A0B14")

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

def page_portfolio():
    _pin_nav("portfolio")
    # Handle remove position via URL action
    _pa = st.query_params.get("port_action","")
    _pt = st.query_params.get("port_ticker","")
    if _pa == "remove" and _pt:
        try:
            delete_holding(uid(), _pt.upper())
        except Exception:
            pass
        st.query_params.pop("port_action", None)
        st.query_params.pop("port_ticker", None)
        st.rerun()
    user = st.session_state.user or {}
    plan = user.get("plan", "free")
    max_h = plan_limit(plan, "max_holdings")
    has_notifs = plan_limit(plan, "notifications")

    page_summary("💼", "My Portfolio", "Position-level conviction scores · signal alerts on change")
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    # ── Ensure scan results ────────────────────────────────────────────────────
    if st.session_state.scan_results is None:
        with st.spinner("Loading model signals..."):
            raw   = _cached_full_scan()
            macro = _live_macro()
            for r in raw:
                if not r.get("sector") or r.get("sector") == "Unknown":
                    r["sector"] = SECTORS.get(r["ticker"], "Unknown")
            results  = apply_macro_overlay(raw, macro)
            enriched = enrich_with_signal_log(results)
            st.session_state.scan_results = finalize_scores_from_signal_log(enriched, macro)
            st.session_state.macro_data   = macro

    score_map = {s["ticker"]: s for s in st.session_state.scan_results}
    holdings  = get_holdings(uid())
    n_holdings = len(holdings)

    # Always pull holdings' scores directly from signal_log so portfolio matches
    # the screener, home, gems, watchlist exactly. scan_results can drift if the
    # session started on a different page that did its own macro fetch — the
    # nightly cron's adj_composite in signal_log is the canonical value.
    if holdings:
        try:
            from data_refresh import _get_supabase as _pf_sb
            _sb_pf = _pf_sb()
            if _sb_pf:
                _pf_tks = [h["ticker"] for h in holdings]
                _pf_resp = _sb_pf.table("signal_log") \
                    .select("ticker,signal_date,adj_composite,composite,signal,"
                            "momentum,quality,volume,value,sentiment,price,"
                            "is_hidden_gem,hidden_gem_reason") \
                    .in_("ticker", _pf_tks) \
                    .order("signal_date", desc=True) \
                    .execute()
                _pf_seen = set()
                _pf_rows = []
                for _row in (_pf_resp.data or []):
                    if _row["ticker"] not in _pf_seen:
                        _pf_seen.add(_row["ticker"])
                        _pf_rows.append(_row)
                # Normalize via the same finalizer so adj_action/score_delta match
                finalize_scores_from_signal_log(_pf_rows, st.session_state.get("macro_data"))
                for _row in _pf_rows:
                    score_map[_row["ticker"]] = _row
        except Exception:
            pass

    # ── Portfolio conviction summary — single primary card ──────────────
    if holdings and score_map:
        _sc = [float(score_map.get(h["ticker"],{}).get("adj_composite",50) or 50) for h in holdings]
        _hi = sum(1 for x in _sc if x>=60)
        _mo = sum(1 for x in _sc if 45<=x<60)
        _lo = sum(1 for x in _sc if x<45)
        _avg = sum(_sc)/len(_sc)
        _conv_label = "High" if _avg>=60 else ("Low" if _avg<45 else "Moderate")
        _conv_color = "#34d399" if _avg>=60 else ("#f87171" if _avg<45 else "#fbbf24")
        # Trend — compare to previous snapshot avg if available
        _trend_html = ""
        _prev_snap = get_signal_snapshot(uid()) or {}
        if _prev_snap:
            _prev_sc = [float((_prev_snap.get(h["ticker"]) or {}).get("adj_composite",50) or 50) for h in holdings]
            _prev_avg = sum(_prev_sc)/len(_prev_sc) if _prev_sc else _avg
            _delta = _avg - _prev_avg
            if abs(_delta) >= 1:
                _tc = "#34d399" if _delta>0 else "#f87171"
                _ta = "↑" if _delta>0 else "↓"
                _trend_html = f'<span style="font-size:13px;color:{_tc};margin-left:8px;">{_ta} {abs(_delta):.1f} pts</span>'
        # Key risks
        _risks = []
        if _lo > 0: _risks.append(f"{_lo} low conviction position{'s' if _lo>1 else ''}")
        _sectors = []
        try:
            from model_engine import SECTORS as _PORT_SEC
            _sec_count = {}
            for h in holdings:
                s = _PORT_SEC.get(h["ticker"],"")
                if s: _sec_count[s] = _sec_count.get(s,0)+1
            _top_sec = max(_sec_count, key=_sec_count.get) if _sec_count else ""
            if _top_sec and _sec_count[_top_sec] >= 3:
                _risks.append(f"concentration in {_top_sec}")
        except Exception:
            pass
        _risk_html = ""
        if _risks:
            _risk_html = (
                '<div style="margin-top:8px;padding-top:8px;border-top:1px solid rgba(255,255,255,.05);">'
                '<span style="font-size:13px;color:#8896ac;letter-spacing:.06em;">KEY RISKS · </span>'
                + " · ".join(f'<span style="font-size:13px;color:#f87171;">⚠ {r}</span>' for r in _risks)
                + '</div>'
            )
        st.markdown(
            f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
            f'border-radius:10px;padding:16px 20px;margin-bottom:16px;">'
            f'<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.1em;margin-bottom:6px;">PORTFOLIO CONVICTION</div>'
            f'<div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;">'
            f'<div>'
            f'<span style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:{_conv_color};white-space:nowrap;">{_conv_label}</span>'
            f'{_trend_html}'
            f'<div style="font-size:13px;color:#8896ac;margin-top:2px;">avg score {_avg:.0f} · {len(holdings)} positions</div>'
            f'</div>'
            f'<div style="display:flex;gap:12px;">'
            f'<div style="text-align:center;"><div style="font-family:DM Mono,monospace;font-size:16px;color:#34d399;">{_hi}</div><div style="font-size:11px;color:#8896ac;">HIGH</div></div>'
            f'<div style="text-align:center;"><div style="font-family:DM Mono,monospace;font-size:16px;color:#fbbf24;">{_mo}</div><div style="font-size:11px;color:#8896ac;">MOD</div></div>'
            f'<div style="text-align:center;"><div style="font-family:DM Mono,monospace;font-size:16px;color:#f87171;">{_lo}</div><div style="font-size:11px;color:#8896ac;">LOW</div></div>'
            f'</div></div>'
            + _risk_html
            + f'</div>',
            unsafe_allow_html=True
        )


    # ── Plan capacity bar ──────────────────────────────────────────────────────
    if plan == "free" and n_holdings >= 6:
        pct = min(100, int(n_holdings / max_h * 100))
        bar_c = "#f87171" if n_holdings >= max_h else "#fbbf24" if n_holdings >= 8 else "#34d399"
        st.markdown(f"""
        <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
             border-radius:6px;padding:12px 16px;margin-bottom:16px;
             display:flex;align-items:center;gap:16px;">
          <div style="flex:1;">
            <div style="display:flex;justify-content:space-between;margin-bottom:6px;">
              <span style="font-size:14px;color:#b3bed0;">Free plan — positions used</span>
              <span style="font-family:'DM Mono',monospace;font-size:13px;color:{bar_c};">
                {n_holdings} / {max_h}
              </span>
            </div>
            <div style="background:rgba(255,255,255,.06);border-radius:3px;height:4px;overflow:hidden;">
              <div style="width:{pct}%;height:100%;background:{bar_c};border-radius:3px;
                   transition:width .3s;"></div>
            </div>
          </div>
          <div style="font-size:13px;color:#9fabc0;flex-shrink:0;">
            {'<span style="color:#f87171;">At limit</span>' if n_holdings >= max_h else f'{max_h - n_holdings} remaining'}
          </div>
        </div>
        """, unsafe_allow_html=True)



    # ── Check for signal changes (pro users get notifications) ─────────────────
    if holdings and score_map:
        prev_signals = get_signal_snapshot(uid())
        signal_changes = check_and_notify_signal_changes(uid(), plan, score_map, prev_signals)
        save_signal_snapshot(uid(), st.session_state.scan_results)

        if signal_changes:
            _chg_items = []
            for chg in signal_changes:
                ct = chg.get("type","action_change")
                if ct == "action_change" and chg["to"] == "SELL":
                    _chg_items.append(f'<span style="color:#f87171;">▼ {chg["ticker"]}</span>')
                elif ct == "action_change" and chg["to"] == "BUY":
                    _chg_items.append(f'<span style="color:#34d399;">▲ {chg["ticker"]}</span>')
                elif ct == "deterioration":
                    _chg_items.append(f'<span style="color:#fbbf24;">⚠ {chg["ticker"]}</span>')
            if _chg_items:
                st.markdown(
                    '<div style="display:flex;align-items:center;gap:8px;flex-wrap:wrap;'
                    'padding:8px 0;border-bottom:1px solid rgba(255,255,255,.05);margin-bottom:8px;">'
                    '<span style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.1em;">SIGNAL CHANGES</span>'
                    + ' '.join(_chg_items)
                    + '</div>',
                    unsafe_allow_html=True)

    # ── SELL / EXIT signals across portfolio ───────────────────────────────────
    exit_signals = []
    for h in holdings:
        sc = score_map.get(h["ticker"])
        if sc:
            act = sc.get("adj_action", sc.get("action", "HOLD"))
            if act == "SELL":
                exit_signals.append((h["ticker"], sc.get("adj_composite", sc.get("composite",0)), sc.get("signal","")))

    # Exit signals shown in summary card risks

    # ── Add position form ──────────────────────────────────────────────────────
    at_limit = n_holdings >= max_h

    with st.expander("➕ Add Position", expanded=(n_holdings == 0)):
        if at_limit and plan == "free":
            st.markdown("""
            <div style="background:rgba(248,113,113,.06);border:1px solid rgba(248,113,113,.2);
                 border-radius:6px;padding:14px;font-size:13px;color:#f87171;">
              Free plan limit reached (10 positions). Upgrade to Pro for unlimited holdings.
            </div>
            """, unsafe_allow_html=True)
            st.markdown(_cta_gold("Upgrade to Pro — Unlimited Holdings",
                                  _upgrade_url("Unlimited Holdings", "portfolio")),
                        unsafe_allow_html=True)
        else:
            @st.fragment
            def _pf_add_form():
                # Fragment: searching, amount/price entry, and the live preview run
                # reload-free. Only a successful Add does a full rerun, to refresh
                # the holdings list and totals below.
                def _pin_portfolio_nav():
                    st.session_state.nav = "portfolio"
                    _v = st.session_state.get("p_tk", "").strip().upper()
                    st.session_state.p_sel_tk = _v if _v in SECTORS else ""

                # Prefill Price from the platform's last price once a ticker is
                # picked; only set on ticker change, so a user override sticks.
                _picked0 = st.session_state.get("p_sel_tk", "").strip().upper()
                if _picked0 and _picked0 in SECTORS:
                    try:
                        _plat_px = float(score_map.get(_picked0, {}).get("price") or 0)
                    except Exception:
                        _plat_px = 0.0
                    if _plat_px > 0 and st.session_state.get("p_price_for") != _picked0:
                        st.session_state["p_price"] = round(_plat_px, 4)
                        st.session_state["p_price_for"] = _picked0

                _mode = st.radio("Add by", ["Dollars", "Shares"], key="p_mode",
                                 horizontal=True, label_visibility="collapsed")
                _dollars = (_mode == "Dollars")

                r1c1, r1c2, r1c3 = st.columns([2, 2, 2])
                with r1c1:
                    tk_query = st.text_input("Ticker / Company Name", key="p_tk",
                        placeholder="🔍  Search ticker or company — NVDA, Apple…",
                        on_change=_pin_portfolio_nav)
                with r1c2:
                    if _dollars:
                        new_amt = st.number_input("Amount ($)", key="p_amt",
                            min_value=0.0, step=50.0, format="%.2f",
                            on_change=_pin_portfolio_nav)
                        new_sh = 0.0
                    else:
                        new_sh = st.number_input("Shares", key="p_sh",
                            min_value=0.0, step=1.0, format="%.4f",
                            on_change=_pin_portfolio_nav)
                        new_amt = 0.0
                with r1c3:
                    new_price = st.number_input("Price ($)", key="p_price",
                        min_value=0.0, step=0.01, format="%.4f",
                        help="Defaults to the platform's last price — edit to set your actual cost.",
                        on_change=_pin_portfolio_nav)

                _p_hold = [h.get("ticker") for h in holdings] if holdings else []
                _p_live = (tk_query or "").strip()
                if _p_live and not st.session_state.get("p_sel_tk"):
                    _render_suggestions(_p_live, "psug", "p_sel_tk", exclude=_p_hold)

                resolved_ticker, resolved_name = "", ""
                _p_picked = st.session_state.get("p_sel_tk", "").strip().upper()
                if _p_picked and _p_picked in SECTORS:
                    resolved_ticker, resolved_name = _p_picked, _SEARCH_NAMES.get(_p_picked, "")
                elif tk_query and tk_query.strip():
                    with st.spinner("Looking up...") if len(tk_query) > 2 and not tk_query.strip().isupper() else contextlib.nullcontext():
                        resolved_ticker, resolved_name = resolve_ticker(tk_query)

                if resolved_ticker and is_valid_universe_ticker(resolved_ticker):
                    _render_stock_result(resolved_ticker, nav="portfolio")
                elif resolved_ticker and resolved_name and resolved_name != resolved_ticker:
                    st.markdown(
                        f'<div style="font-size:14px;color:#34d399;margin-bottom:8px;">'
                        f'✓ {resolved_ticker} — {resolved_name}</div>',
                        unsafe_allow_html=True)
                elif resolved_ticker:
                    st.markdown(
                        f'<div style="font-size:14px;color:#b3bed0;margin-bottom:8px;">'
                        f'Ticker: {resolved_ticker}</div>',
                        unsafe_allow_html=True)

                # Live preview of the resulting position, in whichever mode is active.
                _pv_px = new_price if (new_price and new_price > 0) else (
                    float(score_map.get(resolved_ticker, {}).get("price") or 0) if resolved_ticker else 0.0)
                if _pv_px and _pv_px > 0:
                    if _dollars and new_amt > 0:
                        st.markdown(
                            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
                            f'margin:2px 0 10px;">≈ {new_amt/_pv_px:,.4f} shares @ ${_pv_px:,.2f} '
                            f'· ${new_amt:,.0f} invested</div>',
                            unsafe_allow_html=True)
                    elif (not _dollars) and new_sh > 0:
                        st.markdown(
                            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;'
                            f'margin:2px 0 10px;">≈ ${new_sh*_pv_px:,.2f} invested '
                            f'· {new_sh:,.4f} sh @ ${_pv_px:,.2f}</div>',
                            unsafe_allow_html=True)

                r2c1, r2c2 = st.columns([3, 1])
                with r2c1:
                    new_date = st.date_input("Entry Date", key="p_date", value=date.today())
                with r2c2:
                    st.markdown('<div style="height:28px;"></div>', unsafe_allow_html=True)
                    if st.button("Add", key="p_add", use_container_width=True):
                        new_tk = (resolved_ticker or tk_query.strip().upper())
                        _add_px = new_price if (new_price and new_price > 0) else 0.0
                        if not _add_px and new_tk:
                            try:
                                _add_px = float(score_map.get(new_tk.upper().strip(), {}).get("price") or 0)
                            except Exception:
                                _add_px = 0.0
                        # Resolve share count from the active mode.
                        if _dollars:
                            _shares = (new_amt / _add_px) if (new_amt > 0 and _add_px > 0) else 0.0
                        else:
                            _shares = new_sh
                        if new_tk and _shares > 0 and _add_px > 0:
                            tk_clean = new_tk.upper().strip()
                            if not is_valid_universe_ticker(tk_clean):
                                st.error(f"{tk_clean} is not in the QNTM universe ({_universe_n()} tickers). "
                                         "Holdings must match a tracked ticker so the model can "
                                         "score and price the position.")
                            else:
                                ok = upsert_holding(uid(), tk_clean, _shares, _add_px, new_date)
                                if ok:
                                    _inv = _shares * _add_px
                                    st.success(f"Added {tk_clean} — {_shares:,.4f} sh @ "
                                               f"${_add_px:,.2f} (${_inv:,.0f})")
                                    st.session_state.pop("p_sel_tk", None)
                                    st.session_state.pop("p_price_for", None)
                                    sc = score_map.get(tk_clean)
                                    if sc:
                                        act = sc.get("adj_action", sc.get("action", "HOLD"))
                                        comp = sc.get("adj_composite", sc.get("composite", 50))
                                        if act == "SELL":
                                            st.warning(f"⚠ Note: Model currently shows LOW conviction on {tk_clean} (score {comp:.0f})")
                                        elif act == "BUY":
                                            create_notification(uid(), tk_clean, "buy_signal",
                                                f"HIGH conviction active: {tk_clean}",
                                                f"Score {comp:.0f} — HIGH conviction")
                                    st.rerun()
                                else:
                                    st.error("Failed to add position — check ticker and try again")
                        else:
                            st.warning("Enter a ticker, a price, and "
                                       + ("a dollar amount." if _dollars else "a share quantity."))
            _pf_add_form()

    # ── Empty state ────────────────────────────────────────────────────────────
    if not holdings:
        st.markdown("""
        <div style="text-align:center;padding:24px 16px;max-width:480px;margin:0 auto;">
          <div style="font-size:52px;margin-bottom:16px;">💼</div>
          <div style="font-family:Syne,sans-serif;font-size:22px;font-weight:800;color:#e2e8f0;margin-bottom:12px;">
            Add your first position
          </div>
          <div style="font-size:14px;color:#9fabc0;line-height:1.8;margin-bottom:28px;">
            QNTM will run the full conviction model against every stock you add —
            showing your blended score, pillar breakdown, and whether the signal
            has changed since you entered.
          </div>
          <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:12px;margin-bottom:28px;">
            <div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px 10px;">
              <div style="font-size:20px;margin-bottom:6px;">📊</div>
              <div style="font-size:13px;color:#9fabc0;line-height:1.5;">Conviction<br>score</div>
            </div>
            <div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px 10px;">
              <div style="font-size:20px;margin-bottom:6px;">🎯</div>
              <div style="font-size:13px;color:#9fabc0;line-height:1.5;">Signal<br>changes</div>
            </div>
            <div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:14px 10px;">
              <div style="font-size:20px;margin-bottom:6px;">💰</div>
              <div style="font-size:13px;color:#9fabc0;line-height:1.5;">P&L<br>tracking</div>
            </div>
          </div>
          <div style="font-size:14px;color:#b3bed0;">Use the ＋ Add Position button above to get started</div>
        </div>
        """, unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── Portfolio summary strip ────────────────────────────────────────────────
    port_buys  = sum(1 for h in holdings if score_map.get(h["ticker"],{}).get("adj_action", score_map.get(h["ticker"],{}).get("action")) == "BUY")
    port_holds = sum(1 for h in holdings if score_map.get(h["ticker"],{}).get("adj_action", score_map.get(h["ticker"],{}).get("action")) == "HOLD")
    port_sells = sum(1 for h in holdings if score_map.get(h["ticker"],{}).get("adj_action", score_map.get(h["ticker"],{}).get("action")) == "SELL")
    port_na    = n_holdings - port_buys - port_holds - port_sells

    port_summary_data = [
        ("▲ High Conviction",  port_buys,  "#34d399"),
        ("─ Moderate",         port_holds, "#fbbf24"),
        ("▼ Low Conviction",   port_sells, "#f87171"),
        ("Outside Universe",   port_na,    "#8896ac"),
    ]
    ps_html = "".join([
        f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
        f'border-radius:6px;padding:14px;min-width:0;">'
        f'<div style="font-size:13px;color:#b3bed0;letter-spacing:.08em;margin-bottom:6px;font-family:DM Mono,monospace;">{label}</div>'
        f'<div style="font-size:28px;font-weight:800;color:{color};font-family:Syne,sans-serif;line-height:1;">{int(val)}</div>'
        f'<div style="font-size:13px;color:#b3bed0;margin-top:3px;">position{"s" if val!=1 else ""}</div>'
        f'</div>'
        for label,val,color in port_summary_data
    ])
    st.markdown(
        f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:8px;margin-bottom:16px;">{ps_html}</div>',
        unsafe_allow_html=True)

    # ── Portfolio Value Tracker ───────────────────────────────────────────────
    PERIOD_DATA = [
        ("ACT", "Total Return",  None),
        ("1M",  "1 Month",       30),
        ("3M",  "3 Months",      90),
        ("1Y",  "1 Year",        365),
    ]

    if holdings:
        total_cost_basis = sum(
            float(h.get("avg_cost",0) or 0) * float(h.get("shares",0) or 0)
            for h in holdings
        )
        if "port_period" not in st.session_state:
            st.session_state.port_period = "ACT"

        _uid_val  = (st.session_state.user or {}).get("id", "")
        _plan_val = (st.session_state.user or {}).get("plan", "free")

        # Period selector — URL action links, no selectbox rerun
        _cur_period = st.session_state.port_period
        period_btns = ""
        for _pk, _pl, _ in PERIOD_DATA:
            _active = _pk == _cur_period
            _bg     = "background:#34d399;color:#0a0b14;" if _active else "background:rgba(255,255,255,.04);color:#b3bed0;"
            _url    = f"?qnav=portfolio&uid={_uid_val}&plan={_plan_val}&ck=1&port_period={_pk}&_n=portfolio"
            period_btns += (
                f'<a href="{_url}" target="_self" style="'
                f'padding:7px 14px;border-radius:6px;font-family:DM Mono,monospace;'
                f'font-size:13px;font-weight:700;letter-spacing:.06em;text-decoration:none;'
                f'border:1px solid rgba(255,255,255,.08);{_bg}">{_pl}</a>'
            )
        st.markdown(
            f'<div style="display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px;margin-bottom:8px;">'
            f'<span style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.1em;">PERFORMANCE</span>'
            f'<div style="display:flex;gap:6px;flex-wrap:wrap;">{period_btns}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

        sel = next((p for p in PERIOD_DATA if p[0]==_cur_period), PERIOD_DATA[0])
        pkey, plbl, pdays = sel
        is_actual = (pkey == "ACT")

        # Fetch live + historical prices via yfinance
        tickers = [h["ticker"] for h in holdings]
        live_prices  = {}  # ticker -> current price
        start_prices = {}  # ticker -> price at start of period (or entry date)

        try:
            import yfinance as yf
            from datetime import date as _date, timedelta as _td
            fetch_days = (pdays or 0) + 5  # extra buffer for weekends/holidays
            hist = yf.download(_strip_delisted(tickers), period=f"{max(fetch_days, 10)}d",
                               auto_adjust=True, progress=False, threads=True)
            if not hist.empty:
                close = hist["Close"] if hasattr(hist["Close"], "columns") else hist[["Close"]]
                # Handle single ticker case
                if len(tickers) == 1:
                    close = close.rename(columns={"Close": tickers[0]}) if "Close" in close.columns else close
                for tk in tickers:
                    col = close[tk] if tk in close.columns else None
                    if col is None and len(tickers) == 1:
                        col = close.iloc[:, 0]
                    if col is not None:
                        col = col.dropna()
                        if len(col) >= 1:
                            live_prices[tk] = float(col.iloc[-1])
                        if not is_actual and len(col) >= 2:
                            if pdays and len(col) >= pdays:
                                start_prices[tk] = float(col.iloc[-min(pdays, len(col))])
                            else:
                                # Period longer than position age — use earliest available
                                start_prices[tk] = float(col.iloc[0])
        except Exception:
            pass

        from datetime import date as _date, timedelta as _td
        total_current   = 0.0
        total_start_val = 0.0
        for h in holdings:
            tk     = h["ticker"]
            cost   = float(h.get("avg_cost", 0) or 0)
            shares = float(h.get("shares", 0) or 0)

            # Entry date — cap all lookbacks to this
            try:
                entry_date = _date.fromisoformat(str(h.get("entry_date", ""))[:10])
            except Exception:
                entry_date = _date.today()

            live = live_prices.get(tk, cost)

            if is_actual:
                # Total return: live vs cost basis
                total_current   += live  * shares
                total_start_val += cost  * shares
            else:
                # Period lookback — but never go before entry date
                period_start = _date.today() - _td(days=pdays)
                if period_start <= entry_date:
                    # Position younger than period — use cost basis as start
                    start = cost
                else:
                    start = start_prices.get(tk, cost)
                total_current   += live  * shares
                total_start_val += start * shares

        ref_basis    = total_cost_basis if is_actual else total_start_val
        total_change = total_current - ref_basis
        chg_pct      = (total_change / ref_basis * 100) if ref_basis > 0 else 0
        change_c     = "#34d399" if total_change >= 0 else "#f87171"
        arrow        = "▲" if total_change >= 0 else "▼"
        period_note  = "vs cost basis" if is_actual else f"{plbl} lookback (actual prices)"

        b2    = sum(1 for h in holdings if (score_map.get(h["ticker"],{}) or {}).get("adj_action",(score_map.get(h["ticker"],{}) or {}).get("action","N/A"))=="BUY")
        hold2 = sum(1 for h in holdings if (score_map.get(h["ticker"],{}) or {}).get("adj_action",(score_map.get(h["ticker"],{}) or {}).get("action","N/A"))=="HOLD")
        sell2 = len(holdings) - b2 - hold2

        vc_html = (
            f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
            f'border-left:3px solid #e2e8f0;border-radius:8px;padding:14px;min-width:0;overflow:hidden;">'
            f'<div style="font-size:13px;color:#b3bed0;letter-spacing:.08em;margin-bottom:5px;">TOTAL VALUE</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:clamp(18px,4vw,28px);font-weight:800;color:#e2e8f0;line-height:1;">${total_current:,.0f}</div>'
            f'<div style="font-size:13px;color:#b3bed0;margin-top:4px;">Cost basis ${total_cost_basis:,.0f}</div></div>'

            f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
            f'border-left:3px solid {change_c};border-radius:8px;padding:14px;min-width:0;overflow:hidden;">'
            f'<div style="font-size:13px;color:#b3bed0;letter-spacing:.08em;margin-bottom:5px;">$ CHANGE</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:clamp(16px,4vw,26px);font-weight:800;color:{change_c};line-height:1;">{arrow} ${abs(total_change):,.0f}</div>'
            f'<div style="font-size:13px;color:#b3bed0;margin-top:4px;">{period_note}</div></div>'

            f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
            f'border-left:3px solid {change_c};border-radius:8px;padding:14px;min-width:0;overflow:hidden;">'
            f'<div style="font-size:13px;color:#b3bed0;letter-spacing:.08em;margin-bottom:5px;">% CHANGE</div>'
            f'<div style="font-family:Syne,sans-serif;font-size:clamp(16px,4vw,26px);font-weight:800;color:{change_c};line-height:1;">{arrow} {abs(chg_pct):.1f}%</div>'
            f'<div style="font-size:13px;color:#b3bed0;margin-top:4px;">{period_note}</div></div>'

        )
        st.markdown(
            f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:4px;">{vc_html}</div>',
            unsafe_allow_html=True)

        _disclaimer = (
            'Prices fetched live via yfinance. Total Return compares current price to your cost basis. '
            'Period lookbacks compare current price to price at period start — or entry date if position is newer.'
        )
        st.markdown(f'<div style="font-size:13px;color:#9fabc0;margin:4px 0 20px;">{_disclaimer}</div>', unsafe_allow_html=True)

    # ── Holdings cards — collapsed card pattern (same as screener) ────────────
    st.markdown(
        '<div style="font-family:DM Mono,monospace;font-size:11px;color:#8896ac;letter-spacing:.1em;padding:8px 0 6px;border-top:1px solid rgba(255,255,255,.05);">POSITIONS</div>',
        unsafe_allow_html=True)

    _uid_pv  = (st.session_state.user or {}).get("id","")
    _pln_pv  = (st.session_state.user or {}).get("plan","free")

    # Detect current hidden gems across the held universe so the 💎 emoji is
    # consistent with the screener.
    _port_gems = set()
    try:
        _port_gem_list = detect_hidden_gems(
            list(score_map.values()),
            macro_data=st.session_state.get("macro_data"),
        )
        _port_gems = {g["ticker"] for g in _port_gem_list}
    except Exception:
        pass

    _port_html = ""
    # Fetch today's price change once for all positions so the collapsed cards
    # can show entry date + intraday % without expanding.
    _port_day_change = _fetch_day_change_map(
        [h["ticker"] for h in holdings],
        cache_key="_port_daychange_cache",
    )
    for h in holdings:
        tk     = h["ticker"]
        sc     = dict(score_map.get(tk, {}) or {})  # copy — don't mutate shared dict
        cost   = float(h.get("avg_cost", 0) or 0)
        shares = float(h.get("shares", 0) or 0)
        entry  = h.get("entry_date", "")

        # Get live price
        live_price = None
        if sc.get("price"):
            live_price = float(sc["price"])

        if sc:
            comp = float(sc.get("adj_composite", sc.get("composite", 50)) or 50)
            quant = float(sc.get("composite", comp) or comp)
            sc["adj_action"]    = "BUY" if comp >= 60 else ("SELL" if comp < 45 else "HOLD")
            sc["adj_composite"] = comp
            sc["composite"]     = quant
            # Always recompute macro from the underlying values, since
            # signal_log does not persist score_delta as a column and we want
            # the MACRO box to reflect what the user sees on the screener.
            sc["score_delta"]   = round(comp - quant, 1)
            sc["signal_date"]   = str(entry)[:10] if entry else sc.get("signal_date","")
            if live_price is not None:
                sc["price"] = live_price
        else:
            sc = {"ticker": tk, "adj_action": "N/A", "adj_composite": 0,
                  "composite": 0, "momentum": 0, "quality": 0, "volume": 0,
                  "value": 0, "sentiment": 0, "score_delta": 0, "sector": "Unknown"}

        ci = get_company_info(tk)
        _rm_url = f"?qnav=portfolio&uid={_uid_pv}&plan={_pln_pv}&ck=1&_n=portfolio&port_action=remove&port_ticker={tk}"
        _pbtn = _card_action_button(tk, "portfolio", "portfolio", set(), _uid_pv, _pln_pv, remove_url=_rm_url)

        # Build P&L strip in the same style as the Model Portfolio page so a
        # held position shows entry date, entry price, current, and return.
        _edate   = str(entry)[:10] if entry else "—"
        _ep_str  = f'${cost:,.2f}' if cost else "—"
        _cp_str  = f'${live_price:,.2f}' if live_price else "—"
        if live_price and cost and shares:
            _pnl_d = (live_price - cost) * shares
            _pct_v = ((live_price - cost) / cost * 100) if cost else 0.0
            _rc    = '#34d399' if _pct_v >= 0 else '#f87171'
            _sg    = '+' if _pct_v >= 0 else ''
            _pct_str = f'{_sg}{_pct_v:.2f}%'
            _pnl_str = f'{_sg}${abs(_pnl_d):,.0f}'
        else:
            _rc, _pct_str, _pnl_str = '#9fabc0', "—", "—"
        _shares_str = f'{shares:,.4f}'.rstrip('0').rstrip('.') if shares else "—"
        _pnl_html = (
            f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;'
            f'margin:8px 20px 4px;padding:10px;background:rgba(255,255,255,.02);'
            f'border:1px solid rgba(255,255,255,.05);border-radius:6px;">'
            f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">ENTRY DATE</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;">{_edate}</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;margin-top:2px;">{_shares_str} sh</div></div>'
            f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">ENTRY PRICE</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;">{_ep_str}</div></div>'
            f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">CURRENT</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;">{_cp_str}</div></div>'
            f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">RETURN</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:{_rc};">{_pct_str}</div>'
            f'<div style="font-family:DM Mono,monospace;font-size:13px;color:{_rc};">{_pnl_str}</div></div>'
            f'</div>'
        )
        # Render the card (no iframe): P&L strip + Remove button inside the detail.
        sc["_summary_meta_html"] = _build_summary_meta_html(
            entry_date=entry,
            day_change_entry=_port_day_change.get(tk),
        )
        st.markdown(
            factor_panel_html(sc, tk in _port_gems, company_info=ci,
                              wl_btn=(_pnl_html + _pbtn), as_details=True),
            unsafe_allow_html=True,
        )

    st.markdown('</div>', unsafe_allow_html=True)

        # ── Export to Excel ───────────────────────────────────────────────────────
    if holdings:
        try:
            export_rows = []
            for h in holdings:
                score_data = score_map.get(h["ticker"], {})
                export_rows.append({
                    "Ticker":         h["ticker"],
                    "Entry Date":     h.get("entry_date", ""),
                    "Entry Price":    h.get("entry_price", ""),
                    "Current Price":  score_data.get("price", ""),
                    "Shares":         round(h["pos_size"] / h["entry_price"], 4) if h.get("entry_price") and h["entry_price"] > 0 else "",
                    "Position Value": h.get("pos_size", 10000),
                    "P&L ($)":        round(h.get("pnl", 0), 2),
                    "Return (%)":     round(h.get("pnl_pct", 0), 2),
                    "Score":          round(score_data.get("adj_composite", score_data.get("composite", 0)), 1),
                    "Momentum":       round(score_data.get("momentum", 0), 1),
                    "Quality":        round(score_data.get("quality", 0), 1),
                    "Volume":         round(score_data.get("volume", 0), 1),
                    "Value":          round(score_data.get("value", 0), 1),
                    "Sentiment":      round(score_data.get("sentiment", 0), 1),
                    "Signal":         score_data.get("adj_action", score_data.get("action", "")),
                })
            headers = ["Ticker","Entry Date","Entry Price","Current Price","Shares","Position Value","P&L ($)","Return (%)","Score","Momentum","Quality","Volume","Value","Sentiment","Signal"]
            xl = _make_excel(export_rows, headers, "My Portfolio")
            st.download_button(
                label="⬇ Export to Excel",
                data=xl,
                file_name="qntm_my_portfolio.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="port_export"
            )
        except Exception:
            pass


# ══════════════════════════════════════════════════════════════════════════════
# ALERTS PAGE
# ══════════════════════════════════════════════════════════════════════════════
def page_simulator():
    _pin_nav("simulator")
    page_summary("🧮", "Portfolio Simulator", "Hypothetical portfolios from current signals · nightly scores")
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    if not is_pro():
        st.markdown(
            '<div style="background:rgba(212,168,67,.07);border:1px solid rgba(212,168,67,.25);'
            'border-radius:10px;padding:28px 24px;text-align:center;margin:24px 0;">'
            '<div style="font-size:28px;margin-bottom:12px;">🧮</div>'
            '<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:700;color:#d4a843;margin-bottom:8px;">Portfolio Simulator</div>'
            '<div style="font-size:14px;color:#b3bed0;margin-bottom:20px;">'
            'Build a hypothetical portfolio from current HIGH conviction signals.</div>'
            '<div style="font-size:13px;color:#9fabc0;">Pro feature — upgrade to access</div>'
            '</div>', unsafe_allow_html=True)
        if st.session_state.get("logged_in"):
            st.markdown(_cta_gold("Unlock Simulator — Upgrade to Pro", _upgrade_url("Portfolio Simulator", "simulator")), unsafe_allow_html=True)
        else:
            st.markdown(_cta_gold("Upgrade to Pro — $29/mo →", _upgrade_url("Unlimited Holdings","portfolio")), unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        return

    _uid_val  = (st.session_state.user or {}).get("id", "")
    _plan_val = (st.session_state.user or {}).get("plan", "free")

    scan = st.session_state.get("sim_data") or st.session_state.get("scan_results") or []

    if not scan:
        with st.spinner("Loading signals..."):
            try:
                from data_refresh import _get_supabase as _sim_sb
                _sb = _sim_sb()
                if _sb:
                    _resp = _sb.table("signal_log") \
                        .select("ticker,adj_composite,composite,signal,momentum,quality,volume,value,sentiment,price,signal_date,mktcap,val_low,val_high,value_position,val_basis") \
                        .order("signal_date", desc=True) \
                        .limit(5000) \
                        .execute()
                    _seen = {}
                    for _r in (_resp.data or []):
                        if _r["ticker"] not in _seen:
                            _a = float(_r.get("adj_composite") or _r.get("composite") or 50)
                            _r["adj_action"] = "BUY" if _a >= 60 else ("SELL" if _a < 45 else "HOLD")
                            _seen[_r["ticker"]] = _r
                    # Enrich with sector from universe_data
                    try:
                        from model_engine import SECTORS as _SIM_SECTORS
                        for _tk, _row in _seen.items():
                            _row["sector"] = _SIM_SECTORS.get(_tk, "Unknown")
                    except Exception:
                        pass
                    scan = list(_seen.values())
                    if scan:
                        st.session_state.sim_data = scan
            except Exception as _e:
                st.warning(f"Could not load signals: {_e}")
        if not scan:
            st.warning("No signal data available — run a Rescan on the Screener first.")
            st.markdown('</div>', unsafe_allow_html=True)
            return

    all_buys = sorted(
        [r for r in scan if r.get("adj_action") == "BUY"],
        key=lambda x: float(x.get("adj_composite", x.get("composite", 0)) or 0), reverse=True
    )
    ticker_map = {r["ticker"]: r for r in scan}

    def profile_tickers(profile):
        if profile == "HIGH":
            ranked = sorted(all_buys, key=lambda x: x.get("momentum", 0), reverse=True)
        elif profile == "LOW":
            ranked = sorted(all_buys, key=lambda x: (x.get("quality", 0) + x.get("value", 0)) / 2, reverse=True)
        else:
            ranked = all_buys
        return [r["ticker"] for r in ranked[:20]]

    # Profile from URL param — read without navigation
    _sp = st.query_params.get("_sp", "")
    if _sp in ("HIGH", "MEDIUM", "LOW"):
        if st.session_state.get("sim_profile") != _sp:
            st.session_state.sim_profile = _sp
            st.session_state.sim_weights = {}
            st.session_state.sim_profile_applied = None

    if "sim_profile" not in st.session_state:
        st.session_state.sim_profile = "MEDIUM"
    if ("sim_selected" not in st.session_state
            or st.session_state.get("sim_profile_applied") != st.session_state.sim_profile
            or (not st.session_state.get("sim_selected") and all_buys)):
        st.session_state.sim_selected = profile_tickers(st.session_state.sim_profile)
        st.session_state.sim_weights  = {}
        st.session_state.sim_profile_applied = st.session_state.sim_profile

    available = set(ticker_map.keys())
    st.session_state.sim_selected = [t for t in st.session_state.sim_selected if t in available]

    sim_amount = st.number_input("Investment Amount ($)", min_value=1000, max_value=10000000,
                                  value=50000, step=1000, format="%d", key="sim_amount")
    equal_weight = st.toggle("Equal weight", value=True, key="sim_equal")

    PROFILES = {
        "HIGH":   ("🔥 High Risk",   "Top 20 by momentum. Higher volatility, higher potential return."),
        "MEDIUM": ("⚖️ Medium Risk", "Top 20 by conviction score. Balanced. Model default."),
        "LOW":    ("🛡 Low Risk",    "Top 20 by quality + value. More defensive positioning."),
    }
    st.markdown('<div style="height:8px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:10px;">RISK PROFILE</div>', unsafe_allow_html=True)

    p_cols = st.columns(3)
    for col, (pk, (plbl, pdesc)) in zip(p_cols, PROFILES.items()):
        with col:
            active = st.session_state.sim_profile == pk
            bg     = "rgba(212,168,67,.12)" if pk=="HIGH" else "rgba(52,211,153,.10)" if pk=="LOW" else "rgba(255,255,255,.06)"
            border = "rgba(212,168,67,.6)"  if pk=="HIGH" else "rgba(52,211,153,.5)"  if pk=="LOW" else "rgba(148,163,184,.35)"
            tc     = "#d4a843" if pk=="HIGH" else "#34d399" if pk=="LOW" else "#b3bed0"
            if active:
                bg = bg.replace(",.12",",.2").replace(",.10",",.18").replace(",.06",",.12")
            _prof_url = f"?qnav=simulator&uid={_uid_val}&plan={_plan_val}&ck=1&_sp={pk}&_n=simulator"
            _btn_label = "✓ Selected" if active else "Select"
            st.markdown(
                f'<a href="{_prof_url}" target="_self" style="display:block;text-decoration:none;">'
                f'<div style="background:{bg};border:1px solid {border};border-radius:8px;'
                f'padding:10px 8px;text-align:center;margin-bottom:4px;">'
                f'<div style="font-size:13px;font-weight:700;color:{tc};">{plbl}</div>'
                f'<div style="font-size:13px;color:#9fabc0;margin-top:3px;line-height:1.3;">{pdesc[:55]}</div>'
                f'<div style="font-size:13px;color:{tc};margin-top:6px;font-weight:700;">{_btn_label}</div>'
                f'</div></a>',
                unsafe_allow_html=True)

    st.markdown("""
    <style>
    /* Sim suggestion buttons */
    div[data-testid='stButton'][data-key^='simsug_'] > div > button {
        display:none !important;
    }
    /* Add to sim button */
    div[data-testid='stButton'][data-key='sim_add_sel'] > div > button {
        background:linear-gradient(135deg,#d4a843,#b8922e) !important;
        color:#000 !important; border:none !important;
        font-family:Syne,sans-serif !important; font-weight:800 !important;
        font-size:13px !important; letter-spacing:.08em !important;
        padding:10px !important; border-radius:6px !important;
        margin-top:4px !important;
    }
    /* Remove from sim button */
    div[data-testid='stButton'][data-key='sim_rm_sel'] > div > button,
    div[data-testid='stButton'][data-key^='sim_rm_'] > div > button {
        background:rgba(248,113,113,.08) !important;
        border:1px solid rgba(248,113,113,.3) !important;
        color:#f87171 !important;
        font-family:Syne,sans-serif !important; font-weight:700 !important;
        font-size:13px !important; padding:10px !important;
        border-radius:6px !important; margin-top:4px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:6px;">ADD POSITION</div>', unsafe_allow_html=True)

    # Search with live suggestions + Enter to select first match
    def _on_sim_search():
        _raw = st.session_state.sim_add_query.strip().upper()
        st.session_state._sim_search_live = _raw
        st.session_state.nav = "simulator"
        # Exact ticker + Enter → show its card immediately, like a suggestion tap.
        if _raw in SECTORS:
            st.session_state._sim_selected_tk = _raw

    if "_sim_sug_just_picked" in st.session_state:
        _picked = st.session_state.pop("_sim_sug_just_picked")
        if "sim_add_query" in st.session_state:
            del st.session_state["sim_add_query"]
        st.session_state._sim_search_live = _picked
        st.session_state._sim_selected_tk = _picked

    if "sim_add_query" not in st.session_state:
        st.session_state.sim_add_query = ""
    if "_sim_search_live" not in st.session_state:
        st.session_state._sim_search_live = ""

    st.text_input("Search ticker or company", key="sim_add_query",
                  placeholder="🔍  Search ticker or company — NVDA, Apple…",
                  label_visibility="collapsed", on_change=_on_sim_search)

    _sim_q = st.session_state.get("_sim_search_live", "").strip().upper()

    # Suggestions dropdown — standard ticker/company search (shared platform-wide)
    if _sim_q:
        _render_suggestions(_sim_q, "simsug", "_sim_selected_tk",
                            exclude=st.session_state.get("sim_selected", []))

    # Show selected ticker card with Add/Remove CTA
    _sim_sel_tk = st.session_state.get("_sim_selected_tk", "")
    if _sim_sel_tk and _sim_sel_tk in ticker_map:
        _sel_r = dict(ticker_map[_sim_sel_tk])
        _sel_adj = float(_sel_r.get("adj_composite", _sel_r.get("composite", 50)) or 50)
        _sel_r["adj_action"] = "BUY" if _sel_adj >= 60 else ("SELL" if _sel_adj < 45 else "HOLD")
        _sel_r["adj_composite"] = _sel_adj
        _sel_ci = get_company_info(_sim_sel_tk)
        st.markdown('<div style="margin-top:8px;">', unsafe_allow_html=True)
        st.markdown(factor_panel_html(_sel_r, False, company_info=_sel_ci, suppress_wl_btn=True), unsafe_allow_html=True)
        _render_rel_spy_chart(_sim_sel_tk, 20)
        # Add / Remove CTA
        _in_sim = _sim_sel_tk in st.session_state.get("sim_selected", [])
        if _in_sim:
            if st.button(f"✕ Remove {_sim_sel_tk} from Simulation", key="sim_rm_sel", use_container_width=True):
                st.session_state.sim_selected.remove(_sim_sel_tk)
                st.session_state._sim_selected_tk = ""
                st.session_state.nav = "simulator"
                st.rerun()
        else:
            if st.button(f"+ Add {_sim_sel_tk} to Simulation", key="sim_add_sel", use_container_width=True):
                if "sim_selected" not in st.session_state:
                    st.session_state.sim_selected = []
                st.session_state.sim_selected.append(_sim_sel_tk)
                st.session_state._sim_selected_tk = ""
                st.session_state.nav = "simulator"
                st.rerun()
        try:
            from db import get_watchlist_items as _gwi2, get_watchlists as _gws2
            _u2 = uid()
            _l2 = _gws2(_u2)
            _d2 = next((l["id"] for l in _l2 if l.get("is_default")),
                       _l2[0]["id"] if _l2 else None)
            _w2 = {w["ticker"] for w in _gwi2(_u2, _d2)} if _d2 else set()
            render_watchlist_actions([_sim_sel_tk], nav="simulator", in_list=_w2)
        except Exception:
            pass
        st.markdown('</div>', unsafe_allow_html=True)

    selected_rows = [ticker_map[t] for t in st.session_state.sim_selected if t in ticker_map]
    n_sel = len(selected_rows)

    if n_sel == 0:
        st.info("No positions — select a risk profile or search for a ticker above.")
        st.markdown('</div>', unsafe_allow_html=True)
        return

    weight_map = {r["ticker"]: 100.0 / n_sel for r in selected_rows} if equal_weight else {
        r["ticker"]: st.session_state.sim_weights.get(r["ticker"], 100.0 / n_sel) for r in selected_rows
    }
    if not equal_weight:
        total_w = sum(weight_map.values())
        weight_map = {tk: v / total_w * 100 for tk, v in weight_map.items()} if total_w > 0 else weight_map

    alloc = []
    for r in selected_rows:
        tk    = r["ticker"]
        score = r.get("adj_composite", r.get("composite", 0))
        price = r.get("price")
        pct   = weight_map[tk]
        w_dollar = sim_amount * pct / 100
        shares   = round(w_dollar / price, 4) if price and price > 0 else None
        alloc.append({"ticker": tk, "score": score, "price": price, "allocation": w_dollar,
                       "pct": pct, "shares": shares, "sector": r.get("sector", "Unknown"),
                       "momentum": r.get("momentum", 50), "quality": r.get("quality", 50),
                       "volume": r.get("volume", 50), "value": r.get("value", 50),
                       "sentiment": r.get("sentiment", 50)})

    weighted_score = sum(a["pct"] * a["score"] for a in alloc) / 100
    sc_col = "#34d399" if weighted_score >= 70 else "#fbbf24" if weighted_score >= 55 else "#f87171"

    # Aggregate per-pillar averages — weighted by the SAME pct the score uses,
    # so flipping the equal/dollar toggle reweights pillars and score together.
    agg_pillars = {}
    for _pk in ("momentum", "quality", "volume", "value", "sentiment"):
        agg_pillars[_pk] = sum(a["pct"] * float(a.get(_pk, 50) or 50) for a in alloc) / 100

    st.markdown('<div style="height:6px"></div>', unsafe_allow_html=True)
    st.markdown(
        f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:16px;">'
        f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:12px;text-align:center;">'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:4px;">INVESTED</div>'
        f'<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#d4a843;">${sim_amount:,.0f}</div></div>'
        f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:12px;text-align:center;">'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:4px;">POSITIONS</div>'
        f'<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:#cbd5e1;">{n_sel}</div></div>'
        f'<div style="background:rgba(255,255,255,.03);border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:12px;text-align:center;">'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:4px;">AVG SCORE</div>'
        f'<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:800;color:{sc_col};">{weighted_score:.1f}</div></div>'
        f'</div>', unsafe_allow_html=True)

    # ── Save / Load basket to named watchlists ────────────────────────────────
    from db import (get_watchlists, create_watchlist, add_watchlist_item,
                    get_watchlist_items)
    _sim_uid = uid()
    st.markdown('<div style="height:4px"></div>', unsafe_allow_html=True)
    _save_c, _load_c = st.columns(2)
    with _save_c:
        with st.popover("💾 Save as watchlist", use_container_width=True):
            _new_name = st.text_input("List name", key="sim_save_name",
                                       placeholder="e.g. High Momentum Basket")
            if st.button("Save basket", key="sim_save_go", use_container_width=True):
                _nm = (_new_name or "").strip()
                if not _nm:
                    st.warning("Name the list first.")
                else:
                    _lst = create_watchlist(_sim_uid, _nm)
                    if not _lst:
                        # name may already exist — find it
                        _existing = [w for w in get_watchlists(_sim_uid) if w["name"] == _nm]
                        _lst = _existing[0] if _existing else None
                    if _lst:
                        _ok = 0
                        for _t in st.session_state.sim_selected:
                            _p = (ticker_map.get(_t) or {}).get("price")
                            if add_watchlist_item(_sim_uid, _lst["id"], _t, _p):
                                _ok += 1
                        st.success(f"Saved {_ok} positions to “{_nm}”.")
                    else:
                        st.error("Could not create that list.")
    with _load_c:
        _all_lists = get_watchlists(_sim_uid)
        with st.popover("📂 Load from watchlist", use_container_width=True,
                        disabled=not _all_lists):
            if _all_lists:
                _names = [w["name"] for w in _all_lists]
                _pick = st.selectbox("Choose a list", _names, key="sim_load_pick")
                if st.button("Load into simulator", key="sim_load_go", use_container_width=True):
                    _chosen = next((w for w in _all_lists if w["name"] == _pick), None)
                    if _chosen:
                        _items = get_watchlist_items(_sim_uid, _chosen["id"])
                        _loaded = [it["ticker"] for it in _items if it["ticker"] in ticker_map]
                        if _loaded:
                            st.session_state.sim_selected = _loaded
                            st.session_state.sim_weights = {}
                            st.session_state.sim_profile_applied = "_loaded_"
                            st.session_state.nav = "simulator"
                            st.rerun()
                        else:
                            st.warning("No tickers from that list are in the current universe.")

    sector_totals = {}
    for a in alloc:
        sector_totals[a["sector"]] = sector_totals.get(a["sector"], 0) + a["allocation"]
    SECTOR_CAP_PCT = 30.0
    _over_cap = []
    bars_html = ""
    for sec, val in sorted(sector_totals.items(), key=lambda x: x[1], reverse=True)[:6]:
        pct = val / sim_amount * 100
        _is_over = pct > SECTOR_CAP_PCT
        if _is_over:
            _over_cap.append((sec, pct))
        _bar_col = "#f59e0b" if _is_over else "#d4a843"
        _pct_col = "#f59e0b" if _is_over else "#cbd5e1"
        _flag = ' <span style="font-size:13px;color:#f59e0b;">⚠ over 30%</span>' if _is_over else ''
        bars_html += (f'<div style="margin-bottom:8px;"><div style="display:flex;justify-content:space-between;margin-bottom:3px;">'
                      f'<span style="font-size:13px;color:#b3bed0;">{sec}{_flag}</span>'
                      f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{_pct_col};">{pct:.1f}%</span></div>'
                      f'<div style="background:rgba(255,255,255,.06);border-radius:3px;height:5px;">'
                      f'<div style="width:{min(pct,100):.1f}%;height:100%;background:{_bar_col};border-radius:3px;"></div></div></div>')
    _cap_note = ''
    if _over_cap:
        _cap_note = ('<div style="font-size:13px;color:#f59e0b;margin-top:4px;padding-top:8px;'
                     'border-top:1px solid rgba(245,158,11,.15);">Concentration flag — '
                     + ', '.join(f'{s} {p:.0f}%' for s, p in _over_cap)
                     + ' exceeds the 30% single-sector guideline.</div>')
    st.markdown(
        f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
        f'border-radius:8px;padding:16px 20px;margin-bottom:16px;">'
        f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:12px;">SECTOR EXPOSURE</div>'
        f'{bars_html}{_cap_note}</div>', unsafe_allow_html=True)

    # ── Aggregate pillar profile (weighted, recomputes with the toggle) ────────
    def _agg_bar(label, v):
        c = "#34d399" if v >= 60 else ("#f59e0b" if v >= 45 else "#f87171")
        return (f'<div style="margin-bottom:10px;">'
                f'<div style="display:flex;justify-content:space-between;margin-bottom:3px;">'
                f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;letter-spacing:.06em;">{label}</span>'
                f'<span style="font-family:DM Mono,monospace;font-size:13px;color:{c};">{v:.1f}</span></div>'
                f'<div style="background:rgba(255,255,255,.06);border-radius:3px;height:6px;">'
                f'<div style="width:{min(max(v,2),100):.1f}%;height:100%;background:{c};border-radius:3px;"></div></div></div>')
    _pillar_bars = (
        _agg_bar("MOMENTUM",  agg_pillars["momentum"])
        + _agg_bar("QUALITY",  agg_pillars["quality"])
        + _agg_bar("VOLUME",   agg_pillars["volume"])
        + _agg_bar("VALUE",    agg_pillars["value"])
        + _agg_bar("SENTIMENT", agg_pillars["sentiment"])
    )
    st.markdown(
        f'<div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);'
        f'border-radius:8px;padding:16px 20px;margin-bottom:16px;">'
        f'<div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:14px;">'
        f'<span style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;">PORTFOLIO FACTOR PROFILE</span>'
        f'<span style="font-size:13px;color:#8896ac;">{"equal-weighted" if equal_weight else "dollar-weighted"}</span></div>'
        f'{_pillar_bars}</div>', unsafe_allow_html=True)

    st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:8px;">POSITIONS</div>', unsafe_allow_html=True)

    def pill_bar(v):
        c = "#34d399" if v >= 60 else ("#f59e0b" if v >= 45 else "#f87171")
        return (f'<div style="height:4px;border-radius:2px;background:rgba(255,255,255,.08);margin:1px 0;">'
                f'<div style="width:{max(4,int(v))}%;height:100%;background:{c};border-radius:2px;"></div></div>')

    for a in sorted(alloc, key=lambda x: x["score"], reverse=True):
        sc_color  = "#34d399" if a["score"] >= 70 else "#fbbf24" if a["score"] >= 55 else "#f87171"
        price_str = f'${a["price"]:,.2f}' if a["price"] else "—"
        shares_str = f'{a["shares"]:,.3f}' if a["shares"] else "—"
        with st.expander(f'{a["ticker"]}  ·  ${a["allocation"]:,.0f} ({a["pct"]:.1f}%)  ·  score {a["score"]:.0f}', expanded=False):
            st.markdown(
                f'<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-bottom:12px;">'
                f'<div style="background:rgba(255,255,255,.04);border-radius:6px;padding:10px;text-align:center;">'
                f'<div style="font-size:13px;color:#9fabc0;margin-bottom:3px;">PRICE</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:15px;color:#cbd5e1;">{price_str}</div></div>'
                f'<div style="background:rgba(255,255,255,.04);border-radius:6px;padding:10px;text-align:center;">'
                f'<div style="font-size:13px;color:#9fabc0;margin-bottom:3px;">SHARES</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:15px;color:#b3bed0;">{shares_str}</div></div>'
                f'<div style="background:rgba(255,255,255,.04);border-radius:6px;padding:10px;text-align:center;">'
                f'<div style="font-size:13px;color:#9fabc0;margin-bottom:3px;">CONVICTION</div>'
                f'<div style="font-family:Syne,sans-serif;font-size:16px;font-weight:800;color:{sc_color};">{a["score"]:.0f}</div></div>'
                f'</div>', unsafe_allow_html=True)
            st.markdown(
                f'<div style="margin-bottom:10px;">'
                f'<div style="font-size:13px;color:#9fabc0;margin-bottom:1px;">MOM {a["momentum"]:.0f}</div>{pill_bar(a["momentum"])}'
                f'<div style="font-size:13px;color:#9fabc0;margin-top:4px;margin-bottom:1px;">QUAL {a["quality"]:.0f}</div>{pill_bar(a["quality"])}'
                f'<div style="font-size:13px;color:#9fabc0;margin-top:4px;margin-bottom:1px;">VOL {a["volume"]:.0f}</div>{pill_bar(a["volume"])}'
                f'<div style="font-size:13px;color:#9fabc0;margin-top:4px;margin-bottom:1px;">VAL {a["value"]:.0f}</div>{pill_bar(a["value"])}'
                f'<div style="font-size:13px;color:#9fabc0;margin-top:4px;margin-bottom:1px;">SENT {a["sentiment"]:.0f}</div>{pill_bar(a["sentiment"])}'
                f'</div>', unsafe_allow_html=True)
            st.markdown(f'<div style="font-size:13px;color:#8896ac;margin-bottom:8px;">{a["sector"]}</div>', unsafe_allow_html=True)
            if not equal_weight:
                raw_pct = st.session_state.sim_weights.get(a["ticker"], round(100.0 / n_sel, 1))
                new_pct = st.slider(f"Weight % for {a['ticker']}", min_value=0.5, max_value=50.0,
                                     value=float(raw_pct), step=0.5, key=f"sim_w_{a['ticker']}",
                                     help="Normalised to 100% across all positions")
                if new_pct != raw_pct:
                    st.session_state.sim_weights[a["ticker"]] = new_pct
            # Remove uses a real st.button so the state mutation is server-side
            # and not dependent on URL params surviving an iframe roundtrip.
            if st.button(f"✕ Remove {a['ticker']}",
                          key=f"sim_rm_{a['ticker']}",
                          use_container_width=True):
                if a["ticker"] in st.session_state.sim_selected:
                    st.session_state.sim_selected.remove(a["ticker"])
                st.session_state.get("sim_weights", {}).pop(a["ticker"], None)
                st.session_state.nav = "simulator"
                st.rerun()

    st.markdown(
        f'<div style="font-size:13px;color:#8896ac;padding-top:12px;margin-top:8px;'
        f'border-top:1px solid rgba(255,255,255,.05);">'
        f'{"Equal weight" if equal_weight else "Custom weight (normalised)"} · ${sim_amount:,.0f} across {n_sel} positions · '
        f'Shares at last scan price · Hypothetical — not investment advice.</div>',
        unsafe_allow_html=True)

    # ── Export to Excel ───────────────────────────────────────────────────────
    try:
        export_rows = [{
            "Ticker":       a["ticker"],
            "Sector":       a["sector"],
            "Price":        a["price"] or "",
            "Allocation ($)": round(a["allocation"], 2),
            "Weight (%)":   round(a["pct"], 2),
            "Shares":       a["shares"] or "",
            "Score":        round(a["score"], 1),
            "Momentum":     round(a["momentum"], 1),
            "Quality":      round(a["quality"], 1),
            "Volume":       round(a["volume"], 1),
            "Value":        round(a["value"], 1),
            "Sentiment":    round(a["sentiment"], 1),
        } for a in sorted(alloc, key=lambda x: x["score"], reverse=True)]
        headers = ["Ticker","Sector","Price","Allocation ($)","Weight (%)","Shares","Score","Momentum","Quality","Volume","Value","Sentiment"]
        xl = _make_excel(export_rows, headers, "Simulator")
        st.download_button(
            label="⬇ Export to Excel",
            data=xl,
            file_name="qntm_simulator.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="sim_export"
        )
    except Exception:
        pass
    st.markdown('</div>', unsafe_allow_html=True)


def page_alerts():
    _pin_nav("alerts")
    user = st.session_state.user or {}
    plan = user.get("plan", "free")
    has_alerts = plan_limit(plan, "notifications")

    page_summary(
        "🔔", "Alerts",
        "Signal changes on your holdings — the moment the model issues a HIGH or LOW conviction signal, you'll know. "
        "Macro regime shifts (war, oil spikes, rate changes) trigger alerts too. "
        "Pro members get email notifications on every signal change across their portfolio.",

    )
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    # ── Free tier gate ─────────────────────────────────────────────────────────
    if not has_alerts:
        st.markdown("""
        <div style="background:rgba(212,168,67,.04);border:1px solid rgba(212,168,67,.2);
             border-radius:12px;padding:24px;text-align:center;margin-bottom:16px;">
          <div style="font-size:52px;margin-bottom:16px;">🔔</div>
          <div style="font-family:'Syne',sans-serif;font-size:24px;font-weight:800;
               color:#d4a843;margin-bottom:12px;">Pro Feature — Signal Alerts</div>
          <div style="color:#9fabc0;max-width:520px;margin:0 auto;line-height:1.8;margin-bottom:32px;">
            Get notified the moment the model issues a conviction change on any of
            your holdings. Macro regime changes, hidden gem alerts, and weekly
            performance summaries all included.
          </div>
          <div style="font-family:'DM Mono',monospace;font-size:13px;color:#d4a843;margin-bottom:8px;">
            PRO PLAN — $29/MO · FOUNDING MEMBER — FREE (FIRST 50)
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.session_state.get("logged_in"):
            st.markdown(_cta_gold("Unlock Alerts — Upgrade to Pro", _upgrade_url("Signal Alerts", "alerts")), unsafe_allow_html=True)
        else:
            st.markdown(_cta_gold("Upgrade to Pro — Unlock Alerts", _upgrade_url("Signal Alerts","alerts")), unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        return

    # ── Pro user — show notifications ──────────────────────────────────────────
    notifs = get_notifications(uid())
    unread = sum(1 for n in notifs if not n.get("is_read"))

    # Action bar
    ac1, ac2, ac3 = st.columns([2, 1, 1])
    with ac1:
        st.markdown(f"""
        <div style="padding:8px 0;font-size:13px;color:#b3bed0;">
          {len(notifs)} notifications
          {'· <span style="color:#34d399;">' + str(unread) + ' unread</span>' if unread else ''}
        </div>
        """, unsafe_allow_html=True)
    with ac2:
        if unread > 0 and st.button("✓ Read", key="mark_read", use_container_width=True):
            mark_notifications_read(uid())
            st.rerun()
    with ac3:
        filter_type = st.selectbox("Filter", ["All","HIGH","LOW","Macro","Gems"], key="notif_filter", label_visibility="collapsed")

    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)

    if not notifs:
        st.markdown("""
        <div style="text-align:center;padding:24px 16px;max-width:440px;margin:0 auto;">
          <div style="font-size:48px;margin-bottom:16px;">🔔</div>
          <div style="font-family:Syne,sans-serif;font-size:20px;font-weight:800;color:#e2e8f0;margin-bottom:10px;">
            No alerts yet
          </div>
          <div style="font-size:13px;color:#9fabc0;line-height:1.8;margin-bottom:20px;">
            Alerts fire when the model issues a signal change on one of your holdings,
            or when a macro regime shift affects the market. Add positions in Portfolio
            and the model will watch them every scan.
          </div>
          <div style="font-size:13px;color:#b3bed0;">Macro alerts are always active — portfolio alerts require holdings</div>
        </div>
        """, unsafe_allow_html=True)
    else:
        icon_map = {
            "buy_signal":  ("▲", "#34d399"),
            "sell_signal": ("▼", "#f87171"),
            "hold_alert":  ("─", "#fbbf24"),
            "hidden_gem":  ("💎", "#34d399"),
            "macro_alert": ("⚡", "#d4a843"),
            "system":      ("ℹ", "#8896ac"),
        }
        type_filter_map = {
            "All": None,
            "HIGH": "buy_signal",
            "LOW": "sell_signal",
            "Macro": "macro_alert",
            "Gems": "hidden_gem",
        }
        filter_val = type_filter_map.get(filter_type)

        shown = 0
        for n in notifs[:50]:
            ntype = n.get("notification_type", "system")
            if filter_val and ntype != filter_val:
                continue
            shown += 1
            icon, icolor = icon_map.get(ntype, ("ℹ", "#8896ac"))
            is_read = n.get("is_read", False)
            bg  = "rgba(255,255,255,.015)" if is_read else "rgba(255,255,255,.03)"
            brd = "rgba(255,255,255,.05)"  if is_read else f"{icolor}33"
            opacity = "opacity:.65;" if is_read else ""
            created = str(n.get("created_at", ""))[:16].replace("T", " ")

            st.markdown(f"""
            <div style="background:{bg};border:1px solid {brd};border-left:3px solid {icolor};
                 border-radius:6px;padding:14px 16px;margin-bottom:8px;{opacity}">
              <div style="display:flex;justify-content:space-between;align-items:flex-start;gap:12px;">
                <div style="display:flex;align-items:flex-start;gap:10px;flex:1;">
                  <span style="font-size:14px;color:{icolor};flex-shrink:0;margin-top:1px;">{icon}</span>
                  <div>
                    <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:600;
                         color:{'#e2e8f0' if not is_read else '#b3bed0'};">
                      {n.get('title','').replace('BUY','High Conviction').replace('SELL','Low Conviction').replace('HOLD','Moderate Conviction')}
                    </div>
                    <div style="font-size:13px;color:#b3bed0;margin-top:3px;line-height:1.5;">
                      {n.get('body','').replace('BUY','High Conviction').replace('SELL','Low Conviction').replace('HOLD','Moderate Conviction')}
                    </div>
                  </div>
                </div>
                <div style="font-family:'DM Mono',monospace;font-size:13px;color:#9fabc0;
                     flex-shrink:0;white-space:nowrap;">{created}</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

        if shown == 0:
            st.markdown(f'<div style="color:#9fabc0;padding:24px;text-align:center;font-size:13px;">No {filter_type.lower()} alerts</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# ACCOUNT PAGE
# ══════════════════════════════════════════════════════════════════════════════
def page_account():
    _pin_nav("account")
    if st.session_state.get("_checkout_page") != "account":
        st.session_state["_checkout_page"] = "account"
        st.session_state.pop("_checkout_url", None)
        st.session_state.pop("_checkout_err", None)
    from db import disable_mfa, upgrade_plan, plan_limit, update_email, change_password
    user = st.session_state.user or {}
    plan = user.get("plan", "free")

    page_summary(
        "⚙️", "Account",
        "Manage your profile, secure your account with two-factor authentication, and upgrade your plan. "
        "Founding Member gives you full Pro access free — unlimited holdings, Hidden Gems, and signal alerts — "
        "locked in for the first 50 users.",
    )
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    if st.query_params.get("acct_focus") == "billing":
        # Arrived via the plan badge → surface Plan & Billing first.
        tab_plan, tab_profile, tab_security, tab_notifs = st.tabs([
            "Plan & Billing", "Profile", "Security & MFA", "Notification Prefs"
        ])
    else:
        tab_profile, tab_security, tab_plan, tab_notifs = st.tabs([
            "Profile", "Security & MFA", "Plan & Billing", "Notification Prefs"
        ])

    # ── PROFILE ───────────────────────────────────────────────────────────────
    with tab_profile:
        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
        c1, c2 = st.columns(2)
        with c1:
            new_name = st.text_input("Full name", value=user.get("full_name",""), key="acc_name")
            new_email = st.text_input("Email address", value=user.get("email",""), key="acc_email",
                          help="Updating your email changes the address you sign in with.")
            st.text_input("Member since",
                          value=str(user.get("created_at",""))[:10] or "—",
                          disabled=True)
            plan_display = plan.upper()
            plan_c = "#d4a843" if plan in ("pro","institutional") else "#8896ac"
            st.markdown(f"""
            <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
                 border-radius:4px;padding:10px 14px;margin:8px 0;">
              <span style="font-size:13px;color:#9fabc0;letter-spacing:.1em;">PLAN </span>
              <span style="font-family:'Syne',sans-serif;font-weight:700;color:{plan_c};">
                {plan_display}
              </span>
            </div>
            """, unsafe_allow_html=True)
            st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
            if st.button("Save Profile", key="acc_save"):
                _errs, _changed = [], False
                _nm = (new_name or "").strip()
                if not _nm:
                    _errs.append("Name cannot be blank")
                elif _nm != (user.get("full_name") or ""):
                    if update_preferences(uid(), {"full_name": _nm}):
                        st.session_state.user["full_name"] = _nm
                        _changed = True
                    else:
                        _errs.append("Couldn't save name — try again")
                _em = (new_email or "").strip().lower()
                if _em and _em != (user.get("email") or "").lower():
                    _r = update_email(uid(), _em)
                    if _r.get("success"):
                        st.session_state.user["email"] = _em
                        _changed = True
                    else:
                        _errs.append(_r.get("error", "Couldn't save email"))
                if _errs:
                    for _e in _errs:
                        st.error(_e)
                elif _changed:
                    st.success("Profile saved")
                else:
                    st.info("No changes to save")

    # ── SECURITY & MFA ────────────────────────────────────────────────────────
    with tab_security:
        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)

        # ── Change password ─────────────────────────────────────────────────
        st.markdown('<div style="font-family:Syne,sans-serif;font-size:15px;font-weight:700;'
                    'color:#e2e8f0;margin-bottom:8px;">Change password</div>', unsafe_allow_html=True)
        cp_cur = st.text_input("Current password", type="password", key="cp_cur")
        cp_n1  = st.text_input("New password", type="password", key="cp_n1",
                               placeholder="At least 8 characters")
        cp_n2  = st.text_input("Confirm new password", type="password", key="cp_n2")
        if st.button("Update password", key="cp_btn"):
            if not cp_cur or not cp_n1:
                st.error("Fill in all three fields")
            elif cp_n1 != cp_n2:
                st.error("New passwords don't match")
            elif len(cp_n1) < 8:
                st.error("New password must be at least 8 characters")
            else:
                _cpr = change_password(uid(), cp_cur, cp_n1)
                if _cpr.get("success"):
                    st.success("Password updated")
                else:
                    st.error(_cpr.get("error", "Couldn't update password"))
        st.markdown('<div style="height:14px;border-bottom:1px solid rgba(255,255,255,.07);'
                    'margin-bottom:18px;"></div>', unsafe_allow_html=True)

        mfa_data = get_user_mfa(uid())
        mfa_on   = mfa_data.get("mfa_enabled", False)

        if mfa_on:
            st.markdown("""
            <div style="background:rgba(52,211,153,.06);border:1px solid rgba(52,211,153,.25);
                 border-radius:8px;padding:18px 20px;margin-bottom:20px;
                 display:flex;align-items:center;gap:12px;">
              <span style="font-size:20px;color:#34d399;">✓</span>
              <div>
                <div style="font-family:'Syne',sans-serif;font-size:14px;font-weight:700;color:#34d399;">
                  Two-factor authentication is enabled
                </div>
                <div style="font-size:14px;color:#b3bed0;margin-top:2px;">
                  Your account is protected with TOTP (Google Authenticator / Authy)
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)
            if st.button("Disable MFA", key="dis_mfa"):
                if disable_mfa(uid()):
                    st.session_state.user["mfa_enabled"] = False
                    st.success("MFA disabled")
                    st.rerun()
                else:
                    st.error("Failed to disable MFA")
        else:
            st.markdown("""
            <div style="background:rgba(248,113,113,.05);border:1px solid rgba(248,113,113,.2);
                 border-radius:8px;padding:18px 20px;margin-bottom:20px;">
              <div style="font-family:'Syne',sans-serif;font-size:14px;font-weight:700;
                   color:#f87171;margin-bottom:4px;">⚠ Two-factor authentication is off</div>
              <div style="font-size:14px;color:#b3bed0;">
                We strongly recommend enabling MFA to protect your account.
                Use Google Authenticator, Authy, or any TOTP app.
              </div>
            </div>
            """, unsafe_allow_html=True)

            if st.button("Enable MFA →", key="en_mfa"):
                st.session_state.show_mfa_setup = True

            if st.session_state.get("show_mfa_setup"):
                st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
                if not st.session_state.get("totp_secret_temp"):
                    result = generate_totp_secret(user.get("email", "user"))
                    st.session_state.totp_secret_temp   = result["secret"]
                    st.session_state.totp_qr_bytes_temp = result["qr_bytes"]

                st.markdown("""
                <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
                     border-radius:8px;padding:24px;margin-bottom:16px;">
                  <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;
                       color:#e2e8f0;margin-bottom:16px;">Set up two-factor authentication</div>
                """, unsafe_allow_html=True)

                col_qr, col_inst = st.columns([1, 2])
                with col_qr:
                    st.image(st.session_state.totp_qr_bytes_temp, width=160)
                with col_inst:
                    st.markdown("""
                    <div style="font-size:13px;color:#b3bed0;line-height:1.8;">
                      <strong style="color:#e2e8f0;">Step 1</strong><br>
                      Open Google Authenticator, Authy, or any TOTP app.<br><br>
                      <strong style="color:#e2e8f0;">Step 2</strong><br>
                      Tap + → Scan QR code, or enter the manual key below.<br><br>
                      <strong style="color:#e2e8f0;">Step 3</strong><br>
                      Enter the 6-digit code your app shows to confirm.
                    </div>
                    """, unsafe_allow_html=True)
                    st.code(st.session_state.totp_secret_temp, language=None)

                st.markdown('</div>', unsafe_allow_html=True)

                mfa_code = st.text_input(
                    "Enter 6-digit code from your app",
                    max_chars=6, placeholder="000000", key="mfa_confirm_acc"
                )
                col_confirm, col_cancel = st.columns([1, 1])
                with col_confirm:
                    if st.button("Confirm & Enable MFA", key="confirm_mfa_acc", use_container_width=True):
                        if mfa_code and len(mfa_code) == 6:
                            if verify_totp(st.session_state.totp_secret_temp, mfa_code):
                                enable_mfa(uid(), st.session_state.totp_secret_temp)
                                st.session_state.show_mfa_setup   = False
                                st.session_state.totp_secret_temp = None
                                st.session_state.user["mfa_enabled"] = True
                                st.success("MFA enabled — your account is now protected")
                                st.rerun()
                            else:
                                st.error("Invalid code — check your authenticator app and try again")
                        else:
                            st.warning("Enter the 6-digit code")
                with col_cancel:
                    if st.button("Cancel", key="cancel_mfa_acc", use_container_width=True):
                        st.session_state.show_mfa_setup   = False
                        st.session_state.totp_secret_temp = None
                        st.rerun()

        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
        st.markdown("""
        <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
             border-radius:8px;padding:18px 20px;">
          <div style="font-family:'DM Mono',monospace;font-size:13px;color:#b3bed0;
               letter-spacing:.12em;margin-bottom:10px;">DATA SECURITY</div>
          <div style="font-size:13px;color:#9fabc0;line-height:1.8;">
            Your email and personal data are stored encrypted using Fernet
            authenticated encryption (AES-128-CBC with HMAC-SHA256). Passwords
            are hashed with bcrypt (cost 12) and never stored
            in plaintext. TOTP secrets are encrypted before storage. No sensitive
            data is ever logged or transmitted in plain text.
          </div>
        </div>
        """, unsafe_allow_html=True)

    # ── PLAN & BILLING ────────────────────────────────────────────────────────
    with tab_plan:
        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
        plan_color = "#d4a843" if plan in ("pro","institutional") else "#9fabc0"

        st.markdown(f"""
        <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.08);
             border-left:3px solid {plan_color};border-radius:8px;padding:20px 24px;margin-bottom:24px;">
          <div style="font-family:'DM Mono',monospace;font-size:14px;color:#b3bed0;
               letter-spacing:.12em;margin-bottom:8px;">CURRENT PLAN</div>
          <div style="font-family:'Syne',sans-serif;font-size:30px;font-weight:800;
               color:{plan_color};">{plan.upper()}</div>
          <div style="font-size:13px;color:#b3bed0;margin-top:6px;">
            {"Unlimited holdings · Hidden Gems · Signal alerts · Email notifications"
             if plan in ('pro','institutional')
             else "10 holdings · Market screener · HIGH/MODERATE/LOW conviction signals · Live model portfolio"}
          </div>
        </div>
        """, unsafe_allow_html=True)

        if plan == "free":
            # Comparison table — the "Pro" card shows the founding $0 offer only
            # while first-50 spots remain; once they're gone it reads as the
            # standard $29/mo plan (kept in sync with the upgrade-page gate).
            _acct_founding = _founding_spots_remaining() > 0
            _pro_label    = "FOUNDING MEMBER" if _acct_founding else "QNTM PRO"
            _pro_price    = "$0" if _acct_founding else "$29"
            _pro_sub      = ("first 50 users · then $29/mo" if _acct_founding
                             else "per month · cancel anytime")
            _pro_badge_ln = "✓ Founding member badge<br>" if _acct_founding else ""
            st.markdown(f"""
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;margin-bottom:24px;">

              <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
                   border-radius:10px;padding:24px;">
                <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;
                     color:#9fabc0;letter-spacing:.08em;margin-bottom:6px;">FREE</div>
                <div style="font-family:'Syne',sans-serif;font-size:36px;font-weight:800;
                     color:#e2e4f0;line-height:1;margin-bottom:4px;">$0</div>
                <div style="font-size:13px;color:#b3bed0;margin-bottom:18px;">forever</div>
                <div style="font-size:13px;color:#b3bed0;line-height:2;">
                  ✓ Screener — top 50 of {_universe_n()}<br>
                  ✓ HIGH / MODERATE / LOW conviction signals<br>
                  ✓ 5-pillar factor breakdown<br>
                  ✓ Up to 10 portfolio positions<br>
                  ✓ Live model portfolio (read-only)<br>
                  ✗ Hidden Gems<br>
                  ✗ Signal alerts<br>
                  ✗ Notifications
                </div>
              </div>

              <div style="background:rgba(212,168,67,.04);border:2px solid rgba(212,168,67,.45);
                   border-radius:10px;padding:24px;position:relative;">
                <div style="position:absolute;top:-12px;left:20px;background:#d4a843;color:#000;
                     font-family:'Syne',sans-serif;font-size:13px;font-weight:700;
                     letter-spacing:.1em;padding:3px 12px;border-radius:3px;">RECOMMENDED</div>
                <div style="font-family:'Syne',sans-serif;font-size:13px;font-weight:700;
                     color:#d4a843;letter-spacing:.08em;margin-bottom:6px;">{_pro_label}</div>
                <div style="font-family:'Syne',sans-serif;font-size:36px;font-weight:800;
                     color:#d4a843;line-height:1;margin-bottom:4px;">{_pro_price}</div>
                <div style="font-size:13px;color:#b3bed0;margin-bottom:18px;">
                  {_pro_sub}
                </div>
                <div style="font-size:13px;color:#b3bed0;line-height:2;">
                  ✓ Everything in Free<br>
                  ✓ Full {_universe_n()}-stock screener<br>
                  ✓ Unlimited portfolio positions<br>
                  ✓ 💎 Hidden Gem alerts<br>
                  ✓ Real-time signal notifications<br>
                  ✓ Macro regime change alerts<br>
                  ✓ Email signal summaries<br>
                  {_pro_badge_ln}✓ Priority support
                </div>
              </div>

            </div>
            """, unsafe_allow_html=True)

            st.markdown('</div>', unsafe_allow_html=True)

            st.markdown('<div class="land-btn-primary">', unsafe_allow_html=True)
            if _founding_spots_remaining() > 0:
                if st.button("Join Founding Members — Claim Free Spot", key="upgrade_btn", use_container_width=True):
                    ok = upgrade_plan(uid(), "pro")
                    # Force plan into session state immediately
                    if st.session_state.get("user"):
                        st.session_state.user["plan"] = "pro"
                    # Founder claim overrides any prior paid+canceled cycle: wipe
                    # stale Stripe/cancellation state so the billing reconciler
                    # doesn't downgrade this grant back to free on the next load.
                    if ok:
                        try:
                            from db import clear_stripe_state as _clear_billing
                            _clear_billing(uid())
                        except Exception:
                            pass
                    # Rewrite localStorage token with updated plan so nav restores correctly
                    _write_localstorage_token(uid(), "pro")
                    if ok:
                        # Refresh signed session params and rerun in place so the whole
                        # UI — including the nav badge — reflects Pro right away, with no
                        # navigation needed.
                        try:
                            st.query_params["uid"]  = _sign_token(uid(), "pro")
                            st.query_params["plan"] = "pro"
                        except Exception:
                            pass
                        st.success("✓ Founding Member activated — you now have full Pro access.")
                        st.rerun()
                    else:
                        st.warning("Could not write to DB — contact hello@qntm.live")
            else:
                # Founding window closed — never hand out free Pro from the account
                # page once the first 50 spots are gone. Route through the upgrade
                # page so the paid Stripe trial is the single source of billing,
                # exactly like the feature-gate CTAs.
                st.markdown(_cta_gold("Upgrade to Pro — Start 7-Day Trial",
                                      _upgrade_url("Pro Access", "account")),
                            unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        elif plan in ("pro","institutional"):
            # ── Subscription detail + cancellation flow ───────────────────────
            from db import schedule_cancellation, undo_cancellation
            from datetime import date, timedelta

            _notif = user.get("notifications") or {}
            _cancel_at = _notif.get("cancel_at") if isinstance(_notif, dict) else None
            _is_founding = (plan == "pro" and not _cancel_at and not _notif.get("billing_active"))
            # NB: Stripe wiring will populate `billing_active=True` and a proper
            # `current_period_end` later. For now, paid Pro users without a
            # cancel_at look identical to Founding Members.

            if _cancel_at:
                # Cancellation scheduled — show pending state with Undo
                st.markdown(f"""
                <div style="background:rgba(251,191,36,.05);border:1px solid rgba(251,191,36,.25);
                     border-radius:8px;padding:18px 22px;margin-bottom:12px;">
                  <div style="font-family:'Syne',sans-serif;font-size:14px;font-weight:700;
                       color:#fbbf24;letter-spacing:.05em;margin-bottom:6px;">
                    CANCELLATION SCHEDULED
                  </div>
                  <div style="font-size:13px;color:#b3bed0;line-height:1.7;">
                    Your subscription will end on <strong style="color:#e2e8f0;">{_cancel_at}</strong>.
                    You keep Pro access until that date. No refunds for partial months.
                    After that, your account converts to Free and your data is preserved.
                  </div>
                </div>
                """, unsafe_allow_html=True)
                if st.button("Undo Cancellation — Keep My Subscription",
                              key="undo_cancel_btn", use_container_width=True):
                    if undo_cancellation(uid()):
                        st.success("Cancellation undone — your subscription continues.")
                        st.rerun()
                    else:
                        st.error("Could not undo cancellation — contact billing@qntm.live")
            else:
                # Active subscription panel
                st.markdown(f"""
                <div style="background:rgba(52,211,153,.04);border:1px solid rgba(52,211,153,.15);
                     border-radius:8px;padding:16px 20px;font-size:13px;color:#4ade80;margin-bottom:18px;">
                  ✓ You have full Pro access. All features are enabled.
                </div>
                """, unsafe_allow_html=True)

                # Billing summary panel
                _next_bill = _notif.get("current_period_end") if isinstance(_notif, dict) else None
                _stripe_status = _notif.get("stripe_status") if isinstance(_notif, dict) else None
                _trial_end = _notif.get("trial_end") if isinstance(_notif, dict) else None

                # Trial countdown (status == trialing): compute days remaining.
                _trial_note = None
                if _stripe_status == "trialing" and _trial_end:
                    try:
                        from datetime import datetime as _dtt, timezone as _tz
                        _te = _dtt.fromtimestamp(int(_trial_end), tz=_tz.utc)
                        _now = _dtt.now(_tz.utc)
                        _days_left = max(0, (_te - _now).days)
                        _te_str = _te.strftime("%b %-d, %Y")
                        _trial_note = (f"Free trial — {_days_left} day"
                                       f"{'s' if _days_left != 1 else ''} left. "
                                       f"First charge of $29.00 on {_te_str} unless you cancel before then.")
                    except Exception:
                        _trial_note = "Free trial active."

                _next_bill_display = _next_bill if _next_bill else "—"
                # Convert epoch period-end to a date string if needed.
                if isinstance(_next_bill, (int, float)) or (isinstance(_next_bill, str) and _next_bill.isdigit()):
                    try:
                        from datetime import datetime as _dtb, timezone as _tzb
                        _next_bill_display = _dtb.fromtimestamp(int(_next_bill), tz=_tzb.utc).strftime("%b %-d, %Y")
                    except Exception:
                        pass

                _billing_label = "FOUNDING MEMBER" if _is_founding else (
                    "PRO · FREE TRIAL" if _stripe_status == "trialing" else "PRO · $29/month")
                if _is_founding:
                    _billing_note = "Free forever as a Founding Member. No billing scheduled."
                elif _trial_note:
                    _billing_note = _trial_note
                else:
                    _billing_note = f"Next charge: {_next_bill_display}"
                st.markdown(f"""
                <div style="background:rgba(255,255,255,.02);border:1px solid rgba(255,255,255,.07);
                     border-radius:8px;padding:18px 22px;margin-bottom:16px;">
                  <div style="display:flex;justify-content:space-between;align-items:start;gap:12px;flex-wrap:wrap;">
                    <div>
                      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#9fabc0;
                           letter-spacing:.12em;margin-bottom:4px;">SUBSCRIPTION</div>
                      <div style="font-family:'Syne',sans-serif;font-size:18px;font-weight:700;
                           color:#d4a843;">{_billing_label}</div>
                      <div style="font-size:13px;color:#b3bed0;margin-top:4px;">{_billing_note}</div>
                    </div>
                    <div style="text-align:right;">
                      <div style="font-family:'DM Mono',monospace;font-size:13px;color:#9fabc0;
                           letter-spacing:.12em;margin-bottom:4px;">STATUS</div>
                      <div style="font-family:'Syne',sans-serif;font-size:14px;font-weight:700;
                           color:#34d399;">ACTIVE</div>
                    </div>
                  </div>
                </div>
                """, unsafe_allow_html=True)

                # ── Cancellation ──────────────────────────────────────────────
                # ARL §17602(e)/(f): paid subscribers get TRUE one-click cancel —
                # a single visible click immediately stops the next renewal. No
                # expander maze, no multi-step confirm, no survey. Founding $0
                # members have nothing to cancel (informational only).
                if _is_founding:
                    with st.expander("Cancel subscription", expanded=False):
                        st.markdown(
                            '<div style="font-size:13px;color:#b3bed0;line-height:1.7;">'
                            'You are a Founding Member — no payment is scheduled, '
                            'so there is nothing to cancel. If you delete your account, '
                            'your Founding Member status is forfeited and cannot be restored.'
                            '</div>',
                            unsafe_allow_html=True)

                    # ── Optional: become a paying supporter ───────────────────
                    # A Founding Member already has Pro free. They may CHOOSE to
                    # start the $29/mo subscription to support the product. This
                    # is voluntary and has a one-way consequence: converting gives
                    # up Founding status, so if they later cancel they fall to
                    # regular Free (not free Founding Pro). Because it's an auto-
                    # renewing subscription, the ARL notice + affirmative consent
                    # still apply. FLAG FOR ATTORNEY REVIEW.
                    st.markdown('<div style="height:18px"></div>', unsafe_allow_html=True)
                    with st.expander("💛 Support QNTM — start a paid subscription (optional)", expanded=False):
                        st.markdown(
                            '<div style="background:rgba(212,168,67,.05);border:1px solid rgba(212,168,67,.25);'
                            'border-radius:8px;padding:16px 18px;margin-bottom:14px;">'
                            '<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:800;'
                            'color:#d4a843;margin-bottom:8px;">You don\u2019t need to do this</div>'
                            '<div style="font-size:13px;color:#b3bed0;line-height:1.8;">'
                            'As a Founding Member you already have full Pro access, free, for as long as '
                            'you keep your account. Starting a paid subscription is <strong style="color:#e2e8f0;">'
                            'completely optional</strong> \u2014 a way to support QNTM\u2019s development if you want to. '
                            'It does not unlock anything extra.'
                            '<br><br>'
                            '<strong style="color:#fbbf24;">Important \u2014 this is a one-way change:</strong> if you '
                            'start a paid subscription, you give up your free Founding Member status. '
                            'If you later cancel, your account will convert to the <strong style="color:#e2e8f0;">'
                            'regular Free tier</strong> (limited) at the end of the paid period \u2014 '
                            '<strong style="color:#e2e8f0;">not</strong> back to free Founding Pro. '
                            'Founding status, once given up, cannot be restored.'
                            '</div></div>',
                            unsafe_allow_html=True)

                        # Reuse the ARL initial notice (auto-renewal disclosure).
                        try:
                            import arl as _arl_f
                            _acct_u = f"?qnav=account&uid={uid()}&plan=pro&ck=1&_n=account"
                            st.markdown(_arl_f.initial_notice_html(_acct_u), unsafe_allow_html=True)
                            _f_consent = st.checkbox(
                                "I understand I already have Pro free as a Founding Member, that this paid "
                                "subscription is optional, and that if I cancel later I\u2019ll move to the "
                                "regular Free tier \u2014 not back to free Founding Pro. " + _arl_f.CHECKBOX_TEXT,
                                value=False, key="founder_paid_consent")
                        except Exception:
                            _f_consent = False

                        if _f_consent:
                            import stripe_billing as _sb_f
                            if _sb_f.billing_configured():
                                # Once a checkout URL exists, show ONLY the link
                                # button (hides the st.button so it can't render in
                                # its post-click white state).
                                if st.session_state.get("_checkout_url"):
                                    st.link_button("Continue to secure checkout →",
                                                   st.session_state["_checkout_url"],
                                                   use_container_width=True, type="primary")
                                    st.caption("Opens Stripe's secure checkout in a new tab.")
                                elif st.session_state.get("_checkout_err"):
                                    st.error(f"Could not start checkout: {st.session_state['_checkout_err']}  ·  Contact hello@qntm.live")
                                else:
                                    if st.button("Start $29/mo subscription", key="founder_start_paid",
                                                 use_container_width=True):
                                        import arl as _arl_f2
                                        _ipf = None
                                        try:
                                            _ipf = st.context.headers.get("X-Forwarded-For")
                                        except Exception:
                                            _ipf = None
                                        _arl_f2.log_consent(uid(), plan="pro_supporter", ip_address=_ipf)
                                        from db import get_stripe_billing as _gsbf
                                        _exf = _gsbf(uid()).get("stripe_customer_id")
                                        _basef = "https://qntm.live"
                                        try:
                                            _basef = "https://" + (st.context.headers.get("Host") or "qntm.live")
                                        except Exception:
                                            pass
                                        _emf = (st.session_state.user or {}).get("email", "")
                                        _urlf = _sb_f.create_checkout_url(uid(), _emf, _basef, _exf)
                                        if _urlf:
                                            st.session_state["_checkout_url"] = _urlf
                                            st.session_state["_awaiting_checkout"] = 3
                                            st.session_state["_stripe_polled"] = False
                                        else:
                                            st.session_state["_checkout_err"] = _sb_f.last_error()
                                        st.rerun()
                            else:
                                st.info("Paid subscriptions aren\u2019t enabled yet. Check back soon.")
                else:
                    _proposed_end = (
                        _next_bill if _next_bill
                        else (date.today() + timedelta(days=30)).isoformat()
                    )
                    st.markdown(
                        '<div style="font-size:13px;color:#b3bed0;line-height:1.7;margin-bottom:6px;">'
                        '<strong style="color:#e2e8f0;">Cancel your subscription</strong></div>'
                        '<div style="font-size:13px;color:#9fabc0;line-height:1.7;margin-bottom:12px;">'
                        'Clicking Cancel stops your next charge immediately. You keep Pro access '
                        f'until the end of your current paid period (<strong style="color:#b3bed0;">{_proposed_end}</strong>), '
                        'then your account converts to Free. Your data is preserved. '
                        '<strong style="color:#b3bed0;">If you\u2019re still in your 7-day free trial, '
                        'you won\u2019t be charged \u2014 you keep Pro through the end of your trial.</strong>'
                        '</div>',
                        unsafe_allow_html=True)
                    # TRUE one-click cancel — single button, immediate effect.
                    if st.button("Cancel subscription", key="cancel_sub_btn",
                                 use_container_width=True):
                        # If a real Stripe subscription exists, cancel at period
                        # end there (during trial = no charge ever; after trial =
                        # stops next renewal, access to period end).
                        _stripe_ok = True
                        _cancel_err = ""
                        _was_trial = False
                        _end_iso = _proposed_end
                        try:
                            import stripe_billing as _sbc
                            from db import get_stripe_billing as _gsbc
                            _bsc = _gsbc(uid())
                            _subc = _bsc.get("stripe_subscription_id")
                            if _subc and _sbc.billing_configured():
                                _cres = _sbc.cancel_subscription(_subc)
                                _stripe_ok = _cres.get("ok", False)
                                _was_trial = bool(_cres.get("was_trialing"))
                                _end_ts = _cres.get("end_ts")
                                if _end_ts:
                                    try:
                                        from datetime import datetime as _dt, timezone as _tz
                                        _end_iso = _dt.fromtimestamp(int(_end_ts), _tz.utc).date().isoformat()
                                    except Exception:
                                        pass
                                if not _stripe_ok:
                                    _cancel_err = _cres.get("error", _sbc.last_error())
                        except Exception as _ce:
                            _stripe_ok = True  # don't block local cancel on stripe error
                            _cancel_err = str(_ce)
                        if schedule_cancellation(uid(), _end_iso) and _stripe_ok:
                            # confirmation email (stubbed send + logged).
                            try:
                                import arl as _arl_c
                                _em = user.get("email")
                                if _em:
                                    _arl_c.send_cancellation_confirmation(uid(), _em, _end_iso)
                            except Exception:
                                pass
                            if _was_trial:
                                st.success(
                                    "Your subscription is cancelled — you won\u2019t be charged. "
                                    "You keep Pro access through the end of your free trial on "
                                    f"{_end_iso}, then your account converts to Free."
                                )
                            else:
                                st.success(
                                    "Your subscription is cancelled. Your next charge has been "
                                    f"stopped. You keep Pro access until {_end_iso}, then your "
                                    "account converts to Free. A confirmation email is on its way."
                                )
                            st.rerun()
                        else:
                            st.error(f"Could not cancel: {_cancel_err}  ·  contact billing@qntm.live")
                st.caption(
                    "Billing questions: billing@qntm.live · "
                    "[Billing & Refund Policy](?legal=billing)"
                )

    # ── NOTIFICATION PREFS ────────────────────────────────────────────────────
    with tab_notifs:
        st.markdown('<div style="height:6px;"></div>', unsafe_allow_html=True)
        if not plan_limit(plan, "notifications"):
            st.markdown("""
            <div style="background:rgba(251,191,36,.05);border:1px solid rgba(251,191,36,.2);
                 border-radius:6px;padding:14px 18px;font-size:13px;color:#fbbf24;margin-bottom:16px;">
              Notification preferences require a Pro or Founding Member plan.
            </div>
            """, unsafe_allow_html=True)
        else:
            prefs = user.get("notifications") or {}
            e_on = st.toggle("Email signal summaries (weekly digest)",
                             value=prefs.get("email", False), key="pref_email")
            s_on = st.toggle("In-app signal change alerts",
                             value=prefs.get("signals", True), key="pref_sig")
            a_on = st.toggle("Macro regime change alerts",
                             value=prefs.get("alerts", True), key="pref_alert")
            le_on = st.toggle("Email me when a holding or watchlist stock drops to LOW conviction (intraday, checked ~every 30 min during market hours)",
                              value=prefs.get("low_alert_email", False), key="pref_low_email")

            st.markdown('<div style="height:4px;"></div>', unsafe_allow_html=True)
            if st.button("Save Notification Preferences", key="save_prefs"):
                new_prefs = {**(user.get("notifications") or {}),
                             "email": e_on, "signals": s_on, "alerts": a_on,
                             "low_alert_email": le_on}
                if update_preferences(uid(), {"notifications": new_prefs}):
                    st.session_state.user["notifications"] = new_prefs
                    st.success("Preferences saved")
                else:
                    st.error("Save failed — try again")

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# PLATFORM SHELL
# ══════════════════════════════════════════════════════════════════════════════
def _window_track_series(model_series, spy_series, window_key):
    """Zoom the inception-anchored equity series to a trailing window WITHOUT
    rebasing — both lines keep their real cumulative $ values so the chart shows
    the true divergence. Returns (model_win, spy_win, model_ret_pct, spy_ret_pct,
    label) where the %s are the change over the visible window (= since-inception
    for ALL, matching the headline cards).
    window_key in {'1D','1W','1M','6M','1Y','ALL'}. Any window that predates
    inception (or leaves <2 points) gracefully falls back to all-since-inception."""
    from datetime import date, timedelta
    BASE = 100000.0
    if not model_series or not spy_series or len(model_series) < 2:
        return model_series, spy_series, 0.0, 0.0, "since inception"
    days = {"1D": 1, "1W": 7, "1M": 30, "3M": 90, "6M": 182, "1Y": 365}.get(window_key)
    first_date = model_series[0][0]
    cutoff = None
    if days is not None:
        try:
            cutoff = (date.today() - timedelta(days=days)).isoformat()
        except Exception:
            cutoff = None
    if days is None or cutoff is None or cutoff <= first_date:
        m, s, label = model_series, spy_series, "since inception"
    else:
        m = [pt for pt in model_series if pt[0] >= cutoff]
        s = [pt for pt in spy_series   if pt[0] >= cutoff]
        if len(m) < 2 or len(s) < 2:
            m, s, label = model_series, spy_series, "since inception"
        else:
            label = {"1D": "past day", "1W": "past week", "1M": "past month",
                     "3M": "past 3 months", "6M": "past 6 months",
                     "1Y": "past year"}.get(window_key, "window")
    try:
        # No rebasing — keep the real cumulative dollar values so the chart shows
        # the true divergence (both lines stay anchored to $100K at inception).
        # The reported % is the change over the visible window (start -> now); for
        # ALL this equals the since-inception number on the headline cards.
        m_ret = (m[-1][1] / (m[0][1] or BASE) - 1) * 100
        s_ret = (s[-1][1] / (s[0][1] or BASE) - 1) * 100
        return m, s, m_ret, s_ret, label
    except Exception:
        return m, s, 0.0, 0.0, label


@st.cache_data(ttl=300, show_spinner=False)
def _intraday_prices(tickers_tuple):
    """Today's (or the last session's) intraday 15-minute bars for the given
    tickers plus SPY. Returns (labels, {ticker: [price,...]}) where labels are
    ET clock times aligned to each bar, or (None, None) on failure. Cached 5 min."""
    try:
        import yfinance as yf
        try:
            from zoneinfo import ZoneInfo
            _et = ZoneInfo("America/New_York")
        except Exception:
            _et = None
        syms = sorted({t for t in tickers_tuple if t and t not in DELISTED} | {"SPY"})
        dl = yf.download(syms, period="1d", interval="15m",
                         progress=False, auto_adjust=True)
        if dl is None or dl.empty:
            return None, None
        close = dl["Close"].ffill()
        labels = []
        for ts in close.index:
            t = ts
            if _et is not None:
                try:
                    t = ts.tz_convert(_et)
                except Exception:
                    try:
                        t = ts.tz_localize("UTC").tz_convert(_et)
                    except Exception:
                        t = ts
            try:
                labels.append(t.strftime("%-I:%M %p"))
            except Exception:
                labels.append(str(t)[11:16])
        cols = {}
        multi = hasattr(close, "columns")
        for tk in syms:
            if multi and tk in close.columns:
                cols[tk] = [float(v) if v == v else None for v in close[tk].values]
            elif not multi and len(syms) == 1:
                cols[tk] = [float(v) if v == v else None for v in close.values]
        return labels, cols
    except Exception:
        return None, None


def _intraday_track_series(pt, positions):
    """Build today's intraday equity curve (model vs SPY) by marking the open
    book at each 15-min bar: model(t) = cash + Σ shares·price(t). Cash is held
    constant intraday (no intraday entries/exits). Both series are rebased to
    $100K at the day's first bar so the comparison shows today's move. Returns
    (model, spy, model_ret, spy_ret, label) with HH:MM labels, or None."""
    if not pt or not positions:
        return None
    active = [(p["ticker"], float(p["entry_price"])) for p in positions
              if p.get("entry_price") and float(p["entry_price"]) > 0]
    if not active:
        return None
    labels, cols = _intraday_prices(tuple(sorted({t for t, _ in active})))
    if not labels or not cols or "SPY" not in cols:
        return None
    pm = pt.get("price_map", {})
    def last_close(tk):
        m = pm.get(tk) or {}
        return m[max(m)] if m else None
    shares = {tk: 2000.0 / ep for tk, ep in active}
    base_close = 0.0
    for tk, _ in active:
        lc = last_close(tk)
        if lc:
            base_close += shares[tk] * lc
    cash = (pt.get("model_value") or 100000.0) - base_close
    spy_arr = cols.get("SPY") or []
    model, spy = [], []
    for i in range(len(labels)):
        s = spy_arr[i] if i < len(spy_arr) else None
        if s is None:
            continue
        mtm, ok = cash, True
        for tk, _ in active:
            arr = cols.get(tk) or []
            v = arr[i] if i < len(arr) else None
            if v is None:
                v = last_close(tk)
            if v is None:
                ok = False
                break
            mtm += shares[tk] * v
        if not ok:
            continue
        model.append((labels[i], mtm))
        spy.append((labels[i], s))
    if len(model) < 2:
        return None
    BASE = 100000.0
    m0 = model[0][1] or BASE
    s0 = spy[0][1] or BASE
    m_re = [(l, v / m0 * BASE) for l, v in model]
    s_re = [(l, v / s0 * BASE) for l, v in spy]
    m_ret = (m_re[-1][1] / BASE - 1) * 100
    s_ret = (s_re[-1][1] / BASE - 1) * 100
    return m_re, s_re, m_ret, s_ret, f"today · {labels[0]}\u2013{labels[-1]} ET"


@st.fragment
def _render_track_equity(_pt, positions, day_pct=None, day_spy_pct=None):
    """Equity-curve window selector + chart for the track record, isolated in a
    fragment so changing the time window re-renders ONLY the chart, not the whole
    model-portfolio page. Nothing outside the chart depends on the window."""
    _win = st.radio("Window", ["1D", "1M", "3M", "1Y", "All"],
                    index=0, horizontal=True, key="tr_window",
                    label_visibility="collapsed",
                    format_func=lambda x: "Day" if x == "1D" else x)
    _wk = "ALL" if _win == "All" else _win
    # Anchor both daily curves to the live marks (model_value / spy_ret) so the
    # chart agrees with the headline cards. Without this the series ends at the
    # last daily close and lags the cards by today's intraday move.
    from datetime import date as _date_anchor
    _today_iso = _date_anchor.today().isoformat()
    _mseries = list(_pt.get("model_series") or [])
    _sseries = list(_pt.get("spy_series") or [])
    # Anchor BOTH curves at exactly $100K at inception so the chart's
    # since-inception return matches the headline cards (which measure from the
    # $100K starting capital). The model curve's first daily point is the day-1
    # close mark, already carrying day-1 P&L, so without this the chart measures
    # from ~$100.2K and the ALL % drifts ~0.2pt off the cards. (SPY already starts
    # at $100K; we prepend to both to keep the two series index-aligned.)
    _base0 = 100000.0
    if _mseries and _sseries and _mseries[0][1] != _base0:
        _inc0 = _pt.get("inception") or _mseries[0][0]
        _mseries.insert(0, (_inc0, _base0))
        _sseries.insert(0, (_inc0, _base0))
    _mv_live = _pt.get("model_value")
    _sret_live = _pt.get("spy_ret")
    if _mseries and _mv_live:
        if _mseries[-1][0] == _today_iso:
            _mseries[-1] = (_today_iso, _mv_live)
        else:
            _mseries.append((_today_iso, _mv_live))
    if _sseries and _sret_live is not None:
        _s_live = _sseries[0][1] * (1 + _sret_live / 100.0)
        if _sseries[-1][0] == _today_iso:
            _sseries[-1] = (_today_iso, _s_live)
        else:
            _sseries.append((_today_iso, _s_live))
    _intraday = False
    if _wk == "1D" and day_pct is not None:
        # DAY = today's move. The model line runs prev-close -> live value, so it
        # ends on PORTFOLIO VALUE. SPY is drawn on its REAL cumulative dollar
        # basis (same as every other window and the % vs SPY card) — NOT rebased
        # to the model's start. A SPY that's down since inception therefore stays
        # BELOW the model here too, instead of floating above $100K. Today's move
        # for each line shows as its slope / endpoint %, not as an inflated level.
        _mv = float(_pt.get("model_value") or 100000.0)
        _dp = float(day_pct)
        _sp = float(day_spy_pct) if day_spy_pct is not None else 0.0
        _m_start = _mv / (1 + _dp / 100.0) if _dp != -100 else _mv
        _spy_now = 100000.0 * (1 + float(_pt.get("spy_ret") or 0.0) / 100.0)
        _s_start = _spy_now / (1 + _sp / 100.0) if _sp != -100 else _spy_now
        _ms = [("prev close", _m_start), ("now", _mv)]
        _ss = [("prev close", _s_start), ("now", _spy_now)]
        _mret, _sret = _dp, _sp
        _intraday = True
        _wlabel = "today"
    elif _wk == "1D":
        # No live day-change data — fall back to cumulative since entry. (Also the
        # day-one case, where today's move and since-entry are the same thing.)
        BASE = 100000.0
        _mv = float(_pt.get("model_value") or BASE)
        _mr = float(_pt.get("model_ret") or 0.0)
        _sr = float(_pt.get("spy_ret") or 0.0)
        _ms = [("entry", BASE), ("now", _mv)]
        _ss = [("entry", BASE), ("now", BASE * (1 + _sr / 100.0))]
        _mret, _sret = _mr, _sr
        _intraday = True
        _wlabel = "since entry"
    else:
        _ms, _ss, _mret, _sret, _wlabel = _window_track_series(
            _mseries, _sseries, _wk)
    _eqchart = _tr_line_chart_svg(_ms, _ss, intraday=_intraday)
    if _eqchart:
        _vs = _mret - _sret
        _vs_color = "#34d399" if _vs >= 0 else "#f87171"
        _pd = 2 if _intraday else 1  # DAY window: match the 2dp TODAY card
        # On the DAY window the vs-SPY here is today's race, not the cumulative
        # "% vs SPY" card — tag it so the two don't read as contradictory.
        _vs_tag = f" {_wlabel}" if _wk == "1D" else ""
        st.markdown(f"""
        <div style="background:#0a0b14;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:16px 16px 10px;margin-bottom:20px;">
          <div style="display:flex;gap:16px;align-items:center;flex-wrap:wrap;margin-bottom:10px;font-family:DM Mono,monospace;font-size:13px;">
            <span style="color:#d4a843;">\u2014 QNTM Model {_mret:+.{_pd}f}%</span>
            <span style="color:#7c8aa0;">\u2014 SPY {_sret:+.{_pd}f}%</span>
            <span style="color:{_vs_color};">vs SPY {_vs:+.{_pd}f}%{_vs_tag}</span>
            <span style="color:#8896ac;margin-left:auto;">{_wlabel}</span>
          </div>
          {_eqchart}
        </div>
        """, unsafe_allow_html=True)


@st.cache_data(ttl=30, show_spinner=False)
def _prices_as_of():
    """How fresh the stored intraday prices are: the latest benchmark_price
    .updated_at (SPY is rewritten every intraday cron cycle). In stored-only mode
    this is the true age of the live numbers and doubles as a cron heartbeat.
    Returns (local 'HH:MM TZ' string, minutes_old) or (None, None)."""
    try:
        from data_refresh import _get_supabase
        from datetime import datetime as _dt, timezone as _tz
        sb = _get_supabase()
        if not sb:
            return None, None
        resp = sb.table("benchmark_price").select("updated_at").not_.is_(
            "updated_at", "null").order("updated_at", desc=True).limit(1).execute()
        if not (resp.data and resp.data[0].get("updated_at")):
            return None, None
        raw = str(resp.data[0]["updated_at"]).replace("Z", "+00:00")
        dt_utc = _dt.fromisoformat(raw).astimezone(_tz.utc)
        mins = int((_dt.now(_tz.utc) - dt_utc).total_seconds() // 60)
        try:
            from zoneinfo import ZoneInfo
            loc = dt_utc.astimezone(ZoneInfo("America/Los_Angeles"))
            return loc.strftime("%H:%M %Z"), mins
        except Exception:
            return dt_utc.strftime("%H:%M UTC"), mins
    except Exception:
        return None, None


def page_model_portfolio():
    _pin_nav("model_portfolio")
    # Model portfolio: HIGH conviction positions, exits at score < 45
    from data_refresh import _get_supabase
    import datetime

    page_summary(
        "🏆", "Portfolio & Track Record",
        "Live model portfolio vs SPY · equal-weighted $2K · true P&L · exits at Low Conviction"
    )

    # ── Manual refresh — re-pull the live equity curve without leaving the page ─
    _rc1, _rc2 = st.columns([5, 2])
    with _rc2:
        st.markdown(
            "<style>"
            ".st-key-tr_refresh button{white-space:nowrap;padding:8px 14px;"
            "background:linear-gradient(135deg,rgba(52,211,153,.14),rgba(52,211,153,.04));"
            "border:1px solid rgba(52,211,153,.38);border-radius:10px;color:#6ee7b7;"
            "font-weight:600;letter-spacing:.02em;"
            "box-shadow:inset 0 1px 0 rgba(255,255,255,.05);"
            "transition:background .15s ease,border-color .15s ease,box-shadow .15s ease,color .15s ease;}"
            ".st-key-tr_refresh button:hover{"
            "background:linear-gradient(135deg,rgba(52,211,153,.24),rgba(52,211,153,.09));"
            "border-color:rgba(52,211,153,.65);color:#a7f3d0;"
            "box-shadow:0 0 0 1px rgba(52,211,153,.25),0 4px 14px rgba(52,211,153,.15);}"
            ".st-key-tr_refresh button:active{transform:translateY(1px);}"
            ".st-key-tr_refresh button p,.st-key-tr_refresh button div{"
            "white-space:nowrap;font-size:13px;}"
            "</style>",
            unsafe_allow_html=True)
        if st.button("↻ Refresh", key="tr_refresh", use_container_width=True,
                     help="Re-pull the latest prices and re-mark the equity curve now"):
            # Clear the process-wide ledger cache so the next render recomputes.
            try:
                _track_record_cached.clear()
            except Exception:
                pass
            try:
                _LIVE_QUOTE_CACHE.clear()
            except Exception:
                pass
            try:
                _mini_price_data.clear()
            except Exception:
                pass
            st.rerun()

    sb = _get_supabase()

    # ── Load positions from Supabase ──────────────────────────────────────────
    positions = []
    if sb:
        try:
            resp = sb.table("model_portfolio_positions") \
                .select("*") \
                .eq("is_active", True) \
                .eq("epoch", MODEL_EPOCH) \
                .order("entry_date", desc=False) \
                .execute()
            raw_positions = resp.data or []
            # Defensive dedup: keep only the most recent active row per ticker.
            # Historically the seed + rebuild + intraday cron paths could each
            # insert without coordinating, producing dupes that inflate totals.
            # Newest entry_date wins; if entry_dates tie, newest id wins.
            _dedup = {}
            for p in raw_positions:
                tk = p["ticker"]
                ed = str(p.get("entry_date") or "")
                pid = p.get("id") or 0
                cur = _dedup.get(tk)
                if (cur is None
                    or ed > str(cur.get("entry_date") or "")
                    or (ed == str(cur.get("entry_date") or "") and pid > (cur.get("id") or 0))):
                    _dedup[tk] = p
            positions = list(_dedup.values())
            _n_dupes = len(raw_positions) - len(positions)
            if _n_dupes > 0:
                st.warning(
                    f"⚠ Found {_n_dupes} duplicate position rows in the database — "
                    f"showing the most recent entry per ticker. Run the rebuild "
                    f"script (`python3 rebuild_model_portfolio.py`) to clean up."
                )
        except Exception as e:
            st.warning(f"Could not load positions: {e}")

    scan = st.session_state.get("scan_results") or []
    score_map = {r["ticker"]: r for r in scan}

    # ── Pull latest prices + scores from signal_log (no scan required) ────────
    if sb:
        try:
            tickers = [p["ticker"] for p in positions]
            sig_resp = sb.table("signal_log")                 .select("ticker,price,adj_composite,composite,signal,momentum,quality,volume,value,sentiment,is_hidden_gem,mktcap,val_low,val_high,value_position,val_basis")                 .in_("ticker", tickers)                 .order("signal_date", desc=True)                 .limit(len(tickers) * 3)                 .execute()
            # Take most recent row per ticker
            seen = set()
            for row in (sig_resp.data or []):
                tk = row["ticker"]
                if tk not in seen:
                    seen.add(tk)
                    # Merge into score_map — signal_log wins over stale session state
                    if tk not in score_map:
                        score_map[tk] = {}
                    for field in ["price","adj_composite","composite","signal","momentum","quality","volume","value","sentiment","is_hidden_gem","hidden_gem_reason","mktcap","val_low","val_high","value_position","val_basis"]:
                        if row.get(field) is not None:
                            score_map[tk][field] = row[field]
        except Exception:
            pass  # fall back to session state if query fails


    if not positions:
        # No positions yet — show what would be entered today
        st.markdown(
            '<div style="background:rgba(212,168,67,.06);border:1px solid rgba(212,168,67,.2);'
            'border-radius:8px;padding:20px 24px;margin-bottom:24px;font-size:13px;color:#d4a843;">'
            '⚡ Model portfolio initializes tonight at 2 AM UTC when the nightly cron runs. '
            'Run a Rescan on the Screener first to seed today\'s signals.</div>',
            unsafe_allow_html=True)

        # Preview what would be entered
        buys = sorted(
            [r for r in scan if r.get("adj_composite", r.get("composite", 0)) >= 60],
            key=lambda x: x.get("adj_composite", x.get("composite", 0)),
            reverse=True
        )[:20]

        if buys:
            st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;'
                        'letter-spacing:.1em;margin-bottom:12px;">TONIGHT\'S ENTRIES (PREVIEW)</div>',
                        unsafe_allow_html=True)
            for i, r in enumerate(buys):
                bg = "rgba(255,255,255,.02)" if i % 2 == 0 else "rgba(255,255,255,.008)"
                price_str = f'${r["price"]:,.2f}' if r.get("price") else "—"
                st.markdown(
                    f'<div style="display:grid;grid-template-columns:80px 1fr 80px 60px;'
                    f'gap:4px;padding:8px 12px;background:{bg};'
                    f'border:1px solid rgba(255,255,255,.04);border-radius:4px;margin-bottom:2px;">'
                    f'<div style="font-family:Syne,sans-serif;font-size:13px;font-weight:800;color:#e2e8f0;">{r["ticker"]}</div>'
                    f'<div style="font-size:13px;color:#9fabc0;">Entry today</div>'
                    f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;">{price_str}</div>'
                    f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:#34d399;">{r.get("adj_composite", 0):.0f}</div>'
                    f'</div>', unsafe_allow_html=True)
        return

    # ── Live prices via yfinance — cached per session (5 min TTL) ───────────
    import time as _mp_time
    _mp_cache_age = _mp_time.time() - st.session_state.get("_mp_prices_at", 0)
    tickers_to_fetch = [p["ticker"] for p in positions]
    if _mp_cache_age > 300 or not st.session_state.get("_mp_prices"):
        live_prices = {}
        # First fill from signal_log (instant)
        for tk in tickers_to_fetch:
            if score_map.get(tk, {}).get("price"):
                live_prices[tk] = float(score_map[tk]["price"])
        # Then fetch live prices from yfinance to get current market prices
        try:
            import yfinance as yf
            _px_data = yf.download(_strip_delisted(tickers_to_fetch), period="1d", interval="1d",
                                   auto_adjust=True, progress=False, threads=True)
            if not _px_data.empty:
                _cls = _px_data["Close"]
                if hasattr(_cls, "columns"):
                    for tk in tickers_to_fetch:
                        if tk in _cls.columns:
                            _v = _cls[tk].dropna()
                            if not _v.empty: live_prices[tk] = float(_v.iloc[-1])
                else:
                    _v = _cls.dropna()
                    if not _v.empty and len(tickers_to_fetch)==1:
                        live_prices[tickers_to_fetch[0]] = float(_v.iloc[-1])
        except Exception:
            pass  # fall back to signal_log prices
        st.session_state._mp_prices    = live_prices
        st.session_state._mp_prices_at = _mp_time.time()
    else:
        live_prices = st.session_state._mp_prices

    # ── Calculate portfolio metrics ───────────────────────────────────────────
    today = datetime.date.today().isoformat()
    holdings = []
    total_invested = 0
    total_current  = 0

    for pos in positions:
        tk           = pos["ticker"]
        entry_price  = pos.get("entry_price")
        pos_size     = pos.get("position_size", 2000)
        current_data = score_map.get(tk, {})
        # Prefer live yfinance price, fall back to signal_log
        current_price = live_prices.get(tk) or current_data.get("price")

        if entry_price and current_price and entry_price > 0:
            shares      = pos_size / entry_price
            current_val = shares * current_price
            pnl         = current_val - pos_size
            pnl_pct     = (current_val / pos_size - 1) * 100
        else:
            shares      = None
            current_val = pos_size
            pnl         = 0
            pnl_pct     = 0

        total_invested += pos_size
        total_current  += current_val

        holdings.append({
            "ticker":        tk,
            "entry_date":    pos.get("entry_date", today),
            "entry_price":   entry_price,
            "entry_score":   pos.get("entry_score", 50),
            "current_price": current_price,
            "current_score": current_data.get("adj_composite", current_data.get("composite", pos.get("entry_score", 50))),
            "momentum":      current_data.get("momentum", 50),
            "quality":       current_data.get("quality",  50),
            "volume":        current_data.get("volume",   50),
            "value":         current_data.get("value",    50),
            "sentiment":     current_data.get("sentiment",50),
            "pos_size":      pos_size,
            "current_val":   current_val,
            "pnl":           pnl,
            "pnl_pct":       pnl_pct,
            "is_gem":        current_data.get("is_hidden_gem", False),
        })

    # ── True portfolio P&L — real-account basis. Start $100K; fold in realized
    # P&L from exited positions (e.g. BORR's loss) alongside unrealized P&L on
    # the active book, so this reflects the actual account, not just current
    # holdings. The poisoned-batch exits were reversed in the data (those names
    # are active again with no exit record), so they're correctly excluded.
    # Matches the Track Record page exactly.
    BASE_CAPITAL = 100000.0
    realized_pnl = 0.0
    try:
        _ex = sb.table("model_portfolio_positions") \
            .select("ticker,entry_date,exit_date,entry_price,position_size,exit_price") \
            .eq("is_active", False).eq("epoch", MODEL_EPOCH).execute()
        _seen = {}
        for _r in (_ex.data or []):
            _k = (_r.get("ticker"), str(_r.get("entry_date") or "")[:10], str(_r.get("exit_date") or "")[:10])
            _seen[_k] = _r  # dedup matches Track Record
        for _r in _seen.values():
            _ep, _xp = _r.get("entry_price"), _r.get("exit_price")
            _ps = _r.get("position_size", 2000) or 2000
            if _ep and _xp and float(_ep) > 0:
                realized_pnl += (_ps / float(_ep)) * float(_xp) - _ps
    except Exception:
        pass

    unrealized_pnl  = total_current - total_invested
    portfolio_value = BASE_CAPITAL + unrealized_pnl + realized_pnl
    port_return     = (portfolio_value / BASE_CAPITAL - 1) * 100
    port_pnl        = portfolio_value - BASE_CAPITAL
    sign            = "+" if port_return >= 0 else ""
    ret_color       = "#34d399" if port_return >= 0 else "#f87171"

    # ── SPY benchmark — $100K invested at inception, marked to latest close.
    # Standard lump-at-inception benchmark, matching the Track Record page so
    # the two pages report the same SPY comparison.
    spy_return = 0.0
    spy_pnl    = 0.0
    try:
        import yfinance as yf
        from datetime import date as _dt
        if "_mp_spy" not in st.session_state:
            st.session_state._mp_spy = None  # fetch deferred
        spy_hist = st.session_state._mp_spy
        if spy_hist is None:
            spy_hist = _stored_spy_hist(MODEL_INCEPTION)
            if spy_hist is None:
                spy_hist = yf.download("SPY", start=MODEL_INCEPTION, progress=False, auto_adjust=True)
            st.session_state._mp_spy = spy_hist
        if not spy_hist.empty:
            spy_close = spy_hist["Close"]
            if hasattr(spy_close, "columns"): spy_close = spy_close.iloc[:,0]
            spy_close = spy_close.squeeze().dropna()
            try:
                inception_mp = min(_dt.fromisoformat(str(p.get("entry_date",""))[:10])
                                   for p in positions)
                w = spy_close[spy_close.index.date >= inception_mp]
            except Exception:
                w = spy_close
            if not w.empty:
                spy_now  = float(spy_close.iloc[-1])
                spy_base = float(w.iloc[0])
                spy_return = (spy_now / spy_base - 1) * 100
                spy_pnl    = BASE_CAPITAL * (spy_return / 100)
    except Exception:
        pass

    # ── Single source of truth — override the headline totals with the shared
    # ledger so this page and the Track Record page report IDENTICAL portfolio
    # value, return, and vs-SPY. (The position list below still uses this page's
    # own per-ticker prices for the card detail; only the top-line totals are
    # unified here.) Falls back to this page's own computation if unavailable.
    _pt = _portfolio_truth(sb)
    if _pt:
        portfolio_value = _pt["model_value"]
        port_return     = _pt["model_ret"]
        port_pnl        = portfolio_value - BASE_CAPITAL
        spy_return      = _pt["spy_ret"]
        spy_pnl         = BASE_CAPITAL * (spy_return / 100)
        sign            = "+" if port_return >= 0 else ""
        ret_color       = "#34d399" if port_return >= 0 else "#f87171"

    vs_spy_pct = port_return - spy_return
    vs_spy_pnl = port_pnl - spy_pnl
    vs_color   = "#34d399" if vs_spy_pct >= 0 else "#f87171"
    vs_sign    = "+" if vs_spy_pct >= 0 else ""



    # ── Methodology banner ────────────────────────────────────────────────────
    # Dynamic entry dates from positions
    _entry_dates = sorted(set(str(p.get('entry_date',''))[:10] for p in positions if p.get('entry_date')))
    _start_date  = _entry_dates[0] if _entry_dates else MODEL_INCEPTION
    _end_date    = _entry_dates[-1] if _entry_dates else '2026-05-25'
    st.markdown(
        '<div style="background:rgba(212,168,67,.04);border:1px solid rgba(212,168,67,.15);'
        'border-radius:8px;padding:16px 20px;margin-bottom:20px;">'
        '<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;'
        'letter-spacing:.1em;margin-bottom:8px;">⚡ INVESTMENT METHODOLOGY</div>'
        '<div style="font-size:13px;color:#b3bed0;line-height:1.7;">'
        f'Built from <strong style="color:#cbd5e1;">{_start_date}</strong> — '
        'highest conviction signals entered each day. '
        'Entry threshold: <strong style="color:#34d399;">≥ 67</strong> in HIGH VOLATILITY regime, '
        '<strong style="color:#34d399;">≥ 60</strong> in normal regimes. '
        'Equal-weighted at <strong style="color:#cbd5e1;">$2,000 per position</strong> ($100K total). '
        '30% sector cap enforced at entry.'
        '<br><br>'
        '<strong style="color:#cbd5e1;">Exit discipline:</strong> Positions exit when conviction '
        'drops below <strong style="color:#f87171;">45</strong>. Capital redeploys into next '
        'highest conviction signal. No discretionary overrides.'
        '</div>'
        f'<div style="font-family:DM Mono,monospace;font-size:10px;color:#4b5563;'
        f'letter-spacing:.08em;margin-top:10px;">build {_build_tag()}</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # ── Summary strip — CSS grid wraps to 2-3 cols on mobile ────────────────
    ss = "background:#0d1117;border:1px solid rgba(255,255,255,.07);border-radius:6px;padding:14px 16px;text-align:center;"
    ls = "font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;letter-spacing:.08em;margin-bottom:6px;"

    pnl_sign    = "+" if port_pnl >= 0 else ""
    vs_pnl_sign = "+" if vs_spy_pnl >= 0 else ""

    # Portfolio-level day change vs yesterday's close — value-weighted across
    # positions using the same source the per-card strips use (cached by ticker
    # set, so the per-card fetch later is a cache hit). This is the book's move
    # vs the prior session's close — live through the day, frozen at the close.
    _mp_day_change = _fetch_day_change_map(
        [h["ticker"] for h in holdings] + ["SPY"], cache_key="_mp_daychange_cache")
    _day_today = _day_prev = 0.0
    _day_have = False
    _day_settled = False
    for _h in holdings:
        _dc = _mp_day_change.get(_h["ticker"]) or {}
        _ep = _h.get("entry_price")
        if not _ep or _ep <= 0:
            continue
        _sh = _h.get("pos_size", 2000) / _ep
        if _dc.get("price") and _dc.get("prev_close"):
            _day_today += _sh * float(_dc["price"])
            _day_prev  += _sh * float(_dc["prev_close"])
            _day_have = True
            if _dc.get("settled"):
                _day_settled = True
    if _day_have and _day_prev > 0:
        _day_dollar = _day_today - _day_prev
        _day_pct    = _day_dollar / _day_prev * 100
        # Headline % and $ both come from the live active-book weighting so they
        # agree with each other and with the live 1D chart. (The daily-close
        # ledger move, _pt["day_model"], has no intraday component and would read
        # ~0 here, so it's no longer used for the headline.)
        _day_color  = "#34d399" if _day_pct > 0 else ("#f87171" if _day_pct < 0 else "#b3bed0")
        _day_s      = "+" if _day_pct >= 0 else ""
        _day_val    = f"{_day_s}{_day_pct:.2f}%"
        _day_sub    = f"{_day_s}${_day_dollar:,.0f}" + (" · at close" if _day_settled else "")
    else:
        _day_color, _day_val, _day_sub = "#9fabc0", "—", "no data"

    # Today's move to feed the DAY chart so it matches the TODAY card exactly:
    # the model % is the same active-book day_pct above; SPY's day move comes from
    # the same feed. None when there's no live data (chart falls back to cumulative).
    _chart_day_pct = _day_pct if (_day_have and _day_prev > 0) else None
    _spy_dcd = _mp_day_change.get("SPY") or {}
    _chart_day_spy = (
        (float(_spy_dcd["price"]) / float(_spy_dcd["prev_close"]) - 1) * 100
        if _spy_dcd.get("price") and _spy_dcd.get("prev_close") else None)

    # Portfolio extended-hours move — value-weighted $ summed across holdings,
    # shown as its own summary card in the live session only (pre before open,
    # after-hours after close; nothing extra during regular trading).
    _mp_phase = _market_phase()
    _mp_xk = "pre" if _mp_phase == "pre" else ("post" if _mp_phase in ("post", "closed") else None)
    _xh_card = ""
    if _mp_xk:
        _xh_dollar = 0.0; _xh_have = False
        for _xh_h in holdings:
            _xh_dc = _mp_day_change.get(_xh_h["ticker"]) or {}
            _xh_ep = _xh_h.get("entry_price")
            if not _xh_ep or _xh_ep <= 0:
                continue
            _xh_c = _xh_dc.get(f"{_mp_xk}_chg")
            if _xh_c is not None:
                _xh_dollar += (_xh_h.get("pos_size", 2000) / _xh_ep) * float(_xh_c)
                _xh_have = True
        if _xh_have and portfolio_value:
            _xh_pct = _xh_dollar / portfolio_value * 100
            _xh_col = "#34d399" if _xh_pct > 0 else ("#f87171" if _xh_pct < 0 else "#b3bed0")
            _xh_s   = "+" if _xh_pct >= 0 else ""
            _xh_lbl = "PRE MKT" if _mp_xk == "pre" else "AFTER HRS"
            # SPY's extended-hours move for a side-by-side compare in the same box
            _spy_xh  = _fetch_extended_hours_map(["SPY"]).get("SPY", {})
            _spy_pct = _spy_xh.get(f"{_mp_xk}_pct")
            _spy_line = ""
            if _spy_pct is not None:
                _spy_pct = float(_spy_pct)
                _spy_col = "#34d399" if _spy_pct > 0 else ("#f87171" if _spy_pct < 0 else "#8896ac")
                _spy_s   = "+" if _spy_pct >= 0 else ""
                _spy_line = (
                    f'<div style="font-family:DM Mono,monospace;font-size:12px;color:#8896ac;margin-top:3px;">'
                    f'SPY&nbsp;<span style="color:{_spy_col};">{_spy_s}{_spy_pct:.2f}%</span></div>'
                )
            _xh_card = (
                f'<div style="{ss}"><div style="{ls}">{_xh_lbl}</div>'
                f'<div style="font-size:18px;font-weight:700;color:{_xh_col};">{_xh_s}{_xh_pct:.2f}%</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;margin-top:2px;">{_xh_s}${_xh_dollar:,.0f}</div>'
                f'{_spy_line}</div>'
            )

    st.markdown(f"""
    <div style="display:grid;grid-template-columns:repeat(auto-fit,minmax(120px,1fr));gap:8px;margin-bottom:20px;">
      <div style="{ss}"><div style="{ls}">PORTFOLIO VALUE</div>
        <div style="font-size:18px;font-weight:700;color:#d4a843;">${portfolio_value:,.0f}</div></div>
      <div style="{ss}"><div style="{ls}">TODAY</div>
        <div style="font-size:18px;font-weight:700;color:{_day_color};">{_day_val}</div>
        <div style="font-family:DM Mono,monospace;font-size:13px;color:#8896ac;margin-top:2px;">{_day_sub}</div></div>
      {_xh_card}
      <div style="{ss}"><div style="{ls}">$ CHANGE</div>
        <div style="font-size:18px;font-weight:700;color:{ret_color};">{pnl_sign}${port_pnl:,.0f}</div></div>
      <div style="{ss}"><div style="{ls}">% RETURN</div>
        <div style="font-size:18px;font-weight:700;color:{ret_color};">{sign}{port_return:.1f}%</div></div>
      <div style="{ss}"><div style="{ls}">$ vs SPY</div>
        <div style="font-size:18px;font-weight:700;color:{vs_color};">{vs_pnl_sign}${vs_spy_pnl:,.0f}</div></div>
      <div style="{ss}"><div style="{ls}">% vs SPY</div>
        <div style="font-size:18px;font-weight:700;color:{vs_color};">{vs_sign}{vs_spy_pct:.1f}%</div></div>
    </div>
    """, unsafe_allow_html=True)

    # Price-freshness indicator — true age of the stored intraday marks driving
    # the live numbers (benchmark_price.updated_at). Amber if it's gone stale
    # during regular trading, which means the intraday price cron is lagging.
    _pao, _pao_min = _prices_as_of()
    if _pao:
        _stale = (_market_phase() == "regular" and _pao_min is not None and _pao_min > 20)
        _pao_col = "#f59e0b" if _stale else "#8896ac"
        _pao_note = " \u00b7 stale, refresh cron may be lagging" if _stale else ""
        st.markdown(
            f'<div style="font-family:DM Mono,monospace;font-size:11px;color:{_pao_col};'
            f'margin:-12px 0 18px 2px;">prices as of {_pao}{_pao_note}</div>',
            unsafe_allow_html=True)

    # ── Equity curve (model vs SPY) — folded in from the Track Record view ────
    if _pt:
        _render_track_equity(_pt, positions, day_pct=_chart_day_pct,
                             day_spy_pct=_chart_day_spy)

    # ── Holdings table ────────────────────────────────────────────────────────
    st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;'
                'letter-spacing:.1em;margin-bottom:8px;">▲ ACTIVE POSITIONS</div>',
                unsafe_allow_html=True)

    # ── Detect current hidden gems ────────────────────────────────────────────
    port_gem_tickers = set()
    # First try signal_log.is_hidden_gem (no scan required)
    if sb:
        try:
            tickers_in_port = [h["ticker"] for h in holdings]
            gem_resp = sb.table("signal_log")                 .select("ticker,is_hidden_gem")                 .in_("ticker", tickers_in_port)                 .eq("is_hidden_gem", True)                 .order("signal_date", desc=True)                 .limit(len(tickers_in_port) * 2)                 .execute()
            port_gem_tickers = {row["ticker"] for row in (gem_resp.data or [])}
        except Exception:
            pass
    # Fallback: detect from session scan if available
    if not port_gem_tickers and scan:
        try:
            port_gems = detect_hidden_gems(scan, macro_data=st.session_state.get("macro_data"))
            port_gem_tickers = {g["ticker"] for g in port_gems}
        except Exception:
            pass

    # Render positions as collapsed cards with P&L data
    from model_engine import SECTORS as _MP_SECTORS
    _mp_sorted = sorted(holdings, key=lambda x: x["pnl_pct"], reverse=True)
    # Style the native, fragment-scoped watchlist toggles to match the card ghost
    # links (full-width gold). Targets every st.button keyed mpwl_<ticker>.
    st.markdown(
        "<style>div[class*='st-key-mpwl_'] button{width:100%;"
        "background:rgba(212,168,67,.08);border:1px solid rgba(212,168,67,.3);"
        "border-radius:6px;color:#d4a843;font-family:Syne,sans-serif;font-size:13px;"
        "font-weight:700;letter-spacing:.04em;text-transform:uppercase;"
        "padding:8px;margin:2px 0 14px;}"
        "div[class*='st-key-mpwl_'] button:hover{background:rgba(212,168,67,.16);"
        "border-color:rgba(212,168,67,.5);}</style>",
        unsafe_allow_html=True)
    # Fetch today's price change once for all positions so the collapsed cards
    # can show entry date + intraday % without expanding (cached in session
    # by sorted ticker set; identical to the watchlist pattern).
    _mp_day_change = _fetch_day_change_map(
        [h["ticker"] for h in _mp_sorted] + ["SPY"],
        cache_key="_mp_daychange_cache",
        include_extended=False,
    )
    # Trailing ~20-session vs-SPY mini chart for each holding. Batch-fetch the
    # window once for all holdings (+SPY) in a single call, so a freshly seeded
    # cohort still shows real recent history — a "since entry" window would be
    # empty on day one. Mirrors the screener Top-10 pattern.
    _mp_trail = _trail_start(30)
    _mp_mini_pm, _mp_mini_sm = _mini_price_data(
        tuple(sorted({h["ticker"] for h in _mp_sorted})), _mp_trail)
    @st.fragment
    def _mp_holdings():
        # Isolated in a fragment so the watchlist toggle under each card reruns
        # ONLY this list — no URL navigation, no full-page reload, no main() re-run.
        # Watchlist membership is re-read each run so the label flips immediately.
        _wl_set = {w["ticker"] for w in get_watchlist(uid())} if uid() else set()
        for _mp_i, h in enumerate(_mp_sorted):
            tk    = h["ticker"]
            score = h["current_score"]
            sc    = dict(score_map.get(tk, {}) or {})
            _quant = float(sc.get("composite", score) or score)
            sc["ticker"]        = tk
            sc["adj_composite"] = score
            sc["composite"]     = _quant
            sc["adj_action"]    = "BUY" if score >= 60 else ("SELL" if score < 45 else "HOLD")
            sc["momentum"]      = h["momentum"]
            sc["quality"]       = h["quality"]
            sc["volume"]        = h["volume"]
            sc["value"]         = h["value"]
            sc["sentiment"]     = h["sentiment"]
            sc["price"]         = h["current_price"]
            sc["sector"]        = sc.get("sector") or _MP_SECTORS.get(tk, "")
            sc["signal_date"]   = str(h["entry_date"])[:10]
            sc["score_delta"]   = round(score - _quant, 1)
            _ci_cache_mp = st.session_state.get("company_info_cache", {})
            ci = _ci_cache_mp.get(tk)
            _ep, _cp = h.get('entry_price'), h.get('current_price')
            _pct, _pnl = h.get('pnl_pct', 0), h.get('pnl', 0)
            _edate = str(h.get('entry_date',''))[:10]
            _rc = '#34d399' if _pct >= 0 else '#f87171'
            _sg = '+' if _pct >= 0 else ''
            _ep_str = f'${_ep:,.2f}' if _ep else '—'
            _cp_str = f'${_cp:,.2f}' if _cp else '—'
            _pnl_str = f'{_sg}${abs(_pnl):,.0f}' if _ep and _cp else '—'
            _pct_str = f'{_sg}{_pct:.2f}%' if _ep and _cp else '—'
            _pnl_html = (
                f'<div style="display:grid;grid-template-columns:repeat(4,1fr);gap:6px;'
                f'margin:8px 20px 4px;padding:10px;background:rgba(255,255,255,.02);'
                f'border:1px solid rgba(255,255,255,.05);border-radius:6px;">'
                f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">ENTRY DATE</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;">{_edate}</div></div>'
                f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">ENTRY PRICE</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#b3bed0;">{_ep_str}</div></div>'
                f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">CURRENT</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;">{_cp_str}</div></div>'
                f'<div><div style="font-size:13px;color:#8896ac;letter-spacing:.06em;margin-bottom:3px;">RETURN</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;font-weight:700;color:{_rc};">{_pct_str}</div>'
                f'<div style="font-family:DM Mono,monospace;font-size:13px;color:{_rc};">{_pnl_str}</div></div>'
                f'</div>'
            )
            sc["_summary_meta_html"] = _build_summary_meta_html(
                entry_date=h.get("entry_date"),
                day_change_entry=_mp_day_change.get(tk),
            )
            sc["_mini_chart_html"] = _build_mini_chart_html(
                tk, _mp_trail, _mp_mini_pm, _mp_mini_sm, since_label="vs SPY · 20d")
            st.markdown(
                factor_panel_html(sc, tk in port_gem_tickers, company_info=ci,
                                  wl_btn=_pnl_html, as_details=True),
                unsafe_allow_html=True,
            )
            _in_wl = tk in _wl_set
            if st.button(("\u2715 Remove from Watchlist" if _in_wl else "\u2606 Add to Watchlist"),
                         key=f"mpwl_{tk}", use_container_width=True):
                if _in_wl:
                    remove_from_watchlist(uid(), tk)
                    _evt = "watchlist_removed"
                else:
                    add_to_watchlist(uid(), tk, h.get("current_price"))
                    _evt = "watchlist_added"
                try:
                    analytics.capture(_evt, user=st.session_state.get("user"),
                                      props={"ticker": tk})
                except Exception:
                    pass
                # A fragment rerun does NOT reset _run_cache the way main() does,
                # so bust the watchlist memo here or the re-read below stays stale
                # and the button label won't flip.
                try:
                    st.session_state.get("_run_cache", {}).pop(f"wl:{uid()}", None)
                except Exception:
                    pass
                st.rerun(scope="fragment")
    _mp_holdings()

    # ── Sector concentration + closed positions — folded in from Track Record ─
    if _pt and _pt.get("sector_counts"):
        _top = _pt["sector_counts"][0][1]
        _bars = ""
        for _name, _cnt in _pt["sector_counts"]:
            _pct = _cnt / _top * 100 if _top else 0
            _bars += (f'<div style="margin-bottom:8px;">'
                      f'<div style="display:flex;justify-content:space-between;margin-bottom:3px;font-size:13px;">'
                      f'<span style="color:#b3bed0;">{_name}</span>'
                      f'<span style="font-family:DM Mono,monospace;color:#cbd5e1;">{_cnt}</span></div>'
                      f'<div style="background:rgba(255,255,255,.05);border-radius:3px;height:7px;">'
                      f'<div style="width:{_pct:.0f}%;height:100%;background:#d4a843;border-radius:3px;"></div>'
                      f'</div></div>')
        st.markdown(
            '<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;margin:8px 0 10px;">\u25B2 SECTOR CONCENTRATION</div>'
            f'<div style="background:#0a0b14;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:16px 18px;margin-bottom:20px;">{_bars}</div>',
            unsafe_allow_html=True)

    if _pt:
        st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;letter-spacing:.1em;margin-bottom:10px;">\u25BC CLOSED POSITIONS</div>', unsafe_allow_html=True)
        if _pt.get("exits"):
            _rows = ""
            for _e in _pt["exits"]:
                _rc = "#34d399" if _e["ret"] >= 0 else "#f87171"
                _rows += (
                    f'<div style="display:grid;grid-template-columns:70px 1fr 90px 90px 80px;gap:8px;align-items:center;'
                    f'padding:8px 0;border-bottom:1px solid rgba(255,255,255,.04);font-size:13px;">'
                    f'<span style="font-family:Syne,sans-serif;font-weight:700;color:#e2e8f0;">{_e["ticker"]}</span>'
                    f'<span style="color:#9fabc0;">{_e["sector"]}</span>'
                    f'<span style="font-family:DM Mono,monospace;color:#8896ac;">{_e["entry_date"][5:]}\u2192{_e["exit_date"][5:]}</span>'
                    f'<span style="font-family:DM Mono,monospace;color:{_rc};text-align:right;">{("+" if _e["ret"]>=0 else "")}{_e["ret"]:.1f}%</span>'
                    f'<span style="font-family:DM Mono,monospace;color:#8896ac;text-align:right;font-size:11px;">{_e["reason"]}</span>'
                    f'</div>')
            st.markdown(
                f'<div style="background:#0a0b14;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:8px 18px 12px;margin-bottom:16px;">{_rows}</div>',
                unsafe_allow_html=True)
        else:
            st.markdown('<div style="background:#0a0b14;border:1px solid rgba(255,255,255,.07);border-radius:8px;padding:16px 18px;margin-bottom:16px;font-size:13px;color:#9fabc0;">No positions have exited yet.</div>', unsafe_allow_html=True)
        st.markdown(
            '<div style="font-size:11px;color:#8896ac;line-height:1.7;margin-bottom:8px;">'
            f'Live since {_pt["inception"]}. $100K notional, equal-weighted $2,000 per position, marked daily to '
            'close prices; realized P&amp;L from exits and unrealized P&amp;L on holdings are both included. SPY benchmark '
            'is $100,000 invested at inception. Hypothetical illustration \u2014 ignores slippage, taxes, and commissions. '
            'Past performance does not guarantee future results.</div>',
            unsafe_allow_html=True)

    # Spacer below the iframe so when the last card expands and the iframe
    # grows, the parent page has scroll room to reveal the new content.
    st.markdown('<div style="height:160px;"></div>', unsafe_allow_html=True)

    st.markdown(
        '<div style="font-size:13px;color:#8896ac;padding:6px 8px;background:#050a0f;'
        'border:1px solid rgba(255,255,255,.07);border-radius:0 0 6px 6px;margin-bottom:8px;">'
        '$2,000/position · Equal weighted · Auto-exit score < 45</div>',
        unsafe_allow_html=True)

    # ── Export to Excel ───────────────────────────────────────────────────────
    try:
        export_rows = []
        for h in sorted(holdings, key=lambda x: x["pnl_pct"], reverse=True):
            shares = (h["pos_size"] / h["entry_price"]) if h.get("entry_price") and h["entry_price"] > 0 else ""
            export_rows.append({
                "Ticker":        h["ticker"],
                "Entry Date":    h["entry_date"],
                "Entry Price":   h.get("entry_price", ""),
                "Current Price": h.get("current_price", ""),
                "Shares":        round(shares, 4) if shares else "",
                "Position ($)":  h["pos_size"],
                "Current Value": round(h["current_val"], 2),
                "P&L ($)":       round(h["pnl"], 2),
                "Return (%)":    round(h["pnl_pct"], 2),
                "Score":         round(h["current_score"], 1),
                "Momentum":      round(h["momentum"], 1),
                "Quality":       round(h["quality"], 1),
                "Volume":        round(h["volume"], 1),
                "Value":         round(h["value"], 1),
                "Sentiment":     round(h["sentiment"], 1),
                "Gem":           "💎" if h.get("is_gem") else "",
            })
        headers = ["Ticker","Entry Date","Entry Price","Current Price","Shares","Position ($)","Current Value","P&L ($)","Return (%)","Score","Momentum","Quality","Volume","Value","Sentiment","Gem"]
        xl = _make_excel(export_rows, headers, "Model Portfolio")
        st.download_button(
            label="⬇ Export to Excel",
            data=xl,
            file_name="qntm_model_portfolio.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="model_port_export"
        )
    except Exception:
        pass



    # ── Exit history ──────────────────────────────────────────────────────────
    if sb:
        try:
            exits = sb.table("model_portfolio_positions") \
                .select("ticker,entry_date,entry_price,exit_date,exit_price,exit_score,exit_reason") \
                .eq("is_active", False) \
                .eq("epoch", MODEL_EPOCH) \
                .order("exit_date", desc=True) \
                .limit(20) \
                .execute()
            # Filter out reseeded entries — only show genuine exits
            real_exits = [e for e in (exits.data or [])
                          if e.get("exit_reason","") not in ("reseeded","")]
            if real_exits:
                st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
                st.markdown('<div style="font-family:DM Mono,monospace;font-size:13px;color:#9fabc0;'
                            'letter-spacing:.1em;margin-bottom:12px;">RECENT EXITS</div>',
                            unsafe_allow_html=True)
                for ex in real_exits:
                    ep = ex.get("entry_price")
                    xp = ex.get("exit_price")
                    if ep and xp and ep > 0:
                        ret = (xp / ep - 1) * 100
                        rc  = "#34d399" if ret >= 0 else "#f87171"
                        sg  = "+" if ret >= 0 else ""
                        ret_str = f'{sg}{ret:.1f}%'
                    else:
                        rc = "#9fabc0"
                        ret_str = "—"
                    st.markdown(
                        f'<div style="display:flex;gap:16px;padding:6px 12px;'
                        f'border-bottom:1px solid rgba(255,255,255,.04);font-size:13px;">'
                        f'<span style="font-family:Syne,sans-serif;font-weight:800;color:#b3bed0;width:60px;">{ex["ticker"]}</span>'
                        f'<span style="color:#8896ac;">{ex.get("exit_date","")} · {ex.get("exit_reason","")}</span>'
                        f'<span style="font-family:DM Mono,monospace;color:{rc};margin-left:auto;">{ret_str}</span>'
                        f'</div>', unsafe_allow_html=True)
        except Exception:
            pass

    st.markdown(
        '<div style="font-size:13px;color:#8896ac;padding:16px 0;margin-top:16px;'
        'border-top:1px solid rgba(255,255,255,.05);">'
        '⚠ Model portfolio is hypothetical. $2,000 equal weight per position. '
        'Does not account for slippage, taxes, or transaction costs. For informational purposes only.</div>',
        unsafe_allow_html=True)


def page_methodology():
    _pin_nav("methodology")
    """How QNTM Works — methodology, factor logic, disclaimers."""
    page_summary("📖", "How QNTM Works",
        "Transparent methodology — what the model does, how it scores stocks, and what it doesn't do.")
    st.markdown('<div style="padding:0 32px;">', unsafe_allow_html=True)

    sections = [
        ("Getting Started — Where to Begin", "#34d399",
         "New here? This is the path most users follow:\n\n"
         "1. Screener — your home base. Every stock in the universe, ranked by conviction. Start at the top (High Conviction) and work down.\n"
         "2. Open a stock — tap any row to see its plain-English rationale and the five pillar scores behind the signal.\n"
         "3. Watchlist — star the names you want to follow; they get a tracked view marked against the S&P 500.\n"
         "4. Hidden Gems — strong-scoring names that fly under Wall Street's radar.\n"
         "5. Portfolio Simulator — test a hypothetical allocation and see how the model scores it.\n"
         "6. Portfolio & Track Record — the model's live, rules-based portfolio and its performance since inception.\n"
         "7. Alerts — get notified when a stock's conviction changes (Pro).\n\n"
         "Throughout, the macro overlay at the top of the Screener tells you what regime the market is in "
         "and how it's shaping the scores. Everything below explains how those scores are built."),

        ("The Universe", "#34d399",
         f"QNTM covers {_universe_n()} US stocks: the Russell 1000 (large- and mid-cap) plus a layer "
         "of the largest Russell 2000 small-caps, cleaned of delisted and illiquid tickers. The "
         "Russell 1000 core anchors the Screener and the model portfolio; the small-cap layer feeds "
         "the Hidden Gems screen with genuinely under-followed names. Scores refresh nightly, with "
         "live price updates and hourly full re-scores during market hours.\n\n"
         f"Universe expansion (June 21, 2026): the investable universe was widened from ~830 names to the "
         f"current ~{_universe_n()}. This is a forward methodology change — the model portfolio and its "
         "track record are not restated retroactively. The wider opportunity set applies from the "
         "expansion date forward, and the portfolio migrates into it as positions turn over."),

        ("The Factor Model", "#34d399",
         "Each stock receives a composite score (0–100) built from five weighted pillars:\n\n"
         "• Momentum (30%) — price trend, RSI, MACD, 52-week proximity, rate of change\n"
         "• Quality (25%) — ROE, profit margin, revenue growth, EPS beat rate, FCF yield\n"
         "• Volume (20%) — relative volume, OBV, Chaikin Money Flow, accumulation/distribution\n"
         "• Value (15%) — forward P/E, PEG ratio, EV/EBITDA, Price-to-Sales\n"
         "• Sentiment (10%) — short interest, insider buy ratio, institutional ownership\n\n"
         "Scores are cross-sectional — a score of 75 means the stock ranks stronger than 75% of the universe."),

        ("Conviction Signals", "#34d399",
         "• High Conviction (score ≥ 60) — model sees strong multi-factor alignment. Top 40% of universe.\n"
         "• Moderate Conviction (45–59) — mixed factor signals, neither strong nor deteriorating.\n"
         "• Low Conviction (score < 45) — weakest factor profile in the universe. Elevated model risk.\n\n"
         "Signals are quantitative rankings, not buy/sell/hold recommendations. "
         "Signals update nightly. In HIGH VOLATILITY regimes, conviction thresholds tighten — "
         "only scores ≥ 67 surface as High Conviction."),

        ("Macro Overlay", "#d4a843",
         "A live macro regime overlay adjusts composite scores based on current market conditions:\n\n"
         "• VIX level — real-time fear gauge via yfinance (updates every 15 minutes)\n"
         "• WTI crude price — oil spike detection via CL=F futures\n"
         "• News sentiment — 70+ headlines scanned from Yahoo Finance RSS and FRED\n"
         "• Active events — war escalation, tariff regimes, Fed policy, oil spikes\n\n"
         "Weighting: 75% quant model / 25% macro overlay. In RISK OFF / HIGH VOLATILITY regimes, "
         "macro dampening reduces adj_composite scores to reflect elevated systemic risk. "
         "Regime updates every 15 minutes during your session."),

        ("Hidden Gems", "#34d399",
         "Hidden Gems are mid- and small-cap stocks scoring above conviction threshold that fly under Wall Street's radar. "
         "Detection criteria:\n\n"
         "• Not a mega-cap (excludes NVDA, AAPL, MSFT, etc.)\n"
         "• adj_composite ≥ 65 (67+ in Risk-Off regimes)\n"
         "• Momentum ≥ 58, Quality ≥ 55\n"
         "• At least one fundamental reason: revenue acceleration, earnings beats, low short interest, insider buying\n\n"
         "Gems are identified fresh each scan — the list changes as fundamentals and scores shift."),

        ("Performance & Track Record", "#d4a843",
         "QNTM does not currently publish a historical backtest. A backtest is only credible when every score "
         "is computed from data that was actually available at the time — point-in-time fundamentals, the "
         "universe as it existed then, and macro conditions as they were known. We're building that properly "
         "rather than publishing numbers that can't withstand scrutiny.\n\n"
         "The track record we show is the live Model Portfolio: rules-based entries and exits on the model's "
         "signals, tracked daily from inception forward. A live record is short by nature. "
         "Past model performance does not guarantee future results."),

        ("Scores & Alerts", "#d4a843",
         "• Nightly refresh — full universe rescored each night via automated cron\n"
         "• Daily signals — conviction scores are close-to-close; the macro overlay is the only intraday-moving input\n"
         "• Signal alerts — Pro users receive notifications when watchlist stocks change conviction tier\n"
         "• Macro regime — refreshed every 15 minutes from live VIX, WTI, and news feeds\n"
         "• Platform stats — gem count, high/low conviction counts updated after each refresh"),

        ("What QNTM Does NOT Do", "#f87171",
         "• QNTM does not provide personalized investment advice\n"
         "• QNTM does not account for your individual tax situation, risk tolerance, or financial goals\n"
         "• QNTM does not predict short-term price movements or guarantee future results\n"
         "• QNTM is not a registered investment adviser under the Investment Advisers Act of 1940\n"
         "• Conviction scores are quantitative model outputs — not buy or sell recommendations\n"
         "• Prices shown are indicative snapshots — not real-time execution prices\n\n"
         "Always consult a qualified financial adviser before making investment decisions."),

        ("Billing & Cancellation", "#d4a843",
         "• Pro includes a 7-day free trial — you are not charged during the trial.\n"
         "• After the trial, QNTM Pro automatically renews at $29.00/month until you cancel.\n"
         "• You are charged $29.00 on the same date each month.\n"
         "• Cancel anytime in Account Settings → Subscription with a single click. "
         "Cancelling stops your next charge immediately; you keep Pro access until the end "
         "of your current paid period, then your account converts to Free.\n"
         "• Founding Members pay $0 and have no auto-renewal.\n\n"
         "Full terms are in the Billing & Refund Policy and Terms of Service."),
    ]

    for title, color, body in sections:
        st.markdown(
            f'<div style="border-left:3px solid {color};padding:16px 20px;margin-bottom:16px;"'
            f'background:rgba(255,255,255,.02);border-radius:0 8px 8px 0;">'
            f'<div style="font-family:Syne,sans-serif;font-size:14px;font-weight:700;color:{color};"'
            f'letter-spacing:.06em;margin-bottom:8px;">{title}</div>'
            f'<div style="font-size:13px;color:#b3bed0;line-height:1.8;white-space:pre-line;">{body}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown('</div>', unsafe_allow_html=True)


def _render_whats_new():
    """Login 'What's new' banner. Shows changelog entries newer than the user's
    stored marker. Dismisses when the user presses 'Got it' OR navigates to a
    different tab; either way it records the marker so it won't reappear next
    login. Fragment-scoped so dismissing doesn't reload. Never raises."""
    user = st.session_state.get("user") or {}
    if not user.get("id") or user.get("id") == "demo":
        return
    if st.session_state.get("_wn_dismissed"):
        return
    try:
        import whats_new as _wn
        _seen = (user.get("notifications") or {}).get("whatsnew_seen") or ""
        items = _wn.unseen_entries(_seen)
    except Exception:
        return
    if not items:
        return

    def _mark_seen():
        try:
            _newest = _wn.latest_id()
            _np = dict(st.session_state.user.get("notifications") or {})
            _np["whatsnew_seen"] = _newest
            update_preferences(uid(), {"notifications": _np})
            st.session_state.user["notifications"] = _np
        except Exception:
            pass
        st.session_state["_wn_dismissed"] = True

    # Auto-dismiss on navigation: if the active tab differs from the one the
    # banner first appeared on, treat it as seen and don't render.
    _cur_nav = st.session_state.get("nav", "")
    if "_wn_shown_on" not in st.session_state:
        st.session_state["_wn_shown_on"] = _cur_nav
    elif st.session_state["_wn_shown_on"] != _cur_nav:
        _mark_seen()
        return

    @st.fragment
    def _wn_banner():
        if st.session_state.get("_wn_dismissed"):
            return
        _tagc = {"new": "#34d399", "improved": "#d4a843"}
        _rows = ""
        for e in items:
            _c = _tagc.get(e.get("tag", "new"), "#34d399")
            _rows += (
                f'<div style="padding:10px 0;border-top:1px solid rgba(255,255,255,.06);">'
                f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:3px;flex-wrap:wrap;">'
                f'<span style="font-size:10px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;'
                f'color:{_c};border:1px solid {_c};border-radius:4px;padding:1px 6px;">{e.get("tag","new")}</span>'
                f'<span style="font-family:Syne,sans-serif;font-size:14px;font-weight:700;color:#e2e8f0;">{e.get("title","")}</span>'
                f'<span style="font-size:11px;color:#8896ac;margin-left:auto;">{e.get("date","")}</span>'
                f'</div>'
                f'<div style="font-size:13px;color:#9fabc0;line-height:1.6;">{e.get("body","")}</div>'
                f'</div>'
            )
        st.markdown(
            f'<div style="background:linear-gradient(180deg,rgba(212,168,67,.06),rgba(13,17,23,0));'
            f'border:1px solid rgba(212,168,67,.25);border-radius:12px;padding:14px 18px 6px;margin:0 0 14px;">'
            f'<div style="display:flex;align-items:center;gap:8px;margin-bottom:2px;">'
            f'<span style="font-size:16px;">\u2728</span>'
            f'<span style="font-family:Syne,sans-serif;font-size:15px;font-weight:800;color:#d4a843;'
            f'letter-spacing:.02em;">What\'s new since you were last here</span></div>'
            f'{_rows}</div>',
            unsafe_allow_html=True)
        if st.button("Got it", key="_wn_got"):
            _mark_seen()
            st.rerun(scope="fragment")
    _wn_banner()


def page_platform():
    # ── Accordion behavior for cards (one open at a time) ──────────────────────
    # Cards render as <details name="qntm-cards">; modern browsers make same-named
    # details mutually exclusive natively. This JS is a fallback for browsers that
    # don't yet support the `name` attribute — it closes sibling cards on open.
    qntm_html("""
    <script>
    (function(){
      var pd = parent.document;
      function bind(){
        pd.querySelectorAll('details.qntm-card-details').forEach(function(d){
          if(d._accBound) return;
          d._accBound = true;
          d.addEventListener('toggle', function(){
            if(d.open){
              pd.querySelectorAll('details.qntm-card-details').forEach(function(o){
                if(o!==d && o.open) o.open=false;
              });
            }
          });
        });
      }
      bind();
      var mo = new MutationObserver(bind);
      mo.observe(pd.body, {childList:true, subtree:true});
    })();
    </script>
    """, height=0)

    # ── Force MFA setup on first login ─────────────────────────────────────────
    if st.session_state.get("force_mfa_setup"):
        user = st.session_state.user or {}
        mfa  = get_user_mfa(uid())
        if not mfa.get("mfa_enabled"):
            # Show as a clean centered page — no fixed overlays that cover buttons
            _, mc, _ = st.columns([1, 2, 1])
            with mc:
                st.markdown(
                    '<div style="background:#0d1117;border:1px solid rgba(212,168,67,.4);border-radius:12px;padding:28px 24px;text-align:center;">'
                    '<div style="font-size:28px;margin-bottom:12px;">🔒</div>'
                    '<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:700;color:#d4a843;margin-bottom:12px;">Secure Your Account</div>'
                    '<div style="font-size:13px;color:#b3bed0;line-height:1.7;">'
                    'QNTM holds your portfolio data. We <strong style="color:#e2e8f0;">strongly recommend</strong> enabling 2FA before continuing. Takes 60 seconds.'
                    '</div></div>',
                    unsafe_allow_html=True
                )
                st.markdown("<div style='height:4px'></div>", unsafe_allow_html=True)
                if st.button("⚡ Enable 2FA", key="force_mfa_yes", use_container_width=True):
                    st.session_state.force_mfa_setup = False
                    nav("account")
                    st.session_state.show_mfa_setup = True
                st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)
                if st.button("Skip", key="force_mfa_skip", use_container_width=True):
                    st.session_state.force_mfa_setup = False
                    st.rerun()
            return
        else:
            st.session_state.force_mfa_setup = False

            st.session_state.force_mfa_setup = False

    if "last_refresh" not in st.session_state:
        st.session_state.last_refresh = 0
    import time as _time
    now = int(_time.time())
    # Never clear scan_results while a live refresh is in progress
    if not st.session_state.get("live_refresh_running"):
        if now - st.session_state.last_refresh >= 60:
            st.session_state.last_refresh = now
            # Only clear scan on screener page — other pages don't need it and clearing
            # causes nav drops when buttons trigger reruns
            if st.session_state.get("nav") == "screener":
                st.session_state.scan_results = None
    platform_nav()
    show_onboarding()
    _render_whats_new()


    nav_map = {
        "screener":        page_screener,
        "watchlist":       page_watchlist,
        "gems":            page_gems,
        "backtest":       page_backtest,
        "portfolio":      page_portfolio,
        "simulator":      page_simulator,
        "model_portfolio": page_model_portfolio,
        "methodology":     page_methodology,
        "alerts":         page_alerts,
        "account":        page_account,
    }
    # Persist nav in URL so WebSocket reconnects (mobile blur) can restore it
    _cur_nav = st.session_state.get("nav", "screener")
    st.query_params["_n"] = _cur_nav
    # One-time trial-started confirmation — shows on whatever page Stripe returns
    # the user to (screener for the main upgrade flow, account for the supporter
    # flow). Themed banner instead of the unthemeable white toast.
    if st.session_state.pop("_trial_just_started", False):
        st.markdown("""
        <div style="background:rgba(212,168,67,.08);border:1px solid rgba(212,168,67,.35);
             border-radius:8px;padding:16px 20px;margin:8px 0 18px;">
          <div style="font-family:'Syne',sans-serif;font-size:15px;font-weight:800;
               color:#d4a843;margin-bottom:4px;">Your 7-day free trial has started 🎉</div>
          <div style="font-size:13px;color:#b3bed0;line-height:1.6;">
            You're now on QNTM Pro. You won't be charged during the trial —
            manage or cancel anytime in Account → Plan &amp; Billing.
          </div>
        </div>
        """, unsafe_allow_html=True)
    # ── "What's New" popup — fire once per session for logged-in users who have
    # changelog entries newer than their last acknowledgement. Purely cosmetic,
    # so any failure is swallowed rather than breaking the platform render.
    if not st.session_state.get("_changelog_checked"):
        st.session_state["_changelog_checked"] = True
        try:
            from changelog import maybe_show_whats_new
            maybe_show_whats_new()
        except Exception:
            pass

    # ── Analytics: one-per-navigation view events + admin-only dashboard ──────
    _prev_nav = st.session_state.get("_last_nav")
    if _cur_nav != _prev_nav:
        st.session_state["_last_nav"] = _cur_nav
        _view_evt = {"gems": "hidden_gems_viewed",
                     "simulator": "simulator_viewed"}.get(_cur_nav)
        if _view_evt:
            try:
                analytics.capture(_view_evt, user=st.session_state.get("user"))
            except Exception:
                pass
    if _cur_nav == "analytics":
        try:
            if analytics.is_admin():
                analytics.render_analytics_dashboard()
            else:
                st.session_state.nav = "screener"
                st.rerun()
        except Exception:
            page_screener()
    else:
        nav_map.get(_cur_nav, page_screener)()

    # ── One-at-a-time card collapse script ──────────────────────────────────
    st.markdown("""
    <script>
    (function() {
      function closeOthers(checkedEl) {
        // Uncheck all other qcard checkboxes when one is checked
        var allBoxes = document.querySelectorAll('input[id^="c"]');
        allBoxes.forEach(function(cb) {
          if (cb !== checkedEl && cb.type === 'checkbox' && cb.checked) {
            cb.checked = false;
          }
        });
      }
      function attachListeners() {
        var allBoxes = document.querySelectorAll('input[id^="c"]');
        allBoxes.forEach(function(cb) {
          if (cb.type === 'checkbox' && !cb._qntmBound) {
            cb._qntmBound = true;
            cb.addEventListener('change', function() {
              if (cb.checked) closeOthers(cb);
            });
          }
        });
      }
      attachListeners();
      var obs = new MutationObserver(attachListeners);
      obs.observe(document.body, { childList: true, subtree: true });
    })();
    </script>
    """, unsafe_allow_html=True)

    # ── Persistent disclaimer footer ────────────────────────────────────
    st.markdown(
        '<div style="margin:32px 32px 8px;padding:12px 16px;'
        'background:rgba(255,255,255,.02);border-top:1px solid rgba(255,255,255,.05);'
        'border-radius:6px;font-size:13px;color:#94a3b8;line-height:1.6;text-align:center;">'
        'QNTM provides quantitative signal analysis for informational and educational purposes only. '
        'Conviction scores are model outputs — not personalized investment advice. '
        'Past model performance does not guarantee future results. '
        'Not a registered investment adviser.'
        '</div>',
        unsafe_allow_html=True
    )

    # ── Conflicts of interest footer (Part 2B — verbatim) ───────────────
    st.markdown(
        '<div style="margin:0 32px 8px;padding:8px 16px;font-size:13px;color:#94a3b8;'
        'line-height:1.6;text-align:center;">'
        'QNTM LLC holds no securities. Its principals may personally hold securities the model '
        'scores and trade only on published signals. '
        '<a href="?legal=disclaimer" target="_self" style="color:#8896ac;text-decoration:underline;">'
        'See our Conflicts of Interest disclosure.</a>'
        '</div>',
        unsafe_allow_html=True
    )

    # Platform footer
    st.markdown("""
    <div style="padding:16px 32px;border-top:1px solid rgba(255,255,255,.05);margin-top:20px;">
      <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px;">
        <div style="font-size:13px;color:#8896ac;">
          QNTM · Quantitative research platform · Not investment advice
        </div>
        <div style="font-size:13px;color:#8896ac;">
          <a href="#" style="color:#b3bed0;">Privacy</a> ·
          <a href="#" style="color:#b3bed0;">Terms</a> ·
          <a href="#" style="color:#b3bed0;">Disclaimer</a>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)
    cookie_banner()


# ══════════════════════════════════════════════════════════════════════════════
# ROUTER
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=120, show_spinner=False)
def _founding_spots_remaining() -> int:
    """Founding-member spots left in the first-50 window. While > 0, upgrades are
    a free Pro grant (no card, no Stripe); once exhausted, the paid trial kicks
    in. Fails OPEN (returns spots) so a transient DB hiccup never wrongly charges
    a founder — matches the landing-page counter's behaviour."""
    try:
        from data_refresh import _get_supabase as _fs_sb
        sb = _fs_sb()
        if sb:
            r = sb.table("users").select("id", count="exact").execute()
            n = r.count if getattr(r, "count", None) else len(r.data or [])
            return max(0, 50 - int(n))
    except Exception:
        pass
    return 50

def page_upgrade():
    """Upgrade to Pro page — handles upgrade flow, Stripe when ready."""
    # Keep this as a top-level page (route == "upgrade") across reruns. Do NOT use
    # _pin_nav here: that sets page="platform" + nav="upgrade", but "upgrade" isn't
    # a platform sub-page, so the next rerun (e.g. checking the consent box) would
    # dispatch to page_platform and fall through to the screener.
    st.session_state.page = "upgrade"
    try:
        if not st.session_state.get("_pricing_evt_sent"):
            st.session_state["_pricing_evt_sent"] = True
            analytics.capture("pricing_viewed", user=st.session_state.get("user"))
    except Exception:
        pass
    # Clear any stale checkout URL/error from a previous visit, once per page entry
    # (guarded so we don't wipe a URL created during this visit's reruns).
    if st.session_state.get("_checkout_page") != "upgrade":
        st.session_state["_checkout_page"] = "upgrade"
        st.session_state.pop("_checkout_url", None)
        st.session_state.pop("_checkout_err", None)
    feature    = st.session_state.get("upgrade_feature", "Pro")
    return_nav = st.session_state.get("upgrade_return_nav", "screener")

    # Already pro — redirect back
    if is_pro():
        st.session_state.nav  = return_nav
        st.session_state.page = "platform"
        st.rerun()
        return

    _uid_val  = (st.session_state.user or {}).get("id", "")
    _plan_val = (st.session_state.user or {}).get("plan", "free")
    _back_url    = f"?qnav={return_nav}&uid={_uid_val}&plan={_plan_val}&ck=1&_n={return_nav}"
    _confirm_url = (
        f"?qnav={return_nav}&uid={_uid_val}&plan={_plan_val}"
        f"&ck=1&upgrade=pro&_n={return_nav}"
    )

    st.markdown(_back_btn(_back_url), unsafe_allow_html=True)

    st.markdown(f"""
    <div style="max-width:480px;margin:40px auto;padding:0 16px;text-align:center;">
      <div style="font-size:48px;margin-bottom:12px;">⚡</div>
      <div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;
           color:#d4a843;margin-bottom:8px;">Upgrade to Pro</div>
      <div style="font-size:14px;color:#b3bed0;margin-bottom:4px;">
        Unlocking: <strong style="color:#e2e8f0;">{feature}</strong>
      </div>
      <div style="background:rgba(212,168,67,.06);border:1px solid rgba(212,168,67,.25);
           border-radius:10px;padding:20px;margin:20px 0;">
        <div style="font-family:DM Mono,monospace;font-size:13px;color:#d4a843;
             letter-spacing:.1em;margin-bottom:6px;">FOUNDING MEMBER · FIRST 50 SPOTS</div>
        <div style="font-family:Syne,sans-serif;font-size:36px;font-weight:800;color:#d4a843;line-height:1;">$0</div>
        <div style="font-size:13px;color:#9fabc0;margin-top:4px;">free now · $29/mo after launch</div>
      </div>
      <div style="font-size:13px;color:#8896ac;margin-bottom:20px;line-height:1.6;">
        Hidden Gems · Simulator · Alerts · Unlimited holdings · Full {_universe_n()}-stock universe
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── ARL paid-trial mode ───────────────────────────────────────────────────
    # Activates the full ARL checkout (notice + consent + log + ack email) and
    # the Stripe Checkout redirect. Auto-on when Stripe billing is configured
    # (keys + price ID present); can also be forced via session for testing.
    try:
        import stripe_billing as _sb_cfg
        # Founders (first-50 window) get a free Pro grant — no card, no Stripe.
        # Stripe's paid 7-day trial only takes over once founding spots are gone.
        # _paid_trial_mode stays an explicit test override that ignores the window.
        _force_paid = bool(st.session_state.get("_paid_trial_mode", False))
        _founding_open = _founding_spots_remaining() > 0
        _paid_trial = _force_paid or (_sb_cfg.billing_configured() and not _founding_open)
    except Exception:
        _paid_trial = bool(st.session_state.get("_paid_trial_mode", False))

    if _paid_trial:
        import arl as _arl
        _acct_url = f"?qnav=account&uid={_uid_val}&plan={_plan_val}&ck=1&_n=account"
        # 1A — six-element initial notice, ON the page, before the button.
        st.markdown(
            f'<div style="max-width:480px;margin:0 auto;padding:0 16px;">'
            f'{_arl.initial_notice_html(_acct_url)}</div>',
            unsafe_allow_html=True,
        )
        # 1B — separate affirmative-consent checkbox (unchecked default), gates button.
        _, _cb_col, _ = st.columns([1, 10, 1])
        with _cb_col:
            _arl_consent = st.checkbox(_arl.CHECKBOX_TEXT, value=False, key="arl_consent_cb")
        if _arl_consent:
            import stripe_billing as _sb_pay
            # Once a checkout URL exists, show ONLY the link button (hides the
            # st.button so it can't render in its post-click white state).
            if st.session_state.get("_checkout_url"):
                st.link_button("Continue to secure checkout →",
                               st.session_state["_checkout_url"],
                               use_container_width=True, type="primary")
                st.caption("Opens Stripe's secure checkout in a new tab.")
            elif st.session_state.get("_checkout_err"):
                st.error(f"Could not start checkout: {st.session_state['_checkout_err']}  ·  Contact hello@qntm.live")
            elif not _sb_pay.billing_configured():
                # No Stripe configured — direct upgrade (dev/test only)
                if st.button("Start free trial", key="arl_start_trial", use_container_width=True):
                    _arl.log_consent(_uid_val, plan="pro")
                    _email = (st.session_state.user or {}).get("email")
                    if _email:
                        _arl.send_acknowledgment(_uid_val, _email)
                    ok = upgrade_plan(uid(), "pro")
                    if ok and st.session_state.get("user"):
                        st.session_state.user["plan"] = "pro"
                    st.session_state.nav  = return_nav
                    st.session_state.page = "platform"
                    st.rerun()
            else:
                if st.button("Start free trial", key="arl_start_trial", use_container_width=True):
                    _ip = None
                    try:
                        _ip = st.context.headers.get("X-Forwarded-For")
                    except Exception:
                        _ip = None
                    _arl.log_consent(_uid_val, plan="pro", ip_address=_ip)
                    from db import get_stripe_billing as _gsb
                    _existing = _gsb(_uid_val).get("stripe_customer_id")
                    _base = "https://qntm.live"
                    try:
                        _base = "https://" + (st.context.headers.get("Host") or "qntm.live")
                    except Exception:
                        pass
                    _email = (st.session_state.user or {}).get("email", "")
                    _url = _sb_pay.create_checkout_url(_uid_val, _email, _base, _existing)
                    if _url:
                        st.session_state["_checkout_url"] = _url
                        st.session_state["_awaiting_checkout"] = 3
                        st.session_state["_stripe_polled"] = False
                    else:
                        st.session_state["_checkout_err"] = _sb_pay.last_error()
                    st.rerun()
        else:
            st.markdown(
                '<div style="max-width:480px;margin:8px auto 0;padding:0 16px;'
                'text-align:center;opacity:.5;pointer-events:none;">'
                '<div style="background:rgba(212,168,67,.15);border-radius:8px;'
                'padding:12px;font-family:Syne,sans-serif;font-weight:800;color:#0a0b14;'
                'background:linear-gradient(135deg,#d4a843,#b8922e);">Start free trial</div>'
                '<div style="font-size:13px;color:#9fabc0;margin-top:8px;">'
                'Check the box above to continue.</div></div>',
                unsafe_allow_html=True,
            )
    else:
        # Founding $0 flow — no auto-renewal, no ARL notice/consent required.
        st.markdown(_cta_gold("✓ Claim Founding Member Access", _confirm_url), unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;margin-top:12px;font-size:13px;color:#94a3b8;line-height:1.6;">
      Quantitative research tool — not investment advice.<br>
      Past model performance does not guarantee future results.
    </div>
    """, unsafe_allow_html=True)


def page_verify_email():
    """Email-confirmation landing page. Clicking the emailed link hits this with
    ?verify_token=…; we consume it on load (one-shot) and confirm the address."""
    token = st.query_params.get("verify_token", "")
    res = consume_verify_token(token) if token else {"success": False, "error": "Missing token"}
    # If the just-verified user is the one in this session, clear any stale flag
    if res.get("success"):
        u = st.session_state.get("user")
        if u and u.get("id") == res.get("user_id"):
            u["email_verified"] = True
        # Auto sign-in and drop them straight onto the screener — unless the
        # account has MFA enabled, in which case route through normal sign-in so
        # the second factor is still enforced.
        if not st.session_state.get("logged_in") and res.get("user_id"):
            try:
                _mfa = get_user_mfa(res["user_id"])
            except Exception:
                _mfa = {}
            if not (_mfa.get("mfa_enabled") and _mfa.get("totp_secret")):
                try:
                    _vu = get_user_by_id(res["user_id"])
                    if _vu:
                        _vu["email_verified"] = True
                        st.session_state.logged_in       = True
                        st.session_state.user            = _vu
                        st.session_state.mfa_verified    = True
                        st.session_state.signed_out      = False
                        st.session_state.onboarding_done = True
                        st.session_state.page            = "platform"
                        st.session_state.nav             = "screener"
                        try:
                            _write_localstorage_token(_vu["id"], _vu.get("plan", "free"))
                        except Exception:
                            pass
                        st.query_params.clear()
                        st.query_params["uid"]  = _sign_token(_vu["id"], _vu.get("plan", "free"))
                        st.query_params["plan"] = _vu.get("plan", "free")
                        st.query_params["qnav"] = "screener"
                        st.rerun()
                except Exception:
                    pass
    st.markdown('<div style="max-width:420px;margin:56px auto 0;padding:0 24px;">', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;letter-spacing:.04em;'
        'color:#e2e8f0;margin-bottom:8px;">Q<span style="color:#34d399;">NTM</span></div>'
        '<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:700;color:#e2e8f0;'
        'margin-bottom:14px;">Email confirmation</div>',
        unsafe_allow_html=True,
    )
    if res.get("success"):
        st.success("✓ Your email is confirmed. You're all set.")
        dest = "?nav=landing" if st.session_state.get("logged_in") else "?nav=signin"
        label = "→ Back to QNTM" if st.session_state.get("logged_in") else "→ Go to sign in"
        st.markdown(f'<a href="{dest}" target="_self" style="color:#34d399;font-weight:700;'
                    f'text-decoration:none;">{label}</a>', unsafe_allow_html=True)
    else:
        st.error(f"{res.get('error','This link is invalid or has expired.')} "
                 "You can request a fresh link from the banner inside the app.")
        st.markdown('<a href="?nav=signin" target="_self" style="color:#34d399;font-weight:700;'
                    'text-decoration:none;">← Back to sign in</a>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


def page_reset_password():
    from db import peek_auth_token, consume_auth_token, set_password
    token = st.query_params.get("reset_token", "")
    st.markdown('<div style="max-width:420px;margin:56px auto 0;padding:0 24px;">', unsafe_allow_html=True)
    st.markdown(
        '<div style="font-family:Syne,sans-serif;font-size:26px;font-weight:800;letter-spacing:.04em;'
        'color:#e2e8f0;margin-bottom:8px;">Q<span style="color:#34d399;">NTM</span></div>'
        '<div style="font-family:Syne,sans-serif;font-size:18px;font-weight:700;color:#e2e8f0;'
        'margin-bottom:14px;">Reset your password</div>',
        unsafe_allow_html=True,
    )
    if not token or not peek_auth_token(token, "reset"):
        st.error("This reset link is invalid or has expired. Request a new one from the sign-in page.")
        st.markdown('<a href="?nav=signin" target="_self" style="color:#34d399;font-weight:700;'
                    'text-decoration:none;">← Back to sign in</a>', unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)
        return
    np1 = st.text_input("New password", type="password", key="rp_p1",
                        placeholder="At least 8 characters")
    np2 = st.text_input("Confirm new password", type="password", key="rp_p2")
    st.markdown('<div style="height:8px;"></div>', unsafe_allow_html=True)
    if st.button("Set new password", key="rp_btn", use_container_width=True):
        if not np1 or len(np1) < 8:
            st.error("Password must be at least 8 characters")
        elif np1 != np2:
            st.error("Passwords don't match")
        else:
            uid_ = consume_auth_token(token, "reset")
            if not uid_:
                st.error("This reset link is invalid or has expired. Request a new one.")
            else:
                r = set_password(uid_, np1)
                if r.get("success"):
                    st.success("Password updated. You can now sign in with your new password.")
                    st.markdown('<a href="?nav=signin" target="_self" style="color:#34d399;'
                                'font-weight:700;text-decoration:none;">→ Go to sign in</a>',
                                unsafe_allow_html=True)
                else:
                    st.error(r.get("error", "Couldn't update password"))
    st.markdown('</div>', unsafe_allow_html=True)


# ── PERF INSTRUMENTATION (opt-in via ?debug=timing; invisible to normal users) ─
import contextlib as _contextlib

def _perf_on() -> bool:
    """True only when the session was opened with ?debug=timing. Sticky per
    session (cached) so it survives the app's query-param navigation."""
    try:
        v = st.session_state.get("_perf_on")
        if v is None:
            v = (st.query_params.get("debug") == "timing")
            st.session_state["_perf_on"] = bool(v)
        return bool(st.session_state.get("_perf_on"))
    except Exception:
        return False

@_contextlib.contextmanager
def timed(label: str):
    """Log '[PERF] label: Nms' to stdout (Render logs) when timing is on."""
    if not _perf_on():
        yield
        return
    _t = _time.perf_counter()
    try:
        yield
    finally:
        print(f"[PERF] {label}: {(_time.perf_counter()-_t)*1000:.0f}ms", flush=True)

def _run_once(key, fn):
    """Memoize fn() for the duration of one Streamlit run. The backing dict is
    reset at the top of main(), so any interaction (which triggers a rerun)
    re-reads fresh — safe for user-mutable data like watchlists."""
    rc = st.session_state.get("_run_cache")
    if not isinstance(rc, dict):
        rc = {}
        st.session_state["_run_cache"] = rc
    if key not in rc:
        rc[key] = fn()
    return rc[key]

def _instrument_db(sb):
    """When timing is on, wrap the memoized client's postgrest httpx session so
    every DB round-trip is counted + timed (this captures NETWORK latency, not
    just the 2.4ms server execution). Patches the singleton instance once;
    defensive — any version mismatch silently skips DB timing."""
    if sb is None or not _perf_on() or getattr(sb, "_perf_wrapped", False):
        return sb
    try:
        _sess = sb.postgrest.session            # httpx.Client used for DB queries
        _orig_send = _sess.send
        def _timed_send(*a, **k):
            _t = _time.perf_counter()
            try:
                return _orig_send(*a, **k)
            finally:
                dt = (_time.perf_counter() - _t) * 1000
                st.session_state["_perf_db_n"]  = st.session_state.get("_perf_db_n", 0) + 1
                st.session_state["_perf_db_ms"] = st.session_state.get("_perf_db_ms", 0.0) + dt
        _sess.send = _timed_send
        sb._perf_wrapped = True
    except Exception:
        pass
    return sb


def main():
    _perf_on()  # prime the ?debug=timing flag from the entry URL before nav strips params
    # Fresh per-run read caches (deduped within one render; re-read after any rerun)
    st.session_state["_db_run_cache"] = {}
    st.session_state["_run_cache"] = {}

    # ── Analytics: capture UTM + site_visit (once/session), and login_completed
    # on the logged-in transition. Register does NOT auto-login, so this fires
    # only on real sign-ins (incl. MFA/recovery), never on signup. All wrapped.
    try:
        analytics.init_session()
        if (st.session_state.get("logged_in") and st.session_state.get("user")
                and not st.session_state.get("_login_evt_sent")):
            st.session_state["_login_evt_sent"] = True
            analytics.capture("login_completed", user=st.session_state.get("user"))
    except Exception:
        pass
    # ── Legal page via footer links ───────────────────────────────────────────
    if st.query_params.get("legal") in ("privacy","terms","billing","cookies","disclaimer"):
        st.session_state.legal_doc = st.query_params.get("legal")
        st.session_state.page = "legal"

    # ── Nav link routing ──────────────────────────────────────────────────────
    if st.query_params.get("nav") == "signin":
        st.session_state.auth_tab = "signin"
        st.session_state.page = "auth"
        st.query_params.pop("nav", None)
    if st.query_params.get("nav") == "register":
        st.session_state.auth_tab = "register"
        st.session_state.page = "auth"
        try:
            analytics.capture("signup_started")
        except Exception:
            pass
        # Carry the "I want Pro" intent through to post-registration. The
        # Founding/Pro CTAs append &plan=pro to the register URL; setting
        # auto_upgrade=True here makes register_user() then call upgrade_plan()
        # automatically, so the user lands inside the app as Pro instead of
        # needing extra clicks to claim their Founding spot.
        if st.query_params.get("plan") == "pro":
            st.session_state.auto_upgrade = True
            st.query_params.pop("plan", None)
        st.query_params.pop("nav", None)
    if st.query_params.get("nav") == "landing":
        st.session_state.page = "landing"
        st.session_state._show_landing = True  # flag before param is popped
        st.query_params.pop("nav", None)



    # ── Upgrade page routing ──────────────────────────────────────────────────
    if st.query_params.get("upgrade_page") == "1" and st.session_state.get("logged_in"):
        st.session_state.upgrade_feature    = st.query_params.get("feature", "Pro")
        st.session_state.upgrade_return_nav = st.query_params.get("return_nav", "screener")
        st.session_state.page = "upgrade"
        st.query_params.pop("upgrade_page", None)
        st.query_params.pop("feature", None)
        st.query_params.pop("return_nav", None)

    # ── Universe rescan via URL action ──────────────────────────────────────
    if st.query_params.get("rescan") == "1" and st.session_state.get("logged_in"):
        st.query_params.pop("rescan", None)
        st.session_state.scan_results = None
        st.session_state._live_prices_fetched = False
        _run_full_scan_cached.clear()
        _cached_live_prices.clear()

    # ── Simulator rescan via URL action ──────────────────────────────────────
    if st.query_params.get("sim_rescan") == "1" and st.session_state.get("logged_in"):
        st.query_params.pop("sim_rescan", None)
        try:
            from model_engine import fetch_macro_overlay, apply_macro_overlay, run_full_scan
            from model_engine import SECTORS as _SIM_SECTORS
            _run_full_scan_cached.clear()
            _raw = _cached_full_scan()
            _mac = _live_macro()
            for _r in _raw:
                if not _r.get("sector") or _r.get("sector") == "Unknown":
                    _r["sector"] = _SIM_SECTORS.get(_r["ticker"], "Unknown")
            _scored   = apply_macro_overlay(_raw, _mac)
            _enriched = enrich_with_signal_log(_scored)
            _final    = finalize_scores_from_signal_log(_enriched, _mac)
            st.session_state.scan_results = _final
            st.session_state.sim_data     = _final  # also update sim_data
            st.session_state.macro_data   = _mac
        except Exception:
            pass

    # ── Simulator profile select via URL action ───────────────────────────────
    _sim_profile = st.query_params.get("sim_profile", "")
    if _sim_profile in ("HIGH", "MEDIUM", "LOW") and st.session_state.get("logged_in"):
        st.query_params.pop("sim_profile", None)
        st.session_state.sim_profile = _sim_profile
        st.session_state.sim_weights = {}
        st.session_state.sim_profile_applied = None  # force rebuild of sim_selected in page

        # Ensure scan data is loaded so profile_tickers works immediately
        if not st.session_state.get("scan_results"):
            try:
                from data_refresh import _get_supabase as _sim_sb2
                _sb2 = _sim_sb2()
                if _sb2:
                    _resp2 = _sb2.table("signal_log") \
                        .select("ticker,adj_composite,composite,signal,momentum,quality,volume,value,sentiment,price,signal_date,mktcap,val_low,val_high,value_position,val_basis") \
                        .order("signal_date", desc=True) \
                        .limit(5000) \
                        .execute()
                    _seen2 = {}
                    for _r2 in (_resp2.data or []):
                        if _r2["ticker"] not in _seen2:
                            _seen2[_r2["ticker"]] = _r2
                    try:
                        from model_engine import SECTORS as _SIM_SECTORS2
                        for _tk2, _row2 in _seen2.items():
                            _row2["sector"] = _SIM_SECTORS2.get(_tk2, "Unknown")
                    except Exception:
                        pass
                    _final2 = finalize_scores_from_signal_log(
                        list(_seen2.values()),
                        st.session_state.get("macro_data"),
                    )
                    st.session_state.scan_results = _final2
            except Exception:
                pass

    # ── Simulator add/remove ticker via URL action ────────────────────────────
    _sim_add = st.query_params.get("sim_add", "")
    if _sim_add and st.session_state.get("logged_in"):
        st.query_params.pop("sim_add", None)
        if "sim_selected" not in st.session_state:
            st.session_state.sim_selected = []
        if _sim_add not in st.session_state.sim_selected:
            st.session_state.sim_selected.append(_sim_add)
        st.session_state.nav = "simulator"
        st.rerun()

    _sim_remove = st.query_params.get("sim_remove", "")
    if _sim_remove and st.session_state.get("logged_in"):
        st.query_params.pop("sim_remove", None)
        if "sim_selected" in st.session_state and _sim_remove in st.session_state.sim_selected:
            st.session_state.sim_selected.remove(_sim_remove)
        st.session_state.get("sim_weights", {}).pop(_sim_remove, None)
        st.session_state.nav = "simulator"
        st.rerun()

    # ── Plan upgrade via URL action ───────────────────────────────────────────
    if st.query_params.get("upgrade") == "pro" and st.session_state.get("logged_in"):
        # Only grant a free founder spot while spots remain. A previously
        # paid-then-canceled user is allowed to re-take a spot (cancellation is
        # final → free, but a free user can claim again while the first-50 window
        # is open).
        if _founding_spots_remaining() > 0:
            ok = upgrade_plan(uid(), "pro")
            if ok and st.session_state.get("user"):
                st.session_state.user["plan"] = "pro"
                try:
                    analytics.capture("founder_membership_claimed", user=st.session_state.get("user"))
                except Exception:
                    pass
                # Founder claim overrides any prior paid+canceled cycle: wipe the
                # stale Stripe/cancellation state so the billing reconciler doesn't
                # see a canceled subscription on the next load and downgrade this
                # fresh grant back to free (the claim/cancel tug-of-war loop).
                try:
                    from db import clear_stripe_state as _clear_billing
                    _clear_billing(uid())
                except Exception:
                    pass
                # Persist the new plan into the signed URL token + localStorage so
                # the upgrade survives the _confirm_url redirect and any prod
                # reconnect. Without this the page reloads with the stale free
                # token, the session reverts to free, and the user bounces back to
                # the locked feature — an endless claim loop on production.
                try:
                    st.query_params["uid"]  = _sign_token(uid(), "pro")
                    st.query_params["plan"] = "pro"
                except Exception:
                    pass
                _write_localstorage_token(uid(), "pro")
        st.query_params.pop("upgrade", None)

    # ── Stripe checkout return + status polling ───────────────────────────────
    _checkout = st.query_params.get("checkout", "")
    if _checkout and st.session_state.get("logged_in"):
        if _checkout == "success":
            try:
                import stripe_billing as _sbp
                from db import set_stripe_billing as _ssb, get_stripe_billing as _gsb2
                import arl as _arl_ck
                _ck_email = (st.session_state.user or {}).get("email")
                res = _sbp.finalize_checkout(uid(), _ck_email)
                if res.get("ok"):
                    _grants = _sbp.status_grants_access(res.get("status", ""))
                    _ssb(uid(),
                         customer_id=res.get("customer_id"),
                         subscription_id=res.get("subscription_id"),
                         billing_active=_grants,
                         status=res.get("status"))
                    # store trial_end / period_end for the countdown display
                    try:
                        _nb = (st.session_state.user or {}).get("notifications") or {}
                        if isinstance(_nb, dict):
                            _nb["trial_end"] = res.get("trial_end")
                            _nb["current_period_end"] = res.get("current_period_end")
                            from db import update_preferences as _upd
                            _upd(uid(), {"notifications": _nb})
                            st.session_state.user["notifications"] = _nb
                    except Exception:
                        pass
                    if _grants:
                        upgrade_plan(uid(), "pro")
                        if st.session_state.get("user"):
                            st.session_state.user["plan"] = "pro"
                        _em = (st.session_state.user or {}).get("email")
                        if _em:
                            _arl_ck.send_acknowledgment(uid(), _em)
                        st.session_state["_trial_just_started"] = True
                else:
                    st.warning(
                        "Your payment went through, but we couldn't confirm your "
                        "subscription instantly. It'll sync within a minute — refresh "
                        f"the page shortly. ({res.get('error','')})")
            except Exception as _cke:
                st.warning(f"Checkout return issue: {_cke}")
        st.query_params.pop("checkout", None)
        st.query_params.pop("plan", None)

    # Lightweight status poll: once per session, sync plan from live Stripe state
    # for users with a stored subscription (no webhooks).
    if (st.session_state.get("logged_in")
            and not st.session_state.get("_stripe_polled")):
        st.session_state._stripe_polled = True
        try:
            import stripe_billing as _sbp2
            from db import get_stripe_billing as _gsb3, set_stripe_billing as _ssb2
            _bs = _gsb3(uid())
            _sub_id = _bs.get("stripe_subscription_id")
            if _sub_id and _sbp2.billing_configured():
                _ps = _sbp2.poll_subscription_status(_sub_id)
                if _ps.get("gone"):
                    # Subscription no longer exists in Stripe (deleted/resource_missing).
                    # Clear the dead reference so we stop polling + erroring every
                    # session, and mark billing inactive. Plan is left as-is (an
                    # explicit cancel/downgrade path handles status changes).
                    _ssb2(uid(), subscription_id="", billing_active=False, status="canceled")
                elif _ps.get("ok"):
                    _grant = _sbp2.status_grants_access(_ps.get("status", ""))
                    _ssb2(uid(), billing_active=_grant, status=_ps.get("status"))
                    _cur_plan = (st.session_state.user or {}).get("plan", "free")
                    if _grant and _cur_plan != "pro":
                        upgrade_plan(uid(), "pro")
                        if st.session_state.get("user"):
                            st.session_state.user["plan"] = "pro"
                    elif not _grant and _cur_plan == "pro":
                        # subscription ended/canceled past period — downgrade
                        upgrade_plan(uid(), "free")
                        if st.session_state.get("user"):
                            st.session_state.user["plan"] = "free"
            elif (not _sub_id) and _sbp2.billing_configured() and not _bs.get("billing_active") \
                    and int(st.session_state.get("_awaiting_checkout", 0) or 0) > 0:
                # Self-heal ONLY when we're expecting a checkout to land (counter
                # set when the user clicks Start trial). Without this gate, every
                # free user would trigger a Stripe email lookup on every load,
                # stalling the page. Decrement so it runs a bounded number of times.
                st.session_state["_awaiting_checkout"] = int(st.session_state.get("_awaiting_checkout", 0)) - 1
                _em2 = (st.session_state.user or {}).get("email")
                if _em2:
                    _fr = _sbp2.finalize_checkout(uid(), _em2)
                    if _fr.get("ok") and _sbp2.status_grants_access(_fr.get("status", "")):
                        _ssb2(uid(),
                              customer_id=_fr.get("customer_id"),
                              subscription_id=_fr.get("subscription_id"),
                              billing_active=True,
                              status=_fr.get("status"))
                        try:
                            _nb2 = (st.session_state.user or {}).get("notifications") or {}
                            if isinstance(_nb2, dict):
                                _nb2["trial_end"] = _fr.get("trial_end")
                                _nb2["current_period_end"] = _fr.get("current_period_end")
                                from db import update_preferences as _upd2
                                _upd2(uid(), {"notifications": _nb2})
                                st.session_state.user["notifications"] = _nb2
                        except Exception:
                            pass
                        upgrade_plan(uid(), "pro")
                        if st.session_state.get("user"):
                            st.session_state.user["plan"] = "pro"
        except Exception:
            pass

    # ── Watchlist add/remove via URL action ──────────────────────────────────
    _wl_action = st.query_params.get("wl_action", "")
    _wl_ticker = st.query_params.get("wl_ticker", "")
    if _wl_action and _wl_ticker and st.session_state.get("logged_in"):
        if _wl_action == "add":
            _add_px = None
            try:
                _smap = {r["ticker"]: r for r in (st.session_state.get("scan_results") or [])}
                _add_px = (_smap.get(_wl_ticker) or {}).get("price")
            except Exception:
                _add_px = None
            # Fallback: pull the latest known price from signal_log so the
            # "since added" baseline is captured even if the scan lacks a price.
            if not _add_px:
                try:
                    from db import get_price_on_date_latest as _gpl_add
                    _add_px = _gpl_add(_wl_ticker)
                except Exception:
                    _add_px = None
            add_to_watchlist(uid(), _wl_ticker, _add_px)
            st.session_state.pop("_wl_daychange_cache", None)
            try:
                analytics.capture("watchlist_added", user=st.session_state.get("user"),
                                  props={"ticker": _wl_ticker})
            except Exception:
                pass
        elif _wl_action == "remove":
            remove_from_watchlist(uid(), _wl_ticker)
            st.session_state.pop("_wl_daychange_cache", None)
            try:
                analytics.capture("watchlist_removed", user=st.session_state.get("user"),
                                  props={"ticker": _wl_ticker})
            except Exception:
                pass
        st.query_params.pop("wl_action", None)
        st.query_params.pop("wl_ticker", None)
    _port_action = st.query_params.get("port_action", "")
    _port_ticker = st.query_params.get("port_ticker", "")
    if _port_action == "remove" and _port_ticker and st.session_state.get("logged_in"):
        delete_holding(uid(), _port_ticker)
        st.query_params.pop("port_action", None)
        st.query_params.pop("port_ticker", None)
    _port_period = st.query_params.get("port_period", "")
    if _port_period in ("ACT","1M","3M","1Y"):
        st.session_state.port_period = _port_period
        st.query_params.pop("port_period", None)
    _VALID_TABS = {"screener","watchlist","gems","backtest","portfolio","simulator",
                   "model_portfolio","alerts","account","methodology","analytics"}
    _qnav = st.query_params.get("qnav","")
    if _qnav in _VALID_TABS:
        st.session_state.nav  = _qnav
        st.session_state.page = "platform"
        st.query_params.pop("qnav", None)
    if st.query_params.get("qnav") == "signout":
        for k in ["logged_in","user","mfa_verified","scan_results",
                  "macro_data","mfa_recovery_mode","live_refresh_running"]:
            st.session_state[k] = False if k == "logged_in" else None
        st.session_state.signed_out = True
        st.query_params.clear()
        _clear_localstorage_token()
        go("landing")
    if st.query_params.get("ck") == "1":
        st.session_state.cookies_accepted = True

    # ── Reconnect recovery: restore nav from _n param ───────────────────────
    _saved_nav = st.query_params.get("_n", "")
    _VALID_TABS = {"screener","watchlist","gems","backtest","portfolio","simulator",
                   "model_portfolio","alerts","account","methodology","analytics"}
    if _saved_nav in _VALID_TABS and not st.session_state.get("_show_landing"):
        _cur = st.session_state.get("nav", "screener")
        if _cur == "screener" and _saved_nav != "screener":
            st.session_state.nav  = _saved_nav
            st.session_state.page = "platform"

    # ── Landing page routing — must run last so nothing overrides it ─────────
    if st.session_state.get("_show_landing"):
        st.session_state.page = "landing"
        st.session_state._show_landing = False
    elif st.session_state.page == "landing" and st.session_state.logged_in:
        st.session_state.page = "platform"

    # ── Password-reset deep link — overrides everything, works logged out ─────
    if st.query_params.get("reset_token"):
        st.session_state.page = "reset"

    # ── Email-verify deep link — works logged in or out ───────────────────────
    if st.query_params.get("verify_token"):
        st.session_state.page = "verify"

    _pl = st.empty()
    import random as _rnd
    _load_msgs = [
        f"Scoring all {_universe_n()} tickers…",
        "Weighing the five pillars…",
        "Checking the macro regime…",
        "Sorting conviction, high to low…",
        "Stress-testing against SPY…",
        "Sniffing out hidden gems…",
        "Separating signal from noise…",
        "Letting the quants argue it out…",
        "Consulting the bulls and the bears…",
        "Crunching factor scores…",
        "Reading the tape…",
        "Pricing in the macro…",
        "Marking to market…",
        "Doing the due diligence…",
        "Compounding the pixels…",
        "Waiting for the opening bell…",
    ]
    _per   = 0.85                                  # seconds each phrase is shown
    _msgs  = _rnd.sample(_load_msgs, 6)            # rotate through 6 of them
    _cycle = round(_per * len(_msgs), 2)
    _phrase_stack = "".join(
        f'<div style="position:absolute;left:0;right:0;text-align:center;opacity:0;'
        f'font-family:Inter,sans-serif;font-size:13.5px;color:#b3bed0;'
        f'animation:qntmphrase {_cycle}s ease-in-out {round(_i*_per,2)}s infinite;">{_m}</div>'
        for _i, _m in enumerate(_msgs)
    )
    _pl.markdown(
        '<div style="position:fixed;top:0;left:0;width:100%;height:100%;'
        'display:flex;flex-direction:column;align-items:center;justify-content:center;gap:14px;'
        'z-index:99999;background:rgba(6,7,15,.80);backdrop-filter:blur(2px);pointer-events:none;'
        'opacity:0;animation:qntmload .25s ease .45s forwards;">'
        '<div style="font-family:Syne,sans-serif;font-size:21px;font-weight:800;letter-spacing:.05em;'
        'color:#e2e8f0;">Q<span style="color:#34d399;">NTM</span></div>'
        f'<div style="position:relative;height:20px;min-width:300px;max-width:90vw;">{_phrase_stack}</div>'
        '<div style="width:240px;height:5px;border-radius:99px;background:rgba(52,211,153,.14);'
        'overflow:hidden;">'
        '<div style="height:100%;border-radius:99px;background:linear-gradient(90deg,#34d399,#5eead4);'
        'width:4%;animation:qntmfill 5s cubic-bezier(.2,.08,.6,.78) .45s forwards;'
        'box-shadow:0 0 10px rgba(52,211,153,.5);"></div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    try:
        route = st.session_state.page
        if _perf_on():
            st.session_state["_perf_db_n"] = 0
            st.session_state["_perf_db_ms"] = 0.0
            try:
                from db import get_supabase as _dbsb
                _instrument_db(_dbsb())
            except Exception:
                pass
            try:
                from data_refresh import _get_supabase as _drsb
                _instrument_db(_drsb())
            except Exception:
                pass
        _t_page = _time.perf_counter()
        if   route == "landing":  page_landing()
        elif route == "auth":     page_auth()
        elif route == "mfa":      page_mfa()
        elif route == "upgrade":  page_upgrade()
        elif route == "model":    go("landing")
        elif route == "platform": page_platform()
        elif route == "legal":    page_legal(st.session_state.get("legal_doc","privacy"))
        elif route == "reset":    page_reset_password()
        elif route == "verify":   page_verify_email()
        else:                     page_landing()
        if _perf_on():
            _tot  = (_time.perf_counter() - _t_page) * 1000
            _dbn  = st.session_state.get("_perf_db_n", 0)
            _dbms = st.session_state.get("_perf_db_ms", 0.0)
            _sub  = st.session_state.get("nav", route)
            print(f"[PERF] page:{route}/{_sub} total={_tot:.0f}ms "
                  f"db={_dbn}q/{_dbms:.0f}ms render~={_tot-_dbms:.0f}ms", flush=True)
    finally:
        _pl.empty()

main()
